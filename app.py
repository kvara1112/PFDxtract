import streamlit as st
import pyLDAvis
import pyLDAvis.sklearn 
import pandas as pd 
import numpy as np
from datetime import datetime, timedelta
import re
import requests
from bs4 import BeautifulSoup
import time
import urllib3
import io
import pdfplumber
import logging
import os
import zipfile
import unicodedata
import nltk
import random
import string
import traceback
import torch
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import plotly.express as px
import plotly.graph_objects as go
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.preprocessing import normalize
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import networkx as nx
import pytz
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from collections import Counter, defaultdict
from bs4 import BeautifulSoup, Tag
import json  # Added for JSON export functionality
nltk.download("punkt")
nltk.download("stopwords")
nltk.download("averaged_perceptron_tagger")
from openpyxl.utils import get_column_letter
from sklearn.base import BaseEstimator, TransformerMixin
import scipy.sparse as sp
from typing import Union
from sklearn.metrics import (
    silhouette_score,
    calinski_harabasz_score,
    davies_bouldin_score,
)
from transformers import AutoTokenizer, AutoModel
from tqdm import tqdm 
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import tempfile
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.patches import Patch
  
######################################
class BERTResultsAnalyzer:
    """Enhanced class for merging BERT theme analysis results files with specific column outputs and year extraction."""

    
    def __init__(self):
        """Initialize the analyzer with default settings."""
        self.data = None
        # Define the essential columns you want in the reduced output
        self.essential_columns = [
            "Title",
            "URL",
            "Content",
            "date_of_report",
            "ref", 
            "deceased_name",
            "coroner_name",
            "coroner_area",
            "categories",
            "Report ID",
            "Deceased Name",
            "Death Type",
            "Year",
            "year",
            "Extracted_Concerns",
        ]
        
    def render_analyzer_ui(self):
        """Render the file merger UI."""
        st.subheader("Scraped File Merger")
        st.markdown(
            """
            This tool merges multiple scraped files into a single dataset. It prepares the data for steps (3) - (5).
            
            - Run this step even if you only have one scraped file. This step extracts the year and applies other processing as described in the bullets below. 
            - Combine data from multiple CSV or Excel files (the name of these files starts with pfd_reports_scraped_reportID_ )
            - Extract missing concerns from PDF content and fill empty Content fields
            - Extract year information from date fields
            - Remove duplicate records
            - Export full or reduced datasets with essential columns
            
            Use the options below to control how your files will be processed.
            """
        )

        # File upload section
        self._render_multiple_file_upload()

    def _render_multiple_file_upload(self):
        """Render interface for multiple file upload and merging."""
        # Initialize session state for processed data if not already present
        if "bert_merged_data" not in st.session_state:
            st.session_state.bert_merged_data = None

        # Use a dynamic key for file uploader based on reset counter
        reset_counter = st.session_state.get("reset_counter", 0)
        uploaded_files = st.file_uploader(
            "Upload CSV or Excel files exported from the scraper tool",
            type=["csv", "xlsx"],
            accept_multiple_files=True,
            help="Upload multiple CSV or Excel files to merge them",
            key="bert_multi_uploader_static",
        )

        if uploaded_files and len(uploaded_files) > 0:
            st.info(f"Uploaded {len(uploaded_files)} files")

            # Allow user to specify merge settings
            with st.expander("Merge Settings", expanded=True):
                st.info("These settings control how the files will be merged.")

                # Option to remove duplicates - static key
                drop_duplicates = st.checkbox(
                    "Remove Duplicate Records",
                    value=True,
                    help="If checked, duplicate records will be removed after merging",
                    key="drop_duplicates_static",
                )

                # Option to add year from date_of_report
                extract_year = st.checkbox(
                    "Extract Year from date_of_report",
                    value=True,
                    help="If checked, a 'year' column will be added based on the date_of_report",
                    key="extract_year_static",
                )

                # Option to attempt extraction from PDF content when concerns are missing
                extract_from_pdf = st.checkbox(
                    "Extract Missing Concerns from PDF Content",
                    value=True,
                    help="If checked, will attempt to extract missing concerns from PDF content",
                    key="extract_from_pdf_static",
                )

                # Option to fill empty Content from PDF content
                fill_empty_content = st.checkbox(
                    "Fill Empty Content from PDF Content",
                    value=True,
                    help="If checked, will fill empty Content fields from PDF content",
                    key="fill_empty_content_static",
                )

                duplicate_columns = "Record ID"
                if drop_duplicates:
                    duplicate_columns = st.text_input(
                        "Columns for Duplicate Check",
                        value="ref",
                        help="Comma-separated list of columns to check for duplicates",
                        key="duplicate_columns_static",
                    )

            # Button to process the files - static key
            if st.button("Merge Files", key="merge_files_button_static"):
                try:
                    with st.spinner("Processing and merging files..."):
                        # Stack files
                        duplicate_cols = None
                        if drop_duplicates:
                            duplicate_cols = [
                                col.strip() for col in duplicate_columns.split(",")
                            ]

                        self._merge_files_stack(uploaded_files, duplicate_cols)

                        # Now processed_data is in self.data
                        if self.data is not None:
                            # Fill empty Content from PDF if requested
                            if fill_empty_content:
                                before_count = self.data["Content"].notna().sum()
                                self.data = self._fill_empty_content_from_pdf(self.data)
                                after_count = self.data["Content"].notna().sum()
                                newly_filled = after_count - before_count

                                if newly_filled > 0:
                                    st.success(
                                        f"Filled empty Content from PDF content for {newly_filled} records."
                                    )
                                else:
                                    st.info(
                                        "No Content fields could be filled from PDF content."
                                    )

                            # Extract year from date_of_report if requested
                            if extract_year and "date_of_report" in self.data.columns:
                                self.data = self._add_year_column(self.data)
                                with_year = self.data["year"].notna().sum()
                                st.success(
                                    f"Added year data to {with_year} out of {len(self.data)} reports."
                                )

                            # Extract missing concerns from PDF content if requested
                            if extract_from_pdf:
                                before_count = (
                                    self.data["Extracted_Concerns"].notna().sum()
                                )
                                self.data = self._extract_missing_concerns_from_pdf(
                                    self.data
                                )
                                after_count = (
                                    self.data["Extracted_Concerns"].notna().sum()
                                )
                                newly_extracted = after_count - before_count

                                if newly_extracted > 0:
                                    st.success(
                                        f"Extracted missing concerns from PDF content for {newly_extracted} reports."
                                    )
                                else:
                                    st.info(
                                        "No additional concerns could be extracted from PDF content."
                                    )

                            st.success(
                                f"Files merged successfully! Final dataset has {len(self.data)} records."
                            )

                            # Show a preview of the data
                            st.subheader("Preview of Merged Data")
                            st.dataframe(self.data.head(5))

                            # Save merged data to session state
                            st.session_state.bert_merged_data = self.data.copy()
                        else:
                            st.error(
                                "File merging resulted in empty data. Please check your files."
                            )

                except Exception as e:
                    st.error(f"Error merging files: {str(e)}")
                    logging.error(f"File merging error: {e}", exc_info=True)

        # Show download options if we have processed data
        # Use either the current instance data or data from session state
        show_download_options = False

        if hasattr(self, "data") and self.data is not None and len(self.data) > 0:
            show_download_options = True
        elif st.session_state.bert_merged_data is not None:
            self.data = st.session_state.bert_merged_data
            show_download_options = True

        if show_download_options:
            self._provide_download_options()

    def _fill_empty_content_from_pdf(self, df):
        """
        Fill empty Content fields from PDF_1_Content only.
        
        Args:
            df: DataFrame with merged data
        
        Returns:
            DataFrame with filled Content fields
        """
        if df is None or len(df) == 0:
            return df
        
        # Make a copy to avoid modifying the original
        processed_df = df.copy()
        
        # Identify records with missing Content
        missing_content_mask = processed_df["Content"].isna() | (
            processed_df["Content"].astype(str).str.strip() == ""
        )
        missing_content_count = missing_content_mask.sum()
        
        if missing_content_count == 0:
            return processed_df
        
        # Add a progress bar for processing
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text(
            f"Filling empty Content fields from PDF_1_Content for {missing_content_count} records..."
        )
        
        # Check if PDF_1_Content column exists
        if "PDF_1_Content" not in processed_df.columns:
            progress_bar.empty()
            status_text.empty()
            return processed_df
        
        # Process each record with missing Content
        missing_indices = processed_df[missing_content_mask].index
        filled_count = 0
        
        for i, idx in enumerate(missing_indices):
            # Update progress
            progress = (i + 1) / len(missing_indices)
            progress_bar.progress(progress)
            
            # Check if PDF_1_Content exists and has content
            if pd.notna(processed_df.at[idx, "PDF_1_Content"]) and processed_df.at[idx, "PDF_1_Content"].strip() != "":
                # Use the PDF content as the main Content
                processed_df.at[idx, "Content"] = processed_df.at[idx, "PDF_1_Content"]
                processed_df.at[idx, "Content_Source"] = "PDF_1_Content"  # Track where content came from
                filled_count += 1
            
            # Update status
            status_text.text(
                f"Filled Content for {filled_count} of {i+1}/{missing_content_count} records..."
            )
        
        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
        
        return processed_df
    
    
    def _extract_missing_concerns_from_pdf(self, df):
        """
        Extract concerns from PDF content for records with missing Extracted_Concerns.

        Args:
            df: DataFrame with merged data

        Returns:
            DataFrame with additional extracted concerns
        """
        if df is None or len(df) == 0:
            return df

        # Make a copy to avoid modifying the original
        processed_df = df.copy()

        # Identify records with missing concerns
        missing_concerns = self._identify_missing_concerns(processed_df)

        if len(missing_concerns) == 0:
            return processed_df

        # Add a progress bar for extraction
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text(
            f"Extracting concerns from PDF content for {len(missing_concerns)} records..."
        )

        # Check each PDF content column for missing concerns
        pdf_columns = [
            col
            for col in processed_df.columns
            if col.startswith("PDF_") and col.endswith("_Content")
        ]
        count_extracted = 0

        for i, (idx, row) in enumerate(missing_concerns.iterrows()):
            # Update progress
            progress = (i + 1) / len(missing_concerns)
            progress_bar.progress(progress)

            # Try to extract concerns from each PDF content column
            for pdf_col in pdf_columns:
                if pd.notna(row.get(pdf_col)) and row.get(pdf_col) != "":
                    # Extract concerns using existing function
                    concern_text = extract_concern_text(row[pdf_col])

                    # If we found concerns, update the main dataframe
                    if (
                        concern_text and len(concern_text.strip()) > 20
                    ):  # Ensure meaningful text
                        processed_df.at[idx, "Extracted_Concerns"] = concern_text
                        processed_df.at[
                            idx, "Concern_Source"
                        ] = pdf_col  # Track where concerns came from
                        count_extracted += 1
                        break  # Move to next record once concerns are found

            # Update status
            status_text.text(
                f"Extracted concerns for {count_extracted} of {i+1}/{len(missing_concerns)} records..."
            )

        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()

        return processed_df

    def _extract_report_year(self, date_val):
        """
        Optimized function to extract year from dd/mm/yyyy date format.
        Works with both string representations and datetime objects.
        """
        import re
        import datetime
        import pandas as pd

        # Return None for empty inputs
        if pd.isna(date_val):
            return None

        # Handle datetime objects directly
        if isinstance(date_val, (datetime.datetime, pd.Timestamp)):
            return date_val.year

        # Handle string dates with dd/mm/yyyy format
        if isinstance(date_val, str):
            # Direct regex match for dd/mm/yyyy pattern (faster than datetime parsing)
            match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", date_val)
            if match:
                return int(match.group(3))  # Year is in the third capture group

        # Handle numeric or other non-string types by converting to string
        try:
            date_str = str(date_val)
            match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", date_str)
            if match:
                return int(match.group(3))
        except:
            pass

        return None

    # block3
    def _add_missing_years_from_content(self, concern_sections, df=None):
        """Try to extract years from content when date_of_report is missing."""
        import re
        import datetime
        from collections import Counter

        missing_year_count = 0

        # Try to use dataframe if available for additional data sources
        index_to_date = {}
        if df is not None:
            date_related_columns = []
            for col in df.columns:
                if any(term in col.lower() for term in ["date", "year", "time"]):
                    date_related_columns.append(col)

            if date_related_columns:
                print(
                    f"Checking {len(date_related_columns)} date-related columns for missing years"
                )
                for idx, row in df.iterrows():
                    for col in date_related_columns:
                        if pd.notna(row.get(col)):
                            val = row.get(col)
                            # Convert to string if needed
                            if not isinstance(val, str):
                                val = str(val)
                            year = self._extract_report_year(val)
                            if year:
                                index_to_date[idx] = year
                                break

        for section in concern_sections:
            if section.get("year") is None:
                # Check if we have a year for this index in our lookup
                if (
                    df is not None
                    and "index" in section
                    and section["index"] in index_to_date
                ):
                    section["year"] = index_to_date[section["index"]]
                    section["year_source"] = "other_date_column"
                    missing_year_count += 1
                    continue

                # Try to extract year from title first
                title = section.get("title", "") or section.get("Title", "")

                # Look for years in the title
                title_year_match = re.search(r"\b(19|20)\d{2}\b", title)
                if title_year_match:
                    section["year"] = int(title_year_match.group(0))
                    section["year_source"] = "extracted_from_title"
                    missing_year_count += 1
                    continue

                # Then try content
                content = (
                    section.get("Content", "")
                    or section.get("content", "")
                    or section.get("concern_text", "")
                    or section.get("Extracted_Concerns", "")
                )
                if content:
                    # Look for date patterns first
                    date_matches = re.findall(
                        r"\b\d{1,2}[/-]\d{1,2}[/-](19|20)\d{2}\b", content
                    )
                    if date_matches:
                        # Extract year from the first date pattern found
                        full_date = re.search(
                            r"\b\d{1,2}[/-]\d{1,2}[/-](19|20)\d{2}\b", content
                        ).group(0)
                        year_str = re.search(r"(19|20)\d{2}", full_date).group(0)
                        section["year"] = int(year_str)
                        section["year_source"] = "date_in_content"
                        missing_year_count += 1
                        continue

                    # Otherwise look for any years
                    year_matches = re.findall(r"\b(19|20)\d{2}\b", content)
                    if year_matches:
                        # Get all years mentioned
                        years = [int(y) for y in year_matches]
                        # Use the most frequent year
                        most_common_year = Counter(years).most_common(1)[0][0]
                        section["year"] = most_common_year
                        section["year_source"] = "year_in_content"
                        missing_year_count += 1

                        # Extra check for future years which may be errors
                        current_year = datetime.datetime.now().year
                        if section["year"] > current_year:
                            # If it's a future year, try to find a more plausible one
                            plausible_years = [y for y in years if y <= current_year]
                            if plausible_years:
                                section["year"] = max(plausible_years)

        if missing_year_count > 0:
            print(
                f"Added year data to {missing_year_count} reports using content analysis"
            )

        return concern_sections

    def _add_year_column(self, df):
        """
        Add a 'year' column to the DataFrame extracted from the date_of_report column.
        If date_of_report is missing, try to extract year from content.

        Args:
            df: DataFrame with at least date_of_report column

        Returns:
            DataFrame with additional 'year' column
        """
        # Create a copy to avoid modifying the original DataFrame
        processed_df = df.copy()

        # Check if DataFrame has date_of_report column
        if "date_of_report" not in processed_df.columns:
            st.warning(
                "No 'date_of_report' column found in the data. Year extraction may be incomplete."
            )
            processed_df["year"] = None
            return processed_df

        # Create a new column for year extracted from date_of_report
        processed_df["year"] = processed_df["date_of_report"].apply(
            self._extract_report_year
        )

        # Count how many years were successfully extracted
        extracted_count = processed_df["year"].notna().sum()

        # For rows with missing years, try to extract from content
        if processed_df["year"].isna().any():
            # Create a list of dicts from rows with missing years
            missing_year_rows = []
            for idx, row in processed_df[processed_df["year"].isna()].iterrows():
                section_dict = row.to_dict()
                section_dict["index"] = idx  # Add index for tracking
                missing_year_rows.append(section_dict)

            # Try to extract years from content
            if missing_year_rows:
                missing_year_rows = self._add_missing_years_from_content(
                    missing_year_rows, processed_df
                )

                # Update the original DataFrame with extracted years
                for section in missing_year_rows:
                    if (
                        "year" in section
                        and section["year"] is not None
                        and "index" in section
                    ):
                        processed_df.at[section["index"], "year"] = section["year"]
                        # Optionally add the source of extraction
                        if "year_source" in section:
                            processed_df.at[section["index"], "year_source"] = section[
                                "year_source"
                            ]

        # Count final results
        final_count = processed_df["year"].notna().sum()
        if final_count > extracted_count:
            print(
                f"Added {final_count - extracted_count} more years from content analysis"
            )

        return processed_df


# block4

    def _provide_download_options(self):
        """Provide options to download the current data."""
        if self.data is None or len(self.data) == 0:
            return
        
        st.subheader("Download Merged Data")
        
        # Generate timestamp and random suffix for truly unique keys
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
        unique_id = f"{timestamp}_{random_suffix}"
        
        # Deduplicate data by Record ID before download if requested
        dedup_download = st.checkbox(
            "Remove duplicate Record IDs before download (keep only first occurrence)",
            value=True,
            key=f"dedup_checkbox_{unique_id}",
        )
        
        download_data = self.data
        if dedup_download and "Record ID" in self.data.columns:
            download_data = self.data.drop_duplicates(subset=["Record ID"], keep="first")
            st.info(
                f"Download will contain {len(download_data)} rows after removing duplicate Record IDs (original had {len(self.data)} rows)"
            )
        
        # Prepare the reduced dataset with essential columns
        reduced_data = download_data.copy()
        
        # Get list of available essential columns
        available_essential_cols = [
            col for col in self.essential_columns if col in reduced_data.columns
        ]
        
        # Ensure 'ref' is in the available columns if it exists in the data
        if "ref" in reduced_data.columns and "ref" not in available_essential_cols:
            available_essential_cols.append("ref")
            
        if available_essential_cols:
            reduced_data = reduced_data[available_essential_cols]
            st.success(
                f"Reduced dataset includes these columns: {', '.join(available_essential_cols)}"
            )
        else:
            st.warning(
                "None of the essential columns found in the data. Will provide full dataset only."
            )
            reduced_data = None
        
        # Generate filename prefix
        filename_prefix = f"merged_{timestamp}"
        
        # Full Dataset Section
        st.markdown("### Full Dataset")
        full_col1, full_col2 = st.columns(2)
        
        # CSV download button for full data
        with full_col1:
            try:
                # Create export copy with formatted dates
                df_csv = download_data.copy()
                if (
                    "date_of_report" in df_csv.columns
                    and pd.api.types.is_datetime64_any_dtype(df_csv["date_of_report"])
                ):
                    df_csv["date_of_report"] = df_csv["date_of_report"].dt.strftime(
                        "%d/%m/%Y"
                    )
    
                csv_data = df_csv.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Download Full Dataset (CSV)",
                    data=csv_data,
                    file_name=f"{filename_prefix}_full.csv",
                    mime="text/csv",
                    key=f"download_full_csv_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing CSV export: {str(e)}")
        
        # Excel download button for full data
        with full_col2:
            try:
                excel_buffer_full = io.BytesIO()
                download_data.to_excel(excel_buffer_full, index=False, engine="openpyxl")
                excel_buffer_full.seek(0)
                st.download_button(
                    "📥 Download Full Dataset (Excel)",
                    data=excel_buffer_full,
                    file_name=f"{filename_prefix}_full.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_full_excel_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing Excel export: {str(e)}")
        
        # Only show reduced dataset options if we have essential columns
        if reduced_data is not None:
            st.markdown("### Reduced Dataset (Essential Columns)")
            reduced_col1, reduced_col2 = st.columns(2)
        
            # CSV download button for reduced data
            with reduced_col1:
                try:
                    # Create export copy with formatted dates
                    df_csv_reduced = reduced_data.copy()
                    if (
                        "date_of_report" in df_csv_reduced.columns
                        and pd.api.types.is_datetime64_any_dtype(
                            df_csv_reduced["date_of_report"]
                        )
                    ):
                        df_csv_reduced["date_of_report"] = df_csv_reduced[
                            "date_of_report"
                        ].dt.strftime("%d/%m/%Y")
    
                    reduced_csv_data = df_csv_reduced.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        "📥 Download Reduced Dataset (CSV)",
                        data=reduced_csv_data,
                        file_name=f"{filename_prefix}_reduced.csv",
                        mime="text/csv",
                        key=f"download_reduced_csv_{unique_id}",
                    )
                except Exception as e:
                    st.error(f"Error preparing reduced CSV export: {str(e)}")
        
            # Excel download button for reduced data
            with reduced_col2:
                try:
                    excel_buffer_reduced = io.BytesIO()
                    reduced_data.to_excel(
                        excel_buffer_reduced, index=False, engine="openpyxl"
                    )
                    excel_buffer_reduced.seek(0)
                    st.download_button(
                        "📥 Download Reduced Dataset (Excel)",
                        data=excel_buffer_reduced,
                        file_name=f"{filename_prefix}_reduced.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_reduced_excel_{unique_id}",
                    )
                except Exception as e:
                    st.error(f"Error preparing reduced Excel export: {str(e)}")
        
        # NEW: Add section to display and download reports with missing concerns
        st.markdown("### Reports Without Extracted Concerns")
        
        # Find records without concerns
        missing_concerns_df = self._identify_missing_concerns(download_data)
        
        if len(missing_concerns_df) > 0:
            st.warning(
                f"Found {len(missing_concerns_df)} reports without properly extracted concerns."
            )
        
            # Display the dataframe with missing concerns
            essential_columns_for_display = [
                col
                for col in ["Title", "URL", "date_of_report", "year", "deceased_name", "ref"]
                if col in missing_concerns_df.columns
            ]
            if essential_columns_for_display:
                st.dataframe(
                    missing_concerns_df[essential_columns_for_display],
                    use_container_width=True,
                )
        
            # Download options for missing concerns
            missing_col1, missing_col2 = st.columns(2)
        
            # CSV download
            with missing_col1:
                try:
                    missing_csv = missing_concerns_df.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        "📥 Download Missing Concerns Data (CSV)",
                        data=missing_csv,
                        file_name=f"{filename_prefix}_missing_concerns.csv",
                        mime="text/csv",
                        key=f"download_missing_csv_{unique_id}",
                    )
                except Exception as e:
                    st.error(f"Error preparing missing concerns CSV: {str(e)}")
        
            # Excel download
            with missing_col2:
                try:
                    excel_buffer_missing = io.BytesIO()
                    missing_concerns_df.to_excel(
                        excel_buffer_missing, index=False, engine="openpyxl"
                    )
                    excel_buffer_missing.seek(0)
                    st.download_button(
                        "📥 Download Missing Concerns Data (Excel)",
                        data=excel_buffer_missing,
                        file_name=f"{filename_prefix}_missing_concerns.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_missing_excel_{unique_id}",
                    )
                except Exception as e:
                    st.error(f"Error preparing missing concerns Excel: {str(e)}")
        else:
            st.success("All reports have properly extracted concerns.")
    
        # block5
        def _is_response(self, row):
            """Check if a row represents a response document."""
            # Check title for response indicators
            title = str(row.get("Title", "")).lower()
            title_response = any(
                word in title for word in ["response", "reply", "answered"]
            )
    
            # Check PDF types if available
            for i in range(1, 5):  # Check PDF_1 to PDF_4
                pdf_type = str(row.get(f"PDF_{i}_Type", "")).lower()
                if pdf_type == "response":
                    return True
    
            # Check PDF names as backup
            for i in range(1, 5):
                pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
                if "response" in pdf_name or "reply" in pdf_name:
                    return True
    
            # Check content as final fallback if available
            content = str(row.get("Content", "")).lower()
            content_response = any(
                phrase in content
                for phrase in [
                    "in response to",
                    "responding to",
                    "reply to",
                    "response to",
                    "following the regulation 28",
                    "following receipt of the regulation 28",
                ]
            )
    
            return title_response or content_response

    def _filter_out_responses(self, df):
        """Filter out response documents, keeping only reports."""
        return df[~df.apply(self._is_response, axis=1)]

    def _merge_files_stack(self, files, duplicate_cols=None):
        """Merge multiple files by stacking (appending) them."""
        dfs = []
    
        for file_index, file in enumerate(files):
            try:
                # Read file
                if file.name.endswith(".csv"):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)
    
                # Display file information
                st.info(
                    f"Processing file {file_index+1}: {file.name} ({len(df)} rows, {len(df.columns)} columns)"
                )
    
                # Add source filename
                df["Source File"] = file.name
    
                # Add to the list of dataframes
                dfs.append(df)
    
            except Exception as e:
                st.warning(f"Error processing file {file.name}: {str(e)}")
                continue
    
        if not dfs:
            raise ValueError("No valid files to merge")
    
        # Combine all dataframes
        merged_df = pd.concat(dfs, ignore_index=True)
    
        # Remove duplicates if specified
        if duplicate_cols:
            valid_dup_cols = [col for col in duplicate_cols if col in merged_df.columns]
            if valid_dup_cols:
                before_count = len(merged_df)
                merged_df = merged_df.drop_duplicates(
                    subset=valid_dup_cols, keep="first"
                )
                after_count = len(merged_df)
    
                if before_count > after_count:
                    st.success(
                        f"Removed {before_count - after_count} duplicate records based on {', '.join(valid_dup_cols)}"
                    )
            else:
                st.warning(
                    f"Specified duplicate columns {duplicate_cols} not found in the merged data"
                )
    
        # ALWAYS remove duplicate Record IDs, keeping only the first occurrence
        if "Record ID" in merged_df.columns:
            before_count = len(merged_df)
            merged_df = merged_df.drop_duplicates(subset=["Record ID"], keep="first")
            after_count = len(merged_df)
    
            if before_count > after_count:
                st.success(
                    f"Removed {before_count - after_count} records with duplicate Record IDs (keeping first occurrence)"
                )
    
        # Clean coroner_name column
        if "coroner_name" in merged_df.columns:
            before_cleaning = merged_df["coroner_name"].copy()
            merged_df = self._clean_coroner_names(merged_df)
            
            # Count changes made
            changes_made = sum(before_cleaning != merged_df["coroner_name"])
            if changes_made > 0:
                st.success(f"Cleaned {changes_made} coroner name entries")
                
                # Show the first few changes as an example
                example_changes = []
                for i, (old, new) in enumerate(zip(before_cleaning, merged_df["coroner_name"])):
                    if old != new and len(example_changes) < 3 and isinstance(old, str) and isinstance(new, str):
                        example_changes.append(f"'{old}' → '{new}'")
                
                if example_changes:
                    st.info("Examples of cleaned coroner names:\n" + "\n".join(example_changes))
    
        # Clean coroner_area column
        if "coroner_area" in merged_df.columns:
            before_cleaning = merged_df["coroner_area"].copy()
            merged_df = self._clean_coroner_areas(merged_df)
            
            # Count changes made
            changes_made = sum(before_cleaning != merged_df["coroner_area"])
            if changes_made > 0:
                st.success(f"Cleaned {changes_made} coroner area entries")
                
                # Show the first few changes as an example
                example_changes = []
                for i, (old, new) in enumerate(zip(before_cleaning, merged_df["coroner_area"])):
                    if old != new and len(example_changes) < 3 and isinstance(old, str) and isinstance(new, str):
                        example_changes.append(f"'{old}' → '{new}'")
                
                if example_changes:
                    st.info("Examples of cleaned coroner areas:\n" + "\n".join(example_changes))

        ##her
        # Remove duplicate Record IDs, keeping only the first occurrence
        if "Record ID" in merged_df.columns:
            before_count = len(merged_df)
            merged_df = merged_df.drop_duplicates(subset=["Record ID"], keep="first")
            after_count = len(merged_df)

            if before_count > after_count:
                st.success(
                    f"Removed {before_count - after_count} records with duplicate Record IDs (keeping first occurrence)"
                )
        
        # Clean deceased names
        before_cleaning = merged_df["deceased_name"].copy()
        merged_df = self._clean_deceased_name(merged_df)
        
        # Count changes made to deceased names
        changes_made = sum(before_cleaning != merged_df["deceased_name"])
        if changes_made > 0:
            st.success(f"Cleaned {changes_made} deceased name entries")
            
            # Show the first few changes as an example
            example_changes = []
            for i, (old, new) in enumerate(zip(before_cleaning, merged_df["deceased_name"])):
                if old != new and len(example_changes) < 3 and isinstance(old, str) and isinstance(new, str):
                    example_changes.append(f"'{old}' → '{new}'")
            
            if example_changes:
                st.info("Examples of cleaned deceased names:\n" + "\n".join(example_changes))

        
        # Clean categories column
        if "categories" in merged_df.columns:
            # Save the original values for comparison
            original_categories = merged_df["categories"].copy()
            
            # Apply cleaning
            merged_df = self._clean_categories(merged_df)
            
            # Count changes - this is more complex since categories can be lists
            changes_made = 0
            example_changes = []
            
            # Check each row for changes
            for i, (old, new) in enumerate(zip(original_categories, merged_df["categories"])):
                # Handle list case
                if isinstance(old, list) and isinstance(new, list):
                    # Consider it changed if any element changed
                    if any(o != n for o, n in zip(old, new) if isinstance(o, str) and isinstance(n, str)):
                        changes_made += 1
                        # Add example if we don't have many yet
                        if len(example_changes) < 3:
                            old_str = ", ".join(old) if all(isinstance(x, str) for x in old) else str(old)
                            new_str = ", ".join(new) if all(isinstance(x, str) for x in new) else str(new)
                            example_changes.append(f"'{old_str}' → '{new_str}'")
                # Handle string case
                elif isinstance(old, str) and isinstance(new, str) and old != new:
                    changes_made += 1
                    if len(example_changes) < 3:
                        example_changes.append(f"'{old}' → '{new}'")
            
            # Report changes
            if changes_made > 0:
                st.success(f"Cleaned {changes_made} categories entries")
                if example_changes:
                    st.info("Examples of cleaned categories:\n" + "\n".join(example_changes))
                    
        # Store the result
        self.data = merged_df
    
        # Show summary of the merged data
        st.subheader("Merged Data Summary")
        st.write(f"Total rows: {len(merged_df)}")
        st.write(f"Columns: {', '.join(merged_df.columns)}")

    def _identify_missing_concerns(self, df):
        """
        Identify records without extracted concerns

        Args:
            df: DataFrame with BERT results

        Returns:
            DataFrame containing only records without extracted concerns
        """
        if df is None or len(df) == 0:
            return pd.DataFrame()

        # Check for concerns column
        concerns_column = None
        for col_name in ["Extracted_Concerns", "extracted_concerns", "concern_text"]:
            if col_name in df.columns:
                concerns_column = col_name
                break

        if concerns_column is None:
            st.warning("No concerns column found in the data.")
            return pd.DataFrame()

        # Filter out records with missing or empty concerns
        missing_concerns = df[
            df[concerns_column].isna()
            | (df[concerns_column].astype(str).str.strip() == "")
            | (
                df[concerns_column].astype(str).str.len() < 20
            )  # Very short extracts likely failed
        ].copy()

        return missing_concerns

    #
    def _clean_deceased_name(self, df):
        """
        Clean deceased name column by removing coroner-related text and normalizing
        
        Args:
            df (pd.DataFrame): DataFrame containing 'deceased_name' column
        
        Returns:
            pd.DataFrame: DataFrame with cleaned deceased names
        """
        if df is None or len(df) == 0 or 'deceased_name' not in df.columns:
            return df
        
        # Create a copy to avoid modifying the original
        cleaned_df = df.copy()
        
        def clean_name(name_text):
            # Return if input is not a string or is NaN
            if pd.isna(name_text) or not isinstance(name_text, str):
                return name_text
            
            # Convert to string and strip whitespace
            name = str(name_text).strip()
            
            # Remove common coroner-related prefixes and labels
            coroner_patterns = [
                r'^deceased name:?\s*',  # Remove "Deceased Name:" at start
                r'^coroner\'?s?\s*(name)?:?\s*',  # Remove "Coroner" or "Coroner's Name:" at start
                r'^deceased person:?\s*',  # Remove "Deceased Person:" at start
                r'^name of deceased:?\s*',  # Remove "Name of Deceased:" at start
            ]
            
            for pattern in coroner_patterns:
                name = re.sub(pattern, '', name, flags=re.IGNORECASE)
            
            # Remove any text after known trigger words
            name = re.sub(r'\s*(?:coroner|ref|reference).*$', '', name, flags=re.IGNORECASE)
            
            # Remove common unwanted suffixes or additional text
            name = re.sub(r'\s*\(.*\)$', '', name)  # Remove text in parentheses at end
            
            # Remove multiple spaces and extra whitespace
            name = re.sub(r'\s+', ' ', name).strip()
            
            return name
        
        # Save original values for comparison
        original_names = cleaned_df["deceased_name"].copy()
        
        # Apply cleaning to deceased name column
        cleaned_df['deceased_name'] = cleaned_df['deceased_name'].apply(clean_name)
        
        # Count and report changes
        changes_made = sum(original_names != cleaned_df['deceased_name'])
        
        # Optionally log or display changes (you can customize this part)
        if changes_made > 0:
            # Collect a few example changes
            example_changes = []
            for old, new in zip(original_names, cleaned_df['deceased_name']):
                if old != new and isinstance(old, str) and isinstance(new, str):
                    example_changes.append(f"'{old}' → '{new}'")
                    if len(example_changes) >= 5:  # Limit to 5 examples
                        break
            
            # Log or display changes
            print(f"Cleaned {changes_made} deceased name entries")
            for change in example_changes:
                print(change)
        
        return cleaned_df

    #

    def _clean_coroner_names(self, df): 
        """
        Clean coroner_name column by removing titles/prefixes and standardizing format
        
        Args:
            df (pd.DataFrame): DataFrame containing a 'coroner_name' column
            
        Returns:
            pd.DataFrame: DataFrame with cleaned 'coroner_name' column
        """
        if df is None or len(df) == 0 or 'coroner_name' not in df.columns:
            return df
        
        # Create a copy to avoid modifying the original
        cleaned_df = df.copy()
        
        # Common titles and prefixes to remove
        titles = [
            'Dr\\.?\\s+', 'Doctor\\s+', 
            'Mr\\.?\\s+', 'Mrs\\.?\\s+', 'Ms\\.?\\s+', 'Miss\\.?\\s+',
            'Prof\\.?\\s+', 'Professor\\s+',
            'Sir\\s+', 'Dame\\s+',
            'HM\\s+', 'HM\\s+Senior\\s+',
            'Hon\\.?\\s+', 'Honorable\\s+',
            'Justice\\s+', 'Judge\\s+',
            'QC\\s+', 'KC\\s+'
        ]
        
        import re
    
        def clean_name(name_text):
            if pd.isna(name_text) or not isinstance(name_text, str):
                return name_text
            
            name = name_text.strip()
    
            # Remove any titles/prefixes from the beginning
            pattern = r'^(' + '|'.join(titles) + r')+'
            name = re.sub(pattern, '', name, flags=re.IGNORECASE)
    
            # Remove known suffixes
            name = re.sub(r'\s+(QC|KC|Esq\.?|Jr\.?|Sr\.?)$', '', name, flags=re.IGNORECASE)
    
            # Remove content in parentheses
            name = re.sub(r'\(.*?\)', '', name)
    
            # Remove punctuation at the end and normalize whitespace
            name = re.sub(r'[;:,\.]$', '', name)
            name = re.sub(r'\s+', ' ', name).strip()
    
            # Remove repeated titles like "Dr Dr"
            name = re.sub(r'^(Dr\s+){2,}', '', name, flags=re.IGNORECASE)
    
            # Final cleanup: remove all text starting from "Coroner"
            name = re.sub(r'\s*Coroner.*$', '', name, flags=re.IGNORECASE)
    
            return name.strip()
        
        cleaned_df['coroner_name'] = cleaned_df['coroner_name'].apply(clean_name)
        return cleaned_df

    
    def _clean_coroner_areas(self, df):
        """
        Clean coroner_area column by:
        1. Converting everything to lowercase
        2. Removing brackets (but keeping their content)
        3. Replacing & with the word "and"
        4. Removing hyphens
        5. Removing the word "the"
        6. Making specific replacements for known locations
        
        Args:
            df (pd.DataFrame): DataFrame containing a 'coroner_area' column
            
        Returns:
            pd.DataFrame: DataFrame with cleaned 'coroner_area' column
        """
        if df is None or len(df) == 0 or 'coroner_area' not in df.columns:
            return df
        
        # Create a copy to avoid modifying the original
        cleaned_df = df.copy()
        
        # Define the cleaning function
        def clean_area(area_text):
            if pd.isna(area_text) or not isinstance(area_text, str):
                return area_text
            
            import re
            
            # Convert to lowercase
            area = area_text.lower()
            
            # Remove brackets but keep their content
            # For example, "bbbb (aaa)" becomes "bbbb aaa"
            area = re.sub(r'\(', ' ', area)  # Replace opening brackets with space
            area = re.sub(r'\)', ' ', area)  # Replace closing brackets with space
            area = re.sub(r'\[', ' ', area)  # Replace opening square brackets with space
            area = re.sub(r'\]', ' ', area)  # Replace closing square brackets with space
            
            # Replace & with 'and'
            area = area.replace('&', ' and ')  # This ensures the ampersand is properly replaced
            
            # Remove hyphens
            area = area.replace('-', ' ')  # Replace hyphens with spaces
            
            # Remove the word "the" - both standalone and as part of other words
            area = re.sub(r'\bthe\b', ' ', area)  # Remove standalone "the" with word boundaries
            
            # Replace multiple spaces with a single space
            area = re.sub(r'\s+', ' ', area)
            
            # Specific replacements for known variations - do these AFTER other cleanings
            # so they catch all variations including those with dashes or different spacing
            area = re.sub(r'\bisle of scilly\b', 'isles of scilly', area)  # Change "isle of scilly" to "isles of scilly"
            area = re.sub(r'\beast riding of yorkshire\b', 'east riding', area)  # Change "east riding of yorkshire" to "east riding"
            area = re.sub(r'\b(city of )?kingston upon hull\b', 'kingston upon hull', area)  # Remove "city of" from "kingston upon hull"
            
            # Remove common patterns that indicate the end of the coroner area
            end_patterns = [
                "coroner's concerns", 
                "matters of concern",
                "the matters of concern",
                "this report is being sent to:",
                "these reports are being sent to:",
                "the report is being sent to:",
                "this report",
                "these reports",
                "the report",
                "coroner",
                "category"
            ]
            
            # Find the earliest position of any pattern
            earliest_pos = len(area)
            for pattern in end_patterns:
                pos = area.find(pattern)
                if pos != -1 and pos < earliest_pos:
                    earliest_pos = pos
            
            # If a pattern was found, truncate
            if earliest_pos != len(area):
                area = area[:earliest_pos]
            
            # Find the position of 'Category'
            category_pos = area.find('category')
            if category_pos != -1:
                area = area[:category_pos]
            
            # Try with just a pipe character, which often separates coroner area from categories
            pipe_pos = area.find('|')
            if pipe_pos != -1:
                area = area[:pipe_pos]
                
            # Remove any special characters at the beginning or end, but keep alphanumeric and spaces
            area = re.sub(r'^[^a-z0-9]+|[^a-z0-9]+$', '', area)
            
            # Final cleanup - replace multiple spaces again and strip
            area = re.sub(r'\s+', ' ', area).strip()
            
            return area
        
        # Apply the cleaning function
        cleaned_df['coroner_area'] = cleaned_df['coroner_area'].apply(clean_area)
        
        return cleaned_df
    
        

    def _clean_categories(self, df):
        """
        Clean categories column by removing "These reports are being sent to:" and any text that follows
        
        Args:
            df (pd.DataFrame): DataFrame containing a 'categories' column
            
        Returns:
            pd.DataFrame: DataFrame with cleaned 'categories' column
        """
        if df is None or len(df) == 0 or 'categories' not in df.columns:
            return df
        
        # Create a copy to avoid modifying the original
        cleaned_df = df.copy()
        
        # Define the cleaning function for a single value
        def clean_categories_value(categories_text):
            if pd.isna(categories_text) or not isinstance(categories_text, str):
                return categories_text
            
            # Convert to lowercase for case-insensitive matching
            categories_text_lower = categories_text.lower()
            
            # Patterns to look for and remove
            report_patterns = [
                "these reports are being sent to:",
                "this report is being sent to:",
                "the report is being sent to:",
                "this report",
                "these reports",
                "the report",
                "this is being"
            ]
            
            # Find the earliest position of any report-related pattern
            earliest_pos = len(categories_text)
            for pattern in report_patterns:
                pos = categories_text_lower.find(pattern)
                if pos != -1 and pos < earliest_pos:
                    earliest_pos = pos
            
            # If a report pattern was found, truncate
            if earliest_pos != len(categories_text):
                categories_text = categories_text[:earliest_pos].strip()
            
            # Normalize text: remove brackets, convert '&' and 'and' to a standard form
            # Remove brackets
            categories_text = re.sub(r'\(.*?\)', '', categories_text).strip()
            
            # Replace variations of conjunctions
            categories_text = re.sub(r'\s*&\s*', ' and ', categories_text)
            
            return categories_text.strip()
        
        # Apply the cleaning function to the DataFrame
        # Handle both string values and list values
        before_cleaning = cleaned_df["categories"].copy()
        
        # Process based on data type
        for idx, value in enumerate(cleaned_df["categories"]):
            if isinstance(value, list):
                # For list values, we need to check each element
                cleaned_list = []
                for item in value:
                    if isinstance(item, str):
                        cleaned_list.append(clean_categories_value(item))
                    else:
                        cleaned_list.append(item)
                cleaned_df.at[idx, "categories"] = cleaned_list
            elif isinstance(value, str):
                # For string values, clean directly
                cleaned_df.at[idx, "categories"] = clean_categories_value(value)
        
        return cleaned_df



    # End of BERTResultsAnalyzer class


###########################
class ThemeAnalyzer:
    def __init__(self, model_name="emilyalsentzer/Bio_ClinicalBERT"):
        """Initialize the BERT-based theme analyzer with sentence highlighting capabilities"""
        # Initialize transformer model and tokenizer
        st.info("Loading annotation model and tokenizer... This may take a moment.")
        self.tokenizer = AutoTokenizer.from_pretrained(model_name)
        self.model = AutoModel.from_pretrained(model_name)

        # Configuration settings
        self.config = {
            "base_similarity_threshold": 0.65,
            "keyword_match_weight": 0.3,
            "semantic_similarity_weight": 0.7,
            "max_themes_per_framework": 5,
            "context_window_size": 200,
        }

        # Initialize frameworks with themes
        self.frameworks = {
            "I-SIRch": self._get_isirch_framework(),
            "House of Commons": self._get_house_of_commons_themes(),
            "Extended Analysis": self._get_extended_themes(),
        }

        # Color mapping for themes
        self.theme_color_map = {}
        self.theme_colors = [
            "#FFD580",  # Light orange
            "#FFECB3",  # Light amber
            "#E1F5FE",  # Light blue
            "#E8F5E9",  # Light green
            "#F3E5F5",  # Light purple
            "#FFF3E0",  # Light orange
            "#E0F7FA",  # Light cyan
            "#F1F8E9",  # Light lime
            "#FFF8E1",  # Light yellow
            "#E8EAF6",  # Light indigo
            "#FCE4EC",  # Light pink
            "#F5F5DC",  # Beige
            "#E6E6FA",  # Lavender
            "#FFFACD",  # Lemon chiffon
            "#D1E7DD",  # Mint
            "#F8D7DA",  # Light red
            "#D1ECF1",  # Teal light
            "#FFF3CD",  # Light yellow
            "#D6D8D9",  # Light gray
            "#CFF4FC",  # Info light
        ]

        # Pre-assign colors to frameworks
        self._preassign_framework_colors()

    def _preassign_framework_colors(self):
        """Preassign colors to each framework for consistent coloring"""
        # Create a dictionary to track colors used for each framework
        framework_colors = {}

        # Assign colors to each theme in each framework
        for framework, themes in self.frameworks.items():
            for i, theme in enumerate(themes):
                theme_key = f"{framework}_{theme['name']}"
                # Assign color from the theme_colors list, cycling if needed
                color_idx = i % len(self.theme_colors)
                self.theme_color_map[theme_key] = self.theme_colors[color_idx]

    def get_bert_embedding(self, text, max_length=512):
        """Generate BERT embedding for text"""
        if not isinstance(text, str) or not text.strip():
            return np.zeros(768)

        # Tokenize with truncation
        inputs = self.tokenizer(
            text,
            return_tensors="pt",
            truncation=True,
            max_length=max_length,
            padding=True,
        )

        # Get embeddings
        with torch.no_grad():
            outputs = self.model(**inputs)

        # Use CLS token for sentence representation
        return outputs.last_hidden_state[:, 0, :].squeeze().numpy()

    def _get_contextual_embedding(self, text, keyword, window_size=100):
        """Get embedding for text surrounding the keyword occurrence"""
        if not isinstance(text, str) or not text.strip() or keyword not in text.lower():
            return self.get_bert_embedding(keyword)

        text_lower = text.lower()
        position = text_lower.find(keyword.lower())

        # Get context window
        start = max(0, position - window_size)
        end = min(len(text), position + len(keyword) + window_size)

        # Get contextual text
        context = text[start:end]
        return self.get_bert_embedding(context)

    def _calculate_combined_score(
        self, semantic_similarity, keyword_count, text_length
    ):
        """Calculate combined score that balances semantic similarity and keyword presence"""
        # Normalize keyword count by text length
        normalized_keyword_density = min(1.0, keyword_count / (text_length / 1000))

        # Weighted combination
        keyword_component = (
            normalized_keyword_density * self.config["keyword_match_weight"]
        )
        semantic_component = (
            semantic_similarity * self.config["semantic_similarity_weight"]
        )

        return keyword_component + semantic_component

    def _find_sentence_positions(self, text, keywords):
        """Find sentences containing keywords and return their positions"""
        if not isinstance(text, str) or not text.strip():
            return []

        # Split text into sentences
        sentence_endings = r"(?<=[.!?])\s+(?=[A-Z])"
        sentences = re.split(sentence_endings, text)

        # Track character positions and matched sentences
        positions = []
        current_pos = 0

        for sentence in sentences:
            if not sentence.strip():
                current_pos += len(sentence)
                continue

            # Check if any keyword is in this sentence
            sentence_lower = sentence.lower()
            matched_keywords = []

            for keyword in keywords:
                if keyword and len(keyword) >= 3 and keyword.lower() in sentence_lower:
                    # Check if it's a whole word using word boundaries
                    keyword_lower = keyword.lower()
                    pattern = r"\b" + re.escape(keyword_lower) + r"\b"
                    if re.search(pattern, sentence_lower):
                        matched_keywords.append(keyword)

            # If sentence contains any keywords, add to positions
            if matched_keywords:
                start_pos = current_pos
                end_pos = current_pos + len(sentence)
                # Join all matched keywords
                keywords_str = ", ".join(matched_keywords)
                positions.append((start_pos, end_pos, keywords_str, sentence))

            # Move to next position
            current_pos += len(sentence)

            # Account for sentence ending characters and whitespace
            if current_pos < len(text) and text[current_pos - 1] in ".!?":
                # Check for any whitespace after sentence ending
                space_count = 0
                while (
                    current_pos + space_count < len(text)
                    and text[current_pos + space_count].isspace()
                ):
                    space_count += 1
                current_pos += space_count

        return sorted(positions)
######
    def create_highlighted_html(self, text, theme_highlights):
        """Create HTML with sentences highlighted by theme with improved color consistency"""
        if not text or not theme_highlights:
            return text
        
        # Convert highlights to a flat list of positions
        all_positions = []
        for theme_key, positions in theme_highlights.items():
            theme_color = self._get_theme_color(theme_key)
            for pos_info in positions:
                # position format: (start_pos, end_pos, keywords_str, sentence)
                all_positions.append((
                    pos_info[0],  # start position
                    pos_info[1],  # end position
                    theme_key,    # theme key
                    pos_info[2],  # keywords string
                    pos_info[3],  # original sentence
                    theme_color   # theme color
                ))
        
        # Sort positions by start position
        all_positions.sort()
        
        # Merge overlapping sentences using primary theme's color
        merged_positions = []
        if all_positions:
            current = all_positions[0]
            for i in range(1, len(all_positions)):
                if all_positions[i][0] <= current[1]:  # Overlap
                    # Create a meaningful theme name combination
                    combined_theme = current[2] + " + " + all_positions[i][2]
                    combined_keywords = current[3] + " + " + all_positions[i][3]
                    
                    # Use the first theme's color for consistency
                    combined_color = current[5]
                    
                    # Update current with merged information
                    current = (
                        current[0],                # Keep original start position
                        max(current[1], all_positions[i][1]),  # Take the later end position
                        combined_theme,            # Combined theme names
                        combined_keywords,         # Combined keywords
                        current[4],                # Keep original sentence
                        combined_color             # Use the first theme's color
                    )
                else:
                    merged_positions.append(current)
                    current = all_positions[i]
            merged_positions.append(current)
        
        # Create highlighted text
        result = []
        last_end = 0
        
        for start, end, theme_key, keywords, sentence, color in merged_positions:
            # Add text before this highlight
            if start > last_end:
                result.append(text[last_end:start])
            
            # Simple styling with solid background color
            style = f"background-color:{color}; border:1px solid #666; border-radius:2px; padding:1px 2px;"
            tooltip = f"Theme: {theme_key}\nKeywords: {keywords}"
            result.append(f'<span style="{style}" title="{tooltip}">{text[start:end]}</span>')
            
            last_end = end
        
        # Add remaining text
        if last_end < len(text):
            result.append(text[last_end:])
        
        # Create HTML structure
        html_content = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>Highlighted Document Analysis</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    line-height: 1.6;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 20px;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                .paragraph-container {
                    border: 1px solid #ddd;
                    padding: 15px;
                    margin-bottom: 20px;
                    background-color: #f9f9f9;
                }
            </style>
        </head>
        <body>
            <h2>Document Theme Analysis</h2>
            
            <div class="paragraph-container">
                <h3>Highlighted Text</h3>
                <p>
        """
        
        # Add the highlighted text
        html_content += ''.join(result)
        
        html_content += """
                </p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Framework</th>
                        <th>Theme</th>
                        <th>Color</th>
                        <th>Matched Keywords</th>
                        <th>Extracted Sentence</th>
                    </tr>
                </thead>
                <tbody>
        """
    
        # Add rows to HTML with color-coded backgrounds
        for start, end, theme_key, keywords, sentence, color in merged_positions:
            # Split theme key into framework and theme
            framework, theme = theme_key.split('_', 1) if '_' in theme_key else (theme_key, theme_key)
            
            html_content += f"""
                    <tr>
                        <td>{framework}</td>
                        <td>{theme}</td>
                        <td style="background-color: {color};">{color}</td>
                        <td>{keywords}</td>
                        <td>{sentence}</td>
                    </tr>
            """
    
        # Close HTML structure
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
    
        return html_content
    
    def convert_html_to_pdf(self, html_content, output_filename=None):
        """
        Convert the generated HTML to a PDF file
        
        Args:
            html_content (str): HTML content to convert
            output_filename (str, optional): Filename for the PDF. 
                                             If None, generates a timestamped filename.
        
        Returns:
            str: Path to the generated PDF file
        """
        try:
            # Import weasyprint
            from weasyprint import HTML, CSS
            from datetime import datetime
            import os
            
            # Generate default filename if not provided
            if output_filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"theme_analysis_{timestamp}.pdf"
            
            # Ensure the filename ends with .pdf
            if not output_filename.lower().endswith('.pdf'):
                output_filename += '.pdf'
            
            # Additional CSS to ensure proper PDF rendering
            additional_css = """
            @page {
                size: A4;
                margin: 1cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 12pt;
                line-height: 1.6;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            .paragraph-container {
                page-break-inside: avoid;
            }
            """
            
            # Create PDF
            HTML(string=html_content).write_pdf(
                output_filename, 
                stylesheets=[CSS(string=additional_css)]
            )
            
            return output_filename
        
        except ImportError:
            # Handle case where weasyprint is not installed
            st.error("WeasyPrint is not installed. Please install it using 'pip install weasyprint'")
            return None
        except Exception as e:
            # Handle other potential errors
            st.error(f"Error converting HTML to PDF: {str(e)}")
            return None
        
    def _create_integrated_html_for_pdf(self, results_df, highlighted_texts):
        """
        Create a single integrated HTML file with all highlighted records, themes, and framework information
        that can be easily converted to PDF
        """
    
    
    
        # Map report IDs to their themes
        report_themes = defaultdict(list)
        
        # Ensure all themes have unique colors
        self._ensure_unique_theme_colors(results_df)
    
        # Build the report data with consistent colors
        for _, row in results_df.iterrows():
            if "Record ID" in row and "Theme" in row and "Framework" in row:
                record_id = row["Record ID"]
                framework = row["Framework"]
                theme = row["Theme"]
                confidence = row.get("Confidence", "")
                score = row.get("Combined Score", 0)
                matched_keywords = row.get("Matched Keywords", "")
    
                # Get theme color from our mapping
                theme_key = f"{framework}_{theme}"
                theme_color = self._get_theme_color(theme_key)
    
                report_themes[record_id].append({
                    "framework": framework,
                    "theme": theme,
                    "confidence": confidence,
                    "score": score,
                    "keywords": matched_keywords,
                    "color": theme_color,
                    "theme_key": theme_key
                })
    
        # Create HTML content with modern styling
        html_content = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>Theme Analysis Report</title>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1">
                <style>
                    body { 
                        font-family: Arial, sans-serif; 
                        line-height: 1.6; 
                        margin: 0;
                        padding: 20px;
                        color: #333;
                        background-color: #f9f9f9;
                    }
                    h1 { 
                        color: #2c3e50; 
                        border-bottom: 3px solid #3498db; 
                        padding-bottom: 10px; 
                        margin-top: 30px;
                        font-weight: 600;
                    }
                    h2 { 
                        color: #2c3e50; 
                        margin-top: 30px; 
                        border-bottom: 2px solid #bdc3c7; 
                        padding-bottom: 5px; 
                        font-weight: 600;
                    }
                    h3 {
                        color: #34495e;
                        font-weight: 600;
                        margin-top: 20px;
                    }
                    .record-container { 
                        margin-bottom: 40px; 
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 20px;
                        page-break-after: always; 
                    }
                    .highlighted-text { 
                        margin: 15px 0; 
                        padding: 15px; 
                        border-radius: 4px;
                        border: 1px solid #ddd; 
                        background-color: #fff; 
                        line-height: 1.7;
                    }
                    .theme-info { margin: 15px 0; }
                    .theme-info table { 
                        border-collapse: collapse; 
                        width: 100%; 
                        margin-top: 15px;
                        border-radius: 4px;
                        overflow: hidden;
                    }
                    .theme-info th, .theme-info td { 
                        border: 1px solid #ddd; 
                        padding: 12px; 
                        text-align: left; 
                    }
                    .theme-info th { 
                        background-color: #3498db; 
                        color: white;
                        font-weight: 600;
                    }
                    .theme-info tr:nth-child(even) { background-color: #f9f9f9; }
                    .theme-info tr:hover { background-color: #f1f1f1; }
                    .high-confidence { background-color: #D5F5E3; }  /* Light green */
                    .medium-confidence { background-color: #FCF3CF; } /* Light yellow */
                    .low-confidence { background-color: #FADBD8; }   /* Light red */
                    .report-header {
                        background-color: #3498db;
                        color: white;
                        padding: 30px;
                        text-align: center;
                        border-radius: 8px;
                        margin-bottom: 30px;
                    }
                    .summary-card {
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 20px;
                        margin-bottom: 30px;
                        display: flex;
                        flex-wrap: wrap;
                        justify-content: space-between;
                    }
                    .summary-box {
                        flex: 1;
                        min-width: 200px;
                        padding: 15px;
                        text-align: center;
                        border-right: 1px solid #eee;
                    }
                    .summary-box:last-child {
                        border-right: none;
                    }
                    .summary-number {
                        font-size: 36px;
                        font-weight: bold;
                        color: #3498db;
                        margin-bottom: 10px;
                    }
                    .summary-label {
                        font-size: 14px;
                        color: #7f8c8d;
                        text-transform: uppercase;
                    }
                    .theme-color-box {
                        display: inline-block;
                        width: 20px;
                        height: 20px;
                        margin-right: 5px;
                        vertical-align: middle;
                        border: 1px solid #999;
                    }
                    .legend-container {
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 15px;
                        margin-bottom: 20px;
                    }
                    .legend-title {
                        font-weight: bold;
                        margin-bottom: 10px;
                    }
                    .legend-item {
                        display: flex;
                        align-items: center;
                        margin-bottom: 5px;
                    }
                    @media print {
                        .record-container { page-break-after: always; }
                        body { background-color: white; }
                        .record-container, .summary-card { box-shadow: none; }
                    }
                    
                    /* Define theme-specific CSS classes for consistency */
        """
    
        # Add dynamic CSS classes for each theme
        theme_keys = set()
        for _, row in results_df.iterrows():
            if "Framework" in row and "Theme" in row:
                theme_keys.add(f"{row['Framework']}_{row['Theme']}")
        
        for theme_key in theme_keys:
            color = self._get_theme_color(theme_key)
            safe_class_name = "theme-" + theme_key.replace(" ", "-").replace("(", "").replace(")", "").replace(",", "").replace(".", "").lower()
            html_content += f"""
                    .{safe_class_name} {{
                        background-color: {color} !important;
                    }}
            """
    
        html_content += """
                </style>
            </head>
            <body>
                <div class="report-header">
                    <h1>Theme Analysis Results</h1>
                    <p>Generated on """ + datetime.now().strftime("%d %B %Y, %H:%M") + """</p>
                </div>
                
                <div class="summary-card">
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(highlighted_texts)) + """</div>
                        <div class="summary-label">Documents Analyzed</div>
                    </div>
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(results_df)) + """</div>
                        <div class="summary-label">Theme Identifications</div>
                    </div>
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(results_df["Framework"].unique())) + """</div>
                        <div class="summary-label">Frameworks</div>
                    </div>
                </div>
                
                <!-- Add legend explaining gradients -->
                <div class="legend-container">
                    <div class="legend-title">Theme Color Guide</div>
                    <div>When text contains multiple themes, a gradient background is used to show all applicable themes. Check the tooltip for details.</div>
                </div>
            """
    
        # Add framework summary
        html_content += """
                <h2>Framework Summary</h2>
                <table class="theme-info">
                    <tr>
                        <th>Framework</th>
                        <th>Number of Themes</th>
                        <th>Number of Documents</th>
                    </tr>
            """
    
        for framework in results_df["Framework"].unique():
            framework_results = results_df[results_df["Framework"] == framework]
            num_themes = len(framework_results["Theme"].unique())
            num_docs = len(framework_results["Record ID"].unique())
    
            html_content += f"""
                    <tr>
                        <td>{framework}</td>
                        <td>{num_themes}</td>
                        <td>{num_docs}</td>
                    </tr>
                """
    
        html_content += """
                </table>
            """
    
        # Add each record with its themes and highlighted text
        html_content += "<h2>Document Analysis</h2>"
    
        for record_id, themes in report_themes.items():
            if record_id in highlighted_texts:
                record_title = next(
                    (row["Title"] for _, row in results_df.iterrows() if row.get("Record ID") == record_id),
                    f"Document {record_id}"
                )
    
                html_content += f"""
                    <div class="record-container">
                        <h2>Document: {record_title}</h2>
                        
                        <div class="theme-info">
                            <h3>Identified Themes</h3>
                            <table>
                                <tr>
                                    <th>Framework</th>
                                    <th>Theme</th>
                                    <th>Confidence</th>
                                    <th>Score</th>
                                    <th>Matched Keywords</th>
                                    <th>Color</th>
                                </tr>
                    """
    
                # Add theme rows with consistent styling and no gradients
                for theme_info in sorted(themes, key=lambda x: (x["framework"], -x.get("score", 0))):
                    theme_color = theme_info["color"]
                    
                    html_content += f"""
                                <tr style="background-color: {theme_color};">
                                    <td>{theme_info['framework']}</td>
                                    <td>{theme_info['theme']}</td>
                                    <td>{theme_info.get('confidence', '')}</td>
                                    <td>{round(theme_info.get('score', 0), 3)}</td>
                                    <td>{theme_info.get('keywords', '')}</td>
                                    <td><div class="theme-color-box" style="background-color:{theme_color};"></div></td>
                                </tr>
                        """
    
                html_content += """
                            </table>
                        </div>
                        
                        <div class="highlighted-text">
                            <h3>Text with Highlighted Keywords</h3>
                    """
    
                # Add highlighted text
                html_content += highlighted_texts[record_id]
    
                html_content += """
                        </div>
                    </div>
                    """
    
        html_content += """
            </body>
            </html>
            """
    
        return html_content
    
    def _ensure_unique_theme_colors(self, results_df):
        """Ensure all themes have unique colors by checking and reassigning if needed"""
        from collections import defaultdict
        
        # First collect all theme keys
        theme_keys = set()
        for _, row in results_df.iterrows():
            if "Framework" in row and "Theme" in row:
                theme_key = f"{row['Framework']}_{row['Theme']}"
                theme_keys.add(theme_key)
        
        # Assign colors for any missing themes
        for theme_key in theme_keys:
            if theme_key not in self.theme_color_map:
                self._assign_unique_theme_color(theme_key)
        
        # Check for duplicate colors and fix them
        color_to_themes = defaultdict(list)
        for theme_key in theme_keys:
            color = self.theme_color_map[theme_key]
            color_to_themes[color].append(theme_key)
        
        # Reassign colors for themes with duplicates
        for color, duplicate_themes in color_to_themes.items():
            if len(duplicate_themes) > 1:
                # Keep the first theme's color, reassign others
                for theme_key in duplicate_themes[1:]:
                    self._assign_unique_theme_color(theme_key)
    
    def _assign_unique_theme_color(self, theme_key):
        """Assign a unique color to a theme, ensuring no duplicates"""
        # Get currently used colors
        used_colors = set(self.theme_color_map.values())
        
        # Find an unused color from our palette
        for color in self.theme_colors:
            if color not in used_colors:
                self.theme_color_map[theme_key] = color
                return

        
        def random_hex_color():
            """Generate a random pastel color that's visually distinct"""
            # Higher base value (200) ensures lighter/pastel colors
            r = random.randint(180, 240)
            g = random.randint(180, 240)
            b = random.randint(180, 240)
            return f"#{r:02x}{g:02x}{b:02x}"
        
        # Generate colors until we find one that's not too similar to existing ones
        while True:
            new_color = random_hex_color()
            if new_color not in used_colors:
                self.theme_color_map[theme_key] = new_color
                break

    
    def _create_gradient_css(self, colors):
        """Create a CSS gradient string from a list of colors
        
        For 2 colors: simple diagonal gradient
        For 3+ colors: striped gradient with equal divisions
        """
        if len(colors) == 2:
            # Simple diagonal gradient for 2 colors
            return f"linear-gradient(135deg, {colors[0]} 50%, {colors[1]} 50%)"
        else:
            # Create striped gradient for 3+ colors
            stops = []
            segment_size = 100.0 / len(colors)
            
            for i, color in enumerate(colors):
                start = i * segment_size
                end = (i + 1) * segment_size
                
                # Add color stop
                stops.append(f"{color} {start:.1f}%")
                stops.append(f"{color} {end:.1f}%")
            
            return f"linear-gradient(135deg, {', '.join(stops)})"




######



    

    
    def _get_theme_color(self, theme_key):
        """Get a consistent color for a specific theme"""
        # If this theme already has an assigned color, use it
        if theme_key in self.theme_color_map:
            return self.theme_color_map[theme_key]

        # Extract framework and theme from the theme_key (format: "framework_theme")
        parts = theme_key.split("_", 1)
        framework = parts[0] if len(parts) > 0 else "unknown"

        # Count existing colors for this framework
        framework_count = sum(
            1
            for existing_key in self.theme_color_map
            if existing_key.startswith(framework + "_")
        )

        # Assign the next available color from our palette
        color_idx = framework_count % len(self.theme_colors)
        assigned_color = self.theme_colors[color_idx]

        # Store the assignment for future consistency
        self.theme_color_map[theme_key] = assigned_color
        return assigned_color

    def analyze_document(self, text):
        """Analyze document text for themes and highlight sentences containing theme keywords"""
        if not isinstance(text, str) or not text.strip():
            return {}, {}

        # Get full document embedding
        document_embedding = self.get_bert_embedding(text)
        text_length = len(text.split())

        framework_themes = {}
        theme_highlights = {}

        for framework_name, framework_theme_list in self.frameworks.items():
            # Track keyword matches across the entire document
            all_keyword_matches = []

            # First pass: identify all keyword matches and their contexts
            theme_matches = []
            for theme in framework_theme_list:
                # Find all sentence positions containing any matching keywords
                sentence_positions = self._find_sentence_positions(
                    text, theme["keywords"]
                )

                # Extract keywords from sentence positions
                keyword_matches = []
                match_contexts = []

                for _, _, keywords_str, _ in sentence_positions:
                    for keyword in keywords_str.split(", "):
                        if keyword not in keyword_matches:
                            keyword_matches.append(keyword)

                            # Get contextual embeddings for each keyword occurrence
                            context_embedding = self._get_contextual_embedding(
                                text, keyword, self.config["context_window_size"]
                            )
                            match_contexts.append(context_embedding)

                # Calculate semantic similarity with theme description
                theme_description = theme["name"] + ": " + ", ".join(theme["keywords"])
                theme_embedding = self.get_bert_embedding(theme_description)
                theme_doc_similarity = cosine_similarity(
                    [document_embedding], [theme_embedding]
                )[0][0]

                # Calculate context similarities if available
                context_similarities = []
                if match_contexts:
                    for context_emb in match_contexts:
                        sim = cosine_similarity([context_emb], [theme_embedding])[0][0]
                        context_similarities.append(sim)

                # Use max context similarity if available, otherwise use document similarity
                max_context_similarity = (
                    max(context_similarities) if context_similarities else 0
                )
                semantic_similarity = max(theme_doc_similarity, max_context_similarity)

                # Calculate combined score
                combined_score = self._calculate_combined_score(
                    semantic_similarity, len(keyword_matches), text_length
                )

                if (
                    keyword_matches
                    and combined_score >= self.config["base_similarity_threshold"]
                ):
                    theme_matches.append(
                        {
                            "theme": theme["name"],
                            "semantic_similarity": round(semantic_similarity, 3),
                            "combined_score": round(combined_score, 3),
                            "matched_keywords": ", ".join(keyword_matches),
                            "keyword_count": len(keyword_matches),
                            "sentence_positions": sentence_positions,  # Store sentence positions for highlighting
                        }
                    )

                    all_keyword_matches.extend(keyword_matches)

            # Sort by combined score
            theme_matches.sort(key=lambda x: x["combined_score"], reverse=True)

            # Limit number of themes
            top_theme_matches = theme_matches[: self.config["max_themes_per_framework"]]

            # Store theme matches and their highlighting info
            if top_theme_matches:
                # Count keywords to identify potential overlaps
                keyword_counter = Counter(all_keyword_matches)

                # Filter out themes with high keyword overlap and lower scores
                final_themes = []
                used_keywords = set()

                for theme_match in top_theme_matches:
                    # Check if this theme adds unique keywords
                    theme_keywords = set(theme_match["matched_keywords"].split(", "))
                    unique_keywords = theme_keywords - used_keywords

                    # If theme adds unique keywords or has high score, include it
                    if unique_keywords or theme_match["combined_score"] > 0.75:
                        # Store the theme data
                        theme_match_data = {
                            "theme": theme_match["theme"],
                            "semantic_similarity": theme_match["semantic_similarity"],
                            "combined_score": theme_match["combined_score"],
                            "matched_keywords": theme_match["matched_keywords"],
                            "keyword_count": theme_match["keyword_count"],
                        }
                        final_themes.append(theme_match_data)

                        # Store the highlighting positions separately
                        theme_key = f"{framework_name}_{theme_match['theme']}"
                        theme_highlights[theme_key] = theme_match["sentence_positions"]

                        used_keywords.update(theme_keywords)

                framework_themes[framework_name] = final_themes
            else:
                framework_themes[framework_name] = []

        return framework_themes, theme_highlights

    def _get_isirch_framework(self):
        """I-SIRCh framework themes mapped exactly to the official framework structure"""
        return [
            {
                "name": "External - Policy factor",
                "keywords": ["policy factor", "policy", "factor"],
            },
            {
                "name": "External - Societal factor",
                "keywords": ["societal factor", "societal", "factor"],
            },
            {
                "name": "External - Economic factor",
                "keywords": ["economic factor", "economic", "factor"],
            },
            {"name": "External - COVID ✓", "keywords": ["covid ✓", "covid"]},
            {
                "name": "External - Geographical factor (e.g. Location of patient)",
                "keywords": [
                    "geographical factor",
                    "geographical",
                    "factor",
                    "location of patient",
                ],
            },
            {
                "name": "Internal - Physical layout and Environment",
                "keywords": [
                    "physical layout and environment",
                    "physical",
                    "layout",
                    "environment",
                ],
            },
            {
                "name": "Internal - Acuity (e.g., capacity of the maternity unit as a whole)",
                "keywords": ["acuity", "capacity of the maternity unit as a whole"],
            },
            {
                "name": "Internal - Availability (e.g., operating theatres)",
                "keywords": ["availability", "operating theatres"],
            },
            {
                "name": "Internal - Time of day (e.g., night working or day of the week)",
                "keywords": ["time of day", "time", "night working or day of the week"],
            },
            {
                "name": "Organisation - Team culture factor (e.g., patient safety culture)",
                "keywords": [
                    "team culture factor",
                    "team",
                    "culture",
                    "factor",
                    "patient safety culture",
                ],
            },
            {
                "name": "Organisation - Incentive factor (e.g., performance evaluation)",
                "keywords": [
                    "incentive factor",
                    "incentive",
                    "factor",
                    "performance evaluation",
                ],
            },
            {"name": "Organisation - Teamworking", "keywords": ["teamworking"]},
            {
                "name": "Organisation - Communication factor",
                "keywords": ["communication factor", "communication", "factor"],
            },
            {
                "name": "Organisation - Communication factor - Between staff",
                "keywords": ["between staff", "between", "staff"],
            },
            {
                "name": "Organisation - Communication factor - Between staff and patient (verbal)",
                "keywords": [
                    "between staff and patient",
                    "between",
                    "staff",
                    "patient",
                    "verbal",
                ],
            },
            {"name": "Organisation - Documentation", "keywords": ["documentation"]},
            {
                "name": "Organisation - Escalation/referral factor (including fresh eyes reviews)",
                "keywords": [
                    "escalation/referral factor",
                    "escalation/referral",
                    "factor",
                    "including fresh eyes reviews",
                    "specialist referral",
                    "delay in escalation",
                    "specialist review",
                    "senior input",
                    "interdisciplinary referral",
                    "escalation delay",
                    "consultant opinion",
                ],
            },
            {
                "name": "Organisation - National and/or local guidance",
                "keywords": [
                    "national and/or local guidance",
                    "national",
                    "local",
                    "guidance",
                    "national screening",
                    "screening program",
                    "standard implementation",
                    "standardized screening",
                    "protocol adherence",
                ],
            },
            {
                "name": "Organisation - Language barrier",
                "keywords": ["language barrier", "language", "barrier"],
            },
            {
                "name": "Jobs/Task - Assessment, investigation, testing, screening (e.g., holistic review)",
                "keywords": [
                    "assessment, investigation, testing, screening",
                    "assessment,",
                    "investigation,",
                    "testing,",
                    "screening",
                    "holistic review",
                    "specimen",
                    "sample",
                    "laboratory",
                    "test result",
                    "abnormal finding",
                    "test interpretation",
                ],
            },
            {
                "name": "Jobs/Task - Care planning",
                "keywords": ["care planning", "care", "planning"],
            },
            {
                "name": "Jobs/Task - Dispensing, administering",
                "keywords": [
                    "dispensing, administering",
                    "dispensing,",
                    "administering",
                ],
            },
            {"name": "Jobs/Task - Monitoring", "keywords": ["monitoring"]},
            {
                "name": "Jobs/Task - Risk assessment",
                "keywords": ["risk assessment", "risk", "assessment"],
            },
            {
                "name": "Jobs/Task - Situation awareness (e.g., loss of helicopter view)",
                "keywords": [
                    "situation awareness",
                    "situation",
                    "awareness",
                    "loss of helicopter view",
                ],
            },
            {
                "name": "Jobs/Task - Obstetric review",
                "keywords": ["obstetric review", "obstetric", "review"],
            },
            {"name": "Technologies - Issues", "keywords": ["issues"]},
            {
                "name": "Technologies - Interpretation (e.g., CTG)",
                "keywords": ["interpretation", "ctg"],
            },
            {
                "name": "Person - Patient (characteristics and performance)",
                "keywords": ["patient", "characteristics and performance"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics",
                "keywords": ["characteristics", "patient characteristics"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Physical characteristics",
                "keywords": ["physical characteristics", "physical", "characteristics"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Psychological characteristics (e.g., stress, mental health)",
                "keywords": [
                    "psychological characteristics",
                    "psychological",
                    "characteristics",
                    "stress",
                    "mental health",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Language competence (English)",
                "keywords": [
                    "language competence",
                    "language",
                    "competence",
                    "english",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Disability (e.g., hearing problems)",
                "keywords": ["disability", "hearing problems"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Training and education (e.g., attendance at ante-natal classes)",
                "keywords": [
                    "training and education",
                    "training",
                    "education",
                    "attendance at ante-natal classes",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Record of attendance (e.g., failure to attend antenatal classes)",
                "keywords": [
                    "record of attendance",
                    "record",
                    "attendance",
                    "failure to attend antenatal classes",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Performance",
                "keywords": ["performance", "patient performance"],
            },
            {
                "name": "Person - Staff (characteristics and performance)",
                "keywords": ["staff", "characteristics and performance"],
            },
            {
                "name": "Person - Staff (characteristics and performance) - Characteristics",
                "keywords": ["characteristics", "staff characteristics"],
            },
            {
                "name": "Person - Staff (characteristics and performance) - Performance",
                "keywords": ["performance", "staff performance"],
            },
        ]

    def _get_house_of_commons_themes(self):
        """House of Commons themes mapped exactly to the official document"""
        return [
            {
                "name": "Communication",
                "keywords": [
                    "communication",
                    "dismissed",
                    "listened",
                    "concerns not taken seriously",
                    "concerns",
                    "seriously",
                ],
            },
            {
                "name": "Fragmented care",
                "keywords": [
                    "fragmented care",
                    "fragmented",
                    "care",
                    "spread",
                    "poorly",
                    "communicating",
                    "providers",
                    "no clear coordination",
                    "clear",
                    "coordination",
                ],
            },
            {
                "name": "Guidance gaps",
                "keywords": [
                    "guidance gaps",
                    "guidance",
                    "gaps",
                    "information",
                    "needs",
                    "optimal",
                    "minority",
                ],
            },
            {
                "name": "Pre-existing conditions and comorbidities",
                "keywords": [
                    "pre-existing conditions and comorbidities",
                    "pre-existing",
                    "conditions",
                    "comorbidities",
                    "overrepresented",
                    "ethnic",
                    "minority",
                    "contribute",
                    "higher",
                    "mortality",
                ],
            },
            {
                "name": "Inadequate maternity care",
                "keywords": [
                    "inadequate maternity care",
                    "inadequate",
                    "maternity",
                    "care",
                    "individualized",
                    "culturally",
                    "sensitive",
                ],
            },
            {
                "name": "Care quality and access issues",
                "keywords": [
                    "microaggressions and racism",
                    "microaggressions",
                    "racism",
                    "implicit/explicit",
                    "impacts",
                    "access",
                    "treatment",
                    "quality",
                    "stereotyping",
                ],
            },
            {
                "name": "Socioeconomic factors and deprivation",
                "keywords": [
                    "socioeconomic factors and deprivation",
                    "socioeconomic",
                    "factors",
                    "deprivation",
                    "links to poor outcomes",
                    "links",
                    "outcomes",
                    "minority",
                    "overrepresented",
                    "deprived",
                    "areas",
                ],
            },
            {
                "name": "Biases and stereotyping",
                "keywords": [
                    "biases and stereotyping",
                    "biases",
                    "stereotyping",
                    "perpetuation",
                    "stereotypes",
                    "providers",
                ],
            },
            {
                "name": "Consent/agency",
                "keywords": [
                    "consent/agency",
                    "consent",
                    "agency",
                    "informed consent",
                    "agency over care decisions",
                    "informed",
                    "decisions",
                ],
            },
            {
                "name": "Dignity/respect",
                "keywords": [
                    "dignity/respect",
                    "dignity",
                    "respect",
                    "neglectful",
                    "lacking",
                    "discrimination faced",
                    "discrimination",
                    "faced",
                ],
            },
        ]

    def _get_extended_themes(self):
        """Extended Analysis themes with unique concepts not covered in I-SIRCh or House of Commons frameworks"""
        return [
            {
                "name": "Procedural and Process Failures",
                "keywords": [
                    "procedure failure",
                    "process breakdown",
                    "protocol breach",
                    "standard violation",
                    "workflow issue",
                    "operational failure",
                    "process gap",
                    "procedural deviation",
                    "system failure",
                    "process error",
                    "workflow disruption",
                    "task failure",
                ],
            },
            {
                "name": "Medication safety",
                "keywords": [
                    "medication safety",
                    "medication",
                    "safety",
                    "drug error",
                    "prescription",
                    "drug administration",
                    "medication error",
                    "adverse reaction",
                    "medication reconciliation",
                ],
            },
            {
                "name": "Resource allocation",
                "keywords": [
                    "resource allocation",
                    "resource",
                    "allocation",
                    "resource management",
                    "resource constraints",
                    "prioritisation",
                    "resource distribution",
                    "staffing levels",
                    "staff shortage",
                    "budget constraints",
                ],
            },
            {
                "name": "Facility and Equipment Issues",
                "keywords": [
                    "facility",
                    "equipment",
                    "maintenance",
                    "infrastructure",
                    "device failure",
                    "equipment malfunction",
                    "equipment availability",
                    "technical failure",
                    "equipment maintenance",
                    "facility limitations",
                ],
            },
            {
                "name": "Emergency preparedness",
                "keywords": [
                    "emergency preparedness",
                    "emergency protocol",
                    "emergency response",
                    "crisis management",
                    "contingency planning",
                    "disaster readiness",
                    "emergency training",
                    "rapid response",
                ],
            },
            {
                "name": "Staff Wellbeing and Burnout",
                "keywords": [
                    "burnout",
                    "staff wellbeing",
                    "resilience",
                    "psychological safety",
                    "stress management",
                    "compassion fatigue",
                    "work-life balance",
                    "staff support",
                    "mental health",
                    "emotional burden",
                ],
            },
            {
                "name": "Ethical considerations",
                "keywords": [
                    "ethical dilemma",
                    "ethical decision",
                    "moral distress",
                    "ethical conflict",
                    "value conflict",
                    "ethics committee",
                    "moral judgment",
                    "conscientious objection",
                    "ethical framework",
                ],
            },
            {
                "name": "Diagnostic process",
                "keywords": [
                    "diagnostic error",
                    "misdiagnosis",
                    "delayed diagnosis",
                    "diagnostic uncertainty",
                    "diagnostic reasoning",
                    "differential diagnosis",
                    "diagnostic testing",
                    "diagnostic accuracy",
                    "test interpretation",
                ],
            },
            {
                "name": "Post-Event Learning and Improvement",
                "keywords": [
                    "incident learning",
                    "corrective action",
                    "improvement plan",
                    "feedback loop",
                    "lessons learned",
                    "action tracking",
                    "improvement verification",
                    "learning culture",
                    "incident review",
                    "recommendation implementation",
                    "systemic improvement",
                    "organisational learning",
                ],
            },
            {
                "name": "Electronic Health Record Issues",
                "keywords": [
                    "electronic health record",
                    "ehr issue",
                    "alert fatigue",
                    "interface design",
                    "copy-paste error",
                    "dropdown selection",
                    "clinical decision support",
                    "digital documentation",
                    "system integration",
                    "information retrieval",
                    "data entry error",
                    "electronic alert",
                ],
            },
            {
                "name": "Time-Critical Interventions",
                "keywords": [
                    "time-critical",
                    "delayed intervention",
                    "response time",
                    "golden hour",
                    "deterioration recognition",
                    "rapid response",
                    "timely treatment",
                    "intervention delay",
                    "time sensitivity",
                    "critical timing",
                    "delayed recognition",
                    "prompt action",
                    "urgent intervention",
                    "emergency response",
                    "time-sensitive decision",
                    "immediate action",
                    "rapid assessment",
                ],
            },
            {
                "name": "Human Factors and Cognitive Aspects",
                "keywords": [
                    "cognitive bias",
                    "situational awareness",
                    "attention management",
                    "visual perception",
                    "cognitive overload",
                    "decision heuristic",
                    "tunnel vision",
                    "confirmation bias",
                    "fixation error",
                    "anchoring bias",
                    "memory limitation",
                    "cognitive fatigue",
                    "isolation decision-making",
                    "clinical confidence",
                    "professional authority",
                    "hierarchical barriers",
                    "professional autonomy",
                ],
            },
            {
                "name": "Service Design and Patient Flow",
                "keywords": [
                    "service design",
                    "patient flow",
                    "care pathway",
                    "bottleneck",
                    "patient journey",
                    "waiting time",
                    "system design",
                    "process mapping",
                    "patient transfer",
                    "capacity planning",
                    "workflow design",
                    "service bottleneck",
                ],
            },
            {
                "name": "Maternal and Neonatal Risk Factors",
                "keywords": [
                    "maternal risk",
                    "pregnancy complication",
                    "obstetric risk",
                    "neonatal risk",
                    "fetal risk",
                    "gestational diabetes",
                    "preeclampsia",
                    "placental issue",
                    "maternal age",
                    "parity",
                    "previous cesarean",
                    "multiple gestation",
                    "fetal growth restriction",
                    "prematurity",
                    "congenital anomaly",
                    "birth asphyxia",
                    "maternal obesity",
                    "maternal hypertension",
                    "maternal infection",
                    "obstetric hemorrhage",
                    "maternal cardiac",
                    "thromboembolism",
                ],
            },
            {
                "name": "Private vs. NHS Care Integration",
                "keywords": [
                    "private care",
                    "private midwife",
                    "private provider",
                    "NHS interface",
                    "care transition",
                    "private-public interface",
                    "independent provider",
                    "private consultation",
                    "private-NHS coordination",
                    "privately arranged care",
                    "independent midwife",
                    "cross-system communication",
                ],
            },
            {
                "name": "Peer Support and Supervision",
                "keywords": [
                    "peer support",
                    "collegial support",
                    "professional isolation",
                    "clinical supervision",
                    "peer review",
                    "case discussion",
                    "professional feedback",
                    "unsupported decision",
                    "lack of collegiality",
                    "professional network",
                    "mentoring",
                    "supervision",
                ],
            },
            {
                "name": "Diagnostic Testing and Specimen Handling",
                "keywords": [
                    "specimen",
                    "sample",
                    "test result",
                    "laboratory",
                    "analysis",
                    "interpretation",
                    "abnormal finding",
                    "discolored",
                    "contamination",
                    "collection",
                    "processing",
                    "transportation",
                    "storage",
                    "labeling",
                    "amniocentesis",
                    "blood sample",
                ],
            },
        ]

    # New methods to add

    def _get_confidence_label(self, score):
        """Convert numerical score to confidence label"""
        if score >= 0.7:
            return "High"
        elif score >= 0.5:
            return "Medium"
        else:
            return "Low"


    def create_detailed_results(self, data, content_column="Content"):
        """
        Analyze multiple documents and create detailed results with progress tracking.
        Enhanced to include additional metadata (coroner_name, coroner_area, year).
    
        Args:
            data (pd.DataFrame): DataFrame containing documents
            content_column (str): Name of the column containing text to analyze
    
        Returns:
            Tuple[pd.DataFrame, Dict]: (Results DataFrame, Dictionary of highlighted texts)
        """
        import streamlit as st
    
        results = []
        highlighted_texts = {}
    
        # Create progress tracking elements
        progress_bar = st.progress(0)
        status_text = st.empty()
        doc_count_text = st.empty()
    
        # Calculate total documents to process
        total_docs = len(data)
        doc_count_text.text(f"Processing 0/{total_docs} documents")
    
        # Process each document
        for idx, (i, row) in enumerate(data.iterrows()):
            # Update progress
            progress = (idx + 1) / total_docs
            progress_bar.progress(progress)
            status_text.text(
                f"Analyzing document {idx + 1}/{total_docs}: {row.get('Title', f'Document {i}')}"
            )
    
            # Skip empty content
            if pd.isna(row[content_column]) or row[content_column] == "":
                continue
    
            content = str(row[content_column])
    
            # Analyze themes and get highlights
            framework_themes, theme_highlights = self.analyze_document(content)
    
            # Create highlighted HTML for this document
            highlighted_html = self.create_highlighted_html(content, theme_highlights)
            highlighted_texts[i] = highlighted_html
    
            # Store results for each theme
            theme_count = 0
            for framework_name, themes in framework_themes.items():
                for theme in themes:
                    theme_count += 1
    
                    # Extract matched sentences for this theme
                    matched_sentences = []
                    theme_key = f"{framework_name}_{theme['theme']}"
                    if theme_key in theme_highlights:
                        for (
                            start_pos,
                            end_pos,
                            keywords_str,
                            sentence,
                        ) in theme_highlights[theme_key]:
                            matched_sentences.append(sentence)
    
                    # Join sentences if there are any
                    matched_text = (
                        "; ".join(matched_sentences) if matched_sentences else ""
                    )
    
                    # Prepare the base result dictionary with theme information
                    result_dict = {
                        "Record ID": i,
                        "Title": row.get("Title", f"Document {i}"),
                        "Framework": framework_name,
                        "Theme": theme["theme"],
                        "Confidence": self._get_confidence_label(
                            theme["combined_score"]
                        ),
                        "Combined Score": theme["combined_score"],
                        "Semantic_Similarity": theme["semantic_similarity"],
                        "Matched Keywords": theme["matched_keywords"],
                        "Matched Sentences": matched_text,
                    }
                    
                    # Add additional metadata fields from the original data
                    # Only add if they exist in the input data
                    if "coroner_name" in row:
                        result_dict["coroner_name"] = row.get("coroner_name", "")
                    
                    if "coroner_area" in row:
                        result_dict["coroner_area"] = row.get("coroner_area", "")
                    
                    if "year" in row:
                        result_dict["year"] = row.get("year", "")
                    elif "Year" in row:
                        result_dict["year"] = row.get("Year", "")
                    
                    # Add date_of_report if available
                    if "date_of_report" in row:
                        result_dict["date_of_report"] = row.get("date_of_report", "")
    
                    results.append(result_dict)
    
                # Update documents processed count with theme info
                doc_count_text.text(
                    f"Processed {idx + 1}/{total_docs} documents. Found {theme_count} themes in current document."
                )
    
        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
    
        # Final count update
        if results:
            doc_count_text.text(
                f"Completed analysis of {total_docs} documents. Found {len(results)} total themes."
            )
        else:
            doc_count_text.text(
                f"Completed analysis, but no themes were identified in the documents."
            )
    
        # Create results DataFrame
        results_df = pd.DataFrame(results) if results else pd.DataFrame()
    
        return results_df, highlighted_texts

    def create_comprehensive_pdf(
        self, results_df, highlighted_texts, output_filename=None
    ):
        """
        Create a comprehensive PDF report with analysis results

        Args:
            results_df (pd.DataFrame): Results DataFrame
            highlighted_texts (Dict): Dictionary of highlighted texts
            output_filename (str, optional): Output filename

        Returns:
            str: Path to the created PDF file
        """

        # Generate default filename if not provided
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"theme_analysis_report_{timestamp}.pdf"

        # Use a tempfile for matplotlib to avoid file conflicts
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmpfile:
            temp_pdf_path = tmpfile.name

        # Create PDF with matplotlib
        with PdfPages(temp_pdf_path) as pdf:
            # Title page
            fig = plt.figure(figsize=(12, 10))
            plt.text(
                0.5,
                0.6,
                "Theme Analysis Report",
                fontsize=28,
                ha="center",
                va="center",
                weight="bold",
            )
            plt.text(
                0.5,
                0.5,
                f"Generated on {datetime.now().strftime('%d %B %Y, %H:%M')}",
                fontsize=16,
                ha="center",
                va="center",
            )

            # Add a decorative header bar
            plt.axhline(y=0.75, xmin=0.1, xmax=0.9, color="#3366CC", linewidth=3)
            plt.axhline(y=0.35, xmin=0.1, xmax=0.9, color="#3366CC", linewidth=3)

            # Add framework names
            frameworks = self.frameworks.keys()
            framework_text = "Frameworks analyzed: " + ", ".join(frameworks)
            plt.text(
                0.5,
                0.3,
                framework_text,
                fontsize=14,
                ha="center",
                va="center",
                style="italic",
            )

            plt.axis("off")
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

            # Summary statistics page
            if not results_df.empty:
                # Create a summary page with charts
                fig = plt.figure(figsize=(12, 10))
                gs = gridspec.GridSpec(3, 2, height_ratios=[1, 2, 2])

                # Header
                ax_header = plt.subplot(gs[0, :])
                ax_header.text(
                    0.5,
                    0.5,
                    "Analysis Summary",
                    fontsize=20,
                    ha="center",
                    va="center",
                    weight="bold",
                )
                ax_header.axis("off")

                # Document count and metrics
                ax_metrics = plt.subplot(gs[1, 0])
                doc_count = len(highlighted_texts)
                theme_count = len(results_df)
                frameworks_count = len(results_df["Framework"].unique())

                metrics_text = (
                    f"Total Documents Analyzed: {doc_count}\n"
                    f"Total Theme Predictions: {theme_count}\n"
                    f"Unique Frameworks: {frameworks_count}\n"
                )

                if "Confidence" in results_df.columns:
                    confidence_counts = results_df["Confidence"].value_counts()
                    metrics_text += "\nConfidence Levels:\n"
                    for conf, count in confidence_counts.items():
                        metrics_text += f"  {conf}: {count} themes\n"

                ax_metrics.text(
                    0.1, 0.9, metrics_text, fontsize=12, va="top", linespacing=2
                )
                ax_metrics.axis("off")

                # Framework distribution chart
                ax_framework = plt.subplot(gs[1, 1])
                if not results_df.empty:
                    framework_counts = results_df["Framework"].value_counts()
                    bars = ax_framework.bar(
                        framework_counts.index,
                        framework_counts.values,
                        color=["#3366CC", "#DC3912", "#FF9900"],
                    )
                    ax_framework.set_title(
                        "Theme Distribution by Framework", fontsize=14
                    )
                    ax_framework.set_ylabel("Number of Themes")

                    # Add value labels on bars
                    for bar in bars:
                        height = bar.get_height()
                        ax_framework.text(
                            bar.get_x() + bar.get_width() / 2.0,
                            height + 0.1,
                            f"{height:d}",
                            ha="center",
                            fontsize=10,
                        )

                    ax_framework.spines["top"].set_visible(False)
                    ax_framework.spines["right"].set_visible(False)
                    plt.setp(ax_framework.get_xticklabels(), rotation=30, ha="right")

                # Themes by confidence chart
                ax_confidence = plt.subplot(gs[2, :])
                if "Confidence" in results_df.columns and "Theme" in results_df.columns:
                    # Prepare data for stacked bar chart
                    theme_conf_data = pd.crosstab(
                        results_df["Theme"], results_df["Confidence"]
                    )

                    # Select top themes by total count
                    top_themes = (
                        theme_conf_data.sum(axis=1)
                        .sort_values(ascending=False)
                        .head(10)
                        .index
                    )
                    theme_conf_data = theme_conf_data.loc[top_themes]

                    # Plot stacked bar chart
                    confidence_colors = {
                        "High": "#4CAF50",
                        "Medium": "#FFC107",
                        "Low": "#F44336",
                    }

                    # Get confidence levels present in the data
                    confidence_levels = list(theme_conf_data.columns)
                    colors = [
                        confidence_colors.get(level, "#999999")
                        for level in confidence_levels
                    ]

                    theme_conf_data.plot(
                        kind="barh",
                        stacked=True,
                        ax=ax_confidence,
                        color=colors,
                        figsize=(10, 6),
                    )

                    ax_confidence.set_title(
                        "Top Themes by Confidence Level", fontsize=14
                    )
                    ax_confidence.set_xlabel("Number of Documents")
                    ax_confidence.set_ylabel("Theme")

                    # Create custom legend
                    patches = [
                        Patch(
                            color=confidence_colors.get(level, "#999999"), label=level
                        )
                        for level in confidence_levels
                    ]
                    ax_confidence.legend(
                        handles=patches, title="Confidence", loc="upper right"
                    )

                    ax_confidence.spines["top"].set_visible(False)
                    ax_confidence.spines["right"].set_visible(False)
                else:
                    ax_confidence.axis("off")
                    ax_confidence.text(
                        0.5,
                        0.5,
                        "Confidence data not available",
                        fontsize=14,
                        ha="center",
                        va="center",
                    )

                plt.tight_layout()
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

            # Framework-specific pages
            for framework_name in self.frameworks.keys():
                # Filter results for this framework
                framework_results = results_df[
                    results_df["Framework"] == framework_name
                ]

                if not framework_results.empty:
                    # Create a new page for the framework
                    fig = plt.figure(figsize=(12, 10))

                    # Title
                    plt.suptitle(
                        f"{framework_name} Framework Analysis",
                        fontsize=20,
                        y=0.95,
                        weight="bold", 
                    )

                    # Theme counts
                    theme_counts = framework_results["Theme"].value_counts().head(15)

                    if not theme_counts.empty:
                        plt.subplot(111)
                        bars = plt.barh(
                            theme_counts.index[::-1],
                            theme_counts.values[::-1],
                            color="#5975A4",
                            alpha=0.8,
                        )

                        # Add value labels
                        for i, bar in enumerate(bars):
                            width = bar.get_width()
                            plt.text(
                                width + 0.3,
                                bar.get_y() + bar.get_height() / 2,
                                f"{width:d}",
                                va="center",
                                fontsize=10,
                            )

                        plt.xlabel("Number of Documents")
                        plt.ylabel("Theme")
                        plt.title(f"Top Themes in {framework_name}", pad=20)
                        plt.grid(axis="x", linestyle="--", alpha=0.7)
                        plt.tight_layout(rect=[0, 0, 1, 0.95])  # Adjust for suptitle

                    pdf.savefig(fig, bbox_inches="tight")
                    plt.close(fig)

            # Sample highlighted documents (text descriptions only)
            if highlighted_texts:
                # Create document summaries page
                fig = plt.figure(figsize=(12, 10))
                plt.suptitle(
                    "Document Analysis Summaries", fontsize=20, y=0.95, weight="bold"
                )

                # We'll show summaries of a few documents
                max_docs_to_show = min(3, len(highlighted_texts))
                docs_to_show = list(highlighted_texts.keys())[:max_docs_to_show]

                # Get theme counts for each document
                doc_summaries = []
                for doc_id in docs_to_show:
                    doc_themes = results_df[results_df["Record ID"] == doc_id]
                    theme_count = len(doc_themes)
                    frameworks = doc_themes["Framework"].unique()
                    doc_summaries.append(
                        {
                            "doc_id": doc_id,
                            "theme_count": theme_count,
                            "frameworks": ", ".join(frameworks),
                        }
                    )

                # Format as a table-like display
                plt.axis("off")
                table_text = "Document Analysis Summaries:\n\n"
                for i, summary in enumerate(doc_summaries):
                    doc_id = summary["doc_id"]
                    table_text += f"Document {i+1} (ID: {doc_id}):\n"
                    table_text += f"  • Identified Themes: {summary['theme_count']}\n"
                    table_text += f"  • Frameworks: {summary['frameworks']}\n\n"

                plt.text(0.1, 0.8, table_text, fontsize=12, va="top", linespacing=1.5)

                # Also add a note about the full HTML version
                note_text = (
                    "Note: A detailed HTML report with highlighted text excerpts has been\n"
                    "created alongside this PDF. The HTML version provides interactive\n"
                    "highlighting of theme-related sentences in each document."
                )
                plt.text(
                    0.1,
                    0.3,
                    note_text,
                    fontsize=12,
                    va="top",
                    linespacing=1.5,
                    style="italic",
                    bbox=dict(
                        facecolor="#F0F0F0",
                        edgecolor="#CCCCCC",
                        boxstyle="round,pad=0.5",
                    ),
                )

                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        # Copy the temp file to the desired output filename
        import shutil

        shutil.copy2(temp_pdf_path, output_filename)

        # Clean up temp file
        try:
            os.unlink(temp_pdf_path)
        except:
            pass

        return output_filename

    def _preassign_framework_colors(self):
        """Preassign colors to each framework for consistent coloring"""
        # Create a dictionary to track colors used for each framework
        framework_colors = {}

        # Assign colors to each theme in each framework
        for framework, themes in self.frameworks.items():
            for i, theme in enumerate(themes):
                theme_key = f"{framework}_{theme['name']}"
                # Assign color from the theme_colors list, cycling if needed
                color_idx = i % len(self.theme_colors)
                self.theme_color_map[theme_key] = self.theme_colors[color_idx]


    def export_to_excel(df: pd.DataFrame) -> bytes:
        """
        Export BERT results DataFrame to Excel bytes with proper formatting,
        including additional metadata columns.
        """
        try:
            if df is None or len(df) == 0:
                raise ValueError("No data available to export")
    
            # Create clean copy for export
            df_export = df.copy()
    
            # Format dates to UK format
            if "date_of_report" in df_export.columns:
                df_export["date_of_report"] = df_export["date_of_report"].dt.strftime(
                    "%d/%m/%Y"
                )
    
            # Handle list columns (like categories)
            for col in df_export.columns:
                if df_export[col].dtype == "object":
                    df_export[col] = df_export[col].apply(
                        lambda x: ", ".join(x)
                        if isinstance(x, list)
                        else str(x)
                        if pd.notna(x)
                        else ""
                    )
    
            # Create output buffer
            output = io.BytesIO()
    
            # Write to Excel
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, sheet_name="BERT Results", index=False)
    
                # Get the worksheet
                worksheet = writer.sheets["BERT Results"]
    
                # Auto-adjust column widths
                for idx, col in enumerate(df_export.columns, 1):
                    # Set larger width for Matched Sentences column
                    if col == "Matched Sentences":
                        worksheet.column_dimensions[get_column_letter(idx)].width = 80
                    # Set wider width for coroner_area (often long)
                    elif col == "coroner_area":
                        worksheet.column_dimensions[get_column_letter(idx)].width = 40
                    # Set appropriate width for coroner_name
                    elif col == "coroner_name":
                        worksheet.column_dimensions[get_column_letter(idx)].width = 30
                    else:
                        max_length = max(
                            df_export[col].astype(str).apply(len).max(),
                            len(str(col)),
                        )
                        adjusted_width = min(max_length + 2, 50)
                        column_letter = get_column_letter(idx)
                        worksheet.column_dimensions[
                            column_letter
                        ].width = adjusted_width
    
                # Add filters to header row
                worksheet.auto_filter.ref = worksheet.dimensions
    
                # Freeze the header row
                worksheet.freeze_panes = "A2"
    
                # Set wrap text for Matched Sentences column
                matched_sent_col = next(
                    (
                        idx
                        for idx, col in enumerate(df_export.columns, 1)
                        if col == "Matched Sentences"
                    ),
                    None,
                )
                if matched_sent_col:
                    col_letter = get_column_letter(matched_sent_col)
                    for row in range(2, len(df_export) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = cell.alignment.copy(wrapText=True)
                        # Set row height to accommodate wrapped text
                        worksheet.row_dimensions[row].height = 60
    
            # Get the bytes value
            output.seek(0)
            return output.getvalue()
    
        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}", exc_info=True)
            raise Exception(f"Failed to export data to Excel: {str(e)}")
        

    
class BM25Vectorizer(BaseEstimator, TransformerMixin):
    """BM25 vectorizer implementation"""

    def __init__(
        self,
        k1: float = 1.5,
        b: float = 0.75,
        max_features: Optional[int] = None,
        min_df: Union[int, float] = 1,
        max_df: Union[int, float] = 1.0,
    ):
        self.k1 = k1
        self.b = b
        self.max_features = max_features
        self.min_df = min_df
        self.max_df = max_df

        self.count_vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )

    def fit(self, raw_documents: List[str], y=None):
        X = self.count_vectorizer.fit_transform(raw_documents)

        # Calculate document lengths
        self.doc_lengths = np.array(X.sum(axis=1)).flatten()
        self.avg_doc_length = np.mean(self.doc_lengths)

        # Calculate IDF scores
        n_samples = X.shape[0]
        df = np.bincount(X.indices, minlength=X.shape[1])
        df = np.maximum(df, 1)
        self.idf = np.log((n_samples - df + 0.5) / (df + 0.5) + 1.0)

        return self

    def transform(self, raw_documents: List[str]) -> sp.csr_matrix:
        X = self.count_vectorizer.transform(raw_documents)
        doc_lengths = np.array(X.sum(axis=1)).flatten()

        X = sp.csr_matrix(X)

        # Calculate BM25 scores
        for i in range(X.shape[0]):
            start_idx = X.indptr[i]
            end_idx = X.indptr[i + 1]

            freqs = X.data[start_idx:end_idx]
            length_norm = 1 - self.b + self.b * doc_lengths[i] / self.avg_doc_length

            # BM25 formula
            X.data[start_idx:end_idx] = (
                ((self.k1 + 1) * freqs) / (self.k1 * length_norm + freqs)
            ) * self.idf[X.indices[start_idx:end_idx]]

        return X

    def get_feature_names_out(self):
        return self.count_vectorizer.get_feature_names_out()


class WeightedTfidfVectorizer(BaseEstimator, TransformerMixin):
    """TF-IDF vectorizer with configurable weighting schemes"""

    def __init__(
        self,
        tf_scheme: str = "raw",
        idf_scheme: str = "smooth",
        norm: Optional[str] = "l2",
        max_features: Optional[int] = None,
        min_df: Union[int, float] = 1,
        max_df: Union[int, float] = 1.0,
    ):
        self.tf_scheme = tf_scheme
        self.idf_scheme = idf_scheme
        self.norm = norm
        self.max_features = max_features
        self.min_df = min_df
        self.max_df = max_df

        self.count_vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )

    def _compute_tf(self, X: sp.csr_matrix) -> sp.csr_matrix:
        if self.tf_scheme == "raw":
            return X
        elif self.tf_scheme == "log":
            X.data = np.log1p(X.data)
        elif self.tf_scheme == "binary":
            X.data = np.ones_like(X.data)
        elif self.tf_scheme == "augmented":
            max_tf = X.max(axis=1).toarray().flatten()
            max_tf[max_tf == 0] = 1
            for i in range(X.shape[0]):
                start = X.indptr[i]
                end = X.indptr[i + 1]
                X.data[start:end] = 0.5 + 0.5 * (X.data[start:end] / max_tf[i])
        return X

    def _compute_idf(self, X: sp.csr_matrix) -> np.ndarray:
        n_samples = X.shape[0]
        df = np.bincount(X.indices, minlength=X.shape[1])
        df = np.maximum(df, 1)

        if self.idf_scheme == "smooth":
            return np.log((n_samples + 1) / (df + 1)) + 1
        elif self.idf_scheme == "standard":
            return np.log(n_samples / df) + 1
        elif self.idf_scheme == "probabilistic":
            return np.log((n_samples - df) / df)

    def fit(self, raw_documents: List[str], y=None):
        X = self.count_vectorizer.fit_transform(raw_documents)
        self.idf_ = self._compute_idf(X)
        return self

    def transform(self, raw_documents: List[str]) -> sp.csr_matrix:
        X = self.count_vectorizer.transform(raw_documents)
        X = self._compute_tf(X)
        X = X.multiply(self.idf_)

        if self.norm:
            X = normalize(X, norm=self.norm, copy=False)

        return X

    def get_feature_names_out(self):
        return self.count_vectorizer.get_feature_names_out()


def get_vectorizer(
    vectorizer_type: str, max_features: int, min_df: float, max_df: float, **kwargs
) -> Union[TfidfVectorizer, BM25Vectorizer, WeightedTfidfVectorizer]:
    """Create and configure the specified vectorizer type"""

    if vectorizer_type == "tfidf":
        return TfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )
    elif vectorizer_type == "bm25":
        return BM25Vectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            k1=kwargs.get("k1", 1.5),
            b=kwargs.get("b", 0.75),
        )
    elif vectorizer_type == "weighted":
        return WeightedTfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            tf_scheme=kwargs.get("tf_scheme", "raw"),
            idf_scheme=kwargs.get("idf_scheme", "smooth"),
        )
    else:
        raise ValueError(f"Unknown vectorizer type: {vectorizer_type}")


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s: %(message)s",
    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global headers for all requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Referer": "https://judiciary.uk/",
}

# Core utility functions
def make_request(
    url: str, retries: int = 3, delay: int = 2
) -> Optional[requests.Response]:
    """Make HTTP request with retries and delay"""
    for attempt in range(retries):
        try:
            time.sleep(delay)
            response = requests.get(url, headers=HEADERS, verify=False, timeout=30)
            response.raise_for_status()
            return response
        except Exception as e:
            if attempt == retries - 1:
                st.error(f"Request failed: {str(e)}")
                raise e
            time.sleep(delay * (attempt + 1))
    return None


def combine_document_text(row: pd.Series) -> str:
    """Combine all text content from a document"""
    text_parts = []

    # Add title and content
    if pd.notna(row.get("Title")):
        text_parts.append(str(row["Title"]))
    if pd.notna(row.get("Content")):
        text_parts.append(str(row["Content"]))

    # Add PDF contents
    pdf_columns = [
        col for col in row.index if col.startswith("PDF_") and col.endswith("_Content")
    ]
    for pdf_col in pdf_columns:
        if pd.notna(row.get(pdf_col)):
            text_parts.append(str(row[pdf_col]))

    return " ".join(text_parts)


def clean_text_for_modeling(text: str) -> str:
    """Clean text with enhanced noise removal"""
    if not isinstance(text, str):
        return ""

    try:
        # Convert to lowercase
        text = text.lower()

        # Remove URLs
        text = re.sub(r"http\S+|www\S+|https\S+", "", text)

        # Remove email addresses and phone numbers
        text = re.sub(r"\S+@\S+", "", text)
        text = re.sub(r"\b\d{3}[-.]?\d{3}[-.]?\d{4}\b", "", text)

        # Remove dates and times
        text = re.sub(
            r"\b\d{1,2}(?:st|nd|rd|th)?\s+(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(r"\b\d{1,2}:\d{2}\b", "", text)
        text = re.sub(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", "", text)

        # Remove specific document-related terms
        text = re.sub(
            r"\b(?:ref|reference|case)(?:\s+no)?\.?\s*[-:\s]?\s*\w+[-\d]+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(
            r"\b(regulation|paragraph|section|subsection|article)\s+\d+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )

        # Remove common legal document terms
        legal_terms = r"\b(coroner|inquest|hearing|evidence|witness|statement|report|dated|signed)\b"
        text = re.sub(legal_terms, "", text, flags=re.IGNORECASE)

        # Remove special characters and multiple spaces
        text = re.sub(r"[^a-z\s]", " ", text)
        text = re.sub(r"\s+", " ", text)

        # Remove very short words
        text = " ".join(word for word in text.split() if len(word) > 2)

        # Ensure minimum content length
        cleaned_text = text.strip()
        return cleaned_text if len(cleaned_text.split()) >= 3 else ""

    except Exception as e:
        logging.error(f"Error in text cleaning: {e}")
        return ""


def clean_text(text: str) -> str:
    """Clean text while preserving structure and metadata formatting"""
    if not text:
        return ""

    try:
        text = str(text)
        text = unicodedata.normalize("NFKD", text)

        replacements = {
            "â€™": "'",
            "â€œ": '"',
            "â€": '"',
            "â€¦": "...",
            'â€"': "-",
            "â€¢": "•",
            "Â": " ",
            "\u200b": "",
            "\uf0b7": "",
            "\u2019": "'",
            "\u201c": '"',
            "\u201d": '"',
            "\u2013": "-",
            "\u2022": "•",
        }

        for encoded, replacement in replacements.items():
            text = text.replace(encoded, replacement)

        text = re.sub(r"<[^>]+>", "", text)
        text = "".join(
            char if char.isprintable() or char == "\n" else " " for char in text
        )
        text = re.sub(r" +", " ", text)
        text = re.sub(r"\n+", "\n", text)

        return text.strip()

    except Exception as e:
        logging.error(f"Error in clean_text: {e}")
        return ""

def extract_concern_text(content):
    """Extract complete concern text from PFD report content with robust section handling"""
    if pd.isna(content) or not isinstance(content, str):
        return ""

    # Keep ALL original identifiers (critical for catching variations in reports)
    concern_identifiers = [
        "CORONER'S CONCERNS",
        "MATTERS OF CONCERN",
        "The MATTERS OF CONCERN",
        "CORONER'S CONCERNS are",  
        "MATTERS OF CONCERN are",
        "The MATTERS OF CONCERN are",
        "HEALTHCARE SAFETY CONCERNS",
        "SAFETY CONCERNS",
        "PATIENT SAFETY ISSUES",
        "HSIB FINDINGS",
        "INVESTIGATION FINDINGS",
        "THE CORONER'S MATTER OF CONCERN",
        "CONCERNS AND RECOMMENDATIONS",
        "CONCERNS IDENTIFIED",
        "TheMATTERS OF CONCERN"
    ]

    # Normalize content (remove excessive whitespace but preserve structure)
    content = ' '.join(content.split())  # Collapse multiple spaces
    content_lower = content.lower()

    # Find the start of concerns section (case-insensitive)
    start_idx = -1
    for identifier in concern_identifiers:
        identifier_lower = identifier.lower()
        pos = content_lower.find(identifier_lower)
        if pos != -1:
            # Start after the identifier (handles colons, "are", etc.)
            start_idx = pos + len(identifier)
            # Skip past a colon if present
            if content[start_idx:start_idx+1] == ":":
                start_idx += 1
            break

    if start_idx == -1:
        return ""  # No concerns section found

    # Look for the end of the concerns section (using major section headers)
    end_markers = [
        "ACTION SHOULD BE TAKEN", 
        "CONCLUSIONS", 
        "YOUR RESPONSE",
        "COPIES",
        "SIGNED:",
        "DATED THIS",
        "NEXT STEPS",
        "YOU ARE UNDER A DUTY",
    ]

    # Find the earliest end marker
    end_idx = len(content)
    for marker in end_markers:
        marker_pos = content_lower.find(marker.lower(), start_idx)
        if marker_pos != -1 and marker_pos < end_idx:
            end_idx = marker_pos

    # Extract the full concerns text
    concerns_text = content[start_idx:end_idx].strip()

    # Post-processing: Ensure we don't cut off mid-sentence
    last_period = concerns_text.rfind('.')
    if last_period != -1:
        concerns_text = concerns_text[:last_period + 1]

    return concerns_text


def extract_concern_text2(content):
    """Extract concern text from PFD report content with improved handling for complete text extraction"""
    if pd.isna(content) or not isinstance(content, str):
        return ""
    
    # Keywords to identify sections with concerns - expanded with more variations
    concern_identifiers = [
        "CORONER'S CONCERNS",
        "MATTERS OF CONCERN",
        "The MATTERS OF CONCERN",
        "CORONER'S CONCERNS are",  
        "MATTERS OF CONCERN are",
        "The MATTERS OF CONCERN are",
        "HEALTHCARE SAFETY CONCERNS",
        "SAFETY CONCERNS",
        "PATIENT SAFETY ISSUES",
        "HSIB FINDINGS",
        "INVESTIGATION FINDINGS",
        "THE CORONER'S MATTER OF CONCERN",
        "CONCERNS AND RECOMMENDATIONS",
        "CONCERNS IDENTIFIED",
        "Concerns of the Coroner:"
    ]
    
    # Make case-insensitive search by converting content to lowercase once
    content_lower = content.lower()
    
    for identifier in concern_identifiers:
        identifier_lower = identifier.lower()
        if identifier_lower in content_lower:
            # Find position - use the original content after position is found
            start_pos = content_lower.find(identifier_lower)
            
            # Find the end of the identifier to determine where content starts
            # Look for colon or text like "are" that might occur at the end
            if ":" in content[start_pos:start_pos + len(identifier) + 5]:
                # If there's a colon right after the identifier, start after it
                colon_pos = content.find(":", start_pos)
                start_idx = colon_pos + 1
            elif " are" in content_lower[start_pos:start_pos + len(identifier) + 5]:
                # If "are" appears right after, find the end
                are_pos = content_lower.find(" are", start_pos)
                start_idx = are_pos + 5  # " are " length is 4 + 1 for safety
            else:
                # Otherwise just use end of identifier
                start_idx = start_pos + len(identifier)
            
            # Find end markers - expanded list with more variants
            end_markers = [
                "ACTION SHOULD BE TAKEN", 
                "RECOMMENDATION", 
                "CONCLUSIONS", 
                "NEXT STEPS",
                "YOUR RESPONSE",
                "YOU ARE UNDER A DUTY",
                "RESPONSE",
                "COPIES"
            ]
            
            end_idx = float('inf')
            
            for marker in end_markers:
                marker_lower = marker.lower()
                temp_pos = content_lower.find(marker_lower, start_idx)
                if temp_pos != -1 and temp_pos < end_idx:
                    end_idx = temp_pos
            
            # If identified a proper end marker
            if end_idx != float('inf'):
                # Get the full text between start and end markers
                extracted_text = content[start_idx:end_idx].strip()
                
                # Add extra validation to ensure we got complete paragraphs
                if extracted_text:
                    # Double check if the extracted text is substantial (at least 20 chars)
                    if len(extracted_text) > 20:
                        return extracted_text
                    else:
                        # If too short, try to get more text in case markers weren't correctly identified
                        # Get a larger chunk and try to find paragraph boundaries
                        enlarged_extract = content[start_idx:start_idx + 2000].strip()
                        return enlarged_extract
                
            else:
                # No end marker found, try paragraph-based extraction strategies
                
                # Strategy 1: Look for numbered sections that follow the concerns
                possible_extract = content[start_idx:start_idx + 5000].strip()
                
                # Check for numbered sections (a., b., 1., 2., etc.)
                section_matches = list(re.finditer(r'\n\s*(?:[a-z]|[0-9]+)\.\s+', possible_extract))
                if section_matches:
                    # Found numbered sections - collect all content until we hit a major heading
                    major_heading_pattern = r'\n\n+\s*[A-Z]{2,}'
                    major_heading_match = re.search(major_heading_pattern, possible_extract)
                    
                    if major_heading_match:
                        # If we found a major heading after the sections, use that as endpoint
                        return possible_extract[:major_heading_match.start()].strip()
                    else:
                        # Otherwise just return what we have
                        return possible_extract
                
                # Strategy 2: Look for a chunk of text that might be a complete statement
                # Try to find a natural ending after a reasonable chunk
                chunk = content[start_idx:start_idx + 1500].strip()
                
                # Look for multiple consecutive line breaks which often indicate section changes
                section_breaks = list(re.finditer(r'\n\s*\n\s*\n', chunk))
                if section_breaks:
                    # Return up to the first major section break
                    return chunk[:section_breaks[0].start()].strip()
                
                # If all else fails, just return a reasonable chunk
                return chunk
    
    # If no identifier found but content contains "concern" keyword, try a more liberal approach
    if "concern" in content_lower:
        concern_pos = content_lower.find("concern")
        # Extract a section around the keyword
        start_pos = max(0, concern_pos - 100)
        # Find sentence start
        while start_pos > 0 and content[start_pos] not in ".!?\n":
            start_pos -= 1
        if start_pos > 0:
            start_pos += 1  # Move past the period
            
        # Find a reasonable end point
        end_pos = min(len(content), concern_pos + 1000)
        return content[start_pos:end_pos].strip()
            
    return ""  # No identifier found
    
def extract_metadata(content: str) -> dict:
    """
    Extract structured metadata from report content with improved category handling.

    Args:
        content (str): Raw report content

    Returns:
        dict: Extracted metadata including date, reference, names, categories, etc.
    """
    metadata = {
        "date_of_report": None,
        "ref": None,
        "deceased_name": None,
        "coroner_name": None,
        "coroner_area": None,
        "categories": [],
    }

    if not content:
        return metadata

    try:
        # Extract date patterns
        date_patterns = [
            r"Date of report:?\s*(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+\d{4})",
            r"Date of report:?\s*(\d{1,2}/\d{1,2}/\d{4})",
            r"DATED this (\d{1,2}(?:st|nd|rd|th)?\s+day of [A-Za-z]+\s+\d{4})",
            r"Date:?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        ]

        for pattern in date_patterns:
            date_match = re.search(pattern, content, re.IGNORECASE)
            if date_match:
                date_str = date_match.group(1)
                try:
                    if "/" in date_str:
                        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                    else:
                        date_str = re.sub(r"(?<=\d)(st|nd|rd|th)", "", date_str)
                        date_str = re.sub(r"day of ", "", date_str)
                        try:
                            date_obj = datetime.strptime(date_str, "%d %B %Y")
                        except ValueError:
                            date_obj = datetime.strptime(date_str, "%d %b %Y")

                    metadata["date_of_report"] = date_obj.strftime("%d/%m/%Y")
                    break
                except ValueError as e:
                    logging.warning(f"Invalid date format found: {date_str} - {e}")

        # Extract reference number
        ref_match = re.search(r"Ref(?:erence)?:?\s*([-\d]+)", content)
        if ref_match:
            metadata["ref"] = ref_match.group(1).strip()

        # Extract deceased name
        name_match = re.search(r"Deceased name:?\s*([^\n]+)", content)
        if name_match:
            metadata["deceased_name"] = clean_text(name_match.group(1)).strip()

        # Extract coroner details
        coroner_match = re.search(r"Coroner(?:\'?s)? name:?\s*([^\n]+)", content)
        if coroner_match:
            metadata["coroner_name"] = clean_text(coroner_match.group(1)).strip()

        area_match = re.search(r"Coroner(?:\'?s)? Area:?\s*([^\n]+)", content)
        if area_match:
            metadata["coroner_area"] = clean_text(area_match.group(1)).strip()

        # Extract categories with enhanced handling
        cat_match = re.search(
            r"Category:?\s*(.+?)(?=This report is being sent to:|$)",
            content,
            re.IGNORECASE | re.DOTALL,
        )
        if cat_match:
            category_text = cat_match.group(1).strip()

            # Normalize all possible separators to pipe
            category_text = re.sub(r"\s*[,;]\s*", "|", category_text)
            category_text = re.sub(r"[•·⋅‣⁃▪▫–—-]\s*", "|", category_text)
            category_text = re.sub(r"\s{2,}", "|", category_text)
            category_text = re.sub(r"\n+", "|", category_text)

            # Split into individual categories
            categories = category_text.split("|")
            cleaned_categories = []

            # Get standard categories for matching
            standard_categories = {cat.lower(): cat for cat in get_pfd_categories()}

            for cat in categories:
                # Clean and normalize the category
                cleaned_cat = clean_text(cat).strip()
                cleaned_cat = re.sub(r"&nbsp;", "", cleaned_cat)
                cleaned_cat = re.sub(
                    r"\s*This report.*$", "", cleaned_cat, flags=re.IGNORECASE
                )
                cleaned_cat = re.sub(r"[|,;]", "", cleaned_cat)

                # Only process non-empty categories
                if cleaned_cat and not re.match(r"^[\s|,;]+$", cleaned_cat):
                    # Try to match with standard categories
                    cat_lower = cleaned_cat.lower()

                    # Check for exact match first
                    if cat_lower in standard_categories:
                        cleaned_cat = standard_categories[cat_lower]
                    else:
                        # Try partial matching
                        for std_lower, std_original in standard_categories.items():
                            if cat_lower in std_lower or std_lower in cat_lower:
                                cleaned_cat = std_original
                                break

                    cleaned_categories.append(cleaned_cat)

            # Remove duplicates while preserving order
            seen = set()
            metadata["categories"] = [
                x
                for x in cleaned_categories
                if not (x.lower() in seen or seen.add(x.lower()))
            ]

        return metadata

    except Exception as e:
        logging.error(f"Error extracting metadata: {e}")
        return metadata


def get_pfd_categories() -> List[str]:
    """Get all available PFD report categories"""
    return [
        "Accident at Work and Health and Safety related deaths",
        "Alcohol drug and medication related deaths",
        "Care Home Health related deaths",
        "Child Death from 2015",
        "Community health care and emergency services related deaths",
        "Emergency services related deaths 2019 onwards",
        "Hospital Death Clinical Procedures and medical management related deaths",
        "Mental Health related deaths",
        "Other related deaths",
        "Police related deaths",
        "Product related deaths",
        "Railway related deaths",
        "Road Highways Safety related deaths",
        "Service Personnel related deaths",
        "State Custody related deaths",
        "Suicide from 2015",
        "Wales prevention of future deaths reports 2019 onwards",
    ]


# PDF handling functions
def save_pdf(
    pdf_url: str, base_dir: str = "pdfs"
) -> Tuple[Optional[str], Optional[str]]:
    """Download and save PDF, return local path and filename"""
    try:
        os.makedirs(base_dir, exist_ok=True)

        response = make_request(pdf_url)
        if not response:
            return None, None

        filename = os.path.basename(pdf_url)
        filename = re.sub(r"[^\w\-_\. ]", "_", filename)
        local_path = os.path.join(base_dir, filename)

        with open(local_path, "wb") as f:
            f.write(response.content)

        return local_path, filename

    except Exception as e:
        logging.error(f"Error saving PDF {pdf_url}: {e}")
        return None, None


def extract_pdf_content(pdf_path: str, chunk_size: int = 10) -> str:
    """Extract text from PDF file with memory management"""
    try:
        filename = os.path.basename(pdf_path)
        text_chunks = []

        with pdfplumber.open(pdf_path) as pdf:
            for i in range(0, len(pdf.pages), chunk_size):
                chunk = pdf.pages[i : i + chunk_size]
                chunk_text = "\n\n".join([page.extract_text() or "" for page in chunk])
                text_chunks.append(chunk_text)

        full_content = f"PDF FILENAME: {filename}\n\n{''.join(text_chunks)}"
        return clean_text(full_content)

    except Exception as e:
        logging.error(f"Error extracting PDF text from {pdf_path}: {e}")
        return ""


def get_report_content(url: str) -> Optional[Dict]:
    """Get full content from report page with improved PDF and response handling"""
    try:
        logging.info(f"Fetching content from: {url}")
        response = make_request(url)
        if not response:
            return None

        soup = BeautifulSoup(response.text, "html.parser")
        content = soup.find("div", class_="flow") or soup.find(
            "article", class_="single__post"
        )

        if not content:
            logging.warning(f"No content found at {url}")
            return None

        # Extract main report content
        paragraphs = content.find_all(["p", "table"])
        webpage_text = "\n\n".join(
            p.get_text(strip=True, separator=" ") for p in paragraphs
        )

        pdf_contents = []
        pdf_paths = []
        pdf_names = []
        pdf_types = []  # Track if PDF is main report or response

        # Find all PDF links with improved classification
        pdf_links = soup.find_all("a", href=re.compile(r"\.pdf$"))

        for pdf_link in pdf_links:
            pdf_url = pdf_link["href"]
            pdf_text = pdf_link.get_text(strip=True).lower()

            # Determine PDF type
            is_response = any(
                word in pdf_text.lower() for word in ["response", "reply"]
            )
            pdf_type = "response" if is_response else "report"

            if not pdf_url.startswith(("http://", "https://")):
                pdf_url = (
                    f"https://www.judiciary.uk{pdf_url}"
                    if not pdf_url.startswith("/")
                    else f"https://www.judiciary.uk/{pdf_url}"
                )

            pdf_path, pdf_name = save_pdf(pdf_url)

            if pdf_path:
                pdf_content = extract_pdf_content(pdf_path)
                pdf_contents.append(pdf_content)
                pdf_paths.append(pdf_path)
                pdf_names.append(pdf_name)
                pdf_types.append(pdf_type)

        return {
            "content": clean_text(webpage_text),
            "pdf_contents": pdf_contents,
            "pdf_paths": pdf_paths,
            "pdf_names": pdf_names,
            "pdf_types": pdf_types,
        }

    except Exception as e:
        logging.error(f"Error getting report content: {e}")
        return None


def scrape_page(url: str) -> List[Dict]:
    """Scrape a single page with improved PDF handling"""
    reports = []
    try:
        response = make_request(url)
        if not response:
            return []

        soup = BeautifulSoup(response.text, "html.parser")
        results_list = soup.find("ul", class_="search__list")

        if not results_list:
            logging.warning(f"No results list found on page: {url}")
            return []

        cards = results_list.find_all("div", class_="card")

        for card in cards:
            try:
                title_elem = card.find("h3", class_="card__title")
                if not title_elem:
                    continue

                title_link = title_elem.find("a")
                if not title_link:
                    continue

                title = clean_text(title_link.text)
                card_url = title_link["href"]

                if not card_url.startswith(("http://", "https://")):
                    card_url = f"https://www.judiciary.uk{card_url}"

                logging.info(f"Processing report: {title}")
                content_data = get_report_content(card_url)

                if content_data:
                    report = {
                        "Title": title,
                        "URL": card_url,
                        "Content": content_data["content"],
                    }

                    # Add PDF details with type classification
                    for i, (name, content, path, pdf_type) in enumerate(
                        zip(
                            content_data["pdf_names"],
                            content_data["pdf_contents"],
                            content_data["pdf_paths"],
                            content_data["pdf_types"],
                        ),
                        1,
                    ):
                        report[f"PDF_{i}_Name"] = name
                        report[f"PDF_{i}_Content"] = content
                        report[f"PDF_{i}_Path"] = path
                        report[f"PDF_{i}_Type"] = pdf_type

                    reports.append(report)
                    logging.info(f"Successfully processed: {title}")

            except Exception as e:
                logging.error(f"Error processing card: {e}")
                continue

        return reports

    except Exception as e:
        logging.error(f"Error fetching page {url}: {e}")
        return []


def get_total_pages(url: str) -> Tuple[int, int]:
    """
    Get total number of pages and total results count

    Returns:
        Tuple[int, int]: (total_pages, total_results)
    """
    try:
        response = make_request(url)
        if not response:
            logging.error(f"No response from URL: {url}")
            return 0, 0

        soup = BeautifulSoup(response.text, "html.parser")

        # First check for total results count
        total_results = 0
        results_header = soup.find("div", class_="search__header")
        if results_header:
            results_text = results_header.get_text()
            match = re.search(r"found (\d+) results?", results_text, re.IGNORECASE)
            if match:
                total_results = int(match.group(1))
                total_pages = (total_results + 9) // 10  # 10 results per page
                return total_pages, total_results

        # If no results header, check pagination
        pagination = soup.find("nav", class_="navigation pagination")
        if pagination:
            page_numbers = pagination.find_all("a", class_="page-numbers")
            numbers = [
                int(p.text.strip()) for p in page_numbers if p.text.strip().isdigit()
            ]
            if numbers:
                return max(numbers), len(numbers) * 10  # Approximate result count

        # If no pagination but results exist
        results = soup.find("ul", class_="search__list")
        if results and results.find_all("div", class_="card"):
            cards = results.find_all("div", class_="card")
            return 1, len(cards)

        return 0, 0

    except Exception as e:
        logging.error(f"Error in get_total_pages: {str(e)}")
        return 0, 0


def process_scraped_data(df: pd.DataFrame) -> pd.DataFrame:
    """Process and clean scraped data with metadata extraction and concern extraction"""
    try:
        if df is None or len(df) == 0:
            return pd.DataFrame()

        # Create a copy
        df = df.copy()

        # Extract metadata from Content field if it exists
        if "Content" in df.columns:
            # Process each row
            processed_rows = []
            for _, row in df.iterrows():
                # Start with original row data
                processed_row = row.to_dict()

                # Extract metadata using existing function
                content = str(row.get("Content", ""))
                metadata = extract_metadata(content)

                # Extract concerns text
                processed_row["Extracted_Concerns"] = extract_concern_text(content)

                # Update row with metadata
                processed_row.update(metadata)
                processed_rows.append(processed_row)

            # Create new DataFrame from processed rows
            result = pd.DataFrame(processed_rows)
        else:
            result = df.copy()

        # Convert date_of_report to datetime with UK format handling
        if "date_of_report" in result.columns:

            def parse_date(date_str):
                if pd.isna(date_str):
                    return pd.NaT

                date_str = str(date_str).strip()

                # If already in DD/MM/YYYY format
                if re.match(r"\d{1,2}/\d{1,2}/\d{4}", date_str):
                    return pd.to_datetime(date_str, format="%d/%m/%Y")

                # Remove ordinal indicators
                date_str = re.sub(r"(\d)(st|nd|rd|th)", r"\1", date_str)

                # Try different formats
                formats = ["%Y-%m-%d", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"]

                for fmt in formats:
                    try:
                        return pd.to_datetime(date_str, format=fmt)
                    except ValueError:
                        continue

                try:
                    return pd.to_datetime(date_str)
                except:
                    return pd.NaT

            result["date_of_report"] = result["date_of_report"].apply(parse_date)

        return result

    except Exception as e:
        logging.error(f"Error in process_scraped_data: {e}")
        return df

def get_category_slug(category: str) -> str:
    """Generate proper category slug for the website's URL structure"""
    if not category:
        return None

    # Create a slug exactly matching the website's format
    slug = (
        category.lower()
        .replace(" ", "-")
        .replace("&", "and")
        .replace("--", "-")
        .strip("-")
    )

    logging.info(f"Generated category slug: {slug} from category: {category}")
    return slug

def scrape_pfd_reports(
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    order: str = "relevance",
    start_page: int = 1,
    end_page: Optional[int] = None,
    auto_save_batches: bool = True,
    batch_size: int = 5
) -> List[Dict]:
    """
    Scrape PFD reports with enhanced progress tracking, proper pagination, and automatic batch saving
    
    Args:
        keyword: Optional keyword to search for
        category: Optional category to filter by
        order: Sort order ("relevance", "desc", "asc")
        start_page: First page to scrape
        end_page: Last page to scrape (None for all pages)
        auto_save_batches: Whether to automatically save batches of results
        batch_size: Number of pages per batch
        
    Returns:
        List of report dictionaries
    """
    all_reports = []
    base_url = "https://www.judiciary.uk/"
    batch_number = 1

    try:
        # Initialize progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        report_count_text = st.empty()
        batch_status = st.empty()

        # Validate and prepare category
        category_slug = None
        if category:
            category_slug = (
                category.lower()
                .replace(" ", "-")
                .replace("&", "and")
                .replace("--", "-")
                .strip("-")
            )
            logging.info(f"Using category: {category}, slug: {category_slug}")

        # Construct initial search URL
        base_search_url = construct_search_url(
            base_url=base_url,
            keyword=keyword,
            category=category,
            category_slug=category_slug,
        )

        st.info(f"Searching at: {base_search_url}")

        # Get total pages and results count
        total_pages, total_results = get_total_pages(base_search_url)

        if total_results == 0:
            st.warning("No results found matching your search criteria")
            return []

        st.info(f"Found {total_results} matching reports across {total_pages} pages")

        # Apply page range limits
        start_page = max(1, start_page)  # Ensure start_page is at least 1
        if end_page is None:
            end_page = total_pages
        else:
            end_page = min(
                end_page, total_pages
            )  # Ensure end_page doesn't exceed total_pages

        if start_page > end_page:
            st.warning(f"Invalid page range: {start_page} to {end_page}")
            return []

        st.info(f"Scraping pages {start_page} to {end_page}")
        
        # Variables for batch processing
        batch_reports = []
        current_batch_start = start_page
        batch_end = min(start_page + batch_size - 1, end_page)

        # Process each page in the specified range
        for current_page in range(start_page, end_page + 1):
            try:
                # Check if scraping should be stopped
                if (
                    hasattr(st.session_state, "stop_scraping")
                    and st.session_state.stop_scraping
                ):
                    # Save the current batch before stopping
                    if auto_save_batches and batch_reports:
                        save_batch(
                            batch_reports, 
                            batch_number, 
                            keyword, 
                            category, 
                            current_batch_start, 
                            current_page - 1
                        )
                    st.warning("Scraping stopped by user")
                    break

                # Update progress
                progress = (current_page - start_page) / (end_page - start_page + 1)
                progress_bar.progress(progress)
                status_text.text(
                    f"Processing page {current_page} of {end_page} (out of {total_pages} total pages)"
                )

                # Construct current page URL
                page_url = construct_search_url(
                    base_url=base_url,
                    keyword=keyword,
                    category=category,
                    category_slug=category_slug,
                    page=current_page,
                )

                # Scrape current page
                page_reports = scrape_page(page_url)

                if page_reports:
                    # Deduplicate based on title and URL
                    existing_reports = {(r["Title"], r["URL"]) for r in all_reports}
                    existing_batch_reports = {(r["Title"], r["URL"]) for r in batch_reports}
                    
                    new_reports = [
                        r
                        for r in page_reports
                        if (r["Title"], r["URL"]) not in existing_reports 
                        and (r["Title"], r["URL"]) not in existing_batch_reports
                    ]

                    # Add to both all_reports and batch_reports
                    all_reports.extend(new_reports)
                    batch_reports.extend(new_reports)
                    
                    report_count_text.text(
                        f"Retrieved {len(all_reports)} unique reports so far..."
                    )

                # Check if we've reached the end of a batch
                if auto_save_batches and (current_page == batch_end or current_page == end_page):
                    if batch_reports:
                        # Automatically save the batch
                        saved_file = save_batch(
                            batch_reports, 
                            batch_number, 
                            keyword, 
                            category, 
                            current_batch_start, 
                            current_page
                        )
                        batch_status.success(
                            f"Saved batch #{batch_number} (pages {current_batch_start}-{current_page}) to {saved_file}"
                        )
                        
                        # Reset for next batch
                        batch_reports = []
                        batch_number += 1
                        current_batch_start = current_page + 1
                        batch_end = min(current_batch_start + batch_size - 1, end_page)
                
                # Add delay between pages
                time.sleep(2)

            except Exception as e:
                logging.error(f"Error processing page {current_page}: {e}")
                st.warning(
                    f"Error on page {current_page}. Continuing with next page..."
                )
                continue

        # Sort results if specified
        if order != "relevance":
            all_reports = sort_reports(all_reports, order)

        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
        report_count_text.empty()

        if all_reports:
            st.success(f"Successfully scraped {len(all_reports)} unique reports")
            
            # Final report
            if auto_save_batches:
                st.info(f"Reports were automatically saved in {batch_number} batches")
        else:
            st.warning("No reports were successfully retrieved")

        return all_reports

    except Exception as e:
        logging.error(f"Error in scrape_pfd_reports: {e}")
        st.error(f"An error occurred while scraping reports: {e}")
        # Save any unsaved reports if an error occurs
        if auto_save_batches and batch_reports:
            save_batch(
                batch_reports, 
                batch_number, 
                keyword, 
                category, 
                current_batch_start, 
                "error"
            )
        return []


def save_batch(
    reports: List[Dict], 
    batch_number: int, 
    keyword: Optional[str], 
    category: Optional[str], 
    start_page: int, 
    end_page: Union[int, str]
) -> str:
    """
    Save a batch of reports to Excel file with appropriate naming
    
    Args:
        reports: List of report dictionaries to save
        batch_number: Current batch number
        keyword: Search keyword used (for filename)
        category: Category used (for filename)
        start_page: Starting page of this batch
        end_page: Ending page of this batch (or "error" if saving due to error)
        
    Returns:
        Filename of the saved file
    """
    if not reports:
        return ""
    
    # Process the data
    df = pd.DataFrame(reports)
    df = process_scraped_data(df)
    
    # Create timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create descriptive filename parts
    keyword_part = f"kw_{keyword.replace(' ', '_')}" if keyword else "no_keyword"
    category_part = f"cat_{category.replace(' ', '_')}" if category else "no_category"
    page_part = f"pages_{start_page}_to_{end_page}"
    
    # Generate filename
    filename = f"pfd_reports_scraped_batch{batch_number}_{keyword_part}_{category_part}_{page_part}_{timestamp}.xlsx"
    
    # Ensure filename is valid (remove any problematic characters)
    filename = re.sub(r'[\\/*?:"<>|]', "", filename)
    
    # Create directory if it doesn't exist
    os.makedirs("scraped_reports", exist_ok=True)
    file_path = os.path.join("scraped_reports", filename)
    
    # Save to Excel
    df.to_excel(file_path, index=False, engine="openpyxl")
    
    logging.info(f"Saved batch {batch_number} to {file_path}")
    return filename


def render_scraping_tab():
    """Render the scraping tab with batch saving options"""
    st.subheader("Scrape PFD Reports")

    # Initialize default values if not in session state
    if "init_done" not in st.session_state:
        st.session_state.init_done = True
        st.session_state["search_keyword_default"] = ""  # Changed from "report" to empty string
        st.session_state["category_default"] = ""
        st.session_state["order_default"] = "relevance"
        st.session_state["start_page_default"] = 1
        st.session_state["end_page_default"] = None
        st.session_state["auto_save_batches_default"] = True
        st.session_state["batch_size_default"] = 5

    if "scraped_data" in st.session_state and st.session_state.scraped_data is not None:
        st.success(f"Found {len(st.session_state.scraped_data)} reports")

        st.subheader("Results")
        st.dataframe(
            st.session_state.scraped_data,
            column_config={
                "URL": st.column_config.LinkColumn("Report Link"),
                "date_of_report": st.column_config.DateColumn(
                    "Date of Report", format="DD/MM/YYYY"
                ),
                "categories": st.column_config.ListColumn("Categories"),
            },
            hide_index=True,
        )

        show_export_options(st.session_state.scraped_data, "scraped")

    # Create the search form with page range selection and batch options
    with st.form("scraping_form"):
        # Create two rows with two columns each
        row1_col1, row1_col2 = st.columns(2)
        row2_col1, row2_col2 = st.columns(2)
        row3_col1, row3_col2 = st.columns(2)
        row4_col1, row4_col2 = st.columns(2)

        # First row
        with row1_col1:
            search_keyword = st.text_input(
                "Search keywords:",
                value=st.session_state.get("search_keyword_default", ""),  # Changed default to empty
                key="search_keyword",
                help="Do not leave empty, use 'report' or another search term",
            )

        with row1_col2:
            category = st.selectbox(
                "PFD Report type:",
                [""] + get_pfd_categories(),
                index=0,
                key="category",
                format_func=lambda x: x if x else "Select a category",
            )

        # Second row
        with row2_col1:
            order = st.selectbox(
                "Sort by:",
                ["relevance", "desc", "asc"],
                index=0,
                key="order",
                format_func=lambda x: {
                    "relevance": "Relevance",
                    "desc": "Newest first",
                    "asc": "Oldest first",
                }[x],
            )

        with row2_col2:
            # Get total pages for the query (preview)
            if search_keyword or category:
                base_url = "https://www.judiciary.uk/"

                # Prepare category slug
                category_slug = None
                if category:
                    category_slug = (
                        category.lower()
                        .replace(" ", "-")
                        .replace("&", "and")
                        .replace("--", "-")
                        .strip("-")
                    )

                # Create preview URL
                preview_url = construct_search_url(
                    base_url=base_url,
                    keyword=search_keyword,
                    category=category,
                    category_slug=category_slug,
                )

                try:
                    with st.spinner("Checking total pages..."):
                        total_pages, total_results = get_total_pages(preview_url)
                        if total_pages > 0:
                            st.info(
                                f"This search has {total_pages} pages with {total_results} results"
                            )
                            st.session_state["total_pages_preview"] = total_pages
                        else:
                            st.warning("No results found for this search")
                            st.session_state["total_pages_preview"] = 0
                except Exception as e:
                    st.error(f"Error checking pages: {str(e)}")
                    st.session_state["total_pages_preview"] = 0
            else:
                st.session_state["total_pages_preview"] = 0

        # Third row for page range
        with row3_col1:
            start_page = st.number_input(
                "Start page:",
                min_value=1,
                value=st.session_state.get("start_page_default", 1),
                key="start_page",
                help="First page to scrape (minimum 1)",
            )

        with row3_col2:
            end_page = st.number_input(
                "End page (Optimal: 10 pages per extraction:)",
                min_value=0,
                value=st.session_state.get("end_page_default", 0),
                key="end_page",
                help="Last page to scrape (0 for all pages)",
            )

        # Fourth row for batch options
        with row4_col1:
            auto_save_batches = st.checkbox(
                "Auto-save batches",
                value=st.session_state.get("auto_save_batches_default", True),
                key="auto_save_batches",
                help="Automatically save results in batches as they are scraped",
            )

        with row4_col2:
            batch_size = st.number_input(
                "Pages per batch: (ideally set to 5)",
                min_value=1,
                max_value=10,
                value=st.session_state.get("batch_size_default", 5),
                key="batch_size",
                help="Number of pages to process before saving a batch",
            )

        # Action buttons in a row
        button_col1, button_col2 = st.columns(2)
        with button_col1:
            submitted = st.form_submit_button("Search Reports")
        with button_col2:
            stop_scraping = st.form_submit_button("Stop Scraping")

    # Handle stop scraping
    if stop_scraping:
        st.session_state.stop_scraping = True
        st.warning("Scraping will be stopped after the current page completes...")
        return

    if submitted:
        try:
            # Store search parameters in session state
            st.session_state.last_search_params = {
                "keyword": search_keyword,
                "category": category,
                "order": order,
                "start_page": start_page,
                "end_page": end_page,
                "auto_save_batches": auto_save_batches,
                "batch_size": batch_size,
            }

            # Initialize stop_scraping flag
            st.session_state.stop_scraping = False

            # Convert end_page=0 to None (all pages)
            end_page_val = None if end_page == 0 else end_page

            # Perform scraping with batch options
            reports = scrape_pfd_reports(
                keyword=search_keyword,
                category=category if category else None,
                order=order,
                start_page=start_page,
                end_page=end_page_val,
                auto_save_batches=auto_save_batches,
                batch_size=batch_size,
            )

            if reports:
                # Process the data
                df = pd.DataFrame(reports)
                df = process_scraped_data(df)

                # Store in session state
                st.session_state.scraped_data = df
                st.session_state.data_source = "scraped"
                st.session_state.current_data = df

                # Trigger a rerun to refresh the page
                st.rerun()
            else:
                st.warning("No reports found matching your search criteria")

        except Exception as e:
            st.error(f"An error occurred: {e}")
            logging.error(f"Scraping error: {e}")
            return False

def scrape_pfd_reports2(
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    order: str = "relevance",
    start_page: int = 1,
    end_page: Optional[int] = None,
) -> List[Dict]:
    """
    Scrape PFD reports with enhanced progress tracking and proper pagination
    """
    all_reports = []
    base_url = "https://www.judiciary.uk/"

    try:
        # Initialize progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        report_count_text = st.empty()

        # Validate and prepare category
        category_slug = None
        if category:
            category_slug = (
                category.lower()
                .replace(" ", "-")
                .replace("&", "and")
                .replace("--", "-")
                .strip("-")
            )
            logging.info(f"Using category: {category}, slug: {category_slug}")

        # Construct initial search URL
        base_search_url = construct_search_url(
            base_url=base_url,
            keyword=keyword,
            category=category,
            category_slug=category_slug,
        )

        st.info(f"Searching at: {base_search_url}")

        # Get total pages and results count
        total_pages, total_results = get_total_pages(base_search_url)

        if total_results == 0:
            st.warning("No results found matching your search criteria")
            return []

        st.info(f"Found {total_results} matching reports across {total_pages} pages")

        # Apply page range limits
        start_page = max(1, start_page)  # Ensure start_page is at least 1
        if end_page is None:
            end_page = total_pages
        else:
            end_page = min(
                end_page, total_pages
            )  # Ensure end_page doesn't exceed total_pages

        if start_page > end_page:
            st.warning(f"Invalid page range: {start_page} to {end_page}")
            return []

        st.info(f"Scraping pages {start_page} to {end_page}")

        # Process each page in the specified range
        for current_page in range(start_page, end_page + 1):
            try:
                # Check if scraping should be stopped
                if (
                    hasattr(st.session_state, "stop_scraping")
                    and st.session_state.stop_scraping
                ):
                    st.warning("Scraping stopped by user")
                    break

                # Update progress
                progress = (current_page - start_page) / (end_page - start_page + 1)
                progress_bar.progress(progress)
                status_text.text(
                    f"Processing page {current_page} of {end_page} (out of {total_pages} total pages)"
                )

                # Construct current page URL
                page_url = construct_search_url(
                    base_url=base_url,
                    keyword=keyword,
                    category=category,
                    category_slug=category_slug,
                    page=current_page,
                )

                # Scrape current page
                page_reports = scrape_page(page_url)

                if page_reports:
                    # Deduplicate based on title and URL
                    existing_reports = {(r["Title"], r["URL"]) for r in all_reports}
                    new_reports = [
                        r
                        for r in page_reports
                        if (r["Title"], r["URL"]) not in existing_reports
                    ]

                    all_reports.extend(new_reports)
                    report_count_text.text(
                        f"Retrieved {len(all_reports)} unique reports so far..."
                    )

                # Add delay between pages
                time.sleep(2)

            except Exception as e:
                logging.error(f"Error processing page {current_page}: {e}")
                st.warning(
                    f"Error on page {current_page}. Continuing with next page..."
                )
                continue

        # Sort results if specified
        if order != "relevance":
            all_reports = sort_reports(all_reports, order)

        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
        report_count_text.empty()

        if all_reports:
            st.success(f"Successfully scraped {len(all_reports)} unique reports")
        else:
            st.warning("No reports were successfully retrieved")

        return all_reports

    except Exception as e:
        logging.error(f"Error in scrape_pfd_reports: {e}")
        st.error(f"An error occurred while scraping reports: {e}")
        return []


def construct_search_url(
    base_url: str,
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    category_slug: Optional[str] = None,
    page: Optional[int] = None,
) -> str:
    """Constructs proper search URL with pagination"""
    # Start with base search URL
    url = f"{base_url}?s=&post_type=pfd"

    # Add category filter
    if category and category_slug:
        url += f"&pfd_report_type={category_slug}"

    # Add keyword search
    if keyword:
        url = f"{base_url}?s={keyword}&post_type=pfd"
        if category and category_slug:
            url += f"&pfd_report_type={category_slug}"

    # Add pagination
    if page and page > 1:
        url += f"&paged={page}"  # Changed from &page= to &paged= for proper pagination

    return url


def render_topic_summary_tab(data: pd.DataFrame = None) -> None:
    """Topic analysis with weighting schemes and essential controls"""
    st.subheader("Topic Analysis & Summaries")
    
    # Start with file upload, ignoring any previously loaded data
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for Topic Analysis",
        type=["csv", "xlsx"],
        help="Upload a preprocessed file containing report content",
        key="topic_analysis_uploader"
    )

    # Only proceed with analysis if a file is uploaded
    if uploaded_file is not None:
        try:
            # Process the uploaded file
            if uploaded_file.name.endswith(".csv"):
                data = pd.read_csv(uploaded_file)
            else:
                data = pd.read_excel(uploaded_file)
                
            # Process the data
            data = process_scraped_data(data)
            
            # Validate that we have the needed content column
            if "Content" not in data.columns:
                st.error("The uploaded file does not contain a 'Content' column needed for topic analysis.")
                return
                
            st.success(f"File loaded successfully with {len(data)} rows.")
            
            # Text Processing options
            st.subheader("Analysis Settings")
            col1, col2 = st.columns(2)

            with col1:
                # Vectorization method
                vectorizer_type = st.selectbox(
                    "Vectorization Method",
                    options=["tfidf", "bm25", "weighted"],
                    help="Choose how to convert text to numerical features",
                )

                # Weighting Schemes
                if vectorizer_type == "weighted":
                    tf_scheme = st.selectbox(
                        "Term Frequency Scheme",
                        options=["raw", "log", "binary", "augmented"],
                        help="How to count term occurrences",
                    )
                    idf_scheme = st.selectbox(
                        "Document Frequency Scheme",
                        options=["smooth", "standard", "probabilistic"],
                        help="How to weight document frequencies",
                    )
                elif vectorizer_type == "bm25":
                    k1 = st.slider(
                        "Term Saturation (k1)",
                        min_value=0.5,
                        max_value=3.0,
                        value=1.5,
                        help="Controls term frequency impact",
                    )
                    b = st.slider(
                        "Length Normalization (b)",
                        min_value=0.0,
                        max_value=1.0,
                        value=0.75,
                        help="Document length impact",
                    )

            with col2:
                # Clustering Parameters
                min_cluster_size = st.slider(
                    "Minimum Group Size",
                    min_value=2,
                    max_value=10,
                    value=3,
                    help="Minimum documents per theme",
                )

                max_features = st.slider(
                    "Maximum Features",
                    min_value=1000,
                    max_value=10000,
                    value=5000,
                    step=1000,
                    help="Number of terms to consider",
                )

            # Date range selection
            st.subheader("Date Range")
            date_col1, date_col2 = st.columns(2)
            
            # Only show date selector if date_of_report column exists
            if "date_of_report" in data.columns and pd.api.types.is_datetime64_any_dtype(data["date_of_report"]):
                with date_col1:
                    start_date = st.date_input(
                        "From",
                        value=data["date_of_report"].min().date(),
                        min_value=data["date_of_report"].min().date(),
                        max_value=data["date_of_report"].max().date(),
                    )

                with date_col2:
                    end_date = st.date_input(
                        "To",
                        value=data["date_of_report"].max().date(),
                        min_value=data["date_of_report"].min().date(),
                        max_value=data["date_of_report"].max().date(),
                    )
                
                # Apply date filter
                data = data[
                    (data["date_of_report"].dt.date >= start_date)
                    & (data["date_of_report"].dt.date <= end_date)
                ]
            else:
                st.info("No date column found. Date filtering is not available.")

            # Category selection
            if "categories" in data.columns:
                all_categories = set()
                for cats in data["categories"].dropna():
                    if isinstance(cats, list):
                        all_categories.update(cats)
                    elif isinstance(cats, str):
                        # Handle comma-separated strings
                        all_categories.update(cat.strip() for cat in cats.split(","))

                # Remove any empty strings
                all_categories = {cat for cat in all_categories if cat and isinstance(cat, str)}

                if all_categories:
                    categories = st.multiselect(
                        "Filter by Categories (Optional)",
                        options=sorted(all_categories),
                        help="Select specific categories to analyse",
                    )
                    
                    # Apply category filter if needed
                    if categories:
                        data = filter_by_categories(data, categories)
            else:
                st.info("No categories column found. Category filtering is not available.")

            # Analysis button
            analyze_clicked = st.button(
                "🔍 Analyse Documents", type="primary", use_container_width=True
            )

            # Run analysis if button is clicked
            if analyze_clicked:
                try:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    # Initialize
                    progress_bar.progress(0.2)
                    status_text.text("Processing documents...")
                    initialize_nltk()

                    # Remove empty content
                    filtered_df = data[
                        data["Content"].notna()
                        & (data["Content"].str.strip() != "")
                    ]

                    if len(filtered_df) < min_cluster_size:
                        progress_bar.empty()
                        status_text.empty()
                        st.warning(
                            f"Not enough documents match the criteria. Found {len(filtered_df)}, need at least {min_cluster_size}."
                        )
                        return

                    # Process content
                    progress_bar.progress(0.4)
                    status_text.text("Identifying themes...")

                    processed_df = pd.DataFrame(
                        {
                            "Content": filtered_df["Content"],
                            "Title": filtered_df["Title"],
                            "date_of_report": filtered_df["date_of_report"] if "date_of_report" in filtered_df.columns else None,
                            "URL": filtered_df["URL"] if "URL" in filtered_df.columns else None,
                            "categories": filtered_df["categories"] if "categories" in filtered_df.columns else None,
                        }
                    )

                    progress_bar.progress(0.6)
                    status_text.text("Analyzing patterns...")

                    # Prepare vectorizer parameters
                    vectorizer_params = {}
                    if vectorizer_type == "weighted":
                        vectorizer_params.update(
                            {"tf_scheme": tf_scheme, "idf_scheme": idf_scheme}
                        )
                    elif vectorizer_type == "bm25":
                        vectorizer_params.update({"k1": k1, "b": b})

                    # Store vectorization settings in session state
                    st.session_state.vectorizer_type = vectorizer_type
                    st.session_state.update(vectorizer_params)

                    # Perform clustering
                    cluster_results = perform_semantic_clustering(
                        processed_df,
                        min_cluster_size=min_cluster_size,
                        max_features=max_features,
                        min_df=2 / len(processed_df),
                        max_df=0.95,
                        similarity_threshold=0.3,
                    )

                    progress_bar.progress(0.8)
                    status_text.text("Generating summaries...")

                    # Store results
                    st.session_state.topic_model = cluster_results

                    progress_bar.progress(1.0)
                    status_text.text("Analysis complete!")

                    progress_bar.empty()
                    status_text.empty()

                    # Display results
                    render_summary_tab(cluster_results, processed_df)

                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"Analysis error: {str(e)}")
                    logging.error(f"Analysis error: {e}", exc_info=True)
                    
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
    else:
        # Show instructions when no file is uploaded
        st.info("Please upload a file to begin topic analysis.")
        
        with st.expander("📋 File Requirements"):
            st.markdown("""
            ## Required Columns
            
            For topic analysis, your file should include:
            
            - **Content**: The text content to analyze (required)
            - **Title**: Report titles (recommended)
            - **date_of_report**: Report dates (optional, for filtering)
            - **categories**: Report categories (optional, for filtering)
            
            Files prepared from Step 2 "Scraped File Preparation" are ideal for this analysis.
            """)
            
        # Show a sample of what to expect
        with st.expander("🔍 What to Expect"):
            st.markdown("""
            ## Topic Analysis Results
            
            The analysis will generate:
            
            1. **Topic clusters**: Groups of similar documents
            2. **Key terms**: Important words in each topic
            3. **Topic summaries**: Brief overview of each topic's content
            4. **Network visualizations**: Showing relationships between terms
            
            The quality of results depends on having enough documents with good text content.
            """)

def render_topic_options():
    """Render enhanced topic analysis options in a clear layout"""
    st.header("Topic Analysis & Summaries")

    # Create two columns for main settings
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("##### Text Processing")
        vectorizer_type = st.selectbox(
            "Vectorization Method",
            options=["tfidf", "bm25", "weighted"],
            help="Choose how to convert text to numerical features:\n"
            + "- TF-IDF: Classic term frequency-inverse document frequency\n"
            + "- BM25: Enhanced version of TF-IDF used in search engines\n"
            + "- Weighted: Customizable term and document weighting",
        )

        # Show specific parameters based on vectorizer type
        if vectorizer_type == "bm25":
            st.markdown("##### BM25 Parameters")
            k1 = st.slider(
                "Term Saturation (k1)",
                min_value=0.5,
                max_value=3.0,
                value=1.5,
                step=0.1,
                help="Controls how quickly term frequency saturates (higher = slower)",
            )
            b = st.slider(
                "Length Normalization (b)",
                min_value=0.0,
                max_value=1.0,
                value=0.75,
                step=0.05,
                help="How much to penalize long documents",
            )

        elif vectorizer_type == "weighted":
            st.markdown("##### Weighting Schemes")
            tf_scheme = st.selectbox(
                "Term Frequency Scheme",
                options=["raw", "log", "binary", "augmented"],
                help="How to count term occurrences:\n"
                + "- Raw: Use actual counts\n"
                + "- Log: Logarithmic scaling\n"
                + "- Binary: Just presence/absence\n"
                + "- Augmented: Normalized frequency",
            )
            idf_scheme = st.selectbox(
                "Document Frequency Scheme",
                options=["smooth", "standard", "probabilistic"],
                help="How to weight document frequencies:\n"
                + "- Smooth: With smoothing factor\n"
                + "- Standard: Classic IDF\n"
                + "- Probabilistic: Based on probability",
            )

    with col2:
        st.markdown("##### Clustering Parameters")
        min_cluster_size = st.slider(
            "Minimum Cluster Size",
            min_value=2,
            max_value=10,
            value=3,
            help="Smallest allowed group of similar documents",
        )

        max_features = st.slider(
            "Maximum Features",
            min_value=1000,
            max_value=10000,
            value=5000,
            step=500,
            help="Maximum number of terms to consider",
        )

        min_similarity = st.slider(
            "Minimum Similarity",
            min_value=0.0,
            max_value=1.0,
            value=0.3,
            step=0.05,
            help="How similar documents must be to be grouped together",
        )

    # Advanced options in expander
    with st.expander("Advanced Settings"):
        st.markdown("##### Document Frequency Bounds")
        col3, col4 = st.columns(2)

        with col3:
            min_df = st.number_input(
                "Minimum Document Frequency",
                min_value=1,
                max_value=100,
                value=2,
                help="Minimum number of documents a term must appear in",
            )

        with col4:
            max_df = st.slider(
                "Maximum Document %",
                min_value=0.1,
                max_value=1.0,
                value=0.95,
                step=0.05,
                help="Maximum % of documents a term can appear in",
            )

        st.markdown("##### Visualization Settings")
        network_layout = st.selectbox(
            "Network Layout",
            options=["force", "circular", "random"],
            help="How to arrange nodes in topic networks",
        )

        show_weights = st.checkbox(
            "Show Edge Weights",
            value=True,
            help="Display connection strengths between terms",
        )

    return {
        "vectorizer_type": vectorizer_type,
        "vectorizer_params": {
            "k1": k1 if vectorizer_type == "bm25" else None,
            "b": b if vectorizer_type == "bm25" else None,
            "tf_scheme": tf_scheme if vectorizer_type == "weighted" else None,
            "idf_scheme": idf_scheme if vectorizer_type == "weighted" else None,
        },
        "cluster_params": {
            "min_cluster_size": min_cluster_size,
            "max_features": max_features,
            "min_similarity": min_similarity,
            "min_df": min_df,
            "max_df": max_df,
        },
        "viz_params": {"network_layout": network_layout, "show_weights": show_weights},
    }


def sort_reports(reports: List[Dict], order: str) -> List[Dict]:
    """Sort reports based on specified order"""
    if order == "date_desc":
        return sorted(
            reports,
            key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d"),
            reverse=True,
        )
    elif order == "date_asc":
        return sorted(reports, key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d"))
    return reports


def plot_category_distribution(df: pd.DataFrame) -> None:
    """Plot category distribution"""
    all_cats = []
    for cats in df["categories"].dropna():
        if isinstance(cats, list):
            all_cats.extend(cats)

    cat_counts = pd.Series(all_cats).value_counts()

    fig = px.bar(
        x=cat_counts.index,
        y=cat_counts.values,
        title="Category Distribution",
        labels={"x": "Category", "y": "Count"},
    )
    fig.update_layout(xaxis_title="Category", yaxis_title="Number of Reports", xaxis={"tickangle": 45})

    st.plotly_chart(fig, use_container_width=True)


def plot_coroner_areas(df: pd.DataFrame) -> None:
    """Plot coroner areas distribution"""
    area_counts = df["coroner_area"].value_counts().head(20)

    fig = px.bar(
        x=area_counts.index,
        y=area_counts.values,
        title="Top 20 Coroner Areas",
        labels={"x": "Area", "y": "Count"},
    )

    fig.update_layout(
        xaxis_title="Coroner Area",
        yaxis_title="Number of Reports",
        xaxis={"tickangle": 45},
    )

    st.plotly_chart(fig, use_container_width=True)


def analyze_data_quality(df: pd.DataFrame) -> None:
    """Analyze and display data quality metrics for PFD reports"""

    # Calculate completeness metrics
    total_records = len(df)

    def calculate_completeness(field):
        if field not in df.columns:
            return 0
        non_empty = df[field].notna()
        if field == "categories":
            non_empty = df[field].apply(lambda x: isinstance(x, list) and len(x) > 0)
        return (non_empty.sum() / total_records) * 100

    completeness_metrics = {
        "Title": calculate_completeness("Title"),
        "Content": calculate_completeness("Content"),
        "Date of Report": calculate_completeness("date_of_report"),
        "Deceased Name": calculate_completeness("deceased_name"),
        "Coroner Name": calculate_completeness("coroner_name"),
        "Coroner Area": calculate_completeness("coroner_area"),
        "Categories": calculate_completeness("categories"),
    }

    # Calculate consistency metrics
    consistency_metrics = {
        "Title Format": (df["Title"].str.len() >= 10).mean() * 100,
        "Content Length": (df["Content"].str.len() >= 100).mean() * 100,
        "Date Format": df["date_of_report"].notna().mean() * 100,
        "Categories Format": df["categories"]
        .apply(lambda x: isinstance(x, list))
        .mean()
        * 100,
    }

    # Calculate PDF metrics
    pdf_columns = [
        col for col in df.columns if col.startswith("PDF_") and col.endswith("_Path")
    ]
    reports_with_pdf = df[pdf_columns].notna().any(axis=1).sum()
    reports_with_multiple_pdfs = (df[pdf_columns].notna().sum(axis=1) > 1).sum()

    pdf_metrics = {
        "Reports with PDFs": (reports_with_pdf / total_records) * 100,
        "Reports with Multiple PDFs": (reports_with_multiple_pdfs / total_records)
        * 100,
    }

    # Display metrics using Streamlit
    st.subheader("Data Quality Analysis")

    # Completeness
    completeness_df = pd.DataFrame(
        list(completeness_metrics.items()), columns=["Field", "Completeness %"]
    )
    fig_completeness = px.bar(
        completeness_df,
        x="Field",
        y="Completeness %",
        title="Field Completeness Analysis",
    )
    fig_completeness.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_completeness, use_container_width=True)

    # Consistency
    consistency_df = pd.DataFrame(
        list(consistency_metrics.items()), columns=["Metric", "Consistency %"]
    )
    fig_consistency = px.bar(
        consistency_df, x="Metric", y="Consistency %", title="Data Consistency Analysis"
    )
    fig_consistency.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_consistency, use_container_width=True)

    # PDF Analysis
    pdf_df = pd.DataFrame(list(pdf_metrics.items()), columns=["Metric", "Percentage"])
    fig_pdf = px.bar(pdf_df, x="Metric", y="Percentage", title="PDF Coverage Analysis")
    fig_pdf.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_pdf, use_container_width=True)

    # Summary metrics
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric(
            "Average Completeness",
            f"{np.mean(list(completeness_metrics.values())):.1f}%",
        )

    with col2:
        st.metric(
            "Average Consistency", f"{np.mean(list(consistency_metrics.values())):.1f}%"
        )

    with col3:
        st.metric("PDF Coverage", f"{pdf_metrics['Reports with PDFs']:.1f}%")

    # Detailed quality issues
    st.markdown("### Detailed Quality Issues")

    issues = []

    # Check for missing crucial fields
    for field, completeness in completeness_metrics.items():
        if completeness < 95:  # Less than 95% complete
            issues.append(
                f"- {field} is {completeness:.1f}% complete ({total_records - int(completeness * total_records / 100)} records missing)"
            )

    # Check for consistency issues
    for metric, consistency in consistency_metrics.items():
        if consistency < 90:  # Less than 90% consistent
            issues.append(f"- {metric} shows {consistency:.1f}% consistency")

    # Check PDF coverage
    if pdf_metrics["Reports with PDFs"] < 90:
        issues.append(
            f"- {100 - pdf_metrics['Reports with PDFs']:.1f}% of reports lack PDF attachments"
        )

    if issues:
        st.markdown("The following quality issues were identified:")
        for issue in issues:
            st.markdown(issue)
    else:
        st.success("No significant quality issues found in the dataset.")


def display_topic_network(lda, feature_names):
    """Display word similarity network with interactive filters"""
    # st.markdown("### Word Similarity Network")
    st.markdown(
        "This network shows relationships between words based on their co-occurrence in documents."
    )

    # Store base network data in session state if not already present
    if "network_data" not in st.session_state:
        # Get word counts across all documents
        word_counts = lda.components_.sum(axis=0)
        top_word_indices = word_counts.argsort()[
            : -100 - 1 : -1
        ]  # Store more words initially

        # Create word co-occurrence matrix
        word_vectors = normalize(lda.components_.T[top_word_indices])
        word_similarities = cosine_similarity(word_vectors)

        st.session_state.network_data = {
            "word_counts": word_counts,
            "top_word_indices": top_word_indices,
            "word_similarities": word_similarities,
            "feature_names": feature_names,
        }

    # Network filters with keys to prevent rerun
    col1, col2, col3 = st.columns(3)
    with col1:
        min_similarity = st.slider(
            "Minimum Similarity",
            min_value=0.0,
            max_value=1.0,
            value=0.9,
            step=0.05,
            help="Higher values show stronger connections only",
            key="network_min_similarity",
        )
    with col2:
        max_words = st.slider(
            "Number of Words",
            min_value=10,
            max_value=100,
            value=30,
            step=5,
            help="Number of most frequent words to show",
            key="network_max_words",
        )
    with col3:
        min_connections = st.slider(
            "Minimum Connections",
            min_value=1,
            max_value=10,
            value=5,
            help="Minimum number of connections per word",
            key="network_min_connections",
        )

    # Create network graph based on current filters
    G = nx.Graph()

    # Get stored data
    word_counts = st.session_state.network_data["word_counts"]
    word_similarities = st.session_state.network_data["word_similarities"]
    top_word_indices = st.session_state.network_data["top_word_indices"][:max_words]
    feature_names = st.session_state.network_data["feature_names"]

    # Add nodes
    for idx, word_idx in enumerate(top_word_indices):
        G.add_node(idx, name=feature_names[word_idx], freq=float(word_counts[word_idx]))

    # Add edges based on current similarity threshold
    for i in range(len(top_word_indices)):
        for j in range(i + 1, len(top_word_indices)):
            similarity = word_similarities[i, j]
            if similarity > min_similarity:
                G.add_edge(i, j, weight=float(similarity))

    # Filter nodes by minimum connections
    nodes_to_remove = []
    for node in G.nodes():
        if G.degree(node) < min_connections:
            nodes_to_remove.append(node)
    G.remove_nodes_from(nodes_to_remove)

    if len(G.nodes()) == 0:
        st.warning(
            "No nodes match the current filter criteria. Try adjusting the filters."
        )
        return

    # Create visualization
    pos = nx.spring_layout(G, k=1 / np.sqrt(len(G.nodes())), iterations=50)

    # Create edge traces with varying thickness and color based on weight
    edge_traces = []
    for edge in G.edges(data=True):
        x0, y0 = pos[edge[0]]
        x1, y1 = pos[edge[1]]
        weight = edge[2]["weight"]

        edge_trace = go.Scatter(
            x=[x0, x1, None],
            y=[y0, y1, None],
            line=dict(width=weight * 3, color=f"rgba(100,100,100,{weight})"),
            hoverinfo="none",
            mode="lines",
        )
        edge_traces.append(edge_trace)

    # Create node trace with size based on frequency
    node_x = []
    node_y = []
    node_text = []
    node_size = []

    for node in G.nodes():
        x, y = pos[node]
        node_x.append(x)
        node_y.append(y)
        freq = G.nodes[node]["freq"]
        name = G.nodes[node]["name"]
        connections = G.degree(node)
        node_text.append(
            f"{name}<br>Frequency: {freq:.0f}<br>Connections: {connections}"
        )
        node_size.append(np.sqrt(freq) * 10)

    node_trace = go.Scatter(
        x=node_x,
        y=node_y,
        mode="markers+text",
        hoverinfo="text",
        text=node_text,
        textposition="top center",
        marker=dict(
            size=node_size, line=dict(width=1), color="lightblue", sizemode="area"
        ),
    )

    # Create figure
    fig = go.Figure(
        data=edge_traces + [node_trace],
        layout=go.Layout(
            title=f"Word Network ({len(G.nodes())} words, {len(G.edges())} connections)",
            titlefont_size=16,
            showlegend=False,
            hovermode="closest",
            margin=dict(b=20, l=5, r=5, t=40),
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        ),
    )

    st.plotly_chart(fig, use_container_width=True)

    # Add network statistics
    st.markdown("### Network Statistics")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Number of Words", len(G.nodes()))
    with col2:
        st.metric("Number of Connections", len(G.edges()))
    with col3:
        if len(G.nodes()) > 0:
            density = 2 * len(G.edges()) / (len(G.nodes()) * (len(G.nodes()) - 1))
            st.metric("Network Density", f"{density:.2%}")


def get_top_words(model, feature_names, topic_idx, n_words=10):
    """Get top words for a given topic"""
    return [
        feature_names[i]
        for i in model.components_[topic_idx].argsort()[: -n_words - 1 : -1]
    ]


def render_file_upload():
    """Render file upload section"""
    st.subheader("Upload Existing Data")

    # Generate unique key for the file uploader using reset counter
    reset_counter = st.session_state.get("reset_counter", 0)
    upload_key = f"file_uploader_{int(time.time() * 1000)}"

    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file", type=["csv", "xlsx"], key=upload_key
    )

    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            # Process uploaded data
            df = process_scraped_data(df)

            # Clear any existing data first
            st.session_state.current_data = None
            st.session_state.scraped_data = None
            st.session_state.uploaded_data = None
            st.session_state.data_source = None

            # Then set new data
            st.session_state.uploaded_data = df.copy()
            st.session_state.data_source = "uploaded"
            st.session_state.current_data = df.copy()

            st.success("File uploaded and processed successfully!")

            # Show the uploaded data
            st.subheader("Uploaded Data Preview")
            st.dataframe(
                df,
                column_config={
                    "URL": st.column_config.LinkColumn("Report Link"),
                    "date_of_report": st.column_config.DateColumn("Date of Report"),
                    "categories": st.column_config.ListColumn("Categories"),
                },
                hide_index=True,
            )

            return True

        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            logging.error(f"File upload error: {e}", exc_info=True)
            return False

    return False

# Add this to your initialize_session_state function
def initialize_session_state():
    """Initialize all required session state variables"""
    # Check if we need to perform first-time initialization
    if "initialized" not in st.session_state:
        # Clear all existing session state
        for key in list(st.session_state.keys()):
            del st.session_state[key]
            
        # Set new session state variables
        st.session_state.initialized = True
        st.session_state.data_source = None
        st.session_state.current_data = None
        st.session_state.scraped_data = None
        st.session_state.uploaded_data = None
        st.session_state.topic_model = None
        st.session_state.cleanup_done = False
        st.session_state.last_scrape_time = None
        st.session_state.last_upload_time = None
        st.session_state.reset_counter = 0  # Add this counter for file uploader resets
        
        # BERT-related session state variables
        st.session_state.bert_results = {}
        st.session_state.bert_initialized = False
        st.session_state.bert_merged_data = None
        
        # Analysis filters
        st.session_state.analysis_filters = {
            "date_range": None,
            "selected_categories": None,
            "selected_areas": None,
        }
        
        # Topic model settings
        st.session_state.topic_model_settings = {
            "num_topics": 5,
            "max_features": 1000,
            "similarity_threshold": 0.3,
        }
    
    # Perform PDF cleanup if not done
    if not st.session_state.get("cleanup_done", False):
        try:
            pdf_dir = "pdfs"
            os.makedirs(pdf_dir, exist_ok=True)
            current_time = time.time()
            cleanup_count = 0
            
            for file in os.listdir(pdf_dir):
                file_path = os.path.join(pdf_dir, file)
                try:
                    if os.path.isfile(file_path):
                        if os.stat(file_path).st_mtime < current_time - 86400:
                            os.remove(file_path)
                            cleanup_count += 1
                except Exception as e:
                    logging.warning(f"Error cleaning up file {file_path}: {e}")
                    continue
                    
            if cleanup_count > 0:
                logging.info(f"Cleaned up {cleanup_count} old PDF files")
        except Exception as e:
            logging.error(f"Error during PDF cleanup: {e}")
        finally:
            st.session_state.cleanup_done = True
            
def initialize_session_state2():
    """Initialize all required session state variables"""
    # Initialize basic state variables if they don't exist
    if not hasattr(st.session_state, "initialized"):
        # Clear all existing session state
        for key in list(st.session_state.keys()):
            del st.session_state[key]

        # Set new session state variables
        st.session_state.data_source = None
        st.session_state.current_data = None
        st.session_state.scraped_data = None
        st.session_state.uploaded_data = None
        st.session_state.topic_model = None
        st.session_state.cleanup_done = False
        st.session_state.last_scrape_time = None
        st.session_state.last_upload_time = None
        st.session_state.analysis_filters = {
            "date_range": None,
            "selected_categories": None,
            "selected_areas": None,
        }
        st.session_state.topic_model_settings = {
            "num_topics": 5,
            "max_features": 1000,
            "similarity_threshold": 0.3,
        }
        st.session_state.initialized = True

    # Perform PDF cleanup if not done
    if not st.session_state.cleanup_done:
        try:
            pdf_dir = "pdfs"
            os.makedirs(pdf_dir, exist_ok=True)

            current_time = time.time()
            cleanup_count = 0

            for file in os.listdir(pdf_dir):
                file_path = os.path.join(pdf_dir, file)
                try:
                    if os.path.isfile(file_path):
                        if os.stat(file_path).st_mtime < current_time - 86400:
                            os.remove(file_path)
                            cleanup_count += 1
                except Exception as e:
                    logging.warning(f"Error cleaning up file {file_path}: {e}")
                    continue

            if cleanup_count > 0:
                logging.info(f"Cleaned up {cleanup_count} old PDF files")
        except Exception as e:
            logging.error(f"Error during PDF cleanup: {e}")
        finally:
            st.session_state.cleanup_done = True


def validate_data(data: pd.DataFrame, purpose: str = "analysis") -> Tuple[bool, str]:
    """
    Validate data for different purposes

    Args:
        data: DataFrame to validate
        purpose: Purpose of validation ('analysis' or 'topic_modeling')

    Returns:
        tuple: (is_valid, message)
    """
    if data is None:
        return False, "No data available. Please scrape or upload data first."

    if not isinstance(data, pd.DataFrame):
        return False, "Invalid data format. Expected pandas DataFrame."

    if len(data) == 0:
        return False, "Dataset is empty."

    if purpose == "analysis":
        required_columns = ["date_of_report", "categories", "coroner_area"]
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"

    elif purpose == "topic_modeling":
        if "Content" not in data.columns:
            return False, "Missing required column: Content"

        valid_docs = data["Content"].dropna().str.strip().str.len() > 0
        if valid_docs.sum() < 2:
            return (
                False,
                "Not enough valid documents found. Please ensure you have documents with text content.",
            )

    # Add type checking for critical columns
    if "date_of_report" in data.columns and not pd.api.types.is_datetime64_any_dtype(
        data["date_of_report"]
    ):
        try:
            pd.to_datetime(data["date_of_report"])
        except Exception:
            return False, "Invalid date format in date_of_report column."

    if "categories" in data.columns:
        if (
            not data["categories"]
            .apply(lambda x: isinstance(x, (list, type(None))))
            .all()
        ):
            return False, "Categories must be stored as lists or None values."

    return True, "Data is valid"


def is_response(row: pd.Series) -> bool:
    """
    Check if a report is a response document based on its metadata and content

    Args:
        row: DataFrame row containing report data

    Returns:
        bool: True if document is a response, False otherwise
    """
    try:
        # Check PDF names for response indicators
        pdf_response = False
        for i in range(1, 10):  # Check PDF_1 to PDF_9
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                pdf_response = True
                break

        # Check title for response indicators
        title = str(row.get("Title", "")).lower()
        title_response = any(
            word in title for word in ["response", "reply", "answered"]
        )

        # Check content for response indicators
        content = str(row.get("Content", "")).lower()
        content_response = any(
            phrase in content
            for phrase in [
                "in response to",
                "responding to",
                "reply to",
                "response to",
                "following the regulation 28",
            ]
        )

        return pdf_response or title_response or content_response

    except Exception as e:
        logging.error(f"Error checking response type: {e}")
        return False


def plot_timeline(df: pd.DataFrame) -> None:
    """Plot timeline of reports with improved formatting"""
    timeline_data = (
        df.groupby(pd.Grouper(key="date_of_report", freq="M")).size().reset_index()
    )
    timeline_data.columns = ["Date", "Count"]

    fig = px.line(
        timeline_data,
        x="Date",
        y="Count",
        title="Reports Timeline",
        labels={"Count": "Number of Reports"},
    )

    fig.update_layout(
        xaxis_title="Date",
        yaxis_title="Number of Reports",
        hovermode="x unified",
        yaxis=dict(
            tickmode="linear",
            tick0=0,
            dtick=1,  # Integer steps
            rangemode="nonnegative",  # Ensure y-axis starts at 0 or above
        ),
        xaxis=dict(tickformat="%B %Y", tickangle=45),  # Month Year format
    )

    st.plotly_chart(fig, use_container_width=True)


def plot_monthly_distribution(df: pd.DataFrame) -> None:
    """Plot monthly distribution with improved formatting"""
    # Format dates as Month Year
    df["month_year"] = df["date_of_report"].dt.strftime("%B %Y")
    monthly_counts = df["month_year"].value_counts().sort_index()

    fig = px.bar(
        x=monthly_counts.index,
        y=monthly_counts.values,
        labels={"x": "Month", "y": "Number of Reports"},
        title="Monthly Distribution of Reports",
    )

    fig.update_layout(
        xaxis_title="Month",
        yaxis_title="Number of Reports",
        xaxis_tickangle=45,
        yaxis=dict(
            tickmode="linear",
            tick0=0,
            dtick=1,  # Integer steps
            rangemode="nonnegative",
        ),
        bargap=0.2,
    )

    st.plotly_chart(fig, use_container_width=True)


def plot_yearly_comparison(df: pd.DataFrame) -> None:
    """Plot year-over-year comparison with improved formatting"""
    yearly_counts = df["date_of_report"].dt.year.value_counts().sort_index()

    fig = px.line(
        x=yearly_counts.index.astype(int),  # Convert to integer years
        y=yearly_counts.values,
        markers=True,
        labels={"x": "Year", "y": "Number of Reports"},
        title="Year-over-Year Report Volumes",
    )

    # Calculate appropriate y-axis maximum
    max_count = yearly_counts.max()
    y_max = max_count + (1 if max_count < 10 else 2)  # Add some padding

    fig.update_layout(
        xaxis_title="Year",
        yaxis_title="Number of Reports",
        xaxis=dict(
            tickmode="linear",
            tick0=yearly_counts.index.min(),
            dtick=1,  # Show every year
            tickformat="d",  # Format as integer
        ),
        yaxis=dict(
            tickmode="linear", tick0=0, dtick=1, range=[0, y_max]  # Integer steps
        ),
    )

    st.plotly_chart(fig, use_container_width=True)


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Export DataFrame to Excel bytes with proper formatting
    """
    try:
        if df is None or len(df) == 0:
            raise ValueError("No data available to export")

        # Create clean copy for export
        df_export = df.copy()

        # Format dates to UK format
        if "date_of_report" in df_export.columns:
            df_export["date_of_report"] = df_export["date_of_report"].dt.strftime(
                "%d/%m/%Y"
            )

        # Handle list columns (like categories)
        for col in df_export.columns:
            if df_export[col].dtype == "object":
                df_export[col] = df_export[col].apply(
                    lambda x: ", ".join(x)
                    if isinstance(x, list)
                    else str(x)
                    if pd.notna(x)
                    else ""
                )

        # Create output buffer
        output = io.BytesIO()

        # Write to Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_export.to_excel(writer, sheet_name="Reports", index=False)

            # Get the worksheet
            worksheet = writer.sheets["Reports"]

            # Auto-adjust column widths
            for idx, col in enumerate(df_export.columns, 1):
                # Set larger width for Content and Extracted_Concerns columns
                if col in ["Content", "Extracted_Concerns"]:
                    worksheet.column_dimensions[get_column_letter(idx)].width = 80
                else:
                    max_length = max(
                        df_export[col].astype(str).apply(len).max(),
                        len(str(col)),
                    )
                    adjusted_width = min(max_length + 2, 50)
                    column_letter = get_column_letter(idx)
                    worksheet.column_dimensions[
                        column_letter
                    ].width = adjusted_width

            # Add filters to header row
            worksheet.auto_filter.ref = worksheet.dimensions

            # Freeze the header row
            worksheet.freeze_panes = "A2"

            # Set wrap text for Content and Extracted_Concerns columns
            for column_name in ["Content", "Extracted_Concerns"]:
                if column_name in df_export.columns:
                    col_idx = df_export.columns.get_loc(column_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for row in range(2, len(df_export) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = cell.alignment.copy(wrapText=True)
                        # Set row height to accommodate wrapped text
                        worksheet.row_dimensions[row].height = 60

        # Get the bytes value
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}", exc_info=True)
        raise Exception(f"Failed to export data to Excel: {str(e)}")
        
def show_export_options(df: pd.DataFrame, prefix: str):
    """Show export options for the data with descriptive filename and unique keys"""
    try:
        st.subheader("Export Options")

        # Generate timestamp and random suffix to create unique keys
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
        unique_id = f"{timestamp}_{random_suffix}"
        filename = f"pfd_reports_{prefix}_{timestamp}"

        col1, col2 = st.columns(2)

        # CSV Export
        with col1:
            try:
                # Create export copy with formatted dates
                df_csv = df.copy()
                if "date_of_report" in df_csv.columns:
                    df_csv["date_of_report"] = df_csv["date_of_report"].dt.strftime("%d/%m/%Y")

                csv = df_csv.to_csv(index=False).encode('utf-8')
                csv_key = f"download_csv_{prefix}_{unique_id}"
                st.download_button(
                    "📥 Download Reports (CSV)",
                    csv,
                    f"{filename}.csv",
                    "text/csv",
                    key=csv_key
                )
            except Exception as e:
                st.error(f"Error preparing CSV export: {str(e)}")

        # Excel Export
        with col2:
            try:
                excel_data = export_to_excel(df)
                excel_key = f"download_excel_{prefix}_{unique_id}"
                st.download_button(
                    "📥 Download Reports (Excel)",
                    excel_data,
                    f"{filename}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=excel_key
                )
            except Exception as e:
                st.error(f"Error preparing Excel export: {str(e)}")
        ##
        # PDF Export section
        # PDF Export section
        # PDF Export section
        if any(col.startswith("PDF_") and col.endswith("_Path") for col in df.columns):
            st.subheader("Download PDFs")
            try:
                # Create the ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, "w") as zipf:
                    pdf_columns = [col for col in df.columns if col.startswith("PDF_") and col.endswith("_Path")]
                    added_files = set()
                    pdf_count = 0
                    folders_created = set()
                    
                    # Process each PDF file
                    for idx, row in df.iterrows():
                        # Build folder name using ref and deceased_name
                        folder_parts = []
                        
                        # Add reference number if available
                        if "ref" in row and pd.notna(row["ref"]):
                            folder_parts.append(str(row["ref"]))
                        
                        # Add deceased name if available
                        if "deceased_name" in row and pd.notna(row["deceased_name"]):
                            # Clean up deceased name for folder name
                            deceased_name = str(row["deceased_name"])
                            # Remove invalid characters for folder names
                            clean_name = re.sub(r'[<>:"/\\|?*]', '_', deceased_name)
                            # Limit folder name length
                            clean_name = clean_name[:50].strip()
                            folder_parts.append(clean_name)
                        
                        # Create folder name from parts
                        if folder_parts:
                            folder_name = "_".join(folder_parts)
                        else:
                            # Fallback if no ref or deceased name
                            record_id = str(row.get("Record ID", idx))
                            folder_name = f"report_{record_id}"
                        
                        # Add year to folder if available
                        if "year" in row and pd.notna(row["year"]):
                            folder_name = f"{folder_name}_{row['year']}"
                        
                        # Keep track of created folders
                        folders_created.add(folder_name)
                        
                        # Process each PDF for this row
                        for col in pdf_columns:
                            pdf_path = row.get(col)
                            if pd.isna(pdf_path) or not pdf_path or not os.path.exists(pdf_path) or pdf_path in added_files:
                                continue
                            
                            # Get the original filename without any modifications
                            original_filename = os.path.basename(pdf_path)
                            
                            # Create archive path with folder structure
                            zip_path = f"{folder_name}/{original_filename}"
                            
                            # Read the file content and write it to the ZIP
                            with open(pdf_path, 'rb') as file:
                                zipf.writestr(zip_path, file.read())
                            
                            added_files.add(pdf_path)
                            pdf_count += 1
                
                # Reset buffer position
                zip_buffer.seek(0)
                
                # PDF Download Button with Unique Key
                pdf_key = f"download_pdfs_{prefix}_{unique_id}"
                st.download_button(
                    f"📦 Download All PDFs ({pdf_count} files in {len(folders_created)} case folders)",
                    zip_buffer,
                    f"{filename}_pdfs.zip",
                    "application/zip",
                    key=pdf_key
                )
                    
            except Exception as e:
                st.error(f"Error preparing PDF download: {str(e)}")
                logging.error(f"PDF download error: {e}", exc_info=True)
        
        ##
       
    except Exception as e:
        st.error(f"Error setting up export options: {str(e)}")
        logging.error(f"Export options error: {e}", exc_info=True)

        
def show_export_options2(df: pd.DataFrame, prefix: str):
    """Show export options for the data with descriptive filename and unique keys"""
    try:
        st.subheader("Export Options")

        # Generate timestamp and random suffix for unique keys
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = "".join(
            random.choices(string.ascii_lowercase + string.digits, k=6)
        )
        unique_id = f"{timestamp}_{random_suffix}"
        filename = f"pfd_reports_{prefix}_{timestamp}"

        col1, col2 = st.columns(2)

        # CSV Export
        with col1:
            try:
                # Create export copy with formatted dates
                df_csv = df.copy()
                if "date_of_report" in df_csv.columns:
                    df_csv["date_of_report"] = df_csv["date_of_report"].dt.strftime(
                        "%d/%m/%Y"
                    )

                csv = df_csv.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Download Reports (CSV)",
                    csv,
                    f"{filename}.csv",
                    "text/csv",
                    key=f"download_csv_{prefix}_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing CSV export: {str(e)}")

        # Excel Export
        with col2:
            try:
                excel_data = export_to_excel(df)
                st.download_button(
                    "📥 Download Reports (Excel)",
                    excel_data,
                    f"{filename}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_excel_{prefix}_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing Excel export: {str(e)}")

        # PDF Download
        if any(col.startswith("PDF_") and col.endswith("_Path") for col in df.columns):
            st.subheader("Download PDFs")
            if st.button(f"Download all PDFs", key=f"pdf_button_{prefix}_{unique_id}"):
                with st.spinner("Preparing PDF download..."):
                    try:
                        pdf_zip_path = f"{filename}_pdfs.zip"

                        with zipfile.ZipFile(pdf_zip_path, "w") as zipf:
                            pdf_columns = [
                                col
                                for col in df.columns
                                if col.startswith("PDF_") and col.endswith("_Path")
                            ]
                            added_files = set()

                            for col in pdf_columns:
                                paths = df[col].dropna()
                                for pdf_path in paths:
                                    if (
                                        pdf_path
                                        and os.path.exists(pdf_path)
                                        and pdf_path not in added_files
                                    ):
                                        zipf.write(pdf_path, os.path.basename(pdf_path))
                                        added_files.add(pdf_path)

                        with open(pdf_zip_path, "rb") as f:
                            st.download_button(
                                "📦 Download All PDFs (ZIP)",
                                f.read(),
                                pdf_zip_path,
                                "application/zip",
                                key=f"download_pdfs_zip_{prefix}_{unique_id}",
                            )

                        # Cleanup zip file
                        os.remove(pdf_zip_path)
                    except Exception as e:
                        st.error(f"Error preparing PDF download: {str(e)}")

    except Exception as e:
        st.error(f"Error setting up export options: {str(e)}")
        logging.error(f"Export options error: {e}", exc_info=True)


def extract_advanced_topics(
    data: pd.DataFrame,
    num_topics: int = 5,
    max_features: int = 1000,
    min_df: int = 2,
    n_iterations: int = 20,
    min_similarity: float = 0.9,
) -> Tuple[LatentDirichletAllocation, CountVectorizer, np.ndarray]:
    """
     Topic modeling with comprehensive preprocessing and error handling

    Args:
        data (pd.DataFrame): Input DataFrame containing documents
        num_topics (int): Number of topics to extract
        max_features (int): Maximum number of features to use
        min_df (int): Minimum document frequency for terms
        n_iterations (int): Maximum number of iterations for LDA
        min_similarity (float): Minimum similarity threshold for the word similarity network

    Returns:
        Tuple containing LDA model, vectorizer, and document-topic distribution
    """
    try:
        # Extensive logging
        logging.info(f"Starting topic modeling with {len(data)} documents")
        logging.info(
            f"Parameters: topics={num_topics}, max_features={max_features}, min_df={min_df}, min_similarity={min_similarity}"
        )

        # Validate input data
        if data is None or len(data) == 0:
            raise ValueError("No data provided for topic modeling")

        # Remove duplicate documents based on content
        def prepare_document(doc: str) -> str:
            """Clean and prepare individual documents"""
            if pd.isna(doc):
                return None

            # Aggressive text cleaning
            cleaned_doc = clean_text_for_modeling(str(doc))

            # Minimum length check
            return cleaned_doc if len(cleaned_doc.split()) > 3 else None

        # Process documents
        documents = data["Content"].apply(prepare_document).dropna().unique().tolist()

        logging.info(f"Processed {len(documents)} unique valid documents")

        # Validate document count
        if len(documents) < num_topics:
            adjusted_topics = max(2, len(documents) // 2)
            logging.warning(
                f"Not enough documents for {num_topics} topics. Adjusting to {adjusted_topics}"
            )
            num_topics = adjusted_topics

        # Vectorization with robust settings
        vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min(min_df, max(2, len(documents) // 10)),  # Adaptive min_df
            max_df=0.95,
            stop_words="english",
        )

        # Create document-term matrix
        dtm = vectorizer.fit_transform(documents)
        feature_names = vectorizer.get_feature_names_out()

        logging.info(f"Document-term matrix shape: {dtm.shape}")
        logging.info(f"Number of features: {len(feature_names)}")

        # LDA with robust parameters
        lda_model = LatentDirichletAllocation(
            n_components=num_topics,
            random_state=42,
            learning_method="online",
            learning_offset=50.0,
            max_iter=n_iterations,
            doc_topic_prior=None,  # Let scikit-learn auto-estimate
            topic_word_prior=None,  # Let scikit-learn auto-estimate
        )

        # Fit LDA model
        doc_topics = lda_model.fit_transform(dtm)

        # Add logging of results
        logging.info("Topic modeling completed successfully")
        logging.info(f"Document-topic matrix shape: {doc_topics.shape}")

        return lda_model, vectorizer, doc_topics

    except Exception as e:
        logging.error(f"Topic modeling failed: {e}", exc_info=True)
        raise


def is_response(row: pd.Series) -> bool:
    """
    Check if a document is a response based on its metadata and content
    """
    try:
        # Check PDF types first (most reliable)
        for i in range(1, 5):  # Check PDF_1 to PDF_4
            pdf_type = str(row.get(f"PDF_{i}_Type", "")).lower()
            if pdf_type == "response":
                return True

        # Check PDF names as backup
        for i in range(1, 5):
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                return True

        # Check title and content as final fallback
        title = str(row.get("Title", "")).lower()
        if any(word in title for word in ["response", "reply", "answered"]):
            return True

        content = str(row.get("Content", "")).lower()
        return any(
            phrase in content
            for phrase in [
                "in response to",
                "responding to",
                "reply to",
                "response to",
                "following the regulation 28",
                "following receipt of the regulation 28",
            ]
        )

    except Exception as e:
        logging.error(f"Error checking response type: {e}")
        return False


def normalize_category(category: str) -> str:
    """Normalize category string for consistent matching"""
    if not category:
        return ""
    # Convert to lowercase and remove extra whitespace
    normalized = " ".join(str(category).lower().split())
    # Remove common separators and special characters
    normalized = re.sub(r"[,;|•·⋅‣⁃▪▫–—-]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def match_category(category: str, standard_categories: List[str]) -> Optional[str]:
    """Match a category against standard categories with fuzzy matching"""
    if not category:
        return None

    normalized_category = normalize_category(category)
    normalized_standards = {normalize_category(cat): cat for cat in standard_categories}

    # Try exact match first
    if normalized_category in normalized_standards:
        return normalized_standards[normalized_category]

    # Try partial matching
    for norm_std, original_std in normalized_standards.items():
        if normalized_category in norm_std or norm_std in normalized_category:
            return original_std

    return category  # Return original if no match found


def extract_categories(category_text: str, standard_categories: List[str]) -> List[str]:
    """Extract and normalize categories from raw text"""
    if not category_text:
        return []

    # Replace common separators with a standard one
    category_text = re.sub(r"\s*[|,;]\s*", "|", category_text)
    category_text = re.sub(r"[•·⋅‣⁃▪▫–—-]\s*", "|", category_text)
    category_text = re.sub(r"\s{2,}", "|", category_text)
    category_text = re.sub(r"\n+", "|", category_text)

    # Split and clean categories
    raw_categories = category_text.split("|")
    cleaned_categories = []

    for cat in raw_categories:
        cleaned_cat = clean_text(cat).strip()
        if cleaned_cat and not re.match(r"^[\s|,;]+$", cleaned_cat):
            matched_cat = match_category(cleaned_cat, standard_categories)
            if matched_cat:
                cleaned_categories.append(matched_cat)

    # Remove duplicates while preserving order
    seen = set()
    return [
        x
        for x in cleaned_categories
        if not (normalize_category(x) in seen or seen.add(normalize_category(x)))
    ]

######################
def filter_by_categories(df: pd.DataFrame, selected_categories: List[str]) -> pd.DataFrame:
    """
    Filter DataFrame by categories with case-insensitive matching
    
    Args:
        df: DataFrame containing 'categories' column
        selected_categories: List of categories to filter by
        
    Returns:
        Filtered DataFrame
    """
    if not selected_categories:
        return df
    
    # Convert selected categories to lowercase for case-insensitive comparison
    selected_cats_lower = [cat.lower().strip() for cat in selected_categories if cat]
    
    # Handle both string and list categories
    if "categories" in df.columns:
        # Check if we can determine the data type based on first non-null value
        first_valid_idx = df["categories"].first_valid_index()
        if first_valid_idx is not None:
            first_valid_value = df["categories"].loc[first_valid_idx]
            
            if isinstance(first_valid_value, list):
                # List case - with case-insensitive comparison
                filtered_df = df[
                    df["categories"].apply(
                        lambda x: isinstance(x, list) and any(
                            cat.lower().strip() in selected_cats_lower or
                            any(selected.lower() in cat.lower().strip() for selected in selected_cats_lower)
                            for cat in x if cat and isinstance(cat, str)
                        )
                    )
                ]
                return filtered_df
        
        # String case (default) or mixed types - with case-insensitive comparison
        filtered_df = df[
            df["categories"]
            .fillna("")
            .astype(str)
            .str.lower()  # Convert to lowercase for case-insensitive matching
            .apply(lambda x: any(
                selected.lower() in x or  # Selected category appears in string
                any(cat.lower().strip() in selected.lower() for cat in x.split(','))  # Category part appears in selected
                for selected in selected_cats_lower
            ))
        ]
        return filtered_df
    
    return df  # Return original if no categories column

#####
def filter_by_areas(df: pd.DataFrame, selected_areas: List[str]) -> pd.DataFrame:
    if not selected_areas:
        return df

    # Normalize both selected areas and dataframe areas
    df_normalized = df.copy()
    df_normalized['coroner_area_norm'] = df['coroner_area'].str.lower().str.strip()
    
    # Normalize selected areas
    selected_areas_norm = [str(area).lower().strip() for area in selected_areas]

    # Create a mask for matching
    mask = df_normalized['coroner_area_norm'].apply(
        lambda x: any(area in x or x in area for area in selected_areas_norm)
    )

    return df[mask]


def filter_by_coroner_names(df: pd.DataFrame, selected_names: List[str]) -> pd.DataFrame:
    if not selected_names:
        return df

    # Normalize both selected names and dataframe names
    df_normalized = df.copy()
    df_normalized['coroner_name_norm'] = df['coroner_name'].str.lower().str.strip()
    
    # Normalize selected names
    selected_names_norm = [str(name).lower().strip() for name in selected_names]

    # Create a mask for matching
    mask = df_normalized['coroner_name_norm'].apply(
        lambda x: any(name in x or x in name for name in selected_names_norm)
    )

    return df[mask]

def filter_by_document_type(df: pd.DataFrame, doc_types: List[str]) -> pd.DataFrame:
    """
    Filter DataFrame based on document types
    """
    if not doc_types:
        return df

    filtered_df = df.copy()
    is_response_mask = filtered_df.apply(is_response, axis=1)

    if len(doc_types) == 1:
        if "Response" in doc_types:
            return filtered_df[is_response_mask]
        elif "Report" in doc_types:
            return filtered_df[~is_response_mask]

    return filtered_df


def extract_topic_insights(lda_model, vectorizer, doc_topics, data: pd.DataFrame):
    """Extract insights from topic modeling results with improved error handling"""
    try:
        # Get feature names and initialize results
        feature_names = vectorizer.get_feature_names_out()
        topics_data = []

        # Ensure we have valid data
        valid_data = data[data["Content"].notna()].copy()
        if len(valid_data) == 0:
            raise ValueError("No valid documents found in dataset")

        # Calculate document frequencies with error handling
        doc_freq = {}
        for doc in valid_data["Content"]:
            try:
                words = set(clean_text_for_modeling(str(doc)).split())
                for word in words:
                    doc_freq[word] = doc_freq.get(word, 0) + 1
            except Exception as e:
                logging.warning(f"Error processing document: {str(e)}")
                continue

        # Process each topic
        for idx, topic in enumerate(lda_model.components_):
            try:
                # Get top words
                top_word_indices = topic.argsort()[: -50 - 1 : -1]
                topic_words = []

                for i in top_word_indices:
                    word = feature_names[i]
                    if len(word) > 1:
                        weight = float(topic[i])
                        topic_words.append(
                            {
                                "word": word,
                                "weight": weight,
                                "count": doc_freq.get(word, 0),
                                "documents": doc_freq.get(word, 0),
                            }
                        )

                # Get representative documents
                doc_scores = doc_topics[:, idx]
                top_doc_indices = doc_scores.argsort()[:-11:-1]

                related_docs = []
                for doc_idx in top_doc_indices:
                    if doc_scores[doc_idx] > 0.01:  # At least 1% relevance
                        if doc_idx < len(valid_data):
                            doc_row = valid_data.iloc[doc_idx]
                            doc_content = str(doc_row.get("Content", ""))

                            related_docs.append(
                                {
                                    "title": doc_row.get("Title", ""),
                                    "date": doc_row.get("date_of_report", ""),
                                    "relevance": float(doc_scores[doc_idx]),
                                    "summary": doc_content[:300] + "..."
                                    if len(doc_content) > 300
                                    else doc_content,
                                }
                            )

                # Generate topic description
                meaningful_words = [word["word"] for word in topic_words[:5]]
                label = " & ".join(meaningful_words[:3]).title()

                topic_data = {
                    "id": idx,
                    "label": label,
                    "description": f"Topic frequently mentions: {', '.join(meaningful_words)}",
                    "words": topic_words,
                    "representativeDocs": related_docs,
                    "prevalence": round((doc_scores > 0.05).mean() * 100, 1),
                }

                topics_data.append(topic_data)

            except Exception as e:
                logging.error(f"Error processing topic {idx}: {str(e)}")
                continue

        if not topics_data:
            raise ValueError("No valid topics could be extracted")

        return topics_data

    except Exception as e:
        logging.error(f"Error extracting topic insights: {str(e)}", exc_info=True)
        raise Exception(f"Failed to extract topic insights: {str(e)}")


def display_topic_analysis(topics_data):
    """Display topic analysis results"""
    for topic in topics_data:
        st.markdown(f"## Topic {topic['id'] + 1}: {topic['label']}")
        st.markdown(f"**Prevalence:** {topic['prevalence']}% of documents")
        st.markdown(f"**Description:** {topic['description']}")

        # Display key terms
        st.markdown("### Key Terms")
        terms_data = pd.DataFrame(topic["words"])
        if not terms_data.empty:
            st.dataframe(
                terms_data,
                column_config={
                    "word": st.column_config.TextColumn("Term"),
                    "weight": st.column_config.NumberColumn("Weight", format="%.4f"),
                    "count": st.column_config.NumberColumn("Document Count"),
                },
                hide_index=True,
            )

        # Display representative documents
        st.markdown("### Representative Documents")
        for doc in topic["representativeDocs"]:
            with st.expander(f"{doc['title']} (Relevance: {doc['relevance']:.2%})"):
                st.markdown(f"**Date:** {doc['date']}")
                st.markdown(doc["summary"])

        st.markdown("---")


# Initialize NLTK resources
def initialize_nltk():
    """Initialize required NLTK resources with error handling"""
    try:
        resources = ["punkt", "stopwords", "averaged_perceptron_tagger"]
        for resource in resources:
            try:
                if resource == "punkt":
                    nltk.data.find("tokenizers/punkt")
                elif resource == "stopwords":
                    nltk.data.find("corpora/stopwords")
                elif resource == "averaged_perceptron_tagger":
                    nltk.data.find("taggers/averaged_perceptron_tagger")
            except LookupError:
                nltk.download(resource)
    except Exception as e:
        logging.error(f"Error initializing NLTK resources: {e}")
        raise


def perform_semantic_clustering(
    data: pd.DataFrame,
    min_cluster_size: int = 3,
    max_features: int = 5000,
    min_df: float = 0.01,
    max_df: float = 0.95,
    similarity_threshold: float = 0.3,
) -> Dict:
    """
    Perform semantic clustering with improved cluster selection
    """
    try:
        # Initialize NLTK resources
        initialize_nltk()

        # Validate input data
        if "Content" not in data.columns:
            raise ValueError("Input data must contain 'Content' column")

        processed_texts = data["Content"].apply(clean_text_for_modeling)
        valid_mask = processed_texts.notna() & (processed_texts != "")
        processed_texts = processed_texts[valid_mask]

        if len(processed_texts) == 0:
            raise ValueError("No valid text content found after preprocessing")

        # Keep the original data for display
        display_data = data[valid_mask].copy()

        # Calculate optimal parameters based on dataset size
        n_docs = len(processed_texts)
        min_clusters = max(2, min(3, n_docs // 20))  # More conservative minimum
        max_clusters = max(3, min(8, n_docs // 10))  # More conservative maximum

        # Get vectorization parameters from session state
        vectorizer_type = st.session_state.get("vectorizer_type", "tfidf")
        vectorizer_params = {}

        if vectorizer_type == "bm25":
            vectorizer_params.update(
                {
                    "k1": st.session_state.get("bm25_k1", 1.5),
                    "b": st.session_state.get("bm25_b", 0.75),
                }
            )
        elif vectorizer_type == "weighted":
            vectorizer_params.update(
                {
                    "tf_scheme": st.session_state.get("tf_scheme", "raw"),
                    "idf_scheme": st.session_state.get("idf_scheme", "smooth"),
                }
            )

        # Create the vectorizer
        vectorizer = get_vectorizer(
            vectorizer_type=vectorizer_type,
            max_features=max_features,
            min_df=max(min_df, 3 / len(processed_texts)),
            max_df=min(max_df, 0.7),
            **vectorizer_params,
        )

        # Create document vectors
        tfidf_matrix = vectorizer.fit_transform(processed_texts)
        feature_names = vectorizer.get_feature_names_out()

        # Find optimal number of clusters
        best_n_clusters, best_labels = find_optimal_clusters(
            tfidf_matrix,
            min_clusters=min_clusters,
            max_clusters=max_clusters,
            min_cluster_size=min_cluster_size,
        )

        # Calculate final clustering quality
        silhouette_avg = silhouette_score(
            tfidf_matrix.toarray(), best_labels, metric="euclidean"
        )

        # Calculate similarities using similarity threshold
        similarity_matrix = cosine_similarity(tfidf_matrix)
        similarity_matrix[similarity_matrix < similarity_threshold] = 0

        # Extract cluster information
        clusters = []
        for cluster_id in range(best_n_clusters):
            cluster_indices = np.where(best_labels == cluster_id)[0]

            # Skip if cluster is too small
            if len(cluster_indices) < min_cluster_size:
                continue

            # Calculate cluster terms
            cluster_tfidf = tfidf_matrix[cluster_indices].toarray()
            centroid = np.mean(cluster_tfidf, axis=0)

            # Get important terms with improved distinctiveness
            term_scores = []
            for idx, score in enumerate(centroid):
                if score > 0:
                    term = feature_names[idx]
                    cluster_freq = np.mean(cluster_tfidf[:, idx] > 0)
                    total_freq = np.mean(tfidf_matrix[:, idx].toarray() > 0)
                    distinctiveness = cluster_freq / (total_freq + 1e-10)

                    term_scores.append(
                        {
                            "term": term,
                            "score": float(score * distinctiveness),
                            "cluster_frequency": float(cluster_freq),
                            "total_frequency": float(total_freq),
                        }
                    )

            term_scores.sort(key=lambda x: x["score"], reverse=True)
            top_terms = term_scores[:20]

            # Get representative documents
            doc_similarities = []
            for idx in cluster_indices:
                doc_vector = tfidf_matrix[idx].toarray().flatten()
                sim_to_centroid = cosine_similarity(
                    doc_vector.reshape(1, -1), centroid.reshape(1, -1)
                )[0][0]

                doc_info = {
                    "title": display_data.iloc[idx]["Title"],
                    "date": display_data.iloc[idx]["date_of_report"],
                    "similarity": float(sim_to_centroid),
                    "summary": display_data.iloc[idx]["Content"][:500],
                }
                doc_similarities.append((idx, sim_to_centroid, doc_info))

            # Sort by similarity and get representative docs
            doc_similarities.sort(key=lambda x: x[1], reverse=True)
            representative_docs = [item[2] for item in doc_similarities]

            # Calculate cluster cohesion
            cluster_similarities = similarity_matrix[cluster_indices][
                :, cluster_indices
            ]
            cohesion = float(np.mean(cluster_similarities))

            clusters.append(
                {
                    "id": len(clusters),
                    "size": len(cluster_indices),
                    "cohesion": cohesion,
                    "terms": top_terms,
                    "documents": representative_docs,
                    "balance_ratio": max(
                        len(cluster_indices)
                        for cluster_indices in [
                            np.where(best_labels == i)[0]
                            for i in range(best_n_clusters)
                        ]
                    )
                    / min(
                        len(cluster_indices)
                        for cluster_indices in [
                            np.where(best_labels == i)[0]
                            for i in range(best_n_clusters)
                        ]
                    ),
                }
            )

        # Add cluster quality metrics to results
        metrics = {
            "silhouette_score": float(silhouette_avg),
            "calinski_score": float(
                calinski_harabasz_score(tfidf_matrix.toarray(), best_labels)
            ),
            "davies_score": float(
                davies_bouldin_score(tfidf_matrix.toarray(), best_labels)
            ),
            "balance_ratio": float(
                max(len(c["documents"]) for c in clusters)
                / min(len(c["documents"]) for c in clusters)
            ),
        }

        return {
            "n_clusters": len(clusters),
            "total_documents": len(processed_texts),
            "silhouette_score": float(silhouette_avg),
            "clusters": clusters,
            "vectorizer_type": vectorizer_type,
            "quality_metrics": metrics,
        }

    except Exception as e:
        logging.error(f"Error in semantic clustering: {e}", exc_info=True)
        raise


def create_document_identifier(row: pd.Series) -> str:
    """Create a unique identifier for a document based on its title and reference number"""
    title = str(row.get("Title", "")).strip()
    ref = str(row.get("ref", "")).strip()
    deceased = str(row.get("deceased_name", "")).strip()

    # Combine multiple fields to create a unique identifier
    identifier = f"{title}_{ref}_{deceased}"
    return identifier


def deduplicate_documents(data: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate documents while preserving unique entries"""
    # Create unique identifiers
    data["doc_id"] = data.apply(create_document_identifier, axis=1)

    # Keep first occurrence of each unique document
    deduped_data = data.drop_duplicates(subset=["doc_id"])

    # Drop the temporary identifier column
    deduped_data = deduped_data.drop(columns=["doc_id"])

    return deduped_data


def format_date_uk(date_obj):
    """Convert datetime object to UK date format string"""
    if pd.isna(date_obj):
        return ""
    try:
        if isinstance(date_obj, str):
            # Try to parse string to datetime first
            date_obj = pd.to_datetime(date_obj)
        return date_obj.strftime("%d/%m/%Y")
    except:
        return str(date_obj)


def generate_extractive_summary(documents, max_length=500):
    """Generate extractive summary from cluster documents with traceability"""
    try:
        # Combine all document texts with source tracking
        all_sentences = []
        for doc in documents:
            sentences = sent_tokenize(doc["summary"])
            for sent in sentences:
                all_sentences.append(
                    {
                        "text": sent,
                        "source": doc["title"],
                        "date": format_date_uk(doc["date"]),  # Format date here
                    }
                )

        # Calculate sentence importance using TF-IDF
        vectorizer = TfidfVectorizer(stop_words="english")
        tfidf_matrix = vectorizer.fit_transform([s["text"] for s in all_sentences])

        # Calculate sentence scores
        sentence_scores = []
        for idx, sentence in enumerate(all_sentences):
            score = np.mean(tfidf_matrix[idx].toarray())
            sentence_scores.append((score, sentence))

        # Sort by importance and select top sentences
        sentence_scores.sort(reverse=True)
        summary_length = 0
        summary_sentences = []

        for score, sentence in sentence_scores:
            if summary_length + len(sentence["text"]) <= max_length:
                summary_sentences.append(
                    {
                        "text": sentence["text"],
                        "source": sentence["source"],
                        "date": sentence["date"],
                        "score": float(score),
                    }
                )
                summary_length += len(sentence["text"])
            else:
                break

        return summary_sentences

    except Exception as e:
        logging.error(f"Error in extractive summarization: {e}")
        return []


def generate_abstractive_summary(cluster_terms, documents, max_length=500):
    """Generate abstractive summary from cluster information with improved date handling"""
    try:
        # Extract key themes from terms
        top_themes = [term["term"] for term in cluster_terms[:5]]

        # Get document dates and format them with proper sorting
        dates = []
        for doc in documents:
            try:
                if doc["date"]:
                    date_obj = pd.to_datetime(doc["date"])
                    dates.append(date_obj)
            except:
                continue

        if dates:
            start_date = min(dates).strftime("%d/%m/%Y")
            end_date = max(dates).strftime("%d/%m/%Y")
            date_range = f"from {start_date} to {end_date}"
        else:
            date_range = ""

        # Extract key themes with better formatting
        main_themes = ", ".join(top_themes[:-1])
        if main_themes:
            themes_text = f"{main_themes} and {top_themes[-1]}"
        else:
            themes_text = top_themes[0] if top_themes else ""

        # Build better structured summary
        summary = f"This cluster contains {len(documents)} documents "
        if date_range:
            summary += f"{date_range} "
        summary += f"focused on {themes_text}. "

        # Add key patterns with improved statistics
        term_patterns = []
        for term in cluster_terms[5:8]:  # Get next 3 terms after main themes
            if term["cluster_frequency"] > 0:
                freq = term["cluster_frequency"] * 100
                # Add context based on frequency
                if freq > 75:
                    context = "very commonly"
                elif freq > 50:
                    context = "frequently"
                elif freq > 25:
                    context = "sometimes"
                else:
                    context = "occasionally"
                term_patterns.append(
                    f"{term['term']} ({context} appearing in {freq:.0f}% of documents)"
                )

        if term_patterns:
            summary += f"Common patterns include {', '.join(term_patterns)}. "

        # Add cluster distinctiveness if available
        if any(term["total_frequency"] < 0.5 for term in cluster_terms[:5]):
            distinctive_terms = [
                term["term"]
                for term in cluster_terms[:5]
                if term["total_frequency"] < 0.5
            ]
            if distinctive_terms:
                summary += f"This cluster is particularly distinctive in its discussion of {', '.join(distinctive_terms)}."

        # Truncate to max length while preserving complete sentences
        if len(summary) > max_length:
            summary = summary[:max_length]
            last_period = summary.rfind(".")
            if last_period > 0:
                summary = summary[: last_period + 1]

        return summary

    except Exception as e:
        logging.error(f"Error in abstractive summarization: {e}")
        return "Error generating summary"


def get_optimal_clustering_params(num_docs: int) -> Dict[str, int]:
    """Calculate optimal clustering parameters based on dataset size"""

    # Base parameters
    params = {
        "min_cluster_size": 2,  # Minimum starting point
        "max_features": 5000,  # Maximum vocabulary size
        "min_docs": 2,  # Minimum document frequency
        "max_docs": None,  # Maximum document frequency (will be calculated)
    }

    # Adjust minimum cluster size based on dataset size
    if num_docs < 10:
        params["min_cluster_size"] = 2
    elif num_docs < 20:
        params["min_cluster_size"] = 3
    elif num_docs < 50:
        params["min_cluster_size"] = 4
    else:
        params["min_cluster_size"] = 5

    # Adjust document frequency bounds
    params["min_docs"] = max(2, int(num_docs * 0.05))  # At least 5% of documents
    params["max_docs"] = min(
        int(num_docs * 0.95),  # No more than 95% of documents
        num_docs - params["min_cluster_size"],  # Leave room for at least one cluster
    )

    # Adjust feature count based on dataset size
    if num_docs < 20:
        params["max_features"] = 2000
    elif num_docs < 50:
        params["max_features"] = 3000
    elif num_docs < 100:
        params["max_features"] = 4000
    else:
        params["max_features"] = 5000

    return params


def display_cluster_analysis(cluster_results: Dict) -> None:
    """Display comprehensive cluster analysis results with quality metrics"""
    try:
        st.subheader("Document Clustering Analysis")

        # Overview metrics in two rows
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Number of Clusters", cluster_results["n_clusters"])
        with col2:
            st.metric("Total Documents", cluster_results["total_documents"])
        with col3:
            st.metric(
                "Average Cluster Size",
                round(
                    cluster_results["total_documents"] / cluster_results["n_clusters"],
                    1,
                ),
            )

        # Quality metrics
        st.subheader("Clustering Quality Metrics")
        metrics = cluster_results["quality_metrics"]

        qual_col1, qual_col2, qual_col3, qual_col4 = st.columns(4)

        with qual_col1:
            st.metric(
                "Silhouette Score",
                f"{metrics['silhouette_score']:.3f}",
                help="Measures how similar an object is to its own cluster compared to other clusters. Range: [-1, 1], higher is better.",
            )

        with qual_col2:
            st.metric(
                "Calinski-Harabasz Score",
                f"{metrics['calinski_score']:.0f}",
                help="Ratio of between-cluster to within-cluster dispersion. Higher is better.",
            )

        with qual_col3:
            st.metric(
                "Davies-Bouldin Score",
                f"{metrics['davies_score']:.3f}",
                help="Average similarity measure of each cluster with its most similar cluster. Lower is better.",
            )

        with qual_col4:
            st.metric(
                "Balance Ratio",
                f"{metrics['balance_ratio']:.1f}",
                help="Ratio of largest to smallest cluster size. Closer to 1 is better.",
            )

        # Display each cluster
        for cluster in cluster_results["clusters"]:
            with st.expander(
                f"Cluster {cluster['id']+1} ({cluster['size']} documents)",
                expanded=True,
            ):
                # Cluster metrics
                met_col1, met_col2 = st.columns(2)
                with met_col1:
                    st.metric(
                        "Cohesion Score",
                        f"{cluster['cohesion']:.3f}",
                        help="Average similarity between documents in the cluster",
                    )
                with met_col2:
                    st.metric(
                        "Size Percentage",
                        f"{(cluster['size'] / cluster_results['total_documents'] * 100):.1f}%",
                        help="Percentage of total documents in this cluster",
                    )

                # Terms analysis
                st.markdown("#### Key Terms")
                terms_df = pd.DataFrame(
                    [
                        {
                            "Term": term["term"],
                            "Frequency": f"{term['cluster_frequency']*100:.1f}%",
                            "Distinctiveness": f"{term['score']:.3f}",
                        }
                        for term in cluster["terms"][:10]
                    ]
                )
                st.dataframe(terms_df, hide_index=True)

                # Representative documents with formatted dates
                st.markdown("#### Representative Documents")
                for doc in cluster["documents"]:
                    st.markdown(
                        f"**{doc['title']}** (Similarity: {doc['similarity']:.2f})"
                    )
                    st.markdown(f"**Date**: {format_date_uk(doc['date'])}")
                    st.markdown(f"**Summary**: {doc['summary'][:300]}...")
                    st.markdown("---")

    except Exception as e:
        st.error(f"Error displaying cluster analysis: {str(e)}")
        logging.error(f"Display error: {str(e)}", exc_info=True)


def find_optimal_clusters(
    tfidf_matrix: sp.csr_matrix,
    min_clusters: int = 2,
    max_clusters: int = 10,
    min_cluster_size: int = 3,
) -> Tuple[int, np.ndarray]:
    """Find optimal number of clusters with relaxed constraints"""

    best_score = -1
    best_n_clusters = min_clusters
    best_labels = None

    # Store metrics for each clustering attempt
    metrics = []

    # Try different numbers of clusters
    for n_clusters in range(min_clusters, max_clusters + 1):
        try:
            # Perform clustering
            clustering = AgglomerativeClustering(
                n_clusters=n_clusters, metric="euclidean", linkage="ward"
            )

            labels = clustering.fit_predict(tfidf_matrix.toarray())

            # Calculate cluster sizes
            cluster_sizes = np.bincount(labels)

            # Skip if any cluster is too small
            if min(cluster_sizes) < min_cluster_size:
                continue

            # Calculate balance ratio (smaller is better)
            balance_ratio = max(cluster_sizes) / min(cluster_sizes)

            # Skip only if clusters are extremely imbalanced
            if balance_ratio > 10:  # Relaxed from 5 to 10
                continue

            # Calculate clustering metrics
            sil_score = silhouette_score(
                tfidf_matrix.toarray(), labels, metric="euclidean"
            )

            # Simplified scoring focused on silhouette and basic balance
            combined_score = sil_score * (
                1 - (balance_ratio / 20)
            )  # Relaxed balance penalty

            metrics.append(
                {
                    "n_clusters": n_clusters,
                    "silhouette": sil_score,
                    "balance_ratio": balance_ratio,
                    "combined_score": combined_score,
                    "labels": labels,
                }
            )

        except Exception as e:
            logging.warning(f"Error trying {n_clusters} clusters: {str(e)}")
            continue

    # If no configurations met the strict criteria, try to find the best available
    if not metrics:
        # Try again with minimal constraints
        for n_clusters in range(min_clusters, max_clusters + 1):
            try:
                clustering = AgglomerativeClustering(
                    n_clusters=n_clusters, metric="euclidean", linkage="ward"
                )

                labels = clustering.fit_predict(tfidf_matrix.toarray())
                sil_score = silhouette_score(
                    tfidf_matrix.toarray(), labels, metric="euclidean"
                )

                if sil_score > best_score:
                    best_score = sil_score
                    best_n_clusters = n_clusters
                    best_labels = labels

            except Exception as e:
                continue

        if best_labels is None:
            # If still no valid configuration, use minimum number of clusters
            clustering = AgglomerativeClustering(
                n_clusters=min_clusters, metric="euclidean", linkage="ward"
            )
            best_labels = clustering.fit_predict(tfidf_matrix.toarray())
            best_n_clusters = min_clusters
    else:
        # Use the best configuration from metrics
        best_metric = max(metrics, key=lambda x: x["combined_score"])
        best_n_clusters = best_metric["n_clusters"]
        best_labels = best_metric["labels"]

    return best_n_clusters, best_labels


def export_cluster_results(cluster_results: Dict) -> bytes:
    """Export cluster results with proper timestamp handling"""
    output = io.BytesIO()

    # Prepare export data with timestamp conversion
    export_data = {
        "metadata": {
            "total_documents": cluster_results["total_documents"],
            "number_of_clusters": cluster_results["n_clusters"],
            "silhouette_score": cluster_results["silhouette_score"],
        },
        "clusters": [],
    }

    # Convert cluster data
    for cluster in cluster_results["clusters"]:
        # Create a copy of cluster with converted documents
        cluster_export = cluster.copy()
        for doc in cluster_export["documents"]:
            # Ensure date is a string
            doc["date"] = str(doc["date"])

        export_data["clusters"].append(cluster_export)

    # Write JSON to BytesIO
    json.dump(export_data, io.TextIOWrapper(output, encoding="utf-8"), indent=2)
    output.seek(0)

    return output.getvalue()


def validate_data_state():
    """Check if valid data exists in session state"""
    return (
        "current_data" in st.session_state
        and st.session_state.current_data is not None
        and not st.session_state.current_data.empty
    )


def validate_model_state():
    """Check if valid topic model exists in session state"""
    return (
        "topic_model" in st.session_state and st.session_state.topic_model is not None
    )


def handle_no_data_state(section):
    """Handle state when no data is available"""
    st.warning("No data available. Please scrape reports or upload a file first.")
    uploaded_file = st.file_uploader(
        "Upload existing data file", type=["csv", "xlsx"], key=f"{section}_uploader"
    )

    if uploaded_file:
        try:
            df = (
                pd.read_csv(uploaded_file)
                if uploaded_file.name.endswith(".csv")
                else pd.read_excel(uploaded_file)
            )
            df = process_scraped_data(df)
            st.session_state.current_data = df
            st.rerun()
        except Exception as e:
            st.error(f"Error loading file: {str(e)}")


def handle_no_model_state():
    """Handle state when no topic model is available"""
    st.warning("Please run the clustering analysis first to view summaries.")
    if st.button("Go to Topic Modeling"):
        st.session_state.current_tab = "🔬 Topic Modeling"
        st.rerun()


def handle_error(error):
    """Handle application errors"""
    st.error("An error occurred")
    st.error(str(error))
    logging.error(f"Application error: {error}", exc_info=True)

    with st.expander("Error Details"):
        st.code(traceback.format_exc())

    st.warning("Recovery options:")
    st.markdown(
        """
    1. Clear data and restart
    2. Upload different data
    3. Check filter settings
    """
    )


################
def truncate_axis_labels(fig, max_length=30, axis='both', truncate_func=None):
    """
    Truncate x and/or y axis labels in a Plotly figure
    
    Args:
        fig (go.Figure): Plotly figure object
        max_length (int): Maximum length of label before truncation
        axis (str): Which axis to truncate ('x', 'y', or 'both')
        truncate_func (callable, optional): Custom truncation function. 
                      If None, will use a basic truncation method.
    
    Returns:
        go.Figure: Figure with truncated axis labels
    """
    # Default truncation function if none provided
    def default_truncate(text, max_len):
        """Basic truncation method"""
        if not isinstance(text, str):
            return str(text)
        
        if len(text) > max_len:
            return text[:max_len-3] + "..."
        return text
    
    # Use the provided truncation function or default
    truncate = truncate_func or default_truncate
    
    # Truncate x-axis labels if specified
    if axis in ['x', 'both']:
        # Check if the figure uses Express or Graph Objects
        if hasattr(fig, 'layout') and 'xaxis' in fig.layout:
            # For Graph Objects figures
            if hasattr(fig.layout.xaxis, 'ticktext') and fig.layout.xaxis.ticktext:
                fig.layout.xaxis.ticktext = [
                    truncate(text, max_length) 
                    for text in fig.layout.xaxis.ticktext
                ]
        
        # For Express figures or figures with layout update
        if hasattr(fig, 'update_xaxes'):
            # Get current tick text, fallback to empty list if not available
            current_ticktext = getattr(fig.layout.xaxis, 'get_ticktext', lambda: [])()
            
            # If no current tick text, we might need to rebuild from the existing data
            if not current_ticktext and hasattr(fig, 'data') and fig.data and hasattr(fig.data[0], 'x'):
                current_ticktext = fig.data[0].x
            
            fig.update_xaxes(
                ticktext=[truncate(str(text), max_length) for text in current_ticktext],
                tickmode='array'
            )
    
    # Truncate y-axis labels if specified
    if axis in ['y', 'both']:
        # Check if the figure uses Express or Graph Objects
        if hasattr(fig, 'layout') and 'yaxis' in fig.layout:
            # For Graph Objects figures
            if hasattr(fig.layout.yaxis, 'ticktext') and fig.layout.yaxis.ticktext:
                fig.layout.yaxis.ticktext = [
                    truncate(text, max_length) 
                    for text in fig.layout.yaxis.ticktext
                ]
        
        # For Express figures or figures with layout update
        if hasattr(fig, 'update_yaxes'):
            # Get current tick text, fallback to empty list if not available
            current_ticktext = getattr(fig.layout.yaxis, 'get_ticktext', lambda: [])()
            
            # If no current tick text, we might need to rebuild from the existing data
            if not current_ticktext and hasattr(fig, 'data') and fig.data and hasattr(fig.data[0], 'y'):
                current_ticktext = fig.data[0].y
            
            fig.update_yaxes(
                ticktext=[truncate(str(text), max_length) for text in current_ticktext],
                tickmode='array'
            )
    
    return fig

# Optional: You can add a helper function to apply truncation to charts
def apply_chart_truncation(fig, truncate_func=None, max_length=30, axis='both'):
    """
    Convenience wrapper to apply truncation with a custom truncation function
    
    Args:
        fig (go.Figure): Plotly figure to modify
        truncate_func (callable, optional): Custom truncation function
        max_length (int): Maximum label length
        axis (str): Which axis to truncate
    
    Returns:
        go.Figure: Modified figure with truncated labels
    """
    return truncate_axis_labels(
        fig, 
        max_length=max_length, 
        axis=axis, 
        truncate_func=truncate_func
    )

###############Add this helper function to your app.py file

def truncate_text(text, max_length=30):
    """
    Improved function to handle long text for chart display by breaking into lines
    instead of simple truncation with ellipses
    
    Args:
        text: String to format
        max_length: Maximum length per line
        
    Returns:
        Text with line breaks inserted at appropriate word boundaries
    """
    if not text or len(text) <= max_length:
        return text
    
    # For theme names with ":" in them (framework:theme format)
    if ":" in text:
        parts = text.split(":", 1)
        framework = parts[0].strip()
        theme = parts[1].strip()
        
        # For theme part, break it into lines rather than truncate
        if len(theme) > max_length - len(framework) - 2:  # -2 for ": "
            # Process the theme part with word-aware line breaking
            words = theme.split()
            processed_theme = []
            current_line = []
            current_length = 0
            
            for word in words:
                # If adding this word keeps us under the limit
                if current_length + len(word) + (1 if current_line else 0) <= max_length - len(framework) - 2:
                    current_line.append(word)
                    current_length += len(word) + (1 if current_line else 0)
                else:
                    # Line is full, start a new one
                    processed_theme.append(" ".join(current_line))
                    current_line = [word]
                    current_length = len(word)
            
            # Add the final line if any
            if current_line:
                processed_theme.append(" ".join(current_line))
            
            # If we have more than 2 lines, keep first 2 and add ellipsis
            if len(processed_theme) > 2:
                return f"{framework}: {processed_theme[0]}<br>{processed_theme[1]}..."
            else:
                # Join with line breaks
                return f"{framework}: {('<br>').join(processed_theme)}"
        
        return f"{framework}: {theme}"
    
    # For normal long strings, add line breaks at word boundaries
    words = text.split()
    lines = []
    current_line = []
    current_length = 0
    
    for word in words:
        if current_length + len(word) + (1 if current_line else 0) <= max_length:
            current_line.append(word)
            current_length += len(word) + (1 if current_line else 0)
        else:
            lines.append(" ".join(current_line))
            current_line = [word]
            current_length = len(word)
    
    if current_line:
        lines.append(" ".join(current_line))
    
    # If we have more than 2 lines, keep first 2 and add ellipsis
    if len(lines) > 2:
        return f"{lines[0]}<br>{lines[1]}..."
    
    # For plotly charts, use <br> for line breaks
    return "<br>".join(lines)



####
def save_dashboard_images_as_zip(filtered_df):
    """
    Save all dashboard visualizations as images and package them into a zip file.
    Improved version that properly generates and captures all visualizations from all tabs.
    
    Args:
        filtered_df: Filtered DataFrame containing theme analysis results
        
    Returns:
        Tuple[bytes, int]: ZIP file containing images and number of images
    """
    import io
    import zipfile
    from datetime import datetime
    import plotly.express as px
    import plotly.graph_objects as go
    import pandas as pd
    import numpy as np
    import logging
    import networkx as nx
    
    # Create a buffer for the zip file
    zip_buffer = io.BytesIO()
    
    # Create a timestamp for the filenames
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Track number of images
    image_count = 0
    
    # Create a zipfile
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Helper function to save a figure to the zip
        def add_figure_to_zip(fig, filename):
            nonlocal image_count
            try:
                # Important: Set explicit dimensions for the image export
                fig.update_layout(
                    width=1000,
                    height=700,
                    margin=dict(l=100, r=80, t=80, b=80),
                    paper_bgcolor="white",
                    plot_bgcolor="white",
                    font=dict(color="black", size=12),
                    title_font=dict(size=16, color="black"),
                    legend=dict(font=dict(size=10))
                )
                
                # Update axes for light mode export
                fig.update_xaxes(
                    title_font=dict(color="black", size=14),
                    tickfont=dict(color="black", size=10),
                    gridcolor="rgba(0,0,0,0.1)",
                    automargin=True  # Enable automargin for all exports
                )
                
                fig.update_yaxes(
                    title_font=dict(color="black", size=14),
                    tickfont=dict(color="black", size=10),
                    gridcolor="rgba(0,0,0,0.1)",
                    automargin=True  # Enable automargin for all exports
                )
                
                # Reset colorbar if present
                if hasattr(fig, 'data') and fig.data and hasattr(fig.data[0], 'colorbar'):
                    fig.update_traces(
                        colorbar=dict(
                            title=dict(text="", font=dict(color="black")),
                            tickfont=dict(color="black", size=10)
                        )
                    )
                
                # Export as PNG with higher resolution
                img_bytes = fig.to_image(format="png", scale=2, engine="kaleido")
                
                if img_bytes and len(img_bytes) > 0:
                    zip_file.writestr(filename, img_bytes)
                    image_count += 1
                    logging.info(f"Successfully added {filename} to zip")
                    return True
                else:
                    logging.warning(f"No image bytes generated for {filename}")
                    return False
            except Exception as e:
                logging.error(f"Error saving {filename}: {str(e)}")
                return False
        
        # === TAB 1: FRAMEWORK HEATMAP ===
        try:
            # Framework distribution chart
            framework_counts = filtered_df["Framework"].value_counts()
            fig = px.bar(
                x=framework_counts.index,
                y=framework_counts.values,
                labels={"x": "Framework", "y": "Count"},
                title="Framework Distribution",
                color=framework_counts.index,
                color_discrete_map={
                    "I-SIRch": "orange",
                    "House of Commons": "royalblue",
                    "Extended Analysis": "firebrick"
                }
            )
            add_figure_to_zip(fig, f"framework_distribution_{timestamp}.png")
            
            # Handle framework theme analysis by year
            if "year" in filtered_df.columns and not filtered_df["year"].isna().all():
                if filtered_df["year"].nunique() == 1:
                    # Single year case
                    year_value = filtered_df["year"].iloc[0]
                    theme_counts = filtered_df.groupby(['Framework', 'Theme']).size().reset_index(name='Count')
                    theme_counts = theme_counts.sort_values(['Framework', 'Count'], ascending=[True, False])
                    
                    # Process theme names for better display using improved function
                    theme_counts['Display_Theme'] = theme_counts['Theme'].apply(
                        lambda x: improved_truncate_text(x, max_length=40)
                    )
                    
                    # Recreate the horizontal bar chart
                    fig = px.bar(
                        theme_counts,
                        y='Display_Theme',
                        x='Count',
                        color='Framework',
                        title=f"Theme Distribution for Year {year_value}",
                        height=max(500, len(theme_counts) * 30),
                        color_discrete_map={
                            "I-SIRch": "orange",
                            "House of Commons": "royalblue",
                            "Extended Analysis": "firebrick"
                        }
                    )
                    
                    fig.update_layout(
                        xaxis_title="Number of Reports",
                        yaxis_title="Theme"
                    )
                    
                    add_figure_to_zip(fig, f"theme_distribution_single_year_{timestamp}.png")
                else:
                    # Multiple years case - recreate the heatmap with improved labels
                    # This code regenerates the framework theme heatmap for the zip file
                    
                    # Create combined framework:theme field if not already there
                    if 'Framework_Theme' not in filtered_df.columns:
                        filtered_df['Framework_Theme'] = filtered_df['Framework'] + ': ' + filtered_df['Theme']
                    
                    # Get needed data for the heatmap
                    id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
                    reports_per_year = filtered_df.groupby('year')[id_column].nunique()
                    
                    # Count unique report IDs per theme per year
                    counts = filtered_df.groupby(['year', 'Framework', 'Framework_Theme'])[id_column].nunique().reset_index()
                    counts.columns = ['year', 'Framework', 'Framework_Theme', 'Count']
                    
                    # Calculate percentages
                    counts['Total'] = counts['year'].map(reports_per_year)
                    counts['Percentage'] = (counts['Count'] / counts['Total'] * 100).round(1)
                    
                    # Get frameworks in the filtered data
                    frameworks_present = filtered_df['Framework'].unique()
                    
                    # Get top themes by framework (5 per framework)
                    top_themes = []
                    for framework in frameworks_present:
                        framework_counts = counts[counts['Framework'] == framework]
                        theme_totals = framework_counts.groupby('Framework_Theme')['Count'].sum().sort_values(ascending=False)
                        top_themes.extend(theme_totals.head(5).index.tolist())
                    
                    # Filter to top themes
                    counts = counts[counts['Framework_Theme'].isin(top_themes)]
                    
                    # Create pivot tables
                    pivot = counts.pivot_table(
                        index='Framework_Theme',
                        columns='year',
                        values='Percentage',
                        fill_value=0
                    )
                    
                    count_pivot = counts.pivot_table(
                        index='Framework_Theme',
                        columns='year',
                        values='Count',
                        fill_value=0
                    )
                    
                    # Sort by framework then by total count
                    theme_totals = counts.groupby('Framework_Theme')['Count'].sum()
                    theme_frameworks = {theme: theme.split(':')[0] for theme in theme_totals.index}
                    sorted_themes = sorted(
                        theme_totals.index,
                        key=lambda x: (theme_frameworks[x], -theme_totals[x])
                    )
                    
                    # Apply the sort order
                    pivot = pivot.reindex(sorted_themes)
                    count_pivot = count_pivot.reindex(sorted_themes)
                    
                    # Create formatted theme names
                    theme_display_data = []
                    framework_colors = {
                        "I-SIRch": "orange",
                        "House of Commons": "royalblue",
                        "Extended Analysis": "firebrick"
                    }
                    
                    # Default colors for any frameworks not specifically mapped
                    other_colors = ["forestgreen", "purple", "darkred"]
                    for i, framework in enumerate(frameworks_present):
                        if framework not in framework_colors:
                            framework_colors[framework] = other_colors[i % len(other_colors)]
                    
                    for theme in pivot.index:
                        framework = theme.split(':')[0].strip()
                        theme_name = theme.split(':', 1)[1].strip()
                        formatted_theme = improved_truncate_text(theme_name, max_length=40)
                        
                        theme_display_data.append({
                            'original': theme,
                            'clean_name': formatted_theme,
                            'framework': framework,
                            'color': framework_colors[framework]
                        })
                    
                    theme_display_df = pd.DataFrame(theme_display_data)
                    
                    # Add year count labels
                    year_labels = [f"{year}\nn={reports_per_year[year]}" for year in pivot.columns]
                    
                    # Create heatmap for export
                    fig = px.imshow(
                        pivot.values,
                        labels=dict(x="Year", y="Theme", color="% of Themes"),
                        x=year_labels,
                        y=theme_display_df['clean_name'],
                        color_continuous_scale="Blues",
                        title="Framework Theme Heatmap by Year",
                        text_auto=".1f"
                    )
                    
                    # Update layout for static export
                    fig.update_layout(
                        width=1200,
                        height=max(650, len(pivot.index) * 35),
                        margin=dict(l=300, r=60, t=80, b=80)
                    )
                    
                    # Add to zip
                    add_figure_to_zip(fig, f"framework_theme_heatmap_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating framework heatmap: {str(e)}")

        # === TAB 2: THEME DISTRIBUTION ===
        try:
            # Get top themes by count
            theme_counts = filtered_df["Theme"].value_counts().head(10)  # Use a reasonable number of themes
            
            # Use improved_truncate_text for better label formatting
            formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in theme_counts.index]
            
            # Create a bar chart with formatted theme names
            fig = px.bar(
                x=formatted_themes,
                y=theme_counts.values,
                labels={"x": "Theme", "y": "Count"},
                title="Top Themes by Occurrence",
                height=600,
                color_discrete_sequence=['#4287f5']
            )
            
            add_figure_to_zip(fig, f"theme_distribution_{timestamp}.png")
            
            # Theme by confidence
            theme_confidence = filtered_df.groupby(["Theme", "Confidence"]).size().reset_index(name="Count")
            
            # Filter for top themes only
            top_themes = theme_counts.index.tolist()
            theme_confidence = theme_confidence[theme_confidence["Theme"].isin(top_themes)]
            
            # Create a mapping dictionary for theme display names
            theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
            
            # Apply the formatting to the DataFrame
            theme_confidence["Display_Theme"] = theme_confidence["Theme"].map(theme_display_map)
            
            # Create a grouped bar chart with formatted theme names
            fig = px.bar(
                theme_confidence, 
                x="Display_Theme",
                y="Count", 
                color="Confidence",
                barmode="group",
                color_discrete_map={"High": "#4CAF50", "Medium": "#FFC107", "Low": "#F44336"},
                category_orders={
                    "Confidence": ["High", "Medium", "Low"],
                    "Display_Theme": [theme_display_map[theme] for theme in top_themes]
                },
                title="Confidence Distribution by Theme",
                height=600
            )
            
            add_figure_to_zip(fig, f"theme_confidence_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating theme distribution charts: {str(e)}")
            
        # === TAB 3: TEMPORAL ANALYSIS ===
        try:
                # Theme trends over time
                year_theme_counts = filtered_df.groupby(["year", "Theme"]).size().reset_index(name="Count")
                
                # Get top themes
                all_theme_counts = filtered_df["Theme"].value_counts()
                top_themes = all_theme_counts.head(8).index.tolist()  # Limit to 8 for readability
                
                # Filter for top themes
                year_theme_counts = year_theme_counts[year_theme_counts["Theme"].isin(top_themes)]
                
                # Create formatted theme names
                theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
                year_theme_counts["Display_Theme"] = year_theme_counts["Theme"].map(theme_display_map)
                
                # Convert year to string for categorical plotting
                year_theme_counts['year_str'] = year_theme_counts['year'].astype(str)
                
                # Create line chart for theme trends
                fig = px.line(
                    year_theme_counts,
                    x="year_str",
                    y="Count",
                    color="Display_Theme",
                    markers=True,
                    title="Theme Trends Over Time",
                    labels={"year_str": "Year", "Count": "Number of Occurrences", "Display_Theme": "Theme"}
                )
                
                add_figure_to_zip(fig, f"theme_temporal_trends_{timestamp}.png")
                
                # Create theme prevalence heatmap if multiple years
                if filtered_df["year"].nunique() > 1:
                    # Create a pivot table
                    pivot_df = year_theme_counts.pivot(index="Theme", columns="year_str", values="Count").fillna(0)
                    
                    # Convert to a normalized heatmap (percentage)
                    year_theme_totals = pivot_df.sum(axis=0)
                    normalized_pivot = pivot_df.div(year_theme_totals, axis=1) * 100
                    
                    # Format the theme names
                    formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in normalized_pivot.index]
                    
                    # Create a heatmap
                    year_order = sorted(year_theme_counts['year'].unique())
                    year_order_str = [str(y) for y in year_order]
                    
                    if len(normalized_pivot) > 0 and len(year_order_str) > 0:
                        fig = px.imshow(
                            normalized_pivot[year_order_str],
                            labels=dict(x="Year", y="Theme", color="% of Themes"),
                            x=year_order_str,
                            y=formatted_themes,
                            color_continuous_scale="YlGnBu",
                            title="Theme Prevalence by Year (%)",
                            text_auto=".1f"
                        )
                        
                        add_figure_to_zip(fig, f"theme_prevalence_heatmap_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating temporal analysis charts: {str(e)}")
        
        # === TAB 4: AREA COMPARISON ===
        try:
            if "coroner_area" in filtered_df.columns and not filtered_df["coroner_area"].isna().all():
                # Get the top areas by theme count
                area_counts = filtered_df["coroner_area"].value_counts().head(10)
                
                # Format area names
                formatted_areas = [improved_truncate_text(area, max_length=40) for area in area_counts.index]
                
                # Create a bar chart of top areas
                fig = px.bar(
                    x=formatted_areas,
                    y=area_counts.values,
                    labels={"x": "Coroner Area", "y": "Count"},
                    title="Theme Identifications by Coroner Area",
                    color_discrete_sequence=['#ff9f40']
                )
                
                add_figure_to_zip(fig, f"coroner_area_distribution_{timestamp}.png")
                
                # Create area-theme heatmap
                top_areas = area_counts.index.tolist()
                top_themes = filtered_df["Theme"].value_counts().head(8).index.tolist()  # Limit to 8 themes
                
                # Calculate area theme data
                area_theme_data = []
                for area in top_areas:
                    area_df = filtered_df[filtered_df["coroner_area"] == area]
                    area_totals = len(area_df)
                    
                    area_themes = area_df["Theme"].value_counts()
                    for theme in top_themes:
                        count = area_themes.get(theme, 0)
                        percentage = (count / area_totals * 100) if area_totals > 0 else 0
                        
                        area_theme_data.append({
                            "Coroner Area": area,
                            "Theme": theme,
                            "Count": count,
                            "Percentage": round(percentage, 1)
                        })
                
                area_theme_df = pd.DataFrame(area_theme_data)
                
                # Create formatted names for display
                theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
                area_display_map = {area: improved_truncate_text(area, max_length=40) for area in top_areas}
                
                area_theme_df["Display_Area"] = area_theme_df["Coroner Area"].map(area_display_map)
                area_theme_df["Display_Theme"] = area_theme_df["Theme"].map(theme_display_map)
                
                # Create heatmap if we have data
                if len(area_theme_df) > 0:
                    pivot_df = area_theme_df.pivot(
                        index="Display_Area", 
                        columns="Display_Theme", 
                        values="Percentage"
                    ).fillna(0)
                    
                    # Check if we have valid data for heatmap
                    if pivot_df.shape[0] > 0 and pivot_df.shape[1] > 0:
                        fig = px.imshow(
                            pivot_df,
                            labels=dict(x="Theme", y="Coroner Area", color="Percentage"),
                            x=pivot_df.columns,
                            y=pivot_df.index,
                            color_continuous_scale="YlGnBu",
                            title="Theme Distribution by Coroner Area (%)",
                            text_auto=".1f"
                        )
                        
                        add_figure_to_zip(fig, f"theme_area_heatmap_{timestamp}.png")
                
                # Create radar chart if we have enough areas
                if len(top_areas) >= 2:
                    radar_areas = top_areas[:3]  # Take the top 3 areas or less
                    
                    # Filter data for these areas and top themes
                    radar_data = area_theme_df[
                        (area_theme_df["Coroner Area"].isin(radar_areas)) & 
                        (area_theme_df["Theme"].isin(top_themes[:6]))  # Limit to 6 themes for readability
                    ]
                    
                    if len(radar_data) > 0:
                        # Create radar chart
                        fig = go.Figure()
                        
                        # Add traces for each area
                        for area in radar_areas:
                            area_data = radar_data[radar_data["Coroner Area"] == area]
                            # Sort by theme to ensure consistency
                            area_data = area_data.set_index("Theme").reindex(top_themes[:6]).reset_index()
                            
                            fig.add_trace(go.Scatterpolar(
                                r=area_data["Percentage"],
                                theta=area_data["Display_Theme"],
                                fill="toself",
                                name=area_display_map.get(area, area)
                            ))

                        #
                        fig.update_layout(
                            polar=dict(
                                bgcolor="#f0f0f0",  # Light gray background inside the polar area
                                radialaxis=dict(
                                    visible=True,
                                    range=[0, max(radar_data["Percentage"]) * 1.1],
                                    color="black",                 # Axis lines
                                    tickfont=dict(color="black"), # Tick labels
                                    gridcolor="gray",             # Circular gridlines
                                    linecolor="black"             # Axis line
                                ),
                                angularaxis=dict(
                                    color="black",
                                    tickfont=dict(color="black"),
                                    gridcolor="gray",
                                    linecolor="black"
                                )
                            ),
                            font=dict(
                                color="black"
                            ),
                            paper_bgcolor="#0c1f30",  # Dark blue outer background
                            plot_bgcolor="#0c1f30",   # Match the paper background
                            legend=dict(
                                font=dict(color="white")  # Ensures legend text stays visible
                            ),
                            title=dict(
                                text="Theme Distribution Radar Chart",
                                font=dict(color="white")
                            ),
                            showlegend=True
                        )

                        
                        add_figure_to_zip(fig, f"area_radar_chart_{timestamp}.png")
                        
        except Exception as e:
            logging.error(f"Error creating area comparison charts: {str(e)}")
            
        # === TAB 5: CORRELATION ANALYSIS ===
        try:
            # Calculate correlation between themes
            id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
            
            # Get top themes for correlation
            top_themes = filtered_df["Theme"].value_counts().head(10).index.tolist()
            
            # Create a binary pivot table
            theme_pivot = pd.crosstab(
                index=filtered_df[id_column], 
                columns=filtered_df['Theme'],
                values=filtered_df.get('Combined Score', filtered_df['Theme']),
                aggfunc='max'
            ).fillna(0)
            
            # Convert to binary
            theme_pivot = (theme_pivot > 0).astype(int)
            
            # Calculate correlation
            if len(theme_pivot.columns) > 1:  # Need at least 2 columns for correlation
                theme_corr = theme_pivot.corr()
                
                # Get only the top themes for clarity
                available_themes = [theme for theme in top_themes if theme in theme_corr.index]
                
                if available_themes:
                    top_theme_corr = theme_corr.loc[available_themes, available_themes]
                    
                    # Create a mapping dictionary for theme display names
                    theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in available_themes}
                    
                    # Format column and index labels
                    formatted_themes = [theme_display_map[theme] for theme in top_theme_corr.columns]
                    
                    # Create a heatmap of correlations
                    fig = px.imshow(
                        top_theme_corr.values,
                        color_continuous_scale=px.colors.diverging.RdBu_r,
                        color_continuous_midpoint=0,
                        labels=dict(x="Theme", y="Theme", color="Correlation"),
                        x=formatted_themes,
                        y=formatted_themes,
                        title="Theme Correlation Matrix",
                        text_auto=".2f"
                    )
                    
                    # Update layout
                    fig.update_layout(
                        width=1000,
                        height=800,
                        margin=dict(l=250, r=100, t=80, b=250)
                    )
                    
                    add_figure_to_zip(fig, f"theme_correlation_matrix_{timestamp}.png")
                    
                    # Create network graph
                    # Try different thresholds until we get a reasonable number of edges
                    for threshold in [0.6, 0.5, 0.4, 0.3, 0.2]:
                        G = nx.Graph()
                        
                        # Add nodes
                        for theme in available_themes:
                            G.add_node(theme, display_name=theme_display_map[theme])
                        
                        # Add edges
                        edge_count = 0
                        for i, theme1 in enumerate(available_themes):
                            for j, theme2 in enumerate(available_themes):
                                if i < j:  # Only process each pair once
                                    correlation = top_theme_corr.loc[theme1, theme2]
                                    if correlation >= threshold:
                                        G.add_edge(theme1, theme2, weight=correlation)
                                        edge_count += 1
                        
                        # If we have a reasonable number of edges, create the visualization
                        if edge_count > 0 and edge_count <= 20:
                            pos = nx.spring_layout(G, seed=42)
                            
                            # Create network visualization
                            edge_traces = []
                            
                            # Add edges
                            for edge in G.edges():
                                x0, y0 = pos[edge[0]]
                                x1, y1 = pos[edge[1]]
                                weight = G[edge[0]][edge[1]]['weight']
                                
                                edge_traces.append(
                                    go.Scatter(
                                        x=[x0, x1, None],
                                        y=[y0, y1, None],
                                        line=dict(width=weight*3, color=f'rgba(100,100,100,{weight})'),
                                        hoverinfo='none',
                                        mode='lines'
                                    )
                                )
                            
                            # Add nodes
                            node_x = []
                            node_y = []
                            node_text = []
                            node_size = []
                            
                            for node in G.nodes():
                                x, y = pos[node]
                                node_x.append(x)
                                node_y.append(y)
                                node_text.append(theme_display_map[node])
                                size = len(list(G.neighbors(node))) * 10 + 20
                                node_size.append(size)
                            
                            node_trace = go.Scatter(
                                x=node_x, 
                                y=node_y,
                                mode='markers+text',
                                text=node_text,
                                textposition="top center",
                                marker=dict(
                                    size=node_size,
                                    color='lightblue',
                                    line=dict(width=1)
                                )
                            )
                            
                            # Create figure
                            fig = go.Figure(
                                data=edge_traces + [node_trace],
                                layout=go.Layout(
                                    title=f'Theme Connection Network (r ≥ {threshold})',
                                    showlegend=False,
                                    hovermode='closest',
                                    margin=dict(b=20, l=5, r=5, t=80),
                                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                    width=800,
                                    height=800
                                )
                            )
                            
                            add_figure_to_zip(fig, f"theme_network_{timestamp}.png")
                            break  # We found a good threshold, no need to try lower ones
                
                # Create co-occurrence matrix
                co_occurrence_matrix = np.zeros((len(available_themes), len(available_themes)))
                
                # Count co-occurrences
                for doc_id in theme_pivot.index:
                    doc_themes = theme_pivot.columns[theme_pivot.loc[doc_id] == 1].tolist()
                    doc_themes = [t for t in doc_themes if t in available_themes]
                    
                    # Count pairs
                    for i, theme1 in enumerate(doc_themes):
                        idx1 = available_themes.index(theme1)
                        for theme2 in doc_themes:
                            idx2 = available_themes.index(theme2)
                            co_occurrence_matrix[idx1, idx2] += 1
                
                # Create a heatmap of co-occurrences
                fig = px.imshow(
                    co_occurrence_matrix,
                    labels=dict(x="Theme", y="Theme", color="Co-occurrences"),
                    x=[theme_display_map[theme] for theme in available_themes],
                    y=[theme_display_map[theme] for theme in available_themes],
                    title="Theme Co-occurrence Matrix",
                    color_continuous_scale="Viridis",
                    text_auto=".0f"
                )
                
                fig.update_layout(
                    width=1000,
                    height=800,
                    margin=dict(l=250, r=100, t=80, b=250)
                )
                
                add_figure_to_zip(fig, f"theme_cooccurrence_matrix_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating correlation analysis charts: {str(e)}")
    
    # Reset buffer position
    zip_buffer.seek(0)
    
    # Check if zip is empty
    if image_count == 0:
        raise ValueError("No images were generated for the dashboard.")
    
    return zip_buffer.getvalue(), image_count

                    

def render_footer():
    """Render footer with timestamp in UK time (GMT/BST)."""
    # Get file modification time (UTC by default on Streamlit Cloud)
    file_path = os.path.abspath(__file__)
    last_modified_timestamp = os.path.getmtime(file_path)
    last_modified_datetime_utc = datetime.utcfromtimestamp(last_modified_timestamp)
    
    # Convert to UK time (handles GMT/BST automatically)
    uk_tz = pytz.timezone("Europe/London")
    last_modified_datetime_uk = last_modified_datetime_utc.replace(tzinfo=pytz.utc).astimezone(uk_tz)
    formatted_time = last_modified_datetime_uk.strftime("%d %B %Y %H:%M (UK Time)")

    # Display footer
    st.markdown("---")
    st.markdown(
        f"""<div style='text-align: center'>
        <p>Built with Streamlit • Data Source: UK Judiciary • Copyright © 2025 Loughborough University • Developer: Georgina Cosma • 
        Contact: g.cosma@lboro.ac.uk • All rights reserved. Last update:{formatted_time}</p>
        </div>""",
        unsafe_allow_html=True,
    )


def render_topic_modeling_tab(data: pd.DataFrame):
    """Render the topic modeling tab with enhanced visualization options"""
    st.subheader("Topic Modeling Analysis")

    if data is None or len(data) == 0:
        st.warning("No data available. Please scrape or upload data first.")
        return

    with st.sidebar:
        st.subheader("Vectorization Settings")
        vectorizer_type = st.selectbox(
            "Vectorization Method",
            ["tfidf", "bm25", "weighted"],
            help="Choose how to convert text to numerical features",
        )

        # BM25 specific parameters
        if vectorizer_type == "bm25":
            k1 = st.slider(
                "k1 parameter", 0.5, 3.0, 1.5, 0.1, help="Term saturation parameter"
            )
            b = st.slider(
                "b parameter",
                0.0,
                1.0,
                0.75,
                0.05,
                help="Length normalization parameter",
            )

        # Weighted TF-IDF parameters
        elif vectorizer_type == "weighted":
            tf_scheme = st.selectbox(
                "TF Weighting Scheme",
                ["raw", "log", "binary", "augmented"],
                help="How to weight term frequencies",
            )
            idf_scheme = st.selectbox(
                "IDF Weighting Scheme",
                ["smooth", "standard", "probabilistic"],
                help="How to weight inverse document frequencies",
            )

    # Analysis parameters
    st.subheader("Analysis Parameters")
    col1, col2 = st.columns(2)

    with col1:
        num_topics = st.slider(
            "Number of Topics",
            min_value=2,
            max_value=20,
            value=5,
            help="Number of distinct topics to identify",
        )

    with col2:
        max_features = st.slider(
            "Maximum Features",
            min_value=500,
            max_value=5000,
            value=1000,
            help="Maximum number of terms to consider",
        )

    # Get vectorizer parameters
    vectorizer_params = {}
    if vectorizer_type == "bm25":
        vectorizer_params.update({"k1": k1, "b": b})
    elif vectorizer_type == "weighted":
        vectorizer_params.update({"tf_scheme": tf_scheme, "idf_scheme": idf_scheme})

    if st.button("Run Analysis", type="primary"):
        with st.spinner("Performing topic analysis..."):
            try:
                # Create vectorizer
                vectorizer = get_vectorizer(
                    vectorizer_type=vectorizer_type,
                    max_features=max_features,
                    min_df=2,
                    max_df=0.95,
                    **vectorizer_params,
                )

                # Process text data
                docs = data["Content"].fillna("").apply(clean_text_for_modeling)

                # Create document-term matrix
                dtm = vectorizer.fit_transform(docs)
                feature_names = vectorizer.get_feature_names_out()

                # Fit LDA model
                lda = LatentDirichletAllocation(
                    n_components=num_topics, random_state=42, n_jobs=-1
                )

                doc_topics = lda.fit_transform(dtm)

                # Store model results
                st.session_state.topic_model = {
                    "lda": lda,
                    "vectorizer": vectorizer,
                    "feature_names": feature_names,
                    "doc_topics": doc_topics,
                }

                # Display results
                st.success("Topic analysis complete!")

                # Show topic words
                st.subheader("Topic Keywords")
                for idx, topic in enumerate(lda.components_):
                    top_words = [feature_names[i] for i in topic.argsort()[:-11:-1]]
                    st.markdown(f"**Topic {idx+1}:** {', '.join(top_words)}")

                # Display network visualization
                st.subheader("Topic Similarity Network")
                display_topic_network(lda, feature_names)

                # Show topic distribution
                st.subheader("Topic Distribution")
                topic_dist = doc_topics.mean(axis=0)
                topic_df = pd.DataFrame(
                    {
                        "Topic": [f"Topic {i+1}" for i in range(num_topics)],
                        "Proportion": topic_dist,
                    }
                )

                fig = px.bar(
                    topic_df,
                    x="Topic",
                    y="Proportion",
                    title="Topic Distribution Across Documents",
                )
                st.plotly_chart(fig, use_container_width=True)

                # Export options
                st.subheader("Export Results")
                if st.download_button(
                    "Download Topic Analysis Results",
                    data=export_topic_results(
                        lda, vectorizer, feature_names, doc_topics
                    ).encode(),
                    file_name="topic_analysis_results.json",
                    mime="application/json",
                ):
                    st.success("Results downloaded successfully!")

            except Exception as e:
                st.error(f"Error during analysis: {str(e)}")
                logging.error(f"Topic modeling error: {e}", exc_info=True)


def export_topic_results(lda_model, vectorizer, feature_names, doc_topics) -> str:
    """Export topic modeling results to JSON format"""
    results = {
        "topics": [],
        "model_params": {
            "n_topics": lda_model.n_components,
            "max_features": len(feature_names),
        },
        "topic_distribution": doc_topics.mean(axis=0).tolist(),
    }

    # Add topic details
    for idx, topic in enumerate(lda_model.components_):
        top_indices = topic.argsort()[:-11:-1]

        topic_words = [
            {"word": feature_names[i], "weight": float(topic[i])} for i in top_indices
        ]

        results["topics"].append(
            {"id": idx, "words": topic_words, "total_weight": float(topic.sum())}
        )

    return json.dumps(results, indent=2)


def render_summary_tab(cluster_results: Dict, original_data: pd.DataFrame) -> None:
    """Render cluster summaries and records with flexible column handling"""
    if not cluster_results or "clusters" not in cluster_results:
        st.warning("No cluster results available.")
        return

    st.write(
        f"Found {cluster_results['total_documents']} total documents in {cluster_results['n_clusters']} clusters"
    )

    for cluster in cluster_results["clusters"]:
        st.markdown(f"### Cluster {cluster['id']+1} ({cluster['size']} documents)")

        # Overview
        st.markdown("#### Overview")
        abstractive_summary = generate_abstractive_summary(
            cluster["terms"], cluster["documents"]
        )
        st.write(abstractive_summary)

        # Key terms table
        st.markdown("#### Key Terms")
        terms_df = pd.DataFrame(
            [
                {
                    "Term": term["term"],
                    "Frequency": f"{term['cluster_frequency']*100:.0f}%",
                }
                for term in cluster["terms"][:10]
            ]
        )
        st.dataframe(terms_df, hide_index=True)

        # Records
        st.markdown("#### Records")
        st.success(f"Showing {len(cluster['documents'])} matching documents")

        # Get the full records from original data
        doc_titles = [doc.get("title", "") for doc in cluster["documents"]]
        cluster_docs = original_data[original_data["Title"].isin(doc_titles)].copy()

        # Sort to match the original order
        title_to_position = {title: i for i, title in enumerate(doc_titles)}
        cluster_docs["sort_order"] = cluster_docs["Title"].map(title_to_position)
        cluster_docs = cluster_docs.sort_values("sort_order").drop("sort_order", axis=1)

        # Determine available columns
        available_columns = []
        column_config = {}

        # Always include URL and Title if available
        if "URL" in cluster_docs.columns:
            available_columns.append("URL")
            column_config["URL"] = st.column_config.LinkColumn("Report Link")

        if "Title" in cluster_docs.columns:
            available_columns.append("Title")
            column_config["Title"] = st.column_config.TextColumn("Title")

        # Add date if available
        if "date_of_report" in cluster_docs.columns:
            available_columns.append("date_of_report")
            column_config["date_of_report"] = st.column_config.DateColumn(
                "Date of Report", format="DD/MM/YYYY"
            )

        # Add optional columns if available
        optional_columns = [
            "ref",
            "deceased_name",
            "coroner_name",
            "coroner_area",
            "categories",
        ]
        for col in optional_columns:
            if col in cluster_docs.columns:
                available_columns.append(col)
                if col == "categories":
                    column_config[col] = st.column_config.ListColumn("Categories")
                else:
                    column_config[col] = st.column_config.TextColumn(
                        col.replace("_", " ").title()
                    )

        # Display the dataframe with available columns
        if available_columns:
            st.dataframe(
                cluster_docs[available_columns],
                column_config=column_config,
                hide_index=True,
                use_container_width=True,
            )
        else:
            st.warning("No displayable columns found in the data")

        st.markdown("---")

def render_bert_analysis_tab(data: pd.DataFrame = None):
    """Modified render_bert_analysis_tab function to include framework selection and custom framework upload"""
    
    # Ensure the bert_results dictionary exists in session state
    if "bert_results" not in st.session_state:
        st.session_state.bert_results = {}
    
    # Track if BERT model is initialized
    if "bert_initialized" not in st.session_state:
        st.session_state.bert_initialized = False
    
    # Initialize custom frameworks dictionary if not present
    if "custom_frameworks" not in st.session_state:
        st.session_state.custom_frameworks = {}
        
    # Safer initialization with validation
    if "selected_frameworks" not in st.session_state:
        # Only include frameworks that actually exist
        default_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"]
        st.session_state.selected_frameworks = default_frameworks
    else:
        # Validate existing selections against available options
        available_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"] + list(st.session_state.get("custom_frameworks", {}).keys())
        st.session_state.selected_frameworks = [f for f in st.session_state.selected_frameworks if f in available_frameworks]

    # File upload section
    st.subheader("Upload Data")
    reset_counter = st.session_state.get("reset_counter", 0)
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for BERT Analysis",
        type=["csv", "xlsx"],
        help="Upload a file with reports for theme analysis",
        key="bert_file_uploader",
    )

    # If a file is uploaded, process it
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                uploaded_data = pd.read_csv(uploaded_file)
            else:
                uploaded_data = pd.read_excel(uploaded_file)

            # Process the uploaded data
            uploaded_data = process_scraped_data(uploaded_data)

            # Update the data reference
            data = uploaded_data

            st.success("File uploaded and processed successfully!")
                
        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            return

    # Check if data is available
    if data is None or len(data) == 0:
        st.warning(
            "No data available. Please upload a file or ensure existing data is loaded."
        )
        return

    # Framework selection section
    st.subheader("Select Frameworks")
    
    # Create columns for the framework selection and custom framework upload
    frame_col1, frame_col2 = st.columns([2, 1])
    
    with frame_col1:
        # Get all available framework options
        available_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"]
        if "custom_frameworks" in st.session_state:
            available_frameworks.extend(list(st.session_state.custom_frameworks.keys()))
        
        # Predefined framework selection - use a unique key
        st.session_state.selected_frameworks = st.multiselect(
            "Choose Frameworks to Use",
            options=available_frameworks,
            default=st.session_state.selected_frameworks,
            help="Select which conceptual frameworks to use for theme analysis",
            key=f"framework_select_{reset_counter}"
        )
    
    with frame_col2:
        # Custom framework upload
        custom_framework_file = st.file_uploader(
            "Upload Custom Framework",
            type=["json", "txt"],
            help="Upload a JSON file containing custom framework definitions",
            key=f"custom_framework_uploader_{reset_counter}"
        )
        
        if custom_framework_file is not None:
            try:
                # Read framework definition
                custom_framework_content = custom_framework_file.read().decode("utf-8")
                custom_framework_data = json.loads(custom_framework_content)
                
                # Validate framework structure
                if isinstance(custom_framework_data, list) and all(isinstance(item, dict) and "name" in item and "keywords" in item for item in custom_framework_data):
                    # Framework name input
                    custom_framework_name = st.text_input(
                        "Custom Framework Name", 
                        f"Custom Framework {len(st.session_state.custom_frameworks) + 1}",
                        key=f"custom_framework_name_{reset_counter}"
                    )
                    
                    # Add button for the custom framework
                    if st.button("Add Custom Framework", key=f"add_custom_framework_{reset_counter}"):
                        # Check if name already exists
                        if custom_framework_name in st.session_state.custom_frameworks:
                            st.warning(f"A framework with the name '{custom_framework_name}' already exists. Please choose a different name.")
                        else:
                            # Add to session state
                            st.session_state.custom_frameworks[custom_framework_name] = custom_framework_data
                            
                            # Add to selected frameworks if not already there
                            if custom_framework_name not in st.session_state.selected_frameworks:
                                st.session_state.selected_frameworks.append(custom_framework_name)
                            
                            st.success(f"Custom framework '{custom_framework_name}' with {len(custom_framework_data)} themes added successfully")
                            st.rerun()  # Refresh to update UI
                else:
                    st.error("Invalid framework format. Each item must have 'name' and 'keywords' fields.")
            except json.JSONDecodeError:
                st.error("Invalid JSON format. Please check your file.")
            except Exception as e:
                st.error(f"Error processing custom framework: {str(e)}")
                logging.error(f"Custom framework error: {e}", exc_info=True)
    
    # Display currently loaded custom frameworks
    if "custom_frameworks" in st.session_state and st.session_state.custom_frameworks:
        st.subheader("Loaded Custom Frameworks")
        for name, framework in st.session_state.custom_frameworks.items():
            with st.expander(f"{name} ({len(framework)} themes)"):
                # Display the first few themes as an example
                for i, theme in enumerate(framework[:5]):
                    st.markdown(f"**{theme['name']}**: {', '.join(theme['keywords'][:5])}...")
                    if i >= 4 and len(framework) > 5:
                        st.markdown(f"*... and {len(framework) - 5} more themes*")
                        break
                
                # Add option to remove this framework
                if st.button("Remove Framework", key=f"remove_{name}_{reset_counter}"):
                    del st.session_state.custom_frameworks[name]
                    if name in st.session_state.selected_frameworks:
                        st.session_state.selected_frameworks.remove(name)
                    st.success(f"Removed framework '{name}'")
                    st.rerun()  # Refresh to update UI

    # Column selection for analysis
    st.subheader("Select Analysis Column")

    # Find text columns (object/string type)
    text_columns = data.select_dtypes(include=["object"]).columns.tolist()

    # If no text columns found
    if not text_columns:
        st.error("No text columns found in the dataset.")
        return

    # Column selection with dropdown
    content_column = st.selectbox(
        "Choose the column to analyse:",
        options=text_columns,
        index=text_columns.index("Content") if "Content" in text_columns else 0,
        help="Select the column containing the text you want to analyse",
        key="bert_content_column",
    )

    # Filtering options
    st.subheader("Select Documents to Analyse")

    # Option to select all or specific records
    analysis_type = st.radio(
        "Analysis Type",
        ["All Reports", "Selected Reports"],
        horizontal=True,
        key="bert_analysis_type",
    )

    if analysis_type == "Selected Reports":
        # Multi-select for reports
        selected_indices = st.multiselect(
            "Choose specific reports to analyse",
            options=list(range(len(data))),
            format_func=lambda x: f"{data.iloc[x]['Title']} ({data.iloc[x]['date_of_report'].strftime('%d/%m/%Y') if pd.notna(data.iloc[x]['date_of_report']) else 'No date'})",
            key="bert_selected_indices",
        )
        selected_data = data.iloc[selected_indices] if selected_indices else None
    else:
        selected_data = data

    # Analysis parameters
    st.subheader("Analysis Parameters")
    similarity_threshold = st.slider(
        "Similarity Threshold",
        min_value=0.3,
        max_value=0.9,
        value=0.65,
        step=0.05,
        help="Minimum similarity score for theme detection (higher = more strict)",
        key="bert_similarity_threshold",
    )

    # Analysis button
    run_analysis = st.button(
        "Run Analysis", type="primary", key="bert_run_analysis"
    )

    # Run analysis if button is clicked
    if run_analysis:
        with st.spinner("Performing Theme Analysis..."):
            try:
                # Validate data selection
                if selected_data is None or len(selected_data) == 0:
                    st.warning("No documents selected for analysis.")
                    return

                # Initialize the theme analyzer (with loading message in a spinner)
                with st.spinner("Loading annotation model and tokenizer..."):
                    # Initialize the analyzer
                    theme_analyzer = ThemeAnalyzer(
                        model_name="emilyalsentzer/Bio_ClinicalBERT"
                    )
                    
                    # Mark as initialized
                    st.session_state.bert_initialized = True
                
                # Set custom configuration
                theme_analyzer.config[
                    "base_similarity_threshold"
                ] = similarity_threshold
                
                # Filter frameworks based on user selection
                filtered_frameworks = {}
                
                # Add selected built-in frameworks
                for framework in st.session_state.selected_frameworks:
                    if framework == "I-SIRch":
                        filtered_frameworks["I-SIRch"] = theme_analyzer._get_isirch_framework()
                    elif framework == "House of Commons":
                        filtered_frameworks["House of Commons"] = theme_analyzer._get_house_of_commons_themes()
                    elif framework == "Extended Analysis":
                        filtered_frameworks["Extended Analysis"] = theme_analyzer._get_extended_themes()
                    elif framework in st.session_state.custom_frameworks:
                        # Add custom framework
                        filtered_frameworks[framework] = st.session_state.custom_frameworks[framework]
                
                # Set the filtered frameworks
                theme_analyzer.frameworks = filtered_frameworks
                
                # If no frameworks selected, show error
                if not filtered_frameworks:
                    st.error("Please select at least one framework for analysis.")
                    return

                # Use the enhanced create_detailed_results method
                (
                    results_df,
                    highlighted_texts,
                ) = theme_analyzer.create_detailed_results(
                    selected_data, content_column=content_column
                )

                # Save results to session state to ensure persistence
                st.session_state.bert_results["results_df"] = results_df
                st.session_state.bert_results["highlighted_texts"] = highlighted_texts

                st.success(f"Analysis complete using {len(filtered_frameworks)} frameworks!")

            except Exception as e:
                st.error(f"Error during annotation analysis: {str(e)}")
                logging.error(f"Annotation analysis error: {e}", exc_info=True)

    # Display results if they exist
    if "bert_results" in st.session_state and st.session_state.bert_results.get("results_df") is not None:
        results_df = st.session_state.bert_results["results_df"]
        
        # Summary stats
        st.subheader("Results")
        
        # Show framework distribution
        if "Framework" in results_df.columns:
            framework_counts = results_df["Framework"].value_counts()
            
            # Create columns for framework distribution metrics
            framework_cols = st.columns(len(framework_counts))
            
            for i, (framework, count) in enumerate(framework_counts.items()):
                with framework_cols[i]:
                    st.metric(framework, count, help=f"Number of theme identifications from {framework} framework")
        
        st.write(f"Total Theme Identifications: {len(results_df)}")
        
        # Clean up the results DataFrame to display only the essential columns
        display_cols = ["Record ID", "Title", "Framework", "Theme", "Confidence", "Combined Score", "Matched Keywords"]
        
        # Add metadata columns if available
        for col in ["coroner_name", "coroner_area", "year", "date_of_report"]:
            if col in results_df.columns:
                display_cols.append(col)
        
        # Add matched sentences at the end
        if "Matched Sentences" in results_df.columns:
            display_cols.append("Matched Sentences")
        
        # Create the display DataFrame with only existing columns
        valid_cols = [col for col in display_cols if col in results_df.columns]
        clean_df = results_df[valid_cols].copy()
        
        # Display the results table
        st.dataframe(
            clean_df,
            use_container_width=True,
            column_config={
                "Title": st.column_config.TextColumn("Document Title"),
                "Framework": st.column_config.TextColumn("Framework"),
                "Theme": st.column_config.TextColumn("Theme"),
                "Confidence": st.column_config.TextColumn("Confidence"),
                "Combined Score": st.column_config.NumberColumn("Score", format="%.3f"),
                "Matched Keywords": st.column_config.TextColumn("Keywords"),
                "Matched Sentences": st.column_config.TextColumn("Matched Sentences"),
                "coroner_name": st.column_config.TextColumn("Coroner Name"),
                "coroner_area": st.column_config.TextColumn("Coroner Area"),
                "year": st.column_config.NumberColumn("Year"),
                "date_of_report": st.column_config.DateColumn("Date of Report", format="DD/MM/YYYY")
            }
        )
        
        # Add download options
        st.subheader("Export Results")
        
        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create columns for download buttons
        col1, col2 = st.columns(2)
        
        with col1:
            # Excel download button using the enhanced export_to_excel function
            excel_data = export_to_excel(clean_df)
            st.download_button(
                "📥 Download Results Table",
                data=excel_data,
                file_name=f"annotated_theme_analysis_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="bert_excel_download",
            )
        
        with col2:
            # Always regenerate HTML report when results are available
            if "results_df" in st.session_state.bert_results and "highlighted_texts" in st.session_state.bert_results:
                # Generate fresh HTML report based on current results
                theme_analyzer = ThemeAnalyzer()
                
                # Set custom frameworks if they exist
                if st.session_state.custom_frameworks:
                    for name, framework in st.session_state.custom_frameworks.items():
                        if name in st.session_state.selected_frameworks:
                            theme_analyzer.frameworks[name] = framework
                
                html_content = theme_analyzer._create_integrated_html_for_pdf(
                    results_df, st.session_state.bert_results["highlighted_texts"]
                )
                html_filename = f"theme_analysis_report_{timestamp}.html"
                
                with open(html_filename, "w", encoding="utf-8") as f:
                    f.write(html_content)
                    
                st.session_state.bert_results["html_filename"] = html_filename
                
                # Provide download button for fresh HTML
                with open(html_filename, "rb") as f:
                    html_data = f.read()
                
                st.download_button(
                    "📄 Download Annotated Reports (HTML)",
                    data=html_data,
                    file_name=os.path.basename(html_filename),
                    mime="text/html",
                    key="bert_html_download",
                )
            else:
                st.warning("HTML report not available")

def render_bert_analysis_tabworking(data: pd.DataFrame = None):
    """Modified render_bert_analysis_tab function to include enhanced metadata in results"""
    
    # Ensure the bert_results dictionary exists in session state
    if "bert_results" not in st.session_state:
        st.session_state.bert_results = {}
    
    # Track if BERT model is initialized
    if "bert_initialized" not in st.session_state:
        st.session_state.bert_initialized = False

    # File upload section
    st.subheader("Upload Data")
    reset_counter = st.session_state.get("reset_counter", 0)
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for BERT Analysis",
        type=["csv", "xlsx"],
        help="Upload a file with reports for theme analysis",
        key="bert_file_uploader",
    )

    # If a file is uploaded, process it
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                uploaded_data = pd.read_csv(uploaded_file)
            else:
                uploaded_data = pd.read_excel(uploaded_file)

            # Process the uploaded data
            uploaded_data = process_scraped_data(uploaded_data)

            # Update the data reference
            data = uploaded_data

            st.success("File uploaded and processed successfully!")
                
        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            return

    # Check if data is available
    if data is None or len(data) == 0:
        st.warning(
            "No data available. Please upload a file or ensure existing data is loaded."
        )
        return

    # Column selection for analysis
    st.subheader("Select Analysis Column")

    # Find text columns (object/string type)
    text_columns = data.select_dtypes(include=["object"]).columns.tolist()

    # If no text columns found
    if not text_columns:
        st.error("No text columns found in the dataset.")
        return

    # Column selection with dropdown
    content_column = st.selectbox(
        "Choose the column to analyse:",
        options=text_columns,
        index=text_columns.index("Content") if "Content" in text_columns else 0,
        help="Select the column containing the text you want to analyse",
        key="bert_content_column",
    )

    # Filtering options
    st.subheader("Select Documents to Analyse")

    # Option to select all or specific records
    analysis_type = st.radio(
        "Analysis Type",
        ["All Reports", "Selected Reports"],
        horizontal=True,
        key="bert_analysis_type",
    )

    if analysis_type == "Selected Reports":
        # Multi-select for reports
        selected_indices = st.multiselect(
            "Choose specific reports to analyse",
            options=list(range(len(data))),
            format_func=lambda x: f"{data.iloc[x]['Title']} ({data.iloc[x]['date_of_report'].strftime('%d/%m/%Y') if pd.notna(data.iloc[x]['date_of_report']) else 'No date'})",
            key="bert_selected_indices",
        )
        selected_data = data.iloc[selected_indices] if selected_indices else None
    else:
        selected_data = data

    # Analysis parameters
    st.subheader("Analysis Parameters")
    similarity_threshold = st.slider(
        "Similarity Threshold",
        min_value=0.3,
        max_value=0.9,
        value=0.65,
        step=0.05,
        help="Minimum similarity score for theme detection (higher = more strict)",
        key="bert_similarity_threshold",
    )

    # Analysis button
    run_analysis = st.button(
        "Run Analysis", type="primary", key="bert_run_analysis"
    )

    # Run analysis if button is clicked
    if run_analysis:
        with st.spinner("Performing Theme Analysis..."):
            try:
                # Validate data selection
                if selected_data is None or len(selected_data) == 0:
                    st.warning("No documents selected for analysis.")
                    return

                # Initialize the theme analyzer (with loading message in a spinner)
                with st.spinner("Loading annotation model and tokenizer..."):
                    # Initialize the analyzer
                    theme_analyzer = ThemeAnalyzer(
                        model_name="emilyalsentzer/Bio_ClinicalBERT"
                    )
                    
                    # Mark as initialized
                    st.session_state.bert_initialized = True
                
                # Set custom configuration
                theme_analyzer.config[
                    "base_similarity_threshold"
                ] = similarity_threshold

                # Use the enhanced create_detailed_results method
                (
                    results_df,
                    highlighted_texts,
                ) = theme_analyzer.create_detailed_results(
                    selected_data, content_column=content_column
                )

                # Save results to session state to ensure persistence
                st.session_state.bert_results["results_df"] = results_df
                st.session_state.bert_results["highlighted_texts"] = highlighted_texts

                st.success("Analysis complete!")

            except Exception as e:
                st.error(f"Error during annotation analysis: {str(e)}")
                logging.error(f"Annotation analysis error: {e}", exc_info=True)

    # Display results if they exist
    if "bert_results" in st.session_state and st.session_state.bert_results.get("results_df") is not None:
        results_df = st.session_state.bert_results["results_df"]
        
        # Summary stats
        st.subheader("Results")
        st.write(f"Total Theme Identifications: {len(results_df)}")
        
        # Clean up the results DataFrame to display only the essential columns
        display_cols = ["Record ID", "Title", "Framework", "Theme", "Confidence", "Combined Score", "Matched Keywords"]
        
        # Add metadata columns if available
        for col in ["coroner_name", "coroner_area", "year", "date_of_report"]:
            if col in results_df.columns:
                display_cols.append(col)
        
        # Add matched sentences at the end
        if "Matched Sentences" in results_df.columns:
            display_cols.append("Matched Sentences")
        
        # Create the display DataFrame with only existing columns
        valid_cols = [col for col in display_cols if col in results_df.columns]
        clean_df = results_df[valid_cols].copy()
        
        # Display the results table
        st.dataframe(
            clean_df,
            use_container_width=True,
            column_config={
                "Title": st.column_config.TextColumn("Document Title"),
                "Framework": st.column_config.TextColumn("Framework"),
                "Theme": st.column_config.TextColumn("Theme"),
                "Confidence": st.column_config.TextColumn("Confidence"),
                "Combined Score": st.column_config.NumberColumn("Score", format="%.3f"),
                "Matched Keywords": st.column_config.TextColumn("Keywords"),
                "Matched Sentences": st.column_config.TextColumn("Matched Sentences"),
                "coroner_name": st.column_config.TextColumn("Coroner Name"),
                "coroner_area": st.column_config.TextColumn("Coroner Area"),
                "year": st.column_config.NumberColumn("Year"),
                "date_of_report": st.column_config.DateColumn("Date of Report", format="DD/MM/YYYY")
            }
        )
        
        # Add download options
        st.subheader("Export Results")
        
        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create columns for download buttons
        col1, col2 = st.columns(2)
        
        with col1:
            # Excel download button using the enhanced export_to_excel function
            excel_data = export_to_excel(clean_df)
            st.download_button(
                "📥 Download Results Table",
                data=excel_data,
                file_name=f"annotated_theme_analysis_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="bert_excel_download",
            )
        
        with col2:
            # Always regenerate HTML report when results are available
            if "results_df" in st.session_state.bert_results and "highlighted_texts" in st.session_state.bert_results:
                # Generate fresh HTML report based on current results
                theme_analyzer = ThemeAnalyzer()
                html_content = theme_analyzer._create_integrated_html_for_pdf(
                    results_df, st.session_state.bert_results["highlighted_texts"]
                )
                html_filename = f"theme_analysis_report_{timestamp}.html"
                
                with open(html_filename, "w", encoding="utf-8") as f:
                    f.write(html_content)
                    
                st.session_state.bert_results["html_filename"] = html_filename
                
                # Provide download button for fresh HTML
                with open(html_filename, "rb") as f:
                    html_data = f.read()
                
                st.download_button(
                    "📄 Download Annotated Reports (HTML)",
                    data=html_data,
                    file_name=os.path.basename(html_filename),
                    mime="text/html",
                    key="bert_html_download",
                )
            else:
                st.warning("HTML report not available")

def render_filter_data_tab():
    """Render a filtering tab within the Scraped File Preparation section with layout similar to Scrape Reports tab"""
    st.subheader("Filter Data")

    # File upload section
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file",
        type=["csv", "xlsx"],
        help="Upload your PFD reports dataset",
        key="filter_file_uploader",
    )
    
    # If no file is uploaded, show instructions
    if uploaded_file is None:
        st.info("Please upload a PFD reports dataset (CSV or Excel file) to begin filtering.")
        
        with st.expander("📋 File Requirements", expanded=False):
            st.markdown(
                """
            ## Recommended Columns
            
            For optimal filtering, your file should include these columns:
            
            - **Title**: Report title
            - **URL**: Link to the original report
            - **date_of_report**: Date in format DD/MM/YYYY
            - **ref**: Reference number
            - **deceased_name**: Name of deceased person (if applicable)
            - **coroner_name**: Name of the coroner
            - **coroner_area**: Coroner jurisdiction area
            - **categories**: Report categories
            - **Content**: Report text content
            - **Extracted_Concerns**: Extracted coroner concerns text
            
            Files created from the File Merger tab should contain all these columns.
            """
            )
        return
    
    # Process uploaded file
    try:
        # Read file
        if uploaded_file.name.endswith(".csv"):
            data = pd.read_csv(uploaded_file)
        else:
            data = pd.read_excel(uploaded_file)

        # Basic data cleaning
        data = data.dropna(how="all")  # Remove completely empty rows

        # Convert date_of_report to datetime if it exists
        if "date_of_report" in data.columns:
            try:
                # Try multiple date formats
                data["date_of_report"] = pd.to_datetime(
                    data["date_of_report"], format="%d/%m/%Y", errors="coerce"
                )
                # Fill in any parsing failures with other formats
                mask = data["date_of_report"].isna()
                if mask.any():
                    data.loc[mask, "date_of_report"] = pd.to_datetime(
                        data.loc[mask, "date_of_report"], format="%Y-%m-%d", errors="coerce"
                    )
                # Final attempt with pandas' smart parsing
                mask = data["date_of_report"].isna()
                if mask.any():
                    data.loc[mask, "date_of_report"] = pd.to_datetime(
                        data.loc[mask, "date_of_report"],
                        infer_datetime_format=True,
                        errors="coerce",
                    )
            except Exception as e:
                st.warning(
                    "Some date values could not be converted. Date filtering may not work completely."
                )
        
        # Create a form for filter settings
        with st.form("filter_settings_form"):
            st.subheader("Filter Settings")
            
            # Create rows with two columns each
            row1_col1, row1_col2 = st.columns(2)
            row2_col1, row2_col2 = st.columns(2)
            row3_col1, row3_col2 = st.columns(2)
            
            # First row
            with row1_col1:
                # Date Range Filter
                if "date_of_report" in data.columns and pd.api.types.is_datetime64_any_dtype(data["date_of_report"]):
                    min_date = data["date_of_report"].min().date()
                    max_date = data["date_of_report"].max().date()
                    
                    st.markdown("**Date Range**")
                    start_date = st.date_input(
                        "From",
                        value=min_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="filter_start_date",
                        format="DD/MM/YYYY",
                    )
            
            with row1_col2:
                # Date Range (continued)
                if "date_of_report" in data.columns and pd.api.types.is_datetime64_any_dtype(data["date_of_report"]):
                    st.markdown("&nbsp;")  # Add some spacing to align with "From"
                    end_date = st.date_input(
                        "To",
                        value=max_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="filter_end_date",
                        format="DD/MM/YYYY",
                    )
                # Year Filter if date_of_report not available
                elif "year" in data.columns:
                    years = sorted(data["year"].dropna().unique())
                    if years:
                        min_year, max_year = min(years), max(years)
                        st.markdown("**Year Range**")
                        selected_years = st.slider(
                            "Select years",
                            min_year,
                            max_year,
                            (min_year, max_year),
                            key="filter_years",
                        )
            
            # Second row
            with row2_col1:
                # Coroner Area Filter
                if "coroner_area" in data.columns:
                    st.markdown("**Coroner Areas**")
                    coroner_areas = sorted(data["coroner_area"].dropna().unique())
                    selected_areas = st.multiselect(
                        "Select coroner areas",
                        options=coroner_areas,
                        key="filter_coroner_areas",
                        help="Select coroner areas to include"
                    )
            
            with row2_col2:
                # Coroner Name Filter
                if "coroner_name" in data.columns:
                    st.markdown("**Coroner Names**")
                    coroner_names = sorted(data["coroner_name"].dropna().unique())
                    selected_coroners = st.multiselect(
                        "Select coroner names",
                        options=coroner_names,
                        key="filter_coroner_names",
                        help="Select coroner names to include"
                    )
            
            # Third row
            with row3_col1:
                # Categories Filter
                if "categories" in data.columns:
                    st.markdown("**Categories**")
                    # Extract unique categories by splitting and cleaning
                    all_categories = set()
                    for cats in data["categories"].dropna():
                        # Split by comma and strip whitespace
                        if isinstance(cats, str):
                            split_cats = [cat.strip() for cat in cats.split(",")]
                            all_categories.update(split_cats)
                        elif isinstance(cats, list):
                            all_categories.update(
                                [cat.strip() for cat in cats if isinstance(cat, str)]
                            )

                    # Sort and remove any empty strings
                    sorted_categories = sorted(cat for cat in all_categories if cat)

                    # Create multiselect for categories
                    selected_categories = st.multiselect(
                        "Select categories",
                        options=sorted_categories,
                        key="filter_categories",
                        help="Select categories to include"
                    )
            
            with row3_col2:
                # Advanced content search
                if "content" in data.columns or "Content" in data.columns:
                    content_col = "content" if "content" in data.columns else "Content"
                    st.markdown("**Content Search**")
                    keyword_search = st.text_input(
                        "Search in content",
                        key="filter_keyword_search",
                        help="Use 'term1 and term2' for AND, 'term1 or term2' for OR",
                        placeholder="Enter search terms"
                    )
            
            # Option to exclude records without extracted concerns
            if "extracted_concerns" in data.columns or "Extracted_Concerns" in data.columns:
                concerns_col = (
                    "extracted_concerns"
                    if "extracted_concerns" in data.columns
                    else "Extracted_Concerns"
                )
                exclude_no_concerns = st.checkbox(
                    "Exclude records without extracted concerns",
                    value=False,
                    key="filter_exclude_no_concerns",
                    help="Show only records with extracted coroner concerns",
                )
            
            # Submission buttons
            submitted = st.form_submit_button("Apply Filters & Search", type="primary")
            
        # Reset Filters Button (outside the form)
        if st.button("Reset Filters"):
            for key in list(st.session_state.keys()):
                if key.startswith("filter_"):
                    del st.session_state[key]
            st.rerun()
        
        # Display results if form submitted or if this is a rerun with active filters
        show_results = submitted or any(k.startswith("filter_") for k in st.session_state.keys())
        
        if show_results:
            # Apply filters to data
            filtered_df = data.copy()
            active_filters = []
            
            # Date filter
            if "date_of_report" in filtered_df.columns and "filter_start_date" in st.session_state and "filter_end_date" in st.session_state:
                start_date = st.session_state.filter_start_date
                end_date = st.session_state.filter_end_date
                min_date = filtered_df["date_of_report"].min().date()
                max_date = filtered_df["date_of_report"].max().date()
                
                if start_date != min_date or end_date != max_date:
                    filtered_df = filtered_df[
                        (filtered_df["date_of_report"].dt.date >= start_date)
                        & (filtered_df["date_of_report"].dt.date <= end_date)
                    ]
                    active_filters.append(
                        f"Date: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
                    )

            # Year filter (if date_of_report not available)
            if "year" in filtered_df.columns and "filter_years" in st.session_state:
                selected_years = st.session_state.filter_years
                years = sorted(data["year"].dropna().unique())
                min_year, max_year = min(years), max(years)
                
                if selected_years != (min_year, max_year):
                    filtered_df = filtered_df[
                        (filtered_df["year"] >= selected_years[0])
                        & (filtered_df["year"] <= selected_years[1])
                    ]
                    active_filters.append(f"Year: {selected_years[0]} to {selected_years[1]}")

            # Coroner name filter
            if "coroner_name" in filtered_df.columns and "filter_coroner_names" in st.session_state and st.session_state.filter_coroner_names:
                selected_coroners = st.session_state.filter_coroner_names
                filtered_df = filtered_df[filtered_df["coroner_name"].isin(selected_coroners)]
                
                if len(selected_coroners) <= 3:
                    active_filters.append(f"Coroners: {', '.join(selected_coroners)}")
                else:
                    active_filters.append(f"Coroners: {len(selected_coroners)} selected")

            # Coroner area filter
            if "coroner_area" in filtered_df.columns and "filter_coroner_areas" in st.session_state and st.session_state.filter_coroner_areas:
                selected_areas = st.session_state.filter_coroner_areas
                filtered_df = filtered_df[filtered_df["coroner_area"].isin(selected_areas)]
                
                if len(selected_areas) <= 3:
                    active_filters.append(f"Areas: {', '.join(selected_areas)}")
                else:
                    active_filters.append(f"Areas: {len(selected_areas)} selected")

            # Categories filter
            if "categories" in filtered_df.columns and "filter_categories" in st.session_state and st.session_state.filter_categories:
                selected_categories = st.session_state.filter_categories
                
                # Handle both string and list categories
                if isinstance(filtered_df["categories"].iloc[0] if len(filtered_df) > 0 else "", list):
                    # List case
                    filtered_df = filtered_df[
                        filtered_df["categories"].apply(
                            lambda x: isinstance(x, list)
                            and any(cat in x for cat in selected_categories)
                        )
                    ]
                else:
                    # String case
                    filtered_df = filtered_df[
                        filtered_df["categories"]
                        .fillna("")
                        .astype(str)
                        .apply(lambda x: any(cat in x for cat in selected_categories))
                    ]
                
                if len(selected_categories) <= 3:
                    active_filters.append(f"Categories: {', '.join(selected_categories)}")
                else:
                    active_filters.append(f"Categories: {len(selected_categories)} selected")

            # Content keyword search with advanced AND/OR operators
            if "filter_keyword_search" in st.session_state and st.session_state.filter_keyword_search:
                keyword_search = st.session_state.filter_keyword_search
                content_col = "content" if "content" in filtered_df.columns else "Content"
                
                if content_col in filtered_df.columns:
                    before_count = len(filtered_df)
                    
                    #

                    # Apply the advanced search with AND/OR operators
                    filtered_df = filtered_df[
                        filtered_df[content_col]
                        .fillna("")
                        .astype(str)
                        .apply(lambda x: perform_advanced_keyword_search(x, keyword_search))
                    ]
                    
                    after_count = len(filtered_df)
                    
                    # Add to active filters
                    if " and " in keyword_search.lower():
                        search_desc = f"Content contains all terms: '{keyword_search}'"
                    elif " or " in keyword_search.lower():
                        search_desc = f"Content contains any term: '{keyword_search}'"
                    else:
                        search_desc = f"Content contains: '{keyword_search}'"
                    
                    active_filters.append(f"{search_desc} ({after_count}/{before_count} records)")

            # Apply filter for records with extracted concerns
            if "filter_exclude_no_concerns" in st.session_state and st.session_state.filter_exclude_no_concerns:
                concerns_col = (
                    "extracted_concerns"
                    if "extracted_concerns" in filtered_df.columns
                    else "Extracted_Concerns"
                )
                
                if concerns_col in filtered_df.columns:
                    before_count = len(filtered_df)
                    filtered_df = filtered_df[
                        filtered_df[concerns_col].notna()
                        & (filtered_df[concerns_col].astype(str).str.strip() != "")
                        & (filtered_df[concerns_col].astype(str).str.len() > 20)  # Ensure meaningful content
                    ]
                    after_count = len(filtered_df)
                    removed_count = before_count - after_count
                    active_filters.append(f"Excluding records without concerns (-{removed_count} records)")

            # Display active filters in a clean info box if there are any
            if active_filters:
                st.info(
                    "Active filters:\n" + "\n".join(f"• {filter_}" for filter_ in active_filters)
                )
            
            # Update session state with filtered data
            st.session_state.filtered_data = filtered_df
            
            # Display results
            st.subheader("Results")
            st.write(f"Showing {len(filtered_df)} of {len(data)} reports")
            
            if len(filtered_df) > 0:
                # Display the dataframe with formatted columns
                column_config = {}
                
                # Configure special columns
                if "date_of_report" in filtered_df.columns:
                    column_config["date_of_report"] = st.column_config.DateColumn(
                        "Date of Report", format="DD/MM/YYYY"
                    )
                
                if "URL" in filtered_df.columns:
                    column_config["URL"] = st.column_config.LinkColumn("Report Link")
                
                if "categories" in filtered_df.columns:
                    column_config["categories"] = st.column_config.ListColumn("Categories")
                
                # Display results in a Streamlit dataframe
                st.dataframe(
                    filtered_df,
                    column_config=column_config,
                    hide_index=True,
                    use_container_width=True
                )
                
                # Show export options
                show_export_options(filtered_df, "filtered")
            else:
                st.warning("No reports match your filter criteria. Try adjusting the filters.")
        
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        logging.error(f"File processing error: {e}", exc_info=True)


def render_bert_analysis_tab(data: pd.DataFrame = None):
    """Modified render_bert_analysis_tab function to include framework selection and custom framework upload"""
    
    # Ensure the bert_results dictionary exists in session state
    if "bert_results" not in st.session_state:
        st.session_state.bert_results = {}
    
    # Track if BERT model is initialized
    if "bert_initialized" not in st.session_state:
        st.session_state.bert_initialized = False
    
    # Initialize custom frameworks dictionary if not present
    if "custom_frameworks" not in st.session_state:
        st.session_state.custom_frameworks = {}
        
    # Safer initialization with validation
    if "selected_frameworks" not in st.session_state:
        # Only include frameworks that actually exist
        default_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"]
        st.session_state.selected_frameworks = default_frameworks
    else:
        # Validate existing selections against available options
        available_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"] + list(st.session_state.get("custom_frameworks", {}).keys())
        st.session_state.selected_frameworks = [f for f in st.session_state.selected_frameworks if f in available_frameworks]

    # File upload section
    st.subheader("Upload Data")
    reset_counter = st.session_state.get("reset_counter", 0)
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for BERT Analysis",
        type=["csv", "xlsx"],
        help="Upload a file with reports for theme analysis",
        key="bert_file_uploader",
    )

    # If a file is uploaded, process it
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                uploaded_data = pd.read_csv(uploaded_file)
            else:
                uploaded_data = pd.read_excel(uploaded_file)

            # Process the uploaded data
            uploaded_data = process_scraped_data(uploaded_data)

            # Update the data reference
            data = uploaded_data

            st.success("File uploaded and processed successfully!")
                
        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            return

    # Check if data is available
    if data is None or len(data) == 0:
        st.warning(
            "No data available. Please upload a file or ensure existing data is loaded."
        )
        return

    # Framework selection section
    st.subheader("Select Frameworks")
    
    # Create columns for the framework selection and custom framework upload
    frame_col1, frame_col2 = st.columns([2, 1])
    
    with frame_col1:
        # Get all available framework options
        available_frameworks = ["I-SIRch", "House of Commons", "Extended Analysis"]
        if "custom_frameworks" in st.session_state:
            available_frameworks.extend(list(st.session_state.custom_frameworks.keys()))
        
        # Predefined framework selection - use a unique key
        st.session_state.selected_frameworks = st.multiselect(
            "Choose Frameworks to Use",
            options=available_frameworks,
            default=st.session_state.selected_frameworks,
            help="Select which conceptual frameworks to use for theme analysis",
            key=f"framework_select_{reset_counter}"
        )
    
    with frame_col2:
        # Custom framework upload
        custom_framework_file = st.file_uploader(
            "Upload Custom Framework",
            type=["json", "txt"],
            help="Upload a JSON file containing custom framework definitions",
            key=f"custom_framework_uploader_{reset_counter}"
        )
        
        if custom_framework_file is not None:
            try:
                # Read framework definition
                custom_framework_content = custom_framework_file.read().decode("utf-8")
                custom_framework_data = json.loads(custom_framework_content)
                
                # Validate framework structure
                if isinstance(custom_framework_data, list) and all(isinstance(item, dict) and "name" in item and "keywords" in item for item in custom_framework_data):
                    # Framework name input
                    custom_framework_name = st.text_input(
                        "Custom Framework Name", 
                        f"Custom Framework {len(st.session_state.custom_frameworks) + 1}",
                        key=f"custom_framework_name_{reset_counter}"
                    )
                    
                    # Add button for the custom framework
                    if st.button("Add Custom Framework", key=f"add_custom_framework_{reset_counter}"):
                        # Check if name already exists
                        if custom_framework_name in st.session_state.custom_frameworks:
                            st.warning(f"A framework with the name '{custom_framework_name}' already exists. Please choose a different name.")
                        else:
                            # Add to session state
                            st.session_state.custom_frameworks[custom_framework_name] = custom_framework_data
                            
                            # Add to selected frameworks if not already there
                            if custom_framework_name not in st.session_state.selected_frameworks:
                                st.session_state.selected_frameworks.append(custom_framework_name)
                            
                            st.success(f"Custom framework '{custom_framework_name}' with {len(custom_framework_data)} themes added successfully")
                            st.rerun()  # Refresh to update UI
                else:
                    st.error("Invalid framework format. Each item must have 'name' and 'keywords' fields.")
            except json.JSONDecodeError:
                st.error("Invalid JSON format. Please check your file.")
            except Exception as e:
                st.error(f"Error processing custom framework: {str(e)}")
                logging.error(f"Custom framework error: {e}", exc_info=True)
    
    # Display currently loaded custom frameworks
    if "custom_frameworks" in st.session_state and st.session_state.custom_frameworks:
        st.subheader("Loaded Custom Frameworks")
        for name, framework in st.session_state.custom_frameworks.items():
            with st.expander(f"{name} ({len(framework)} themes)"):
                # Display the first few themes as an example
                for i, theme in enumerate(framework[:5]):
                    st.markdown(f"**{theme['name']}**: {', '.join(theme['keywords'][:5])}...")
                    if i >= 4 and len(framework) > 5:
                        st.markdown(f"*... and {len(framework) - 5} more themes*")
                        break
                
                # Add option to remove this framework
                if st.button("Remove Framework", key=f"remove_{name}_{reset_counter}"):
                    del st.session_state.custom_frameworks[name]
                    if name in st.session_state.selected_frameworks:
                        st.session_state.selected_frameworks.remove(name)
                    st.success(f"Removed framework '{name}'")
                    st.rerun()  # Refresh to update UI

    # Column selection for analysis
    st.subheader("Select Analysis Column")

    # Find text columns (object/string type)
    text_columns = data.select_dtypes(include=["object"]).columns.tolist()

    # If no text columns found
    if not text_columns:
        st.error("No text columns found in the dataset.")
        return

    # Column selection with dropdown
    content_column = st.selectbox(
        "Choose the column to analyse:",
        options=text_columns,
        index=text_columns.index("Content") if "Content" in text_columns else 0,
        help="Select the column containing the text you want to analyse",
        key="bert_content_column",
    )

    # Filtering options
    st.subheader("Select Documents to Analyse")

    # Option to select all or specific records
    analysis_type = st.radio(
        "Analysis Type",
        ["All Reports", "Selected Reports"],
        horizontal=True,
        key="bert_analysis_type",
    )

    if analysis_type == "Selected Reports":
        # Multi-select for reports
        selected_indices = st.multiselect(
            "Choose specific reports to analyse",
            options=list(range(len(data))),
            format_func=lambda x: f"{data.iloc[x]['Title']} ({data.iloc[x]['date_of_report'].strftime('%d/%m/%Y') if pd.notna(data.iloc[x]['date_of_report']) else 'No date'})",
            key="bert_selected_indices",
        )
        selected_data = data.iloc[selected_indices] if selected_indices else None
    else:
        selected_data = data

    # Analysis parameters
    st.subheader("Analysis Parameters")
    similarity_threshold = st.slider(
        "Similarity Threshold",
        min_value=0.3,
        max_value=0.9,
        value=0.65,
        step=0.05,
        help="Minimum similarity score for theme detection (higher = more strict)",
        key="bert_similarity_threshold",
    )

    # Analysis button
    run_analysis = st.button(
        "Run Analysis", type="primary", key="bert_run_analysis"
    )

    # Run analysis if button is clicked
    if run_analysis:
        with st.spinner("Performing Theme Analysis..."):
            try:
                # Validate data selection
                if selected_data is None or len(selected_data) == 0:
                    st.warning("No documents selected for analysis.")
                    return

                # Initialize the theme analyzer (with loading message in a spinner)
                with st.spinner("Loading annotation model and tokenizer..."):
                    # Initialize the analyzer
                    theme_analyzer = ThemeAnalyzer(
                        model_name="emilyalsentzer/Bio_ClinicalBERT"
                    )
                    
                    # Mark as initialized
                    st.session_state.bert_initialized = True
                
                # Set custom configuration
                theme_analyzer.config[
                    "base_similarity_threshold"
                ] = similarity_threshold
                
                # Filter frameworks based on user selection
                filtered_frameworks = {}
                
                # Add selected built-in frameworks
                for framework in st.session_state.selected_frameworks:
                    if framework == "I-SIRch":
                        filtered_frameworks["I-SIRch"] = theme_analyzer._get_isirch_framework()
                    elif framework == "House of Commons":
                        filtered_frameworks["House of Commons"] = theme_analyzer._get_house_of_commons_themes()
                    elif framework == "Extended Analysis":
                        filtered_frameworks["Extended Analysis"] = theme_analyzer._get_extended_themes()
                    elif framework in st.session_state.custom_frameworks:
                        # Add custom framework
                        filtered_frameworks[framework] = st.session_state.custom_frameworks[framework]
                
                # Set the filtered frameworks
                theme_analyzer.frameworks = filtered_frameworks
                
                # If no frameworks selected, show error
                if not filtered_frameworks:
                    st.error("Please select at least one framework for analysis.")
                    return

                # Use the enhanced create_detailed_results method
                (
                    results_df,
                    highlighted_texts,
                ) = theme_analyzer.create_detailed_results(
                    selected_data, content_column=content_column
                )

                # Save results to session state to ensure persistence
                st.session_state.bert_results["results_df"] = results_df
                st.session_state.bert_results["highlighted_texts"] = highlighted_texts

                st.success(f"Analysis complete using {len(filtered_frameworks)} frameworks!")

            except Exception as e:
                st.error(f"Error during annotation analysis: {str(e)}")
                logging.error(f"Annotation analysis error: {e}", exc_info=True)

    # Display results if they exist
    if "bert_results" in st.session_state and st.session_state.bert_results.get("results_df") is not None:
        results_df = st.session_state.bert_results["results_df"]
        
        # Summary stats
        st.subheader("Results")
        
        # Show framework distribution
        if "Framework" in results_df.columns:
            framework_counts = results_df["Framework"].value_counts()
            
            # Create columns for framework distribution metrics
            framework_cols = st.columns(len(framework_counts))
            
            for i, (framework, count) in enumerate(framework_counts.items()):
                with framework_cols[i]:
                    st.metric(framework, count, help=f"Number of theme identifications from {framework} framework")
        
        st.write(f"Total Theme Identifications: {len(results_df)}")
        
        # Clean up the results DataFrame to display only the essential columns
        display_cols = ["Record ID", "Title", "Framework", "Theme", "Confidence", "Combined Score", "Matched Keywords"]
        
        # Add metadata columns if available
        for col in ["coroner_name", "coroner_area", "year", "date_of_report"]:
            if col in results_df.columns:
                display_cols.append(col)
        
        # Add matched sentences at the end
        if "Matched Sentences" in results_df.columns:
            display_cols.append("Matched Sentences")
        
        # Create the display DataFrame with only existing columns
        valid_cols = [col for col in display_cols if col in results_df.columns]
        clean_df = results_df[valid_cols].copy()
        
        # Display the results table
        st.dataframe(
            clean_df,
            use_container_width=True,
            column_config={
                "Title": st.column_config.TextColumn("Document Title"),
                "Framework": st.column_config.TextColumn("Framework"),
                "Theme": st.column_config.TextColumn("Theme"),
                "Confidence": st.column_config.TextColumn("Confidence"),
                "Combined Score": st.column_config.NumberColumn("Score", format="%.3f"),
                "Matched Keywords": st.column_config.TextColumn("Keywords"),
                "Matched Sentences": st.column_config.TextColumn("Matched Sentences"),
                "coroner_name": st.column_config.TextColumn("Coroner Name"),
                "coroner_area": st.column_config.TextColumn("Coroner Area"),
                "year": st.column_config.NumberColumn("Year"),
                "date_of_report": st.column_config.DateColumn("Date of Report", format="DD/MM/YYYY")
            }
        )
        
        # Add download options
        st.subheader("Export Results")
        
        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create columns for download buttons
        col1, col2 = st.columns(2)
        
        with col1:
            # Excel download button using the enhanced export_to_excel function
            excel_data = export_to_excel(clean_df)
            st.download_button(
                "📥 Download Results Table",
                data=excel_data,
                file_name=f"annotated_theme_analysis_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="bert_excel_download",
            )
        
        with col2:
            # Always regenerate HTML report when results are available
            if "results_df" in st.session_state.bert_results and "highlighted_texts" in st.session_state.bert_results:
                # Generate fresh HTML report based on current results
                theme_analyzer = ThemeAnalyzer()
                
                # Set custom frameworks if they exist
                if st.session_state.custom_frameworks:
                    for name, framework in st.session_state.custom_frameworks.items():
                        if name in st.session_state.selected_frameworks:
                            theme_analyzer.frameworks[name] = framework
                
                html_content = theme_analyzer._create_integrated_html_for_pdf(
                    results_df, st.session_state.bert_results["highlighted_texts"]
                )
                html_filename = f"theme_analysis_report_{timestamp}.html"
                
                with open(html_filename, "w", encoding="utf-8") as f:
                    f.write(html_content)
                    
                st.session_state.bert_results["html_filename"] = html_filename
                
                # Provide download button for fresh HTML
                with open(html_filename, "rb") as f:
                    html_data = f.read()
                
                st.download_button(
                    "📄 Download Annotated Reports (HTML)",
                    data=html_data,
                    file_name=os.path.basename(html_filename),
                    mime="text/html",
                    key="bert_html_download",
                )
            else:
                st.warning("HTML report not available")


def check_app_password():
    """Check if user has entered the correct password to access the app"""
    # Initialize session state for authentication
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    
    # If already authenticated, continue
    if st.session_state.authenticated:
        return True
    
    # Otherwise show login screen
    st.title("UK Judiciary PFD Reports Analysis")
    st.markdown("### Authentication Required")
    st.markdown("Please enter the password to access the application.")
    
    # Password input
    password = st.text_input("Password", type="password")
    
    # Submit button
    if st.button("Login"):
        # Get correct password from secrets.toml
        correct_password = st.secrets.get("app_password")
        
        if password == correct_password:
            st.session_state.authenticated = True
            st.success("Login successful!")
            st.rerun()
            return True
        else:
            st.error("Incorrect password. Please try again.")
            return False
    
    return False

def render_bert_file_merger():
    """Render the BERT file merger tab in the Streamlit app with custom initialization."""
    # Create an instance of the analyzer
    analyzer = BERTResultsAnalyzer()
    
    # Add tabs for Merger and Filter functionality
    merger_tab, filter_tab = st.tabs(["Merge & Process Files", "Filter & Explore Data"])
    
    with merger_tab:
        # Skip the standard render_analyzer_ui and call the file upload directly
        analyzer._render_multiple_file_upload()
    
    with filter_tab:
        render_filter_data_tab()



# Helper function for advanced keyword search
def perform_advanced_keyword_search(text, search_query):
    """
    Perform advanced keyword search with AND/OR operators
    
    Args:
        text (str): Text to search within
        search_query (str): Search query with AND/OR operators
    
    Returns:
        bool: True if text matches search criteria, False otherwise
    """
    if not text or not search_query:
        return False
    
    # Convert text to lowercase for case-insensitive search
    text = str(text).lower()
    
    # Check if we have an AND search (contains all keywords)
    if " and " in search_query.lower():
        keywords = [k.strip() for k in search_query.lower().split(" and ")]
        return all(keyword.lower() in text for keyword in keywords if keyword)
    
    # Check if we have an OR search (contains any keyword)
    elif " or " in search_query.lower():
        keywords = [k.strip() for k in search_query.lower().split(" or ")]
        return any(keyword.lower() in text for keyword in keywords if keyword)
    
    # Default to exact match for single keywords
    else:
        return search_query.lower() in text

def improved_truncate_text(text, max_length=30):
    """
    Improved function to handle long text for chart display by breaking into lines
    instead of simple truncation with ellipses
    
    Args:
        text: String to format
        max_length: Maximum length per line
        
    Returns:
        Text with line breaks inserted at appropriate word boundaries
    """
    if not text or len(text) <= max_length:
        return text
    
    # For theme names with ":" in them (framework:theme format)
    if ":" in text:
        parts = text.split(":", 1)
        framework = parts[0].strip()
        theme = parts[1].strip()
        
        # For theme part, break it into lines rather than truncate
        if len(theme) > max_length - len(framework) - 2:  # -2 for ": "
            # Process the theme part with word-aware line breaking
            words = theme.split()
            processed_theme = []
            current_line = []
            current_length = 0
            
            for word in words:
                # If adding this word keeps us under the limit
                if current_length + len(word) + (1 if current_line else 0) <= max_length - len(framework) - 2:
                    current_line.append(word)
                    current_length += len(word) + (1 if current_line else 0)
                else:
                    # Line is full, start a new one
                    processed_theme.append(" ".join(current_line))
                    current_line = [word]
                    current_length = len(word)
            
            # Add the final line if any
            if current_line:
                processed_theme.append(" ".join(current_line))
            
            # If we have more than 2 lines, keep first 2 and add ellipsis
            if len(processed_theme) > 2:
                return f"{framework}: {processed_theme[0]}<br>{processed_theme[1]}..."
            else:
                # Join with line breaks
                return f"{framework}: {('<br>').join(processed_theme)}"
        
        return f"{framework}: {theme}"
    
    # For normal long strings, add line breaks at word boundaries
    words = text.split()
    lines = []
    current_line = []
    current_length = 0
    
    for word in words:
        if current_length + len(word) + (1 if current_line else 0) <= max_length:
            current_line.append(word)
            current_length += len(word) + (1 if current_line else 0)
        else:
            lines.append(" ".join(current_line))
            current_line = [word]
            current_length = len(word)
    
    if current_line:
        lines.append(" ".join(current_line))
    
    # If we have more than 2 lines, keep first 2 and add ellipsis
    if len(lines) > 2:
        return f"{lines[0]}<br>{lines[1]}..."
    
    # For plotly charts, use <br> for line breaks
    return "<br>".join(lines)
    
def render_theme_analysis_dashboard(data: pd.DataFrame = None):
    """
    Render a comprehensive dashboard for analyzing themes by various metadata fields
    
    Args:
        data: Optional DataFrame containing theme analysis results
    """
    #st.title("Theme Analysis Dashboard")
    
    # Check for existing data in session state
    if data is None:
        if "dashboard_data" in st.session_state:
            data = st.session_state.dashboard_data
    
    # File upload section with persistent state
    upload_key = "theme_analysis_dashboard_uploader"
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for Dashboard Analysis",
        type=["csv", "xlsx"],
        key=upload_key
    )
    
    # Process uploaded file
    if uploaded_file is not None:
        try:
            # Load the file based on its type
            if uploaded_file.name.endswith('.csv'):
                data = pd.read_csv(uploaded_file)
            else:
                data = pd.read_excel(uploaded_file)
            
            # Process the data to ensure it's clean
            data = process_scraped_data(data)
            
            # Store in session state
            st.session_state.dashboard_data = data
            
            st.success(f"File uploaded successfully! Found {len(data)} records.")
        except Exception as e:
            st.error(f"Error processing file: {e}")
            return
    
    # If no data is available after upload
    if data is None or len(data) == 0:
        with st.expander("💡 How to get theme analysis data?"):
            st.markdown("""
            #### To get theme analysis data:
        
            1. **Upload Existing Results**
               - Use the file uploader above to load previously saved theme analysis results
            
            2. **Run New Theme Analysis**
               - Go to the 'Concept Annotation' tab 
               - Upload your merged PFD reports file
               - Run a new theme analysis
            """)
            
        return  # Exit the function if no data
    
    # Validate required columns
    required_cols = ["Framework", "Theme"]
    recommended_cols = ["coroner_area", "coroner_name", "year"]
    
    missing_required = [col for col in required_cols if col not in data.columns]
    missing_recommended = [col for col in recommended_cols if col not in data.columns]
    
    if missing_required:
        st.error(f"Missing required columns: {', '.join(missing_required)}")
        return
    
    if missing_recommended:
        st.warning(f"Some recommended columns are missing: {', '.join(missing_recommended)}")
    
    # Data Overview
    st.subheader("Data Overview")
    metrics_col1, metrics_col2, metrics_col3, metrics_col4 = st.columns(4)
    
    with metrics_col1:
        st.metric("Total Theme Identifications", len(data))
    with metrics_col2:
        st.metric("Unique Themes", data["Theme"].nunique())
    
    with metrics_col3:
        if "coroner_area" in data.columns and not data["coroner_area"].isna().all():
            st.metric("Coroner Areas", data["coroner_area"].nunique())
        else:
            st.metric("Coroner Areas", "N/A")
    
    with metrics_col4:
        if "year" in data.columns and not data["year"].isna().all():
            years_count = data["year"].dropna().nunique()
            year_text = f"{years_count}" if years_count > 0 else "N/A"
            st.metric("Years Covered", year_text)
        else:
            st.metric("Years Covered", "N/A")
            
    results_df = data.copy() if data is not None else pd.DataFrame()
    
    #
    
    # Sidebar filters
    # Find the sidebar filter section in the render_theme_analysis_dashboard function 
    # and replace it with this improved version:
    
    # Sidebar filters
    st.sidebar.header("Dashboard Filters")
    
    # Framework filter
    frameworks = ["All"] + sorted(results_df["Framework"].unique().tolist())
    selected_framework = st.sidebar.selectbox("Filter by Framework", frameworks)
    
    # Year filter - Modified to handle single year selection properly
    years = sorted(results_df["year"].dropna().unique().tolist())
    if years:
        min_year, max_year = min(years), max(years)
        
        # If only one year available, provide a checkbox instead of slider
        if min_year == max_year:
            include_year = st.sidebar.checkbox(f"Include year {min_year}", value=True)
            if include_year:
                selected_years = (min_year, max_year)
            else:
                selected_years = None
        else:
            # For multiple years, keep using the slider
            selected_years = st.sidebar.slider(
                "Year Range", min_year, max_year, (min_year, max_year)
            )
    else:
        selected_years = None
    
    # Coroner area filter - MODIFIED: Multi-select instead of single select
    areas = sorted(results_df["coroner_area"].dropna().unique().tolist())
    area_options = ["All Areas"] + areas
    # Default to "All Areas" if no specific selection
    area_filter_type = st.sidebar.radio("Coroner Area Filter Type", ["All Areas", "Select Specific Areas"])
    if area_filter_type == "All Areas":
        selected_areas = areas  # Include all areas
    else:
        # Multi-select for specific areas
        selected_areas = st.sidebar.multiselect(
            "Select Areas", 
            options=areas,
            default=None,
            help="Select one or more specific coroner areas to include"
        )
        # If nothing selected, default to all areas
        if not selected_areas:
            st.sidebar.warning("No areas selected. Showing all areas.")
            selected_areas = areas
    
    # Coroner name filter - MODIFIED: Multi-select instead of single select
    names = sorted(results_df["coroner_name"].dropna().unique().tolist())
    name_filter_type = st.sidebar.radio("Coroner Name Filter Type", ["All Coroners", "Select Specific Coroners"])
    if name_filter_type == "All Coroners":
        selected_names = names  # Include all coroner names
    else:
        # Multi-select for specific coroner names
        selected_names = st.sidebar.multiselect(
            "Select Coroners", 
            options=names,
            default=None,
            help="Select one or more specific coroners to include"
        )
        # If nothing selected, default to all names
        if not selected_names:
            st.sidebar.warning("No coroners selected. Showing all coroners.")
            selected_names = names
    
    # Number of top themes to display
    top_n_themes = st.sidebar.slider("Number of Top Themes", 5, 20, 10)
    
    # Confidence filter - MODIFIED: Multi-select instead of minimum
    confidence_levels = ["High", "Medium", "Low"]
    confidence_filter_type = st.sidebar.radio("Confidence Filter Type", ["All Confidence Levels", "Select Specific Levels"])
    if confidence_filter_type == "All Confidence Levels":
        selected_confidence_levels = confidence_levels  # Include all confidence levels
    else:
        # Multi-select for specific confidence levels
        selected_confidence_levels = st.sidebar.multiselect(
            "Select Confidence Levels", 
            options=confidence_levels,
            default=["High", "Medium"],  # Default to high and medium
            help="Select one or more confidence levels to include"
        )
        # If nothing selected, default to all confidence levels
        if not selected_confidence_levels:
            st.sidebar.warning("No confidence levels selected. Showing all levels.")
            selected_confidence_levels = confidence_levels
    
    # Apply filters - UPDATED to handle new multi-select filters
    filtered_df = results_df.copy()
    
    if selected_framework != "All":
        filtered_df = filtered_df[filtered_df["Framework"] == selected_framework]
    
    if selected_years:
        # Handle edge case of single year (where both values are the same)
        if selected_years[0] == selected_years[1]:
            filtered_df = filtered_df[filtered_df["year"] == selected_years[0]]
        else:
            filtered_df = filtered_df[(filtered_df["year"] >= selected_years[0]) & 
                                    (filtered_df["year"] <= selected_years[1])]
    
    # Apply multi-select area filter
    filtered_df = filtered_df[filtered_df["coroner_area"].isin(selected_areas)]
    
    # Apply multi-select coroner name filter
    filtered_df = filtered_df[filtered_df["coroner_name"].isin(selected_names)]
    
    # Apply multi-select confidence level filter
    filtered_df = filtered_df[filtered_df["Confidence"].isin(selected_confidence_levels)]
    
    # Display filter summary
    active_filters = []
    if selected_framework != "All":
        active_filters.append(f"Framework: {selected_framework}")
    if selected_years:
        if selected_years[0] == selected_years[1]:
            active_filters.append(f"Year: {selected_years[0]}")
        else:
            active_filters.append(f"Years: {selected_years[0]}-{selected_years[1]}")
    if area_filter_type == "Select Specific Areas" and selected_areas:
        if len(selected_areas) <= 3:
            active_filters.append(f"Areas: {', '.join(selected_areas)}")
        else:
            active_filters.append(f"Areas: {len(selected_areas)} selected")
    if name_filter_type == "Select Specific Coroners" and selected_names:
        if len(selected_names) <= 3:
            active_filters.append(f"Coroners: {', '.join(selected_names)}")
        else:
            active_filters.append(f"Coroners: {len(selected_names)} selected")
    if confidence_filter_type == "Select Specific Levels" and selected_confidence_levels:
        active_filters.append(f"Confidence: {', '.join(selected_confidence_levels)}")
    
    if active_filters:
        st.info("Active filters: " + " | ".join(active_filters))
    
    if len(filtered_df) == 0:
        st.warning("No data matches the selected filters. Please adjust your filters.")
        return
        
    # Continue with the rest of the dashboard code...
    # Create tabs for different visualizations
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Framework Heatmap", 
        "Theme Distribution", 
        "Temporal Analysis", 
        "Area Comparison",
        "Correlation Analysis"
    ])
    
    # === TAB 1: FRAMEWORK HEATMAP ===
    with tab1:
        st.subheader("Framework Theme Heatmap by Year")
        
        if "year" not in filtered_df.columns or filtered_df["year"].isna().all():
            st.warning("No year data available for temporal analysis.")
        else:
            # Handle special case where we only have one year
            if filtered_df["year"].nunique() == 1:
                st.info(f"Showing data for year {filtered_df['year'].iloc[0]}")
                
                # Create a simplified categorical count visualization for single year
                theme_counts = filtered_df.groupby(['Framework', 'Theme']).size().reset_index(name='Count')
                
                # Sort by framework and count
                theme_counts = theme_counts.sort_values(['Framework', 'Count'], ascending=[True, False])
                
                # Process theme names for better display using improved function
                theme_counts['Display_Theme'] = theme_counts['Theme'].apply(
                    lambda x: improved_truncate_text(x, max_length=40)
                )
                
                # Display as a horizontal bar chart grouped by framework
                fig = px.bar(
                    theme_counts,
                    y='Display_Theme',  # Use formatted theme names
                    x='Count',
                    color='Framework',
                    title=f"Theme Distribution for Year {filtered_df['year'].iloc[0]}",
                    height=max(500, len(theme_counts) * 30),
                    color_discrete_map={
                        "I-SIRch": "orange",
                        "House of Commons": "royalblue",
                        "Extended Analysis": "firebrick"
                    }
                )
                
                fig.update_layout(
                    xaxis_title="Number of Reports",
                    yaxis_title="Theme",
                    yaxis={'categoryorder': 'total ascending'},
                    font=dict(family="Arial, sans-serif", color="white"),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    margin=dict(l=250, r=40, t=80, b=60)  # Increased left margin for theme labels
                )
                
                # Update axes for dark mode
                fig.update_xaxes(
                    title_font=dict(color="white"),
                    tickfont=dict(color="white"),
                    gridcolor="rgba(255,255,255,0.1)"
                )
                
                fig.update_yaxes(
                    title_font=dict(color="white"),
                    tickfont=dict(color="white"),
                    automargin=True  # Enable automargin to ensure labels fit
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                # Regular heatmap code for multiple years
                # Create combined framework:theme field
                filtered_df['Framework_Theme'] = filtered_df['Framework'] + ': ' + filtered_df['Theme']
                
                # Count reports per year (for denominator)
                # Assuming Record ID is the unique identifier for reports
                id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
                reports_per_year = filtered_df.groupby('year')[id_column].nunique()
                
                # Count unique report IDs per theme per year
                counts = filtered_df.groupby(['year', 'Framework', 'Framework_Theme'])[id_column].nunique().reset_index()
                counts.columns = ['year', 'Framework', 'Framework_Theme', 'Count']
                
                # Calculate percentages
                counts['Total'] = counts['year'].map(reports_per_year)
                counts['Percentage'] = (counts['Count'] / counts['Total'] * 100).round(1)
                
                # Get frameworks in the filtered data
                frameworks_present = filtered_df['Framework'].unique()
                
                # Get top themes by framework (5 per framework)
                top_themes = []
                for framework in frameworks_present:
                    framework_counts = counts[counts['Framework'] == framework]
                    theme_totals = framework_counts.groupby('Framework_Theme')['Count'].sum().sort_values(ascending=False)
                    top_themes.extend(theme_totals.head(5).index.tolist())
                
                # Filter to top themes
                counts = counts[counts['Framework_Theme'].isin(top_themes)]
                
                # Create pivot table for heatmap
                pivot = counts.pivot_table(
                    index='Framework_Theme',
                    columns='year',
                    values='Percentage',
                    fill_value=0
                )
                
                # Create pivot for counts
                count_pivot = counts.pivot_table(
                    index='Framework_Theme',
                    columns='year',
                    values='Count',
                    fill_value=0
                )
                
                # Sort by framework then by total count
                theme_totals = counts.groupby('Framework_Theme')['Count'].sum()
                theme_frameworks = {theme: theme.split(':')[0] for theme in theme_totals.index}
                
                # Sort first by framework, then by count within framework
                sorted_themes = sorted(
                    theme_totals.index,
                    key=lambda x: (theme_frameworks[x], -theme_totals[x])
                )
                
                # Apply the sort order
                pivot = pivot.reindex(sorted_themes)
                count_pivot = count_pivot.reindex(sorted_themes)
                
                # Create color mapping for frameworks
                framework_colors = {
                    "I-SIRch": "orange",
                    "House of Commons": "royalblue",
                    "Extended Analysis": "firebrick"
                }
                
                # Default colors for any frameworks not specifically mapped
                other_colors = ["forestgreen", "purple", "darkred"]
                for i, framework in enumerate(frameworks_present):
                    if framework not in framework_colors:
                        framework_colors[framework] = other_colors[i % len(other_colors)]
                
                # Create a visually distinctive dataframe for plotting
                # For each theme, create a dict with clean name and framework
                theme_display_data = []
                
                for theme in pivot.index:
                    framework = theme.split(':')[0].strip()
                    theme_name = theme.split(':', 1)[1].strip()
                    
                    # Use improved_truncate_text for line breaking instead of manual handling
                    formatted_theme = improved_truncate_text(theme_name, max_length=40)
                    
                    theme_display_data.append({
                        'original': theme,
                        'clean_name': formatted_theme,
                        'framework': framework,
                        'color': framework_colors[framework]
                    })
                    
                theme_display_df = pd.DataFrame(theme_display_data)
                
                # Add year count labels
                year_labels = [f"{year}<br>n={reports_per_year[year]}" for year in pivot.columns]
                
                # Create heatmap using plotly
                fig = go.Figure()
                
                # Add heatmap
                heatmap = go.Heatmap(
                    z=pivot.values,
                    x=year_labels,
                    y=theme_display_df['clean_name'],
                    colorscale=[
                        [0, '#f7fbff'],      # Lightest blue (almost white) for zero values
                        [0.2, '#deebf7'],    # Very light blue
                        [0.4, '#9ecae1'],    # Light blue
                        [0.6, '#4292c6'],    # Medium blue
                        [0.8, '#2171b5'],    # Deep blue
                        [1.0, '#084594']     # Darkest blue
                    ],
                    zmin=0,
                    zmax=min(100, pivot.values.max() * 1.2),  # Cap at 100% or 20% higher than max
                    colorbar=dict(
                        title=dict(text="Percentage (%)", font=dict(color="white", size=12)),
                        tickfont=dict(color="white", size=10),
                        outlinecolor="rgba(255,255,255,0.3)",
                        outlinewidth=1
                    ),
                    hoverongaps=False,
                    text=count_pivot.values,  # This will show the count in the hover
                    hovertemplate='Year: %{x}<br>Theme: %{y}<br>Percentage: %{z}%<br>Count: %{text}<extra></extra>'
                )
                
                fig.add_trace(heatmap)
                
                # Add count annotations
                for i in range(len(pivot.index)):
                    for j in range(len(pivot.columns)):
                        if pivot.iloc[i, j] > 0:
                            count = count_pivot.iloc[i, j]
                            
                            # Calculate appropriate text color based on cell value
                            bg_intensity = pivot.iloc[i, j] / 100  # Normalize to 0-1
                            text_color = 'white' if bg_intensity > 0.4 else 'black'
                            
                            # Add an annotation with outline for better visibility
                            fig.add_annotation(
                                x=j,
                                y=i,
                                text=f"({count})",
                                font=dict(size=9, color=text_color),
                                showarrow=False,
                                xanchor="center",
                                yanchor="middle",
                                bordercolor="rgba(0,0,0,0.2)",
                                borderwidth=1,
                                borderpad=2
                            )
                
                # Improved framework indicator styling
                for framework, color in framework_colors.items():
                    # Find rows corresponding to this framework
                    framework_rows = theme_display_df[theme_display_df['framework'] == framework]
                    
                    if len(framework_rows) > 0:
                        # Get the indices in the sorted display order
                        indices = framework_rows.index.tolist()
                        min_idx = min(indices)
                        max_idx = max(indices)
                        
                        # Add a colored rectangle with higher contrast
                        fig.add_shape(
                            type="rect",
                            x0=-1.3,  # Further to the left
                            x1=-0.7,  # Wider rectangle
                            y0=min_idx - 0.5,  # Align with the heatmap cells
                            y1=max_idx + 0.5,
                            fillcolor=color,
                            opacity=0.85,  # More visible
                            layer="below",
                            line=dict(width=1, color='rgba(255,255,255,0.5)')  # White border
                        )
                        
                        # Add framework label with black or white text based on background color
                        if len(framework_rows) > 1:  # Only add text if multiple themes in framework
                            # Determine text color (black for light backgrounds, white for dark)
                            text_color = "black" if framework in ["I-SIRch"] else "white"
                            
                            fig.add_annotation(
                                x=-1.0,
                                y=(min_idx + max_idx) / 2,
                                text=framework,
                                showarrow=False,
                                textangle=90,  # Vertical text
                                font=dict(size=12, color=text_color, family="Arial, sans-serif"),
                                xanchor="center",
                                yanchor="middle"
                            )
                
                # Update layout for better readability on dark background
                fig.update_layout(
                    title="Framework Theme Heatmap by Year",
                    font=dict(family="Arial, sans-serif", color="white"),  # White font for dark background
                    title_font=dict(size=16, color="white"),  # Larger title with white color (fixed small size)
                    xaxis_title="Year (number of reports)",
                    yaxis_title="Theme",
                    height=max(650, len(pivot.index) * 35),  # Increased height
                    width=900,  # Set explicit width
                    margin=dict(l=250, r=60, t=80, b=80),  # Increased left margin further
                    paper_bgcolor="rgba(0,0,0,0)",  # Transparent background
                    plot_bgcolor="rgba(0,0,0,0)",  # Transparent background
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.05,  # Moved up for more space
                        xanchor="center",
                        x=0.5,
                        bgcolor="rgba(50,50,50,0.8)",  # Dark semi-transparent background
                        bordercolor="rgba(255,255,255,0.3)",
                        borderwidth=1,
                        font=dict(color="white")  # White text for legend
                    )
                )
                
                # Set y-axis formatting for dark mode
                fig.update_layout(
                    yaxis=dict(
                        tickmode='array',
                        tickvals=list(range(len(theme_display_df))),
                        ticktext=theme_display_df['clean_name'],
                        tickfont=dict(
                            size=11,
                            color='white'  # Changed to white from black for consistency
                        ),
                        automargin=True  # Added to ensure labels fit properly
                    )
                )
                # Improve x-axis formatting for dark mode
                fig.update_xaxes(
                    tickangle=-0,  # Horizontal labels
                    title_font=dict(size=14, color="white"),  # White color for axis title
                    tickfont=dict(size=12, color="white"),  # White color for tick labels
                    gridcolor="rgba(255,255,255,0.1)"  # Very subtle grid
                )
                
                # Improve y-axis formatting for dark mode
                fig.update_yaxes(
                    title_font=dict(size=14, color="white"),  # White color for axis title
                    tickfont=dict(size=11, color="white"),  # White color for tick labels
                    automargin=True  # Ensure labels fit properly
                )
                
                # Add framework legend
                for i, (framework, color) in enumerate(framework_colors.items()):
                    if framework in frameworks_present:  # Only show legends for frameworks present in data
                        fig.add_trace(go.Scatter(
                            x=[None],
                            y=[None],
                            mode='markers',
                            marker=dict(size=10, color=color),
                            name=framework,
                            showlegend=True
                        ))
                
                # Use st.plotly_chart's config parameter for better sizing
                st.plotly_chart(
                    fig, 
                    use_container_width=True,
                    config={
                        'displayModeBar': True,
                        'responsive': True,
                        'toImageButtonOptions': {
                            'format': 'png',
                            'filename': 'theme_heatmap',
                            'height': 800,
                            'width': 1200,
                            'scale': 2  # Higher resolution
                        }
                    }
                )
                
    # === TAB 2: THEME DISTRIBUTION ===
    with tab2:
        st.subheader("Theme Distribution Analysis")
        
        # Get top themes by count
        theme_counts = filtered_df["Theme"].value_counts().head(top_n_themes)
        
        # Use improved_truncate_text for better label formatting
        formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in theme_counts.index]
        
        # Create a bar chart with formatted theme names
        fig = px.bar(
            x=formatted_themes,
            y=theme_counts.values,
            labels={"x": "Theme", "y": "Count"},
            title="Top 10 Themes by Occurrence",
            height=600,  # Increased height to accommodate multi-line labels
            color_discrete_sequence=['#4287f5']  # Consistent blue color
        )
        
        # Improve layout with better spacing for multi-line text
        fig.update_layout(
            xaxis_title="Theme",
            yaxis_title="Number of Occurrences",
            xaxis={'categoryorder':'total descending'},
            xaxis_tickangle=-30,  # Less extreme angle for better readability with multi-line text
            margin=dict(l=50, r=50, b=150, t=80),  # Increased bottom margin for labels
            bargap=0.2,  # Add some gap between bars
            font=dict(family="Arial, sans-serif", color="white"),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)"
        )
        
        # Update axes for dark mode
        fig.update_xaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white"),
            gridcolor="rgba(255,255,255,0.1)",
            automargin=True  # Ensure labels fit properly without overlap
        )
        
        fig.update_yaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white"),
            gridcolor="rgba(255,255,255,0.1)"
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
        # Theme by confidence
        st.subheader("Theme Confidence Breakdown")
        
        # Group by theme and confidence
        theme_confidence = filtered_df.groupby(["Theme", "Confidence"]).size().reset_index(name="Count")
        
        # Filter for top themes only
        top_themes = theme_counts.index.tolist()
        theme_confidence = theme_confidence[theme_confidence["Theme"].isin(top_themes)]
        
        # Create a mapping dictionary for theme display names
        theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
        
        # Apply the formatting to the DataFrame
        theme_confidence["Display_Theme"] = theme_confidence["Theme"].map(theme_display_map)
        
        # Create a grouped bar chart with formatted theme names
        fig = px.bar(
            theme_confidence, 
            x="Display_Theme",  # Use the formatted theme names
            y="Count", 
            color="Confidence",
            barmode="group",
            color_discrete_map={"High": "#4CAF50", "Medium": "#FFC107", "Low": "#F44336"},
            category_orders={
                "Confidence": ["High", "Medium", "Low"],
                "Display_Theme": [theme_display_map[theme] for theme in top_themes]
            },
            title="Confidence Distribution by Theme",
            height=600  # Increased height for better readability
        )
        
        # Improve layout for multi-line labels
        fig.update_layout(
            xaxis_title="Theme",
            yaxis_title="Number of Reports",
            xaxis_tickangle=-30,  # Less extreme angle for better readability
            margin=dict(l=50, r=50, b=150, t=80),  # Increased bottom margin
            legend=dict(
                orientation="h", 
                yanchor="bottom", 
                y=1.02, 
                xanchor="center", 
                x=0.5,
                title=None,
                font=dict(color="white")
            ),
            font=dict(family="Arial, sans-serif", color="white"),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)"
        )
        
        # Update axes for dark mode and ensure labels fit
        fig.update_xaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white"),
            gridcolor="rgba(255,255,255,0.1)",
            automargin=True  # Ensure labels fit properly without overlap
        )
        
        fig.update_yaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white"),
            gridcolor="rgba(255,255,255,0.1)"
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # === TAB 3: TEMPORAL ANALYSIS ===
    with tab3:
        st.subheader("Temporal Analysis")
        
        if "year" not in filtered_df.columns or filtered_df["year"].isna().all():
            st.warning("No year data available for temporal analysis.")
        else:
            # Create a time series of themes by year
            year_theme_counts = filtered_df.groupby(["year", "Theme"]).size().reset_index(name="Count")
            
            # Filter for top themes only - get top themes by total count
            all_theme_counts = filtered_df["Theme"].value_counts()
            top_themes = all_theme_counts.head(top_n_themes).index.tolist()
            
            year_theme_counts = year_theme_counts[year_theme_counts["Theme"].isin(top_themes)]
            
            # Create a mapping dictionary for formatted theme names
            theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
            
            # Apply the formatting to the DataFrame
            year_theme_counts["Display_Theme"] = year_theme_counts["Theme"].map(theme_display_map)
            
            # Create a line chart - important fix: convert year to string to treat as categorical
            year_theme_counts['year_str'] = year_theme_counts['year'].astype(str)
            
            fig = px.line(
                year_theme_counts,
                x="year_str",  # Use string version of year
                y="Count",
                color="Display_Theme",  # Use formatted theme names
                markers=True,
                title="Theme Trends Over Time",
                height=600,  # Increased height
            )
            
            # Improve layout
            fig.update_layout(
                xaxis_title="Year",
                yaxis_title="Number of Occurrences",
                xaxis=dict(
                    type='category',  # Force categorical x-axis
                    tickmode="array",
                    tickvals=sorted(year_theme_counts['year_str'].unique()),
                    ticktext=sorted(year_theme_counts['year_str'].unique()),
                ),
                # Move legend below the chart for more horizontal space and prevent overlap
                legend=dict(
                    orientation="h", 
                    yanchor="top", 
                    y=-0.2,  # Position below the chart
                    xanchor="center", 
                    x=0.5,
                    title=None  # Remove legend title
                ),
                margin=dict(l=50, r=50, b=150, t=80),  # Increase bottom margin for legend
                font=dict(color="white"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
            )
            
            # Update axes for dark mode
            fig.update_xaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                gridcolor="rgba(255,255,255,0.1)"
            )
            
            fig.update_yaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                gridcolor="rgba(255,255,255,0.1)"
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Heatmap of themes by year
            st.subheader("Theme Prevalence by Year")
            
            # Create a pivot table
            pivot_df = year_theme_counts.pivot(index="Theme", columns="year_str", values="Count").fillna(0)
            
            # Convert to a normalized heatmap (percentage)
            # Calculate the total themes per year
            year_theme_totals = pivot_df.sum(axis=0)
            normalized_pivot = pivot_df.div(year_theme_totals, axis=1) * 100
            
            # Format the theme names for better display
            formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in normalized_pivot.index]
            
            # Create a heatmap - ensure years are in correct order
            year_order = sorted(year_theme_counts['year'].unique())
            year_order_str = [str(y) for y in year_order]
            
            fig = px.imshow(
                normalized_pivot[year_order_str],  # Ensure columns are in correct order
                labels=dict(x="Year", y="Theme", color="% of Themes"),
                x=year_order_str,  # Use sorted string years
                y=formatted_themes,  # Use formatted theme names
                color_continuous_scale=[
                    [0, '#f7fbff'],      # Lightest blue (almost white) for zero values
                    [0.2, '#deebf7'],    # Very light blue
                    [0.4, '#9ecae1'],    # Light blue
                    [0.6, '#4292c6'],    # Medium blue
                    [0.8, '#2171b5'],    # Deep blue
                    [1.0, '#084594']     # Darkest blue
                ],
                title="Theme Prevalence by Year (%)",
                height=600,
                aspect="auto",
                text_auto=".1f"  # Show percentage to 1 decimal place
            )
            
            fig.update_layout(
                xaxis=dict(
                    type='category',  # Force categorical x-axis
                    tickmode="array",
                    tickvals=year_order_str,
                    ticktext=year_order_str,
                ),
                margin=dict(l=250, r=50, t=80, b=50),  # Increased left margin for theme labels
                font=dict(color="white"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
            )
            
            # Update axes and colorbar for dark mode
            fig.update_xaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                automargin=True  # Ensure labels don't overlap
            )
            
            fig.update_yaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                automargin=True  # Ensure y-axis labels fit
            )
            
            fig.update_traces(
                colorbar=dict(
                    title=dict(text="% of Themes", font=dict(color="white")),
                    tickfont=dict(color="white")
                )
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    # === TAB 4: AREA COMPARISON ===
    with tab4:
        st.subheader("Coroner Area Comparison")
        
        if "coroner_area" not in filtered_df.columns or filtered_df["coroner_area"].isna().all():
            st.warning("No coroner area data available for area comparison.")
        else:
            # Get the top areas by theme count
            area_counts = filtered_df["coroner_area"].value_counts().head(10)
            top_areas = area_counts.index.tolist()
            
            # Format area names for better display
            formatted_areas = [improved_truncate_text(area, max_length=40) for area in area_counts.index]
            
            # Create a mapping for display names
            area_display_map = dict(zip(area_counts.index, formatted_areas))
            
            # Create a bar chart of top areas with formatted names
            fig = px.bar(
                x=formatted_areas,
                y=area_counts.values,
                labels={"x": "Coroner Area", "y": "Count"},
                title="Theme Identifications by Coroner Area",
                height=500,
                color_discrete_sequence=['#ff9f40']  # Orange color for areas
            )
            
            fig.update_layout(
                xaxis_title="Coroner Area",
                yaxis_title="Number of Theme Identifications",
                xaxis_tickangle=-30,  # Less extreme angle for readability
                margin=dict(l=50, r=50, b=150, t=80),  # Increased bottom margin
                font=dict(color="white"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
            )
            
            # Update axes for dark mode and ensure labels fit
            fig.update_xaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                gridcolor="rgba(255,255,255,0.1)",
                automargin=True  # Ensure labels fit without overlap
            )
            
            fig.update_yaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white"),
                gridcolor="rgba(255,255,255,0.1)"
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Multi-area theme comparison
            st.subheader("Theme Distribution Across Top Coroner Areas")
            
            # Calculate total records per area (for normalization)
            area_totals = {}
            for area in top_areas:
                area_totals[area] = len(filtered_df[filtered_df["coroner_area"] == area])
            
            # Get theme distribution for each area
            area_theme_data = []
            
            # Get top themes overall for comparison
            all_theme_counts = filtered_df["Theme"].value_counts()
            top_themes = all_theme_counts.head(top_n_themes).index.tolist()
            
            # Create a mapping for formatted theme names
            theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
            
            for area in top_areas:
                area_df = filtered_df[filtered_df["coroner_area"] == area]
                area_themes = area_df["Theme"].value_counts()
                
                # Calculate percentage for each top theme
                for theme in top_themes:
                    count = area_themes.get(theme, 0)
                    percentage = (count / area_totals[area] * 100) if area_totals[area] > 0 else 0
                    
                    area_theme_data.append({
                        "Coroner Area": area,
                        "Display_Area": area_display_map[area],
                        "Theme": theme,
                        "Display_Theme": theme_display_map[theme],
                        "Count": count,
                        "Percentage": round(percentage, 1)
                    })
            
            area_theme_df = pd.DataFrame(area_theme_data)
            
            # Create heatmap using formatted names
            pivot_df = area_theme_df.pivot(
                index="Display_Area", 
                columns="Display_Theme", 
                values="Percentage"
            ).fillna(0)
            
            # Ensure we have data to display
            if not pivot_df.empty:
                fig = px.imshow(
                    pivot_df,
                    labels=dict(x="Theme", y="Coroner Area", color="Percentage"),
                    x=pivot_df.columns,
                    y=pivot_df.index,
                    color_continuous_scale="YlGnBu",
                    title="Theme Distribution by Coroner Area (%)",
                    height=700,  # Increased height
                    aspect="auto",
                    text_auto=".1f"  # Show to 1 decimal place
                )
                
                fig.update_layout(
                    xaxis_title="Theme",
                    yaxis_title="Coroner Area",
                    xaxis_tickangle=-30,  # Reduce angle
                    coloraxis_colorbar=dict(
                        title=dict(text="% of Cases", font=dict(color="white")),
                        tickfont=dict(color="white")
                    ),
                    margin=dict(l=250, r=70, b=180, t=80),  # Increased margins
                    font=dict(color="white"),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                )
                
                # Enable automargin to ensure labels fit
                fig.update_xaxes(automargin=True)
                fig.update_yaxes(automargin=True)
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("Not enough data to create area-theme heatmap.")
                
            # Radar chart option for areas
            st.subheader("Theme Radar Comparison")
            
            # Select areas for radar chart
            radar_areas = st.multiselect(
                "Select Areas to Compare (2-5 recommended)",
                options=top_areas,
                default=top_areas[:3] if len(top_areas) >= 3 else top_areas
            )
            
            if radar_areas and len(radar_areas) >= 2:
                # Filter data for selected areas and top themes
                radar_data = area_theme_df[
                    (area_theme_df["Coroner Area"].isin(radar_areas)) & 
                    (area_theme_df["Theme"].isin(top_themes[:8]))  # Limit to 8 themes for readability
                ]
                
                # Create radar chart
                fig = go.Figure()
                
                # Add traces for each area
                for area in radar_areas:
                    area_data = radar_data[radar_data["Coroner Area"] == area]
                    # Sort by theme to ensure consistency
                    area_data = area_data.set_index("Theme").reindex(top_themes[:8]).reset_index()
                    
                    fig.add_trace(go.Scatterpolar(
                        r=area_data["Percentage"],
                        theta=area_data["Display_Theme"],  # Use formatted theme names
                        fill="toself",
                        name=area_display_map.get(area, area)  # Use formatted area names
                    ))
                
                fig.update_layout(
                    polar=dict(
                        radialaxis=dict(
                            visible=True,
                            range=[0, max(radar_data["Percentage"]) * 1.1],
                            tickfont=dict(color="white")
                        ),
                        angularaxis=dict(
                            tickfont=dict(color="white")
                        )
                    ),
                    showlegend=True,
                    legend=dict(font=dict(color="white")),
                    title=dict(
                        text="Theme Distribution Radar Chart",
                        font=dict(color="white")
                    ),
                    height=700,  # Increased height
                    margin=dict(l=80, r=80, t=100, b=80),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Please select at least 2 areas for radar comparison.")
            
    # === TAB 5: CORRELATION ANALYSIS ===
    with tab5:
        st.subheader("Theme Correlation Analysis")
        
        # Create correlation explanation
        st.markdown("""
        This analysis reveals relationships between themes. A high correlation suggests that when one theme appears, 
        the other is likely to appear in the same documents as well.
        """)
        
        # Calculate correlation between themes
        # First, pivot the data to get a binary matrix of themes by report
        id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
        
        # Create a binary pivot table: 1 if theme exists for a report, 0 otherwise
        theme_pivot = pd.crosstab(
            index=filtered_df[id_column], 
            columns=filtered_df['Theme'],
            values=filtered_df.get('Combined Score', 1),  # Use the score if we want weighted values
            aggfunc='max'  # Take the maximum score for each theme in a report
        ).fillna(0)
        
        # Convert to binary (1 if theme exists, 0 otherwise)
        theme_pivot = (theme_pivot > 0).astype(int)
        
        # Calculate correlation between themes
        theme_corr = theme_pivot.corr()
        
        # Get only the top themes for clarity
        top_theme_corr = theme_corr.loc[top_themes, top_themes]
        
        # Create a mapping dictionary for theme display names
        theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
        
        # Format column and index labels
        formatted_themes = [theme_display_map[theme] for theme in top_theme_corr.columns]
        
        # Create the correlation matrix visualization
        fig_corr_matrix = px.imshow(
            top_theme_corr,
            color_continuous_scale=px.colors.diverging.RdBu_r,  # Red-Blue diverging colorscale
            color_continuous_midpoint=0,
            labels=dict(x="Theme", y="Theme", color="Correlation"),
            title="Theme Correlation Matrix",
            height=800,  # Increased height
            width=850,   # Increased width
            text_auto=".2f",  # Show correlation values with 2 decimal places
            x=formatted_themes,
            y=formatted_themes
        )
        
        # Improved layout with better label positioning
        fig_corr_matrix.update_layout(
            margin=dict(l=200, r=80, b=220, t=80),  # Dramatically increased bottom margin
            font=dict(color="white"),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(
                side="bottom",  # Place labels at the bottom
                tickangle=90,   # Vertical text instead of angled
                automargin=True # Auto-adjust margins
            ),
            yaxis=dict(
                automargin=True # Auto-adjust margins
            )
        )
        
        # Update axes for dark mode and ensure labels fit
        fig_corr_matrix.update_xaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white", size=11),
            gridcolor="rgba(255,255,255,0.1)",
            automargin=True  # Ensure labels fit properly
        )
        
        fig_corr_matrix.update_yaxes(
            title_font=dict(color="white"),
            tickfont=dict(color="white", size=11),
            gridcolor="rgba(255,255,255,0.1)",
            automargin=True  # Ensure labels fit properly
        )
        
        # Update colorbar for dark mode
        fig_corr_matrix.update_traces(
            colorbar=dict(
                title=dict(text="Correlation", font=dict(color="white")),
                tickfont=dict(color="white")
            )
        )
        
        # Display the correlation matrix with a unique key
        st.plotly_chart(fig_corr_matrix, use_container_width=True, key="theme_correlation_matrix")
        
        # Network graph of correlations
        st.subheader("Theme Connection Network (THIS WILL BE IMPROVED)")
        
        # Correlation threshold slider
        corr_threshold = st.slider(
            "Correlation Threshold", 
            min_value=0.0, 
            max_value=1.0, 
            value=0.3, 
            step=0.05,
            help="Minimum correlation value to show connections between themes",
            key="correlation_threshold_slider"
        )
        
        # Create network from correlation matrix
        G = nx.Graph()
        
        # Add nodes (themes) with formatted display names
        for theme in top_theme_corr.columns:
            G.add_node(theme, display_name=theme_display_map[theme])
        
        # Add edges (correlations above threshold)
        for i, theme1 in enumerate(top_theme_corr.columns):
            for j, theme2 in enumerate(top_theme_corr.columns):
                if i < j:  # Only process each pair once
                    correlation = top_theme_corr.loc[theme1, theme2]
                    if correlation >= corr_threshold:
                        G.add_edge(theme1, theme2, weight=correlation)
        
        # Check if we have any edges
        if len(G.edges()) == 0:
            st.warning(f"No connections found with correlation threshold of {corr_threshold}. Try lowering the threshold.")
        else:
            # Calculate positions using the Fruchterman-Reingold force-directed algorithm
            pos = nx.spring_layout(G, seed=42)  # For reproducibility
            
            # Create a network visualization
            edge_trace = []
            
            # Add edges with width proportional to correlation
            for edge in G.edges():
                x0, y0 = pos[edge[0]]
                x1, y1 = pos[edge[1]]
                weight = G[edge[0]][edge[1]]['weight']
                
                edge_trace.append(
                    go.Scatter(
                        x=[x0, x1, None],
                        y=[y0, y1, None],
                        line=dict(width=weight*5, color=f'rgba(150,150,150,{weight})'),
                        hoverinfo='none',
                        mode='lines'
                    )
                )
            
            # Add nodes
            node_x = []
            node_y = []
            node_text = []
            node_size = []
            node_hover = []
            
            for node in G.nodes():
                x, y = pos[node]
                node_x.append(x)
                node_y.append(y)
                
                # Use formatted name for display
                display_name = improved_truncate_text(node.split(':')[0] if ':' in node else node, max_length=20)
                node_text.append(display_name)
                
                # Calculate node size based on number of connections
                size = len(list(G.neighbors(node))) * 10 + 20
                node_size.append(size)
                
                # Create node text for hover
                neighbors = list(G.neighbors(node))
                connections = [f"{theme_display_map[neighbor]} (r={G[node][neighbor]['weight']:.2f})" for neighbor in neighbors]
                connection_text = "<br>".join(connections)
                node_hover.append(f"{theme_display_map[node]}<br>Connections: {len(connections)}<br>{connection_text}")
            
            node_trace = go.Scatter(
                x=node_x, y=node_y,
                mode='markers+text',
                text=node_text,
                textposition="top center",
                marker=dict(
                    size=node_size,
                    color='skyblue',
                    line=dict(width=1, color='white')
                ),
                hoverinfo='text',
                hovertext=node_hover,
                textfont=dict(color="white")
            )
            
            # Create the figure
            fig_network = go.Figure(
                data=edge_trace + [node_trace],
                layout=go.Layout(
                    title=dict(
                        text=f'Theme Connection Network (r ≥ {corr_threshold})',
                        font=dict(color="white")
                    ),
                    showlegend=False,
                    hovermode='closest',
                    margin=dict(b=20, l=5, r=5, t=40),
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    width=800,
                    height=800,
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)"
                )
            )
            
            # Display the network graph with a unique key
            st.plotly_chart(fig_network, key="theme_network_graph")
        
            # Create a co-occurrence frequency table
            st.subheader("Theme Co-occurrence Table")
            
            # Create a matrix to count co-occurrences
            co_occurrence_matrix = np.zeros((len(top_themes), len(top_themes)))
            
            # Iterate through each document to count co-occurrences
            for doc_id in theme_pivot.index:
                # Get themes present in this document
                doc_themes = theme_pivot.columns[theme_pivot.loc[doc_id] == 1].tolist()
                # Only consider top themes
                doc_themes = [t for t in doc_themes if t in top_themes]
                
                # Count pairs of co-occurring themes
                for i, theme1 in enumerate(doc_themes):
                    idx1 = top_themes.index(theme1)
                    for theme2 in doc_themes:
                        idx2 = top_themes.index(theme2)
                        co_occurrence_matrix[idx1, idx2] += 1
        
            # Create a formatted DataFrame for display in the UI
            display_co_occurrence = pd.DataFrame(
                co_occurrence_matrix,
                index=[theme_display_map[theme] for theme in top_themes],
                columns=[theme_display_map[theme] for theme in top_themes]
            )
            
            # Keep original for CSV export
            co_occurrence_df = pd.DataFrame(
                co_occurrence_matrix,
                index=top_themes,
                columns=top_themes
            )
            
            # Display the co-occurrence table with formatted names
            st.dataframe(
                display_co_occurrence,
                use_container_width=True,
                height=400,
                key="co_occurrence_table"
            )
            
            # Add explanation
            st.markdown("""
            This table shows the number of documents where each pair of themes co-occurs. 
            The diagonal represents the total count of each theme.
            """)
    
            # Create a heatmap of the co-occurrence matrix
            fig_cooccur = px.imshow(
                co_occurrence_matrix,
                x=[theme_display_map[theme] for theme in top_themes],
                y=[theme_display_map[theme] for theme in top_themes],
                labels=dict(x="Theme", y="Theme", color="Co-occurrences"),
                title="Theme Co-occurrence Heatmap",
                color_continuous_scale="Viridis",
                text_auto=".0f"  # Show integer values
            )
            
            fig_cooccur.update_layout(
                margin=dict(l=200, r=80, b=220, t=80),
                font=dict(color="white"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(
                    side="bottom",
                    tickangle=90,
                    automargin=True
                ),
                yaxis=dict(
                    automargin=True
                )
            )
            
            # Update axes for dark mode
            fig_cooccur.update_xaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white", size=11),
                gridcolor="rgba(255,255,255,0.1)",
                automargin=True
            )
            
            fig_cooccur.update_yaxes(
                title_font=dict(color="white"),
                tickfont=dict(color="white", size=11),
                gridcolor="rgba(255,255,255,0.1)",
                automargin=True
            )
            
            # Update colorbar for dark mode
            fig_cooccur.update_traces(
                colorbar=dict(
                    title=dict(text="Co-occurrences", font=dict(color="white")),
                    tickfont=dict(color="white")
                )
            )
            
            # Display the co-occurrence heatmap with a unique key
            st.plotly_chart(fig_cooccur, use_container_width=True, key="cooccurrence_heatmap")
        
        # Show detailed data table
        with st.expander("View Detailed Data"):
            st.dataframe(
                filtered_df,
                column_config={
                    "Title": st.column_config.TextColumn("Document Title"),
                    "Framework": st.column_config.TextColumn("Framework"),
                    "Theme": st.column_config.TextColumn("Theme"),
                    "Confidence": st.column_config.TextColumn("Confidence"),
                    "Combined Score": st.column_config.NumberColumn("Score", format="%.3f"),
                    "Matched Keywords": st.column_config.TextColumn("Keywords"),
                    "coroner_name": st.column_config.TextColumn("Coroner Name"),
                    "coroner_area": st.column_config.TextColumn("Coroner Area"),
                    "year": st.column_config.NumberColumn("Year"),
                },
                use_container_width=True,
                key="detailed_data_table"
            )
            
        # Export options
        st.subheader("Export Filtered Data")
        
        # Generate timestamp for filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create columns for download buttons
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # CSV Export
            csv = filtered_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "📥 Download Filtered Data (CSV)",
                data=csv,
                file_name=f"theme_analysis_export_{timestamp}.csv",
                mime="text/csv",
                key=f"download_csv_{timestamp}",
            )
        
        with col2:
            # Excel Export
            excel_data = export_to_excel(filtered_df)
            st.download_button(
                "📥 Download Filtered Data (Excel)",
                data=excel_data,
                file_name=f"theme_analysis_export_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{timestamp}",
            )
    
        with col3:
            # All Images Export
            try:
                # Get the zip file and image count
                images_zip, image_count = save_dashboard_images_as_zip(filtered_df)
                
                # Update button text to show number of images
                st.download_button(
                    f"📥 Download {image_count} Visualizations (ZIP)",
                    data=images_zip,
                    file_name=f"theme_analysis_images_{timestamp}.zip",
                    mime="application/zip",
                    key=f"download_images_{timestamp}",
                )
            except Exception as e:
                st.error(f"Error creating visualization zip: {e}")
                logging.error(f"Visualization zip error: {e}", exc_info=True)

def render_analysis_tab(data: pd.DataFrame = None):
    """Render the analysis tab with improved filters, file upload functionality, and analysis sections"""

    # Add file upload section at the top
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file", 
        type=['csv', 'xlsx'],
        help="Upload previously exported data"
    )
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                data = pd.read_csv(uploaded_file)
            else:
                data = pd.read_excel(uploaded_file)
            
            # Process uploaded data
            data = process_scraped_data(data)
            st.success("File uploaded and processed successfully!")
            
            # Update session state
            st.session_state.uploaded_data = data.copy()
            st.session_state.data_source = 'uploaded'
            st.session_state.current_data = data.copy()
        
        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            logging.error(f"File upload error: {e}", exc_info=True)
            return
    
    # Use either uploaded data or passed data
    if data is None:
        data = st.session_state.get('current_data')
    
    if data is None or len(data) == 0:
        st.warning("No data available. Please upload a file or scrape reports first.")
        return
        
    try:
        # Get date range for the data
        min_date = data['date_of_report'].min().date()
        max_date = data['date_of_report'].max().date()
        
        # Filters sidebar
        with st.sidebar:
            st.header("Filters")
            
            # Date Range
            with st.expander("📅 Date Range", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "From",
                        value=min_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="start_date_filter",
                        format="DD/MM/YYYY"
                    )
                with col2:
                    end_date = st.date_input(
                        "To",
                        value=max_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="end_date_filter",
                        format="DD/MM/YYYY"
                    )
            
            # Document Type Filter
            doc_type = st.multiselect(
                "Document Type",
                ["Report", "Response"],
                default=[],
                key="doc_type_filter",
                help="Filter by document type"
            )
            
            # Reference Number
            ref_numbers = sorted(data['ref'].dropna().unique())
            selected_refs = st.multiselect(
                "Reference Numbers",
                options=ref_numbers,
                key="ref_filter"
            )
            
            # Deceased Name - Changed to dropdown instead of text input
            deceased_names = sorted(set(
                str(name).strip() for name in data['deceased_name'].dropna().unique()
            ))
            selected_deceased = st.multiselect(
                "Deceased Name",
                options=deceased_names,
                key="deceased_filter",
                help="Select one or more deceased names"
            )
            
            # Coroner Name
            # Normalize coroner names for selection and ensure uniqueness
            coroner_names = sorted(set(
                str(name).strip() for name in data['coroner_name'].dropna().unique()
            ))
            selected_coroners = st.multiselect(
                "Coroner Names",
                options=coroner_names,
                key="coroner_filter"
            )
            
            # Coroner Area
            # Normalize coroner areas for selection and ensure uniqueness
            coroner_areas = sorted(set(
                str(area).strip() for area in data['coroner_area'].dropna().unique()
            ))
            selected_areas = st.multiselect(
                "Coroner Areas",
                options=coroner_areas,
                key="areas_filter"
            )
            
            # Categories - Improved handling of both list and string formats
            all_categories = set()
            for cats in data['categories'].dropna():
                if isinstance(cats, list):
                    all_categories.update(str(cat).strip() for cat in cats if cat)
                elif isinstance(cats, str):
                    # Handle comma-separated strings
                    all_categories.update(str(cat).strip() for cat in cats.split(',') if cat)
            
            # Remove any empty strings
            all_categories = {cat for cat in all_categories if cat.strip()}
            
            selected_categories = st.multiselect(
                "Categories",
                options=sorted(all_categories),
                key="categories_filter"
            )
            
            # Reset Filters Button
            if st.button("🔄 Reset Filters"):
                for key in st.session_state:
                    if key.endswith('_filter'):
                        del st.session_state[key]
                st.rerun()

        # Apply filters
        filtered_df = data.copy()

        # Date filter
        if start_date and end_date:
            filtered_df = filtered_df[
                (filtered_df['date_of_report'].dt.date >= start_date) &
                (filtered_df['date_of_report'].dt.date <= end_date)
            ]

        # Document type filter
        if doc_type:
            if "Response" in doc_type and "Report" not in doc_type:
                # Only responses
                filtered_df = filtered_df[filtered_df.apply(is_response, axis=1)]
            elif "Report" in doc_type and "Response" not in doc_type:
                # Only reports
                filtered_df = filtered_df[~filtered_df.apply(is_response, axis=1)]

        # Reference number filter
        if selected_refs:
            filtered_df = filtered_df[filtered_df['ref'].isin(selected_refs)]

        # Deceased name filter - changed to use dropdown selection
        if selected_deceased:
            # Normalize selected deceased names and create a case-insensitive filter
            selected_deceased_norm = [str(name).lower().strip() for name in selected_deceased]
            filtered_df = filtered_df[
                filtered_df['deceased_name'].fillna('').str.lower().apply(
                    lambda x: any(selected_name in x or x in selected_name for selected_name in selected_deceased_norm)
                )
            ]

        # Coroner name filter - case-insensitive partial match
        if selected_coroners:
            # Normalize selected coroners and create a case-insensitive filter
            selected_coroners_norm = [str(name).lower().strip() for name in selected_coroners]
            filtered_df = filtered_df[
                filtered_df['coroner_name'].fillna('').str.lower().apply(
                    lambda x: any(selected_name in x or x in selected_name for selected_name in selected_coroners_norm)
                )
            ]

        # Coroner area filter - case-insensitive partial match
        if selected_areas:
            # Normalize selected areas and create a case-insensitive filter
            selected_areas_norm = [str(area).lower().strip() for area in selected_areas]
            filtered_df = filtered_df[
                filtered_df['coroner_area'].fillna('').str.lower().apply(
                    lambda x: any(selected_area in x or x in selected_area for selected_area in selected_areas_norm)
                )
            ]

        # Categories filter - use the improved filter_by_categories function
        if selected_categories:
            filtered_df = filter_by_categories(filtered_df, selected_categories)

        # Show active filters
        active_filters = []
        if start_date != min_date or end_date != max_date:
            active_filters.append(f"Date: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}")
        if doc_type:
            active_filters.append(f"Document Types: {', '.join(doc_type)}")
        if selected_refs:
            active_filters.append(f"References: {', '.join(selected_refs)}")
        if selected_deceased:
            if len(selected_deceased) <= 3:
                active_filters.append(f"Deceased Names: {', '.join(selected_deceased)}")
            else:
                active_filters.append(f"Deceased Names: {len(selected_deceased)} selected")
        if selected_coroners:
            if len(selected_coroners) <= 3:
                active_filters.append(f"Coroners: {', '.join(selected_coroners)}")
            else:
                active_filters.append(f"Coroners: {len(selected_coroners)} selected")
        if selected_areas:
            if len(selected_areas) <= 3:
                active_filters.append(f"Areas: {', '.join(selected_areas)}")
            else:
                active_filters.append(f"Areas: {len(selected_areas)} selected")
        if selected_categories:
            if len(selected_categories) <= 3:
                active_filters.append(f"Categories: {', '.join(selected_categories)}")
            else:
                active_filters.append(f"Categories: {len(selected_categories)} selected")

        if active_filters:
            st.info("Active filters:\n" + "\n".join(f"• {filter_}" for filter_ in active_filters))

        # Display results
        st.subheader("Results")
        st.write(f"Showing {len(filtered_df)} of {len(data)} reports")

        if len(filtered_df) > 0:
            # Display the dataframe
            st.dataframe(
                filtered_df,
                column_config={
                    "URL": st.column_config.LinkColumn("Report Link"),
                    "date_of_report": st.column_config.DateColumn(
                        "Date of Report",
                        format="DD/MM/YYYY"
                    ),
                    "categories": st.column_config.ListColumn("Categories"),
                    "Document Type": st.column_config.TextColumn(
                        "Document Type",
                        help="Type of document based on PDF filename"
                    )
                },
                hide_index=True
            )

            # Create tabs for different analyses
            st.markdown("---")
            quality_tab, temporal_tab, distribution_tab = st.tabs([
                "📊 Data Quality Analysis",
                "📅 Temporal Analysis", 
                "📍 Distribution Analysis"
            ])

            # Data Quality Analysis Tab
            with quality_tab:
                analyze_data_quality(filtered_df)

            # Temporal Analysis Tab
            with temporal_tab:
                # Timeline of reports
                st.subheader("Reports Timeline")
                plot_timeline(filtered_df)
                
                # Monthly distribution
                st.subheader("Monthly Distribution")
                plot_monthly_distribution(filtered_df)
                
                # Year-over-year comparison
                st.subheader("Year-over-Year Comparison")
                plot_yearly_comparison(filtered_df)
                
                # Seasonal patterns
                st.subheader("Seasonal Patterns")
                seasonal_counts = filtered_df['date_of_report'].dt.month.value_counts().sort_index()
                month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                fig = px.line(
                    x=month_names,
                    y=[seasonal_counts.get(i, 0) for i in range(1, 13)],
                    markers=True,
                    labels={'x': 'Month', 'y': 'Number of Reports'},
                    title='Seasonal Distribution of Reports'
                )
                st.plotly_chart(fig, use_container_width=True)

            # Distribution Analysis Tab
            with distribution_tab:
                st.subheader("Reports by Category")
                plot_category_distribution(filtered_df)
  
                st.subheader("Reports by Coroner Area")
                plot_coroner_areas(filtered_df)

            # Export options
            st.markdown("---")
            show_export_options(filtered_df, "analysis")

        else:
            st.warning("No reports match your filter criteria. Try adjusting the filters.")

    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        logging.error(f"Analysis error: {e}", exc_info=True)

def render_analysis_tab2(data: pd.DataFrame = None):
    """Render the analysis tab with improved filters, file upload functionality, and analysis sections"""

    # Add file upload section at the top
    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file", 
        type=['csv', 'xlsx'],
        help="Upload previously exported data"
    )
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                data = pd.read_csv(uploaded_file)
            else:
                data = pd.read_excel(uploaded_file)
            
            # Process uploaded data
            data = process_scraped_data(data)
            st.success("File uploaded and processed successfully!")
            
            # Update session state
            st.session_state.uploaded_data = data.copy()
            st.session_state.data_source = 'uploaded'
            st.session_state.current_data = data.copy()
        
        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            logging.error(f"File upload error: {e}", exc_info=True)
            return
    
    # Use either uploaded data or passed data
    if data is None:
        data = st.session_state.get('current_data')
    
    if data is None or len(data) == 0:
        st.warning("No data available. Please upload a file or scrape reports first.")
        return
        
    try:
        # Get date range for the data
        min_date = data['date_of_report'].min().date()
        max_date = data['date_of_report'].max().date()
        
        # Filters sidebar
        with st.sidebar:
            st.header("Filters")
            
            # Date Range
            with st.expander("📅 Date Range", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "From",
                        value=min_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="start_date_filter",
                        format="DD/MM/YYYY"
                    )
                with col2:
                    end_date = st.date_input(
                        "To",
                        value=max_date,
                        min_value=min_date,
                        max_value=max_date,
                        key="end_date_filter",
                        format="DD/MM/YYYY"
                    )
            
            # Document Type Filter
            doc_type = st.multiselect(
                "Document Type",
                ["Report", "Response"],
                default=[],
                key="doc_type_filter",
                help="Filter by document type"
            )
            
            # Reference Number
            ref_numbers = sorted(data['ref'].dropna().unique())
            selected_refs = st.multiselect(
                "Reference Numbers",
                options=ref_numbers,
                key="ref_filter"
            )
            
            # Deceased Name
            deceased_search = st.text_input(
                "Deceased Name",
                key="deceased_filter",
                help="Enter partial or full name (case-insensitive)"
            )
            
            # Coroner Name
            # Normalize coroner names for selection and ensure uniqueness
            coroner_names = sorted(set(
                str(name).strip() for name in data['coroner_name'].dropna().unique()
            ))
            selected_coroners = st.multiselect(
                "Coroner Names",
                options=coroner_names,
                key="coroner_filter"
            )
            
            # Coroner Area
            # Normalize coroner areas for selection and ensure uniqueness
            coroner_areas = sorted(set(
                str(area).strip() for area in data['coroner_area'].dropna().unique()
            ))
            selected_areas = st.multiselect(
                "Coroner Areas",
                options=coroner_areas,
                key="areas_filter"
            )
            
            # Categories
            all_categories = set()
            for cats in data['categories'].dropna():
                if isinstance(cats, list):
                    all_categories.update(str(cat).strip() for cat in cats)
                elif isinstance(cats, str):
                    all_categories.update(str(cat).strip() for cat in cats.split(','))
            
            selected_categories = st.multiselect(
                "Categories",
                options=sorted(all_categories),
                key="categories_filter"
            )
            
            # Reset Filters Button
            if st.button("🔄 Reset Filters"):
                for key in st.session_state:
                    if key.endswith('_filter'):
                        del st.session_state[key]
                st.rerun()

        # Apply filters
        filtered_df = data.copy()

        # Date filter
        if start_date and end_date:
            filtered_df = filtered_df[
                (filtered_df['date_of_report'].dt.date >= start_date) &
                (filtered_df['date_of_report'].dt.date <= end_date)
            ]

        # Document type filter
        if doc_type:
            if "Response" in doc_type and "Report" not in doc_type:
                # Only responses
                filtered_df = filtered_df[filtered_df.apply(is_response, axis=1)]
            elif "Report" in doc_type and "Response" not in doc_type:
                # Only reports
                filtered_df = filtered_df[~filtered_df.apply(is_response, axis=1)]

        # Reference number filter
        if selected_refs:
            filtered_df = filtered_df[filtered_df['ref'].isin(selected_refs)]

        # Deceased name filter - case-insensitive partial match
        if deceased_search:
            search_lower = deceased_search.lower().strip()
            filtered_df = filtered_df[
                filtered_df['deceased_name'].fillna('').str.lower().str.contains(
                    search_lower, 
                    case=False, 
                    na=False
                )
            ]

        # Coroner name filter - case-insensitive partial match
        if selected_coroners:
            # Normalize selected coroners and create a case-insensitive filter
            selected_coroners_norm = [str(name).lower().strip() for name in selected_coroners]
            filtered_df = filtered_df[
                filtered_df['coroner_name'].fillna('').str.lower().apply(
                    lambda x: any(selected_name in x or x in selected_name for selected_name in selected_coroners_norm)
                )
            ]

        # Coroner area filter - case-insensitive partial match
        if selected_areas:
            # Normalize selected areas and create a case-insensitive filter
            selected_areas_norm = [str(area).lower().strip() for area in selected_areas]
            filtered_df = filtered_df[
                filtered_df['coroner_area'].fillna('').str.lower().apply(
                    lambda x: any(selected_area in x or x in selected_area for selected_area in selected_areas_norm)
                )
            ]

        # Categories filter - handle both list and string types with case-insensitive partial match
        if selected_categories:
            # Normalize selected categories
            selected_cats_norm = [str(cat).lower().strip() for cat in selected_categories]
            
            def category_matches(row_cats):
                # Handle both list and string types
                if pd.isna(row_cats):
                    return False
                
                # Convert to list if it's a string
                if isinstance(row_cats, str):
                    row_cats = [cat.strip() for cat in row_cats.split(',')]
                
                # Normalize row categories
                row_cats_norm = [str(cat).lower().strip() for cat in row_cats]
                
                # Check for partial matches
                return any(
                    any(selected_cat in row_cat or row_cat in selected_cat 
                        for row_cat in row_cats_norm)
                    for selected_cat in selected_cats_norm
                )
            
            filtered_df = filtered_df[filtered_df['categories'].apply(category_matches)]

        # Show active filters
        active_filters = []
        if start_date != min_date or end_date != max_date:
            active_filters.append(f"Date: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}")
        if doc_type:
            active_filters.append(f"Document Types: {', '.join(doc_type)}")
        if selected_refs:
            active_filters.append(f"References: {', '.join(selected_refs)}")
        if deceased_search:
            active_filters.append(f"Deceased name contains: {deceased_search}")
        if selected_coroners:
            active_filters.append(f"Coroners: {', '.join(selected_coroners)}")
        if selected_areas:
            active_filters.append(f"Areas: {', '.join(selected_areas)}")
        if selected_categories:
            active_filters.append(f"Categories: {', '.join(selected_categories)}")

        if active_filters:
            st.info("Active filters:\n" + "\n".join(f"• {filter_}" for filter_ in active_filters))

        # Display results
        st.subheader("Results")
        st.write(f"Showing {len(filtered_df)} of {len(data)} reports")

        if len(filtered_df) > 0:
            # Display the dataframe
            st.dataframe(
                filtered_df,
                column_config={
                    "URL": st.column_config.LinkColumn("Report Link"),
                    "date_of_report": st.column_config.DateColumn(
                        "Date of Report",
                        format="DD/MM/YYYY"
                    ),
                    "categories": st.column_config.ListColumn("Categories"),
                    "Document Type": st.column_config.TextColumn(
                        "Document Type",
                        help="Type of document based on PDF filename"
                    )
                },
                hide_index=True
            )

            # Create tabs for different analyses
            st.markdown("---")
            quality_tab, temporal_tab, distribution_tab = st.tabs([
                "📊 Data Quality Analysis",
                "📅 Temporal Analysis", 
                "📍 Distribution Analysis"
            ])

            # Data Quality Analysis Tab
            with quality_tab:
                analyze_data_quality(filtered_df)

            # Temporal Analysis Tab
            with temporal_tab:
                # Timeline of reports
                st.subheader("Reports Timeline")
                plot_timeline(filtered_df)
                
                # Monthly distribution
                st.subheader("Monthly Distribution")
                plot_monthly_distribution(filtered_df)
                
                # Year-over-year comparison
                st.subheader("Year-over-Year Comparison")
                plot_yearly_comparison(filtered_df)
                
                # Seasonal patterns
                st.subheader("Seasonal Patterns")
                seasonal_counts = filtered_df['date_of_report'].dt.month.value_counts().sort_index()
                month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                fig = px.line(
                    x=month_names,
                    y=[seasonal_counts.get(i, 0) for i in range(1, 13)],
                    markers=True,
                    labels={'x': 'Month', 'y': 'Number of Reports'},
                    title='Seasonal Distribution of Reports'
                )
                st.plotly_chart(fig, use_container_width=True)

            # Distribution Analysis Tab
            with distribution_tab:
        
       
                    st.subheader("Reports by Category")
                    plot_category_distribution(filtered_df)
  
                    st.subheader("Reports by Coroner Area")
                    plot_coroner_areas(filtered_df)

            # Export options
            st.markdown("---")
            show_export_options(filtered_df, "analysis")

        else:
            st.warning("No reports match your filter criteria. Try adjusting the filters.")

    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        logging.error(f"Analysis error: {e}", exc_info=True)

                
def render_framework_heatmap(filtered_df, top_n_themes=5):
    """
    Create a framework-based heatmap of theme distribution by year with framework coloring
    
    Args:
        filtered_df: Filtered DataFrame containing theme analysis results
        top_n_themes: Number of top themes to show per framework
        
    Returns:
        Plotly figure object
    """
    if "year" not in filtered_df.columns or filtered_df["year"].isna().all():
        return None
    
    # Create combined framework:theme field
    filtered_df['Framework_Theme'] = filtered_df['Framework'] + ': ' + filtered_df['Theme']
    
    # Count reports per year (for denominator)
    # Assuming Record ID is the unique identifier for reports
    id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
    reports_per_year = filtered_df.groupby('year')[id_column].nunique()
    
    # Count unique report IDs per theme per year
    counts = filtered_df.groupby(['year', 'Framework', 'Framework_Theme'])[id_column].nunique().reset_index()
    counts.columns = ['year', 'Framework', 'Framework_Theme', 'Count']
    
    # Calculate percentages
    counts['Total'] = counts['year'].map(reports_per_year)
    counts['Percentage'] = (counts['Count'] / counts['Total'] * 100).round(1)
    
    # Get frameworks in the filtered data
    frameworks_present = filtered_df['Framework'].unique()
    
    # Get top themes by framework (N per framework)
    top_themes = []
    for framework in frameworks_present:
        framework_counts = counts[counts['Framework'] == framework]
        theme_totals = framework_counts.groupby('Framework_Theme')['Count'].sum().sort_values(ascending=False)
        top_themes.extend(theme_totals.head(top_n_themes).index.tolist())
    
    # Filter to top themes
    counts = counts[counts['Framework_Theme'].isin(top_themes)]
    
    # Create pivot table for heatmap
    pivot = counts.pivot_table(
        index='Framework_Theme',
        columns='year',
        values='Percentage',
        fill_value=0
    )
    
    # Create pivot for counts
    count_pivot = counts.pivot_table(
        index='Framework_Theme',
        columns='year',
        values='Count',
        fill_value=0
    )
    
    # Sort by framework then by total count
    theme_totals = counts.groupby('Framework_Theme')['Count'].sum()
    theme_frameworks = {theme: theme.split(':')[0] for theme in theme_totals.index}
    
    # Sort first by framework, then by count within framework
    sorted_themes = sorted(
        theme_totals.index,
        key=lambda x: (theme_frameworks[x], -theme_totals[x])
    )
    
    # Apply the sort order
    pivot = pivot.reindex(sorted_themes)
    count_pivot = count_pivot.reindex(sorted_themes)
    
    # Create color mapping for frameworks
    framework_colors = {}
    for i, framework in enumerate(frameworks_present):
        if framework == 'I-SIRch':
            framework_colors[framework] = "orange"  # Orange for I-SIRch
        elif framework == 'House of Commons':
            framework_colors[framework] = "royalblue"  # Blue for House of Commons
        else:
            # Use other colors for other frameworks
            other_colors = ["forestgreen", "purple", "darkred"]
            framework_colors[framework] = other_colors[i % len(other_colors)]
    
    # Create a visually distinctive dataframe for plotting
    # For each theme, create a dict with clean name and framework
    theme_display_data = []
    
    for theme in pivot.index:
        framework = theme.split(':')[0].strip()
        theme_name = theme.split(':', 1)[1].strip()
        
        # Insert line breaks for long theme names
        if len(theme_name) > 30:
            # Try to break at a space near the middle
            words = theme_name.split()
            if len(words) > 1:
                # Find a breaking point near the middle
                mid_point = len(words) // 2
                first_part = ' '.join(words[:mid_point])
                second_part = ' '.join(words[mid_point:])
                theme_name = f"{first_part}<br>{second_part}"
        
        theme_display_data.append({
            'original': theme,
            'clean_name': theme_name,
            'framework': framework,
            'color': framework_colors[framework]
        })
        
    theme_display_df = pd.DataFrame(theme_display_data)
    
    # Add year count labels
    year_labels = [f"{year}<br>n={reports_per_year[year]}" for year in pivot.columns]
    
    # Create heatmap using plotly
    fig = go.Figure()
    
    # Add heatmap
    heatmap = go.Heatmap(
        z=pivot.values,
        x=year_labels,
        y=theme_display_df['clean_name'],
        colorscale='Blues',
        zmin=0,
        zmax=min(100, pivot.values.max() * 1.2),  # Cap at 100% or 20% higher than max
        colorbar=dict(title='Percentage (%)'),
        hoverongaps=False,
        text=count_pivot.values,  # This will show the count in the hover
        hovertemplate='Year: %{x}<br>Theme: %{y}<br>Percentage: %{z}%<br>Count: %{text}<extra></extra>'
    )
    
    fig.add_trace(heatmap)
    
    # Add count annotations
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            if pivot.iloc[i, j] > 0:
                count = count_pivot.iloc[i, j]
                
                # Add an annotation for the actual count
                fig.add_annotation(
                    x=j,
                    y=i,
                    text=f"({count})",
                    font=dict(size=8, color='black' if pivot.iloc[i, j] < 50 else 'white'),
                    showarrow=False,
                    xanchor='center',
                    yanchor='top',
                    yshift=-12
                )
    
    # Set y-axis ordering and color-coding
    fig.update_layout(
        yaxis=dict(
            tickmode='array',
            tickvals=list(range(len(theme_display_df))),
            ticktext=theme_display_df['clean_name'],
            tickfont=dict(
                size=11,
                color='black'
            ),
        )
    )
    
    # Add colored framework indicators
    for framework, color in framework_colors.items():
        # Count themes for this framework
        framework_theme_count = theme_display_df[theme_display_df['framework'] == framework].shape[0]
        
        if framework_theme_count > 0:
            # Add a shape for the framework indicator
            fig.add_shape(
                type="rect",
                x0=-1,  # Slightly to the left of the y-axis
                x1=-0.5,
                y0=len(theme_display_df) - theme_display_df[theme_display_df['framework'] == framework].index[0] - framework_theme_count,
                y1=len(theme_display_df) - theme_display_df[theme_display_df['framework'] == framework].index[0],
                fillcolor=color,
                opacity=0.6,
                layer="below",
                line=dict(width=0)
            )
    
    # Add framework legend
    for i, (framework, color) in enumerate(framework_colors.items()):
        fig.add_trace(go.Scatter(
            x=[None],
            y=[None],
            mode='markers',
            marker=dict(size=10, color=color),
            name=framework,
            showlegend=True
        ))
    
    # Update layout
    fig.update_layout(
        title="Theme Distribution by Year",
        xaxis_title="Year (number of reports)",
        yaxis_title="Theme",
        height=len(pivot.index) * 30 + 200,  # Adjust height based on number of themes
        margin=dict(l=200, r=20, t=60, b=60),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5
        )
    )
    
    return fig

def main():
    """Updated main application entry point."""
    initialize_session_state()
    
    # Check authentication first
    if not check_app_password():
        # Render the footer even when not authenticated
        render_footer()
        return
    
    # Only show the main app content if authenticated
    st.title("UK Judiciary PFD Reports Analysis")
    
    # Add main description
    st.markdown(
        """
        This application analyses Prevention of Future Deaths (PFD) reports from the UK Judiciary website to uncover patterns, themes, and insights.
        """
    )
    
    # Add collapsible help section
    with st.expander("💡 How to Use This Tool"):
        st.markdown(
            """
            ### Complete Analysis Pipeline:
            
            1. **(1) 🔍 Scrape Reports**: Start by collecting PFD reports from the UK Judiciary website
            2. **(2) 📂 Scraped File Preparation**: Process and merge your scraped reports
            3. **(3) 📊 Scraped File Analysis**: Visualise and analyse basic report patterns
            4. **(4) 📝 Topic Analysis & Summaries**: Generate basic themes from report content
            5. **(5) 🔬 Concept Annotation**: Conduct advanced theme analysis with AI
            6. **(6) 📈 Theme Analysis Dashboard**: Explore comprehensive theme visualisations
            
            Select each numbered tab in sequence to move through the complete analysis pipeline.
            
            ### Tips:
            - Each step generates outputs that can be used in subsequent steps
            - You can upload previously generated files at any step
            - Use the "Clear All Data" button in the sidebar to reset the application
            """
        )

    # The radio button selection remains outside the expander
    current_tab = st.radio(
        "Select section:",
        [
            "(1)🔍 Scrape Reports",
            "(2)📂 Scraped File Preparation",
            "(3)📊 Scraped File Analysis",
            "(4)📝 Topic Analysis & Summaries", 
            "(5)🔬 Concept Annotation",
            "(6)📈 Theme Analysis Dashboard",
        ],
        label_visibility="collapsed",
        horizontal=True,
        key="main_tab_selector",
    )
    st.markdown("---")

    try:
        if current_tab == "(1)🔍 Scrape Reports":
            # Add tab-specific description here
            st.markdown(
                """
                Search tool for Prevention of Future Deaths (PFD) reports from the UK Judiciary website.

                - Extract detailed PFD reports with metadata, full content, and associated PDFs
                - Filtering by keywords, categories, and date ranges
                - Export options in CSV and Excel formats

                Handling Large Result Sets: For extensive search results, use the 'Start page' and 'End page' number inputs to download reports in manageable batches.
                """
            )
            render_scraping_tab()
        
        elif current_tab == "(2)📂 Scraped File Preparation":
            # Add tab-specific description here
            st.markdown(
                """
                This tool merges multiple scraped files into a single dataset. It prepares the data for steps (3) - (5).
                
                - Run this step even if you only have one scraped file. This step extracts the year and applies other processing.
                - Combine data from multiple CSV or Excel files (files starting with pfd_reports_scraped_reportID_)
                - Extract missing concerns from PDF content and fill empty Content fields
                - Extract year information from date fields
                - Remove duplicate records
                - Export full or reduced datasets with essential columns
                """
            )
            render_bert_file_merger()
        
        elif current_tab == "(3)📊 Scraped File Analysis":
            # Add tab-specific description here
            st.markdown(
                """
                Analyse and explore your prepared Prevention of Future Deaths (PFD) reports.
                - Upload processed files from Scraped File Preparation (file starting with merged_)
                - Data visualisation
                - Report insights and export options

                Upload your prepared CSV or Excel file from Step 2 to begin analysis.
                """
            )
            if not validate_data_state():
                handle_no_data_state("analysis")
            else:
                render_analysis_tab(st.session_state.current_data)
        
        elif current_tab == "(4)📝 Topic Analysis & Summaries":
            # Add tab-specific description here
            st.markdown(
                """
                Basic thematic analysis of Prevention of Future Deaths (PFD) reports.
                - Automatically identify key themes across document collections
                - Cluster similar documents (adjust the parameters to identify optimal clusters)
                - Generate summaries for each identified theme
                - Visualise relationships between key concepts and topics
                """
            )
            if not validate_data_state():
                handle_no_data_state("topic_summary")
            else:
                render_topic_summary_tab(st.session_state.current_data)
        
        elif current_tab == "(5)🔬 Concept Annotation":
            # Add tab-specific description here
            st.markdown(
                """
                Advanced AI-powered thematic analysis.
                - Upload the merged Prevention of Future Deaths (PFD) reports file from step (2) 
                - Automatic extraction of important themes from Prevention of Future Deaths (PFD) reports (using 4 frameworks)
                - Download detailed results in structured tables
                - Download colour highlighted sentences based on theme colours in a html report
                """
            )
            render_bert_analysis_tab(st.session_state.current_data)
            
        elif current_tab == "(6)📈 Theme Analysis Dashboard":
            # Add tab-specific description here
            st.markdown(
                """
                Interactive Theme Analysis Dashboard
                
                - Upload theme analysis results from step (5) (file named annotated_theme_analysis_*.xlsx)
                - Navigate through multiple visualization tabs: framework heatmaps, distribution charts, temporal analysis and more
                - Filter results by framework, year, coroner area, and confidence level
                - Discover relationships between themes using correlation analysis
                """
            )
            render_theme_analysis_dashboard(st.session_state.current_data)

        # Sidebar data management
        with st.sidebar:
            st.header("Data Management")
        
            if hasattr(st.session_state, "data_source"):
                st.info(f"Current data: {st.session_state.data_source}")

            #
            if st.button("Clear All Data", key="clear_all_data_button"):
                # Define a comprehensive list of keys to clear
                keys_to_clear = [
                    # Core data keys
                    "current_data",
                    "scraped_data", 
                    "uploaded_data",
                    "topic_model",
                    "data_source",
                    
                    # BERT-specific keys
                    "bert_results",
                    "bert_initialized",
                    "bert_merged_data",
                    
                    # Dashboard specific keys
                    "dashboard_data",
                    "theme_analysis_dashboard_uploader",
                    
                    # File upload keys
                    "bert_file_uploader",
                    "bert_content_column",
                    "bert_analysis_type",
                    "bert_selected_indices",
                    "bert_similarity_threshold",
                    
                    # BERT merger settings keys
                    "drop_duplicates_static",
                    "extract_year_static",
                    "extract_from_pdf_static",
                    "fill_empty_content_static",
                    "duplicate_columns_static",
                    "merge_files_button_static",
                ]
                
                # Clear each key if it exists
                for key in keys_to_clear:
                    if key in st.session_state:
                        del st.session_state[key]
                
                # Force re-initialization of key values
                # These explicit resets ensure clean state
                st.session_state.current_data = None
                st.session_state.uploaded_data = None
                st.session_state.scraped_data = None
                st.session_state.data_source = None
                st.session_state.bert_results = {}
                st.session_state.bert_initialized = False
                st.session_state.bert_merged_data = None
                st.session_state.dashboard_data = None
                
                # Clear all filter-related keys
                for key in list(st.session_state.keys()):
                    if key.startswith('filter_'):
                        del st.session_state[key]
                
                # Clear any cached file information
                if 'last_uploaded_file_hash' in st.session_state:
                    del st.session_state.last_uploaded_file_hash
                
                # Generate a unique key for file uploaders to force reload
                if "reset_counter" not in st.session_state:
                    st.session_state.reset_counter = 0
                st.session_state.reset_counter += 1
                
                # Give feedback and rerun
                st.success("All data cleared successfully")
                time.sleep(0.5)  # Brief pause to ensure UI updates
                st.rerun()

    

            # Add logout button
            if st.button("Logout"):
                st.session_state.authenticated = False
                st.rerun()

        render_footer()

    except Exception as e:
        handle_error(e)
        
        # Render footer even when an exception occurs
        render_footer()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error("Critical Error")
        st.error(str(e))
        logging.critical(f"Application crash: {e}", exc_info=True)
