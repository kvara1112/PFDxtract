import re
import logging
import unicodedata
from datetime import datetime
from typing import Dict
import pandas as pd
import nltk # type: ignore
from typing import Union
import urllib3
from sklearn.feature_extraction.text import TfidfVectorizer
from vectorizer_utils import BM25Vectorizer, WeightedTfidfVectorizer
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import math
import plotly.express as px
import plotly.graph_objects as go


# Configure logging (can be centralized in the main app file later)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s: %(message)s",
    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
)

def clean_text(text: str) -> str:
    """Clean text while preserving structure and metadata formatting"""
    if not text:
        return ""

    try:
        text = str(text)
        text = unicodedata.normalize("NFKD", text)

        replacements = {
            "â€™": "'",
            "â€œ": '"',
            "â€": '"',
            "â€¦": "...",
            'â€"': "-",
            "â€¢": "•",
            "Â": " ",
            "\u200b": "",
            "\uf0b7": "",
            "\u2019": "'",
            "\u201c": '"',
            "\u201d": '"',
            "\u2013": "-",
            "\u2022": "•",
        }

        for encoded, replacement in replacements.items():
            text = text.replace(encoded, replacement)

        text = re.sub(r"<[^>]+>", "", text)
        text = "".join(
            char if char.isprintable() or char == "\n" else " " for char in text
        )
        text = re.sub(r" +", " ", text)
        text = re.sub(r"\n+", "\n", text)

        return text.strip()

    except Exception as e:
        logging.error(f"Error in clean_text: {e}")
        return ""

def clean_text_for_modeling(text: str) -> str:
    """Clean text with enhanced noise removal for modeling purposes"""
    if not isinstance(text, str):
        return ""

    try:
        text = text.lower()
        text = re.sub(r"http\S+|www\S+|https\S+", "", text)
        text = re.sub(r"\S+@\S+", "", text)
        text = re.sub(r"\b\d{3}[-.]?\d{3}[-.]?\d{4}\b", "", text)
        text = re.sub(
            r"\b\d{1,2}(?:st|nd|rd|th)?\s+(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(r"\b\d{1,2}:\d{2}\b", "", text)
        text = re.sub(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", "", text)
        text = re.sub(
            r"\b(?:ref|reference|case)(?:\s+no)?\.?\s*[-:\s]?\s*\w+[-\d]+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(
            r"\b(regulation|paragraph|section|subsection|article)\s+\d+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        legal_terms = r"\b(coroner|inquest|hearing|evidence|witness|statement|report|dated|signed)\b"
        text = re.sub(legal_terms, "", text, flags=re.IGNORECASE)
        text = re.sub(r"[^a-z\s]", " ", text)
        text = re.sub(r"\s+", " ", text)
        text = " ".join(word for word in text.split() if len(word) > 2)
        cleaned_text = text.strip()
        return cleaned_text if len(cleaned_text.split()) >= 3 else ""
    except Exception as e:
        logging.error(f"Error in text cleaning for modeling: {e}")
        return ""

def extract_concern_text(content: str) -> str:
    """Extract complete concern text from PFD report content with robust section handling"""
    if pd.isna(content) or not isinstance(content, str):
        return ""

    concern_identifiers = [
        "CORONER'S CONCERNS", "MATTERS OF CONCERN", "The MATTERS OF CONCERN",
        "CORONER'S CONCERNS are", "MATTERS OF CONCERN are", "The MATTERS OF CONCERN are",
        "HEALTHCARE SAFETY CONCERNS", "SAFETY CONCERNS", "PATIENT SAFETY ISSUES",
        "HSIB FINDINGS", "INVESTIGATION FINDINGS", "THE CORONER'S MATTER OF CONCERN",
        "CONCERNS AND RECOMMENDATIONS", "CONCERNS IDENTIFIED", "TheMATTERS OF CONCERN"
    ]
    content_norm = ' '.join(content.split())
    content_lower = content_norm.lower()
    start_idx = -1
    for identifier in concern_identifiers:
        identifier_lower = identifier.lower()
        pos = content_lower.find(identifier_lower)
        if pos != -1:
            start_idx = pos + len(identifier)
            if start_idx < len(content_norm) and content_norm[start_idx] == ":":
                start_idx += 1
            break
    if start_idx == -1:
        return ""

    end_markers = [
        "ACTION SHOULD BE TAKEN", "CONCLUSIONS", "YOUR RESPONSE", "COPIES",
        "SIGNED:", "DATED THIS", "NEXT STEPS", "YOU ARE UNDER A DUTY",
    ]
    end_idx = len(content_norm)
    for marker in end_markers:
        marker_pos = content_lower.find(marker.lower(), start_idx)
        if marker_pos != -1 and marker_pos < end_idx:
            end_idx = marker_pos
    concerns_text = content_norm[start_idx:end_idx].strip()
    last_period = concerns_text.rfind('.')
    if last_period != -1:
        concerns_text = concerns_text[:last_period + 1]
    return concerns_text

def get_pfd_categories() -> list[str]:
    """Get all available PFD report categories"""
    return [
        "Accident at Work and Health and Safety related deaths",
        "Alcohol drug and medication related deaths",
        "Care Home Health related deaths", "Child Death from 2015",
        "Community health care and emergency services related deaths",
        "Emergency services related deaths 2019 onwards",
        "Hospital Death Clinical Procedures and medical management related deaths",
        "Mental Health related deaths", "Other related deaths", "Police related deaths",
        "Product related deaths", "Railway related deaths", "Road Highways Safety related deaths",
        "Service Personnel related deaths", "State Custody related deaths", "Suicide from 2015",
        "Wales prevention of future deaths reports 2019 onwards",
    ]

def extract_metadata(content: str) -> dict:
    """Extract structured metadata from report content."""
    metadata = {
        "date_of_report": None, "ref": None, "deceased_name": None,
        "coroner_name": None, "coroner_area": None, "categories": [],
    }
    if not content:
        return metadata
    try:
        date_patterns = [
            r"Date of report:?\s*(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+\d{4})",
            r"Date of report:?\s*(\d{1,2}/\d{1,2}/\d{4})",
            r"DATED this (\d{1,2}(?:st|nd|rd|th)?\s+day of [A-Za-z]+\s+\d{4})",
            r"Date:?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        ]
        for pattern in date_patterns:
            date_match = re.search(pattern, content, re.IGNORECASE)
            if date_match:
                date_str = date_match.group(1)
                try:
                    if "/" in date_str:
                        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                    else:
                        date_str = re.sub(r"(?<=\d)(st|nd|rd|th)", "", date_str)
                        date_str = re.sub(r"day of ", "", date_str)
                        try:
                            date_obj = datetime.strptime(date_str, "%d %B %Y")
                        except ValueError:
                            date_obj = datetime.strptime(date_str, "%d %b %Y")
                    metadata["date_of_report"] = date_obj.strftime("%d/%m/%Y")
                    break
                except ValueError as e:
                    logging.warning(f"Invalid date format: {date_str} - {e}")

        ref_match = re.search(r"Ref(?:erence)?:?\s*([-\d]+)", content)
        if ref_match:
            metadata["ref"] = ref_match.group(1).strip()

        name_match = re.search(r"Deceased name:?\s*([^\n]+)", content)
        if name_match:
            metadata["deceased_name"] = clean_text(name_match.group(1)).strip()

        coroner_match = re.search(r"Coroner(?:\'?s)? name:?\s*([^\n]+)", content)
        if coroner_match:
            metadata["coroner_name"] = clean_text(coroner_match.group(1)).strip()

        area_match = re.search(r"Coroner(?:\'?s)? Area:?\s*([^\n]+)", content)
        if area_match:
            metadata["coroner_area"] = clean_text(area_match.group(1)).strip()

        cat_match = re.search(
            r"Category:?\s*(.+?)(?=This report is being sent to:|$)",
            content, re.IGNORECASE | re.DOTALL,
        )
        if cat_match:
            category_text = cat_match.group(1).strip()
            category_text = re.sub(r"\s*[,;]\s*", "|", category_text)
            category_text = re.sub(r"[•·⋅‣⁃▪▫–—-]\s*", "|", category_text)
            category_text = re.sub(r"\s{2,}", "|", category_text)
            category_text = re.sub(r"\n+", "|", category_text)
            categories = category_text.split("|")
            cleaned_categories = []
            standard_categories_map = {cat.lower(): cat for cat in get_pfd_categories()}
            for cat in categories:
                cleaned_cat = clean_text(cat).strip()
                cleaned_cat = re.sub(r"&nbsp;", "", cleaned_cat)
                cleaned_cat = re.sub(r"\s*This report.*$", "", cleaned_cat, flags=re.IGNORECASE)
                cleaned_cat = re.sub(r"[|,;]", "", cleaned_cat)
                if cleaned_cat and not re.match(r"^[\s|,;]+$", cleaned_cat):
                    cat_lower = cleaned_cat.lower()
                    if cat_lower in standard_categories_map:
                        cleaned_cat = standard_categories_map[cat_lower]
                    else:
                        for std_lower, std_original in standard_categories_map.items():
                            if cat_lower in std_lower or std_lower in cat_lower:
                                cleaned_cat = std_original
                                break
                    cleaned_categories.append(cleaned_cat)
            seen = set()
            metadata["categories"] = [
                x for x in cleaned_categories if not (x.lower() in seen or seen.add(x.lower()))
            ]
        return metadata
    except Exception as e:
        logging.error(f"Error extracting metadata: {e}")
        return metadata

def process_scraped_data(df: pd.DataFrame) -> pd.DataFrame:
    """Process and clean scraped data with metadata extraction and concern extraction"""
    try:
        if df is None or len(df) == 0:
            return pd.DataFrame()

        # Create a copy
        df = df.copy()

        # Extract metadata from Content field if it exists
        if "Content" in df.columns:
            # Process each row
            processed_rows = []
            for _, row in df.iterrows():
                # Start with original row data
                processed_row = row.to_dict()

                # Extract metadata using existing function
                content = str(row.get("Content", ""))
                metadata = extract_metadata(content)

                # Extract concerns text
                processed_row["Extracted_Concerns"] = extract_concern_text(content)

                # Update row with metadata
                processed_row.update(metadata)
                processed_rows.append(processed_row)

            # Create new DataFrame from processed rows
            result = pd.DataFrame(processed_rows)
        else:
            result = df.copy()

        # Convert date_of_report to datetime with UK format handling
        if "date_of_report" in result.columns:

            def parse_date(date_str):
                if pd.isna(date_str):
                    return pd.NaT

                date_str = str(date_str).strip()

                # If already in DD/MM/YYYY format
                if re.match(r"\d{1,2}/\d{1,2}/\d{4}", date_str):
                    return pd.to_datetime(date_str, format="%d/%m/%Y")

                # Remove ordinal indicators
                date_str = re.sub(r"(\d)(st|nd|rd|th)", r"\1", date_str)

                # Try different formats
                formats = ["%Y-%m-%d", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"]

                for fmt in formats:
                    try:
                        return pd.to_datetime(date_str, format=fmt)
                    except ValueError:
                        continue

                try:
                    return pd.to_datetime(date_str)
                except:
                    return pd.NaT

            result["date_of_report"] = result["date_of_report"].apply(parse_date)

        return result

    except Exception as e:
        logging.error(f"Error in process_scraped_data: {e}")
        return df

def is_response_document(row: pd.Series) -> bool:
    """Check if a document is a response based on its metadata and content."""
    try:
        for i in range(1, 10):  # Check PDF_1 to PDF_9
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                return True
            pdf_type = str(row.get(f"PDF_{i}_Type", "")).lower() # Check type too
            if pdf_type == "response":
                return True

        title = str(row.get("Title", "")).lower()
        if any(word in title for word in ["response", "reply", "answered"]):
            return True

        content = str(row.get("Content", "")).lower()
        return any(
            phrase in content for phrase in [
                "in response to", "responding to", "reply to", "response to",
                "following the regulation 28", "following receipt of the regulation 28"
            ]
        )
    except Exception as e:
        logging.error(f"Error checking response type: {e}")
        return False

def format_date_uk(date_obj):
    """Convert datetime object to UK date format string"""
    if pd.isna(date_obj): return ""
    try:
        if isinstance(date_obj, str):
            date_obj = pd.to_datetime(date_obj, errors='coerce')
        if pd.isna(date_obj): return "" # if conversion failed
        return date_obj.strftime("%d/%m/%Y")
    except:
        return str(date_obj) # fallback to string representation

def create_document_identifier(row: pd.Series) -> str:
    """Create a unique identifier for a document."""
    title = str(row.get("Title", "")).strip()
    ref = str(row.get("ref", "")).strip()
    deceased = str(row.get("deceased_name", "")).strip()
    return f"{title}_{ref}_{deceased}"

def deduplicate_documents(data: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate documents while preserving unique entries."""
    if 'Title' not in data.columns or 'ref' not in data.columns: # Ensure necessary columns exist
        return data
    data["doc_id"] = data.apply(create_document_identifier, axis=1)
    deduped_data = data.drop_duplicates(subset=["doc_id"], keep="first")
    return deduped_data.drop(columns=["doc_id"])

def truncate_text(text, max_length=30):
    if not text or len(text) <= max_length:
        return text
    if ":" in text: # Special handling for "Framework: Theme"
        parts = text.split(":", 1)
        framework = parts[0].strip()
        theme_part = parts[1].strip()
        if len(theme_part) > max_length - len(framework) - 2:
             # Try to break theme part
            words = theme_part.split()
            processed_theme = []
            current_line = []
            current_len = 0
            for word in words:
                if current_len + len(word) + (1 if current_line else 0) <= max_length - len(framework) -2:
                    current_line.append(word)
                    current_len += len(word) + (1 if current_line else 0)
                else:
                    processed_theme.append(" ".join(current_line))
                    current_line = [word]
                    current_len = len(word)
            if current_line: processed_theme.append(" ".join(current_line))
            
            if len(processed_theme) > 2 : return f"{framework}: {processed_theme[0]}<br>{processed_theme[1]}..."
            return f"{framework}: {('<br>').join(processed_theme)}"
        return f"{framework}: {theme_part}"

    words = text.split()
    lines = []
    current_line = []
    current_len = 0
    for word in words:
        if current_len + len(word) + (1 if current_line else 0) <= max_length:
            current_line.append(word)
            current_len += len(word) + (1 if current_line else 0)
        else:
            lines.append(" ".join(current_line))
            current_line = [word]
            current_len = len(word)
    if current_line: lines.append(" ".join(current_line))
    if len(lines) > 2 : return f"{lines[0]}<br>{lines[1]}..."
    return "<br>".join(lines)

improved_truncate_text = truncate_text # Alias for now, can be differentiated later if needed

def perform_advanced_keyword_search(text: str, search_query: str) -> bool:
    """Perform advanced keyword search with AND/OR operators."""
    if not text or not search_query: return False
    text_lower = str(text).lower()
    query_lower = search_query.lower()
    if " and " in query_lower:
        keywords = [k.strip() for k in query_lower.split(" and ")]
        return all(keyword in text_lower for keyword in keywords if keyword)
    elif " or " in query_lower:
        keywords = [k.strip() for k in query_lower.split(" or ")]
        return any(keyword in text_lower for keyword in keywords if keyword)
    else:
        return query_lower.strip() in text_lower

def initialize_nltk():
    """Initialize required NLTK resources with error handling."""
    try:
        resources = ["punkt", "stopwords", "averaged_perceptron_tagger"]
        for resource in resources:
            try:
                if resource == "punkt": nltk.data.find("tokenizers/punkt")
                elif resource == "stopwords": nltk.data.find("corpora/stopwords")
                elif resource == "averaged_perceptron_tagger": nltk.data.find("taggers/averaged_perceptron_tagger")
            except LookupError:
                nltk.download(resource, quiet=True)
    except Exception as e:
        logging.error(f"Error initializing NLTK resources: {e}")
        # Consider raising an error or handling it gracefully in the app
        # For now, just log it. The app might still function for some parts.

def filter_by_categories(df: pd.DataFrame, selected_categories: list[str]) -> pd.DataFrame:
    """Filter DataFrame by categories, handling both list and string formats in the 'categories' column."""
    if not selected_categories: return df
    if "categories" not in df.columns: return df

    selected_categories_lower = [str(sc).lower().strip() for sc in selected_categories]

    def check_match(row_categories):
        if pd.isna(row_categories): return False
        
        current_cats_lower = []
        if isinstance(row_categories, list):
            current_cats_lower = [str(c).lower().strip() for c in row_categories if c]
        elif isinstance(row_categories, str):
            current_cats_lower = [c.strip() for c in str(row_categories).lower().split(',') if c.strip()]
        
        return any(sc in current_cats_lower for sc in selected_categories_lower)

    return df[df["categories"].apply(check_match)]

def filter_by_areas(df: pd.DataFrame, selected_areas: list[str]) -> pd.DataFrame:
    if not selected_areas: return df
    if "coroner_area" not in df.columns: return df
    
    selected_areas_lower = [str(area).lower().strip() for area in selected_areas]
    
    return df[df["coroner_area"].fillna("").str.lower().str.strip().apply(
        lambda x: any(sa == x for sa in selected_areas_lower) # Exact match after normalization
    )]

def filter_by_coroner_names(df: pd.DataFrame, selected_names: list[str]) -> pd.DataFrame:
    if not selected_names: return df
    if "coroner_name" not in df.columns: return df

    selected_names_lower = [str(name).lower().strip() for name in selected_names]
    
    return df[df["coroner_name"].fillna("").str.lower().str.strip().apply(
        lambda x: any(sn == x for sn in selected_names_lower) # Exact match after normalization
    )]

def filter_by_document_type(df: pd.DataFrame, doc_types: list[str]) -> pd.DataFrame:
    """Filter DataFrame based on document types (Report or Response)."""
    if not doc_types: return df
    
    is_response_series = df.apply(is_response_document, axis=1)
    
    if "Report" in doc_types and "Response" not in doc_types:
        return df[~is_response_series]
    elif "Response" in doc_types and "Report" not in doc_types:
        return df[is_response_series]
    # If both are selected or neither (though 'if not doc_types' handles neither), return all
    return df

def combine_document_text(row: pd.Series) -> str:
    """Combine all text content from a document row for analysis."""
    text_parts = []
    if pd.notna(row.get("Title")): text_parts.append(str(row["Title"]))
    if pd.notna(row.get("Content")): text_parts.append(str(row["Content"]))
    
    # Iterate through potential PDF content columns
    for i in range(1, 10): # Assuming up to PDF_9_Content
        pdf_col_name = f"PDF_{i}_Content"
        if pdf_col_name in row.index and pd.notna(row.get(pdf_col_name)):
            text_parts.append(str(row[pdf_col_name]))
            
    return " ".join(text_parts)

def export_to_excel(df: pd.DataFrame, filename: str = None) -> bytes:
    """
    Export DataFrame to Excel format with proper formatting
    
    Args:
        df: DataFrame to export
        filename: Optional filename (not used, kept for compatibility)
        
    Returns:
        Excel file as bytes
    """
    try:
        
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "PFD Reports"
        
        # Add DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable width limits
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        # Set row height for header
        ws.row_dimensions[1].height = 20
        
        # Add data formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Save to buffer
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        # Fallback to simple Excel export
        try:
            excel_buffer = io.BytesIO()
            df.to_excel(excel_buffer, index=False, engine="openpyxl")
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
        except Exception as fallback_error:
            logging.error(f"Fallback Excel export also failed: {fallback_error}")
            raise Exception("Excel export failed")


def validate_data(df: pd.DataFrame) -> Dict[str, any]:
    """
    Validate DataFrame and return validation results
    
    Args:
        df: DataFrame to validate
        
    Returns:
        Dictionary containing validation results
    """
    validation_results = {
        "is_valid": True,
        "errors": [],
        "warnings": [],
        "row_count": len(df),
        "column_count": len(df.columns),
        "missing_required_columns": [],
        "data_quality_score": 0.0
    }
    
    try:
        # Check for required columns
        required_columns = ["Title", "Content"]
        for col in required_columns:
            if col not in df.columns:
                validation_results["missing_required_columns"].append(col)
                validation_results["errors"].append(f"Missing required column: {col}")
                validation_results["is_valid"] = False
        
        # Check for empty DataFrame
        if len(df) == 0:
            validation_results["errors"].append("DataFrame is empty")
            validation_results["is_valid"] = False
            return validation_results
        
        # Calculate data quality metrics
        quality_scores = []
        
        # Title completeness
        if "Title" in df.columns:
            title_completeness = df["Title"].notna().mean()
            quality_scores.append(title_completeness)
            if title_completeness < 0.9:
                validation_results["warnings"].append(f"Title completeness: {title_completeness:.1%}")
        
        # Content completeness
        if "Content" in df.columns:
            content_completeness = df["Content"].notna().mean()
            quality_scores.append(content_completeness)
            if content_completeness < 0.8:
                validation_results["warnings"].append(f"Content completeness: {content_completeness:.1%}")
        
        # Date completeness
        if "date_of_report" in df.columns:
            date_completeness = df["date_of_report"].notna().mean()
            quality_scores.append(date_completeness)
            if date_completeness < 0.7:
                validation_results["warnings"].append(f"Date completeness: {date_completeness:.1%}")
        
        # Calculate overall quality score
        if quality_scores:
            validation_results["data_quality_score"] = sum(quality_scores) / len(quality_scores)
        
        # Check for duplicates
        if "Title" in df.columns and "ref" in df.columns:
            duplicate_count = df.duplicated(subset=["Title", "ref"]).sum()
            if duplicate_count > 0:
                validation_results["warnings"].append(f"Found {duplicate_count} potential duplicates")
        
        # Check data types
        if "date_of_report" in df.columns:
            if not pd.api.types.is_datetime64_any_dtype(df["date_of_report"]):
                validation_results["warnings"].append("Date column is not in datetime format")
        
        return validation_results
        
    except Exception as e:
        validation_results["errors"].append(f"Validation error: {str(e)}")
        validation_results["is_valid"] = False
        logging.error(f"Data validation error: {e}")
        return validation_results


# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global headers for all requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Referer": "https://judiciary.uk/",
}

def get_vectorizer(
    vectorizer_type: str, max_features: int, min_df: float, max_df: float, **kwargs
) -> Union[TfidfVectorizer, any, any]:  # Using 'any' to avoid circular import in type hints
    """
    Get vectorizer instance based on type and parameters.
    Uses lazy imports to avoid circular dependencies.
    """
    
    if vectorizer_type == "bm25":
        return BM25Vectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            **kwargs
        )
    elif vectorizer_type == "weighted":
        return WeightedTfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            **kwargs
        )
    else:  # Default to TfidfVectorizer
        return TfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            ngram_range=(1, 2),
            stop_words="english",
            **kwargs
        )

def export_topic_results(lda_model, vectorizer, feature_names, doc_topics) -> str:
    """Export topic modeling results to JSON format"""
    results = {
        "topics": [],
        "model_params": {
            "n_topics": lda_model.n_components,
            "max_features": len(feature_names),
        },
        "topic_distribution": doc_topics.mean(axis=0).tolist(),
    }

    # Add topic details
    for idx, topic in enumerate(lda_model.components_):
        top_indices = topic.argsort()[:-11:-1]

        topic_words = [
            {"word": feature_names[i], "weight": float(topic[i])} for i in top_indices
        ]

        results["topics"].append(
            {"id": idx, "words": topic_words, "total_weight": float(topic.sum())}
        )

    return json.dumps(results, indent=2)


def save_dashboard_images_as_zip(filtered_df):
    """
    Save all dashboard visualizations as images and package them into a zip file.
    Improved version that properly generates and captures all visualizations from all tabs.
    
    Args:
        filtered_df: Filtered DataFrame containing theme analysis results
        
    Returns:
        Tuple[bytes, int]: ZIP file containing images and number of images
    """
    import io
    import zipfile
    from datetime import datetime
    import plotly.express as px
    import plotly.graph_objects as go
    import pandas as pd
    import numpy as np
    import logging
    import networkx as nx
    
    # Create a buffer for the zip file
    zip_buffer = io.BytesIO()
    
    # Create a timestamp for the filenames
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Track number of images
    image_count = 0
    
    # Create a zipfile
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Helper function to save a figure to the zip
        def add_figure_to_zip(fig, filename):
            nonlocal image_count
            try:
                # Important: Set explicit dimensions for the image export
                fig.update_layout(
                    width=1000,
                    height=700,
                    margin=dict(l=100, r=80, t=80, b=80),
                    paper_bgcolor="white",
                    plot_bgcolor="white",
                    font=dict(color="black", size=12),
                    title_font=dict(size=16, color="black"),
                    legend=dict(font=dict(size=10))
                )
                
                # Update axes for light mode export
                fig.update_xaxes(
                    title_font=dict(color="black", size=14),
                    tickfont=dict(color="black", size=10),
                    gridcolor="rgba(0,0,0,0.1)",
                    automargin=True  # Enable automargin for all exports
                )
                
                fig.update_yaxes(
                    title_font=dict(color="black", size=14),
                    tickfont=dict(color="black", size=10),
                    gridcolor="rgba(0,0,0,0.1)",
                    automargin=True  # Enable automargin for all exports
                )
                
                # Reset colorbar if present
                if hasattr(fig, 'data') and fig.data and hasattr(fig.data[0], 'colorbar'):
                    fig.update_traces(
                        colorbar=dict(
                            title=dict(text="", font=dict(color="black")),
                            tickfont=dict(color="black", size=10)
                        )
                    )
                
                # Export as PNG with higher resolution
                img_bytes = fig.to_image(format="png", scale=2, engine="kaleido")
                
                if img_bytes and len(img_bytes) > 0:
                    zip_file.writestr(filename, img_bytes)
                    image_count += 1
                    logging.info(f"Successfully added {filename} to zip")
                    return True
                else:
                    logging.warning(f"No image bytes generated for {filename}")
                    return False
            except Exception as e:
                logging.error(f"Error saving {filename}: {str(e)}")
                return False
        
        # === TAB 1: FRAMEWORK HEATMAP ===
        try:
            # Framework distribution chart
            framework_counts = filtered_df["Framework"].value_counts()
            fig = px.bar(
                x=framework_counts.index,
                y=framework_counts.values,
                labels={"x": "Framework", "y": "Count"},
                title="Framework Distribution",
                color=framework_counts.index,
                color_discrete_map={
                    "I-SIRch": "orange",
                    "House of Commons": "royalblue",
                    "Extended Analysis": "firebrick"
                }
            )
            add_figure_to_zip(fig, f"framework_distribution_{timestamp}.png")
            
            # Handle framework theme analysis by year
            if "year" in filtered_df.columns and not filtered_df["year"].isna().all():
                if filtered_df["year"].nunique() == 1:
                    # Single year case
                    year_value = filtered_df["year"].iloc[0]
                    theme_counts = filtered_df.groupby(['Framework', 'Theme']).size().reset_index(name='Count')
                    theme_counts = theme_counts.sort_values(['Framework', 'Count'], ascending=[True, False])
                    
                    # Process theme names for better display using improved function
                    theme_counts['Display_Theme'] = theme_counts['Theme'].apply(
                        lambda x: improved_truncate_text(x, max_length=40)
                    )
                    
                    # Recreate the horizontal bar chart
                    fig = px.bar(
                        theme_counts,
                        y='Display_Theme',
                        x='Count',
                        color='Framework',
                        title=f"Theme Distribution for Year {year_value}",
                        height=max(500, len(theme_counts) * 30),
                        color_discrete_map={
                            "I-SIRch": "orange",
                            "House of Commons": "royalblue",
                            "Extended Analysis": "firebrick"
                        }
                    )
                    
                    fig.update_layout(
                        xaxis_title="Number of Reports",
                        yaxis_title="Theme"
                    )
                    
                    add_figure_to_zip(fig, f"theme_distribution_single_year_{timestamp}.png")
                else:
                    # Multiple years case - recreate the heatmap with improved labels
                    # This code regenerates the framework theme heatmap for the zip file
                    
                    # Create combined framework:theme field if not already there
                    if 'Framework_Theme' not in filtered_df.columns:
                        filtered_df['Framework_Theme'] = filtered_df['Framework'] + ': ' + filtered_df['Theme']
                    
                    # Get needed data for the heatmap
                    id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
                    reports_per_year = filtered_df.groupby('year')[id_column].nunique()
                    
                    # Count unique report IDs per theme per year
                    counts = filtered_df.groupby(['year', 'Framework', 'Framework_Theme'])[id_column].nunique().reset_index()
                    counts.columns = ['year', 'Framework', 'Framework_Theme', 'Count']
                    
                    # Calculate percentages
                    counts['Total'] = counts['year'].map(reports_per_year)
                    counts['Percentage'] = (counts['Count'] / counts['Total'] * 100).round(1)
                    
                    # Get frameworks in the filtered data
                    frameworks_present = filtered_df['Framework'].unique()
                    
                    # Get top themes by framework (5 per framework)
                    top_themes = []
                    for framework in frameworks_present:
                        framework_counts = counts[counts['Framework'] == framework]
                        theme_totals = framework_counts.groupby('Framework_Theme')['Count'].sum().sort_values(ascending=False)
                        top_themes.extend(theme_totals.head(5).index.tolist())
                    
                    # Filter to top themes
                    counts = counts[counts['Framework_Theme'].isin(top_themes)]
                    
                    # Create pivot tables
                    pivot = counts.pivot_table(
                        index='Framework_Theme',
                        columns='year',
                        values='Percentage',
                        fill_value=0
                    )
                    
                    count_pivot = counts.pivot_table(
                        index='Framework_Theme',
                        columns='year',
                        values='Count',
                        fill_value=0
                    )
                    
                    # Sort by framework then by total count
                    theme_totals = counts.groupby('Framework_Theme')['Count'].sum()
                    theme_frameworks = {theme: theme.split(':')[0] for theme in theme_totals.index}
                    sorted_themes = sorted(
                        theme_totals.index,
                        key=lambda x: (theme_frameworks[x], -theme_totals[x])
                    )
                    
                    # Apply the sort order
                    pivot = pivot.reindex(sorted_themes)
                    count_pivot = count_pivot.reindex(sorted_themes)
                    
                    # Create formatted theme names
                    theme_display_data = []
                    framework_colors = {
                        "I-SIRch": "orange",
                        "House of Commons": "royalblue",
                        "Extended Analysis": "firebrick"
                    }
                    
                    # Default colors for any frameworks not specifically mapped
                    other_colors = ["forestgreen", "purple", "darkred"]
                    for i, framework in enumerate(frameworks_present):
                        if framework not in framework_colors:
                            framework_colors[framework] = other_colors[i % len(other_colors)]
                    
                    for theme in pivot.index:
                        framework = theme.split(':')[0].strip()
                        theme_name = theme.split(':', 1)[1].strip()
                        formatted_theme = improved_truncate_text(theme_name, max_length=40)
                        
                        theme_display_data.append({
                            'original': theme,
                            'clean_name': formatted_theme,
                            'framework': framework,
                            'color': framework_colors[framework]
                        })
                    
                    theme_display_df = pd.DataFrame(theme_display_data)
                    
                    # Add year count labels
                    year_labels = [f"{math.floor(year)}\nn={reports_per_year[year]}" for year in pivot.columns]
                    
                    # Create heatmap for export
                    fig = px.imshow(
                        pivot.values,
                        labels=dict(x="Year", y="Theme", color="% of Themes"),
                        x=year_labels,
                        y=theme_display_df['clean_name'],
                        color_continuous_scale="Blues",
                        title="Framework Theme Heatmap by Year",
                        text_auto=".1f"
                    )
                    
                    # Update layout for static export
                    fig.update_layout(
                        width=1200,
                        height=max(650, len(pivot.index) * 35),
                        margin=dict(l=300, r=60, t=80, b=80)
                    )
                    
                    # Add to zip
                    add_figure_to_zip(fig, f"framework_theme_heatmap_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating framework heatmap: {str(e)}")

        # === TAB 2: THEME DISTRIBUTION ===
        try:
            # Get top themes by count
            theme_counts = filtered_df["Theme"].value_counts().head(10)  # Use a reasonable number of themes
            
            # Use improved_truncate_text for better label formatting
            formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in theme_counts.index]
            
            # Create a bar chart with formatted theme names
            fig = px.bar(
                x=formatted_themes,
                y=theme_counts.values,
                labels={"x": "Theme", "y": "Count"},
                title="Top Themes by Occurrence",
                height=600,
                color_discrete_sequence=['#4287f5']
            )
            
            add_figure_to_zip(fig, f"theme_distribution_{timestamp}.png")
            
            # Theme by confidence
            theme_confidence = filtered_df.groupby(["Theme", "Confidence"]).size().reset_index(name="Count")
            
            # Filter for top themes only
            top_themes = theme_counts.index.tolist()
            theme_confidence = theme_confidence[theme_confidence["Theme"].isin(top_themes)]
            
            # Create a mapping dictionary for theme display names
            theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
            
            # Apply the formatting to the DataFrame
            theme_confidence["Display_Theme"] = theme_confidence["Theme"].map(theme_display_map)
            
            # Create a grouped bar chart with formatted theme names
            fig = px.bar(
                theme_confidence, 
                x="Display_Theme",
                y="Count", 
                color="Confidence",
                barmode="group",
                color_discrete_map={"High": "#4CAF50", "Medium": "#FFC107", "Low": "#F44336"},
                category_orders={
                    "Confidence": ["High", "Medium", "Low"],
                    "Display_Theme": [theme_display_map[theme] for theme in top_themes]
                },
                title="Confidence Distribution by Theme",
                height=600
            )
            
            add_figure_to_zip(fig, f"theme_confidence_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating theme distribution charts: {str(e)}")
            
        # === TAB 3: TEMPORAL ANALYSIS ===
        try:
                # Theme trends over time
                year_theme_counts = filtered_df.groupby(["year", "Theme"]).size().reset_index(name="Count")
                
                # Get top themes
                all_theme_counts = filtered_df["Theme"].value_counts()
                top_themes = all_theme_counts.head(8).index.tolist()  # Limit to 8 for readability
                
                # Filter for top themes
                year_theme_counts = year_theme_counts[year_theme_counts["Theme"].isin(top_themes)]
                
                # Create formatted theme names
                theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
                year_theme_counts["Display_Theme"] = year_theme_counts["Theme"].map(theme_display_map)
                
                # Convert year to string for categorical plotting
                year_theme_counts['year_str'] = year_theme_counts['year'].astype(str)
                plotly_colors = px.colors.qualitative.Plotly
                theme_color_map = {
                    theme_display_map[theme]: plotly_colors[i % len(plotly_colors)]
                    for i, theme in enumerate(top_themes)
                }
                # Create line chart for theme trends
                fig = px.line(
                    year_theme_counts,
                    x="year_str",
                    y="Count",
                    color="Display_Theme",
                    color_discrete_map=theme_color_map,
                    markers=True,
                    title="Theme Trends Over Time",
                    labels={"year_str": "Year", "Count": "Number of Occurrences", "Display_Theme": "Theme"}
                )
                
                add_figure_to_zip(fig, f"theme_temporal_trends_{timestamp}.png")
                
                # Create theme prevalence heatmap if multiple years
                if filtered_df["year"].nunique() > 1:
                    # Create a pivot table
                    pivot_df = year_theme_counts.pivot(index="Theme", columns="year_str", values="Count").fillna(0)
                    
                    # Convert to a normalized heatmap (percentage)
                    year_theme_totals = pivot_df.sum(axis=0)
                    normalized_pivot = pivot_df.div(year_theme_totals, axis=1) * 100
                    
                    # Format the theme names
                    formatted_themes = [improved_truncate_text(theme, max_length=40) for theme in normalized_pivot.index]
                    
                    # Create a heatmap
                    year_order = sorted(year_theme_counts['year'].unique())
                    year_order_str = [str(y) for y in year_order]
                    
                    if len(normalized_pivot) > 0 and len(year_order_str) > 0:
                        fig = px.imshow(
                            normalized_pivot[year_order_str],
                            labels=dict(x="Year", y="Theme", color="% of Themes"),
                            x=year_order_str,
                            y=formatted_themes,
                            color_continuous_scale="YlGnBu",
                            title="Theme Prevalence by Year (%)",
                            text_auto=".1f"
                        )
                        
                        add_figure_to_zip(fig, f"theme_prevalence_heatmap_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating temporal analysis charts: {str(e)}")
        
        # === TAB 4: AREA COMPARISON ===
        try:
            if "coroner_area" in filtered_df.columns and not filtered_df["coroner_area"].isna().all():
                # Get the top areas by theme count
                area_counts = filtered_df["coroner_area"].value_counts().head(10)
                
                # Format area names
                formatted_areas = [improved_truncate_text(area, max_length=40) for area in area_counts.index]
                
                # Create a bar chart of top areas
                fig = px.bar(
                    x=formatted_areas,
                    y=area_counts.values,
                    labels={"x": "Coroner Area", "y": "Count"},
                    title="Theme Identifications by Coroner Area",
                    color_discrete_sequence=['#ff9f40']
                )
                
                add_figure_to_zip(fig, f"coroner_area_distribution_{timestamp}.png")
                
                # Create area-theme heatmap
                top_areas = area_counts.index.tolist()
                top_themes = filtered_df["Theme"].value_counts().head(8).index.tolist()  # Limit to 8 themes
                
                # Calculate area theme data
                area_theme_data = []
                for area in top_areas:
                    area_df = filtered_df[filtered_df["coroner_area"] == area]
                    area_totals = len(area_df)
                    
                    area_themes = area_df["Theme"].value_counts()
                    for theme in top_themes:
                        count = area_themes.get(theme, 0)
                        percentage = (count / area_totals * 100) if area_totals > 0 else 0
                        
                        area_theme_data.append({
                            "Coroner Area": area,
                            "Theme": theme,
                            "Count": count,
                            "Percentage": round(percentage, 1)
                        })
                
                area_theme_df = pd.DataFrame(area_theme_data)
                
                # Create formatted names for display
                theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in top_themes}
                area_display_map = {area: improved_truncate_text(area, max_length=40) for area in top_areas}
                
                area_theme_df["Display_Area"] = area_theme_df["Coroner Area"].map(area_display_map)
                area_theme_df["Display_Theme"] = area_theme_df["Theme"].map(theme_display_map)
                
                # Create heatmap if we have data
                if len(area_theme_df) > 0:
                    pivot_df = area_theme_df.pivot(
                        index="Display_Area", 
                        columns="Display_Theme", 
                        values="Percentage"
                    ).fillna(0)
                    
                    # Check if we have valid data for heatmap
                    if pivot_df.shape[0] > 0 and pivot_df.shape[1] > 0:
                        fig = px.imshow(
                            pivot_df,
                            labels=dict(x="Theme", y="Coroner Area", color="Percentage"),
                            x=pivot_df.columns,
                            y=pivot_df.index,
                            color_continuous_scale="YlGnBu",
                            title="Theme Distribution by Coroner Area (%)",
                            text_auto=".1f"
                        )
                        
                        add_figure_to_zip(fig, f"theme_area_heatmap_{timestamp}.png")
                
                # Create radar chart if we have enough areas
                if len(top_areas) >= 2:
                    radar_areas = top_areas[:3]  # Take the top 3 areas or less
                    
                    # Filter data for these areas and top themes
                    radar_data = area_theme_df[
                        (area_theme_df["Coroner Area"].isin(radar_areas)) & 
                        (area_theme_df["Theme"].isin(top_themes[:6]))  # Limit to 6 themes for readability
                    ]
                    
                    
                    if len(radar_data) > 0:
                        # Create radar chart
                        fig = go.Figure()
                        colors = px.colors.qualitative.Plotly
                        # Add traces for each area
                        for i, area in enumerate(radar_areas):
                            area_data = radar_data[radar_data["Coroner Area"] == area]
                            # Sort by theme to ensure consistency
                            area_data = area_data.set_index("Theme").reindex(top_themes[:6]).reset_index()
                            colorx = colors[i % len(colors)]
                            fig.add_trace(go.Scatterpolar(
                                r=area_data["Percentage"],
                                theta=area_data["Display_Theme"],
                                fill="toself",
                                name=area_display_map.get(area, area),
                                line = dict(color=colorx),
                                fillcolor=colorx,
                                opacity=0.5
                            ))

                        #
                        fig.update_layout(
                            polar=dict(
                                bgcolor="#f0f0f0",  # Light gray background inside the polar area
                                radialaxis=dict(
                                    visible=True,
                                    range=[0, max(radar_data["Percentage"]) * 1.1],
                                    color="black",                 # Axis lines
                                    tickfont=dict(color="black"), # Tick labels
                                    gridcolor="gray",             # Circular gridlines
                                    linecolor="black"             # Axis line
                                ),
                                angularaxis=dict(
                                    color="black",
                                    tickfont=dict(color="black"),
                                    gridcolor="gray",
                                    linecolor="black"
                                )
                            ),
                            font=dict(
                                color="black"
                            ),
                            paper_bgcolor="#0c1f30",  # Dark blue outer background
                            plot_bgcolor="#0c1f30",   # Match the paper background
                            showlegend=True,
                            legend=dict(
                                font=dict(color="white")  # Ensures legend text stays visible
                            ),
                            title=dict(
                                text="Theme Distribution Radar Chart",
                                font=dict(color="white")
                            )
                            
                        )

                        
                        add_figure_to_zip(fig, f"area_radar_chart_{timestamp}.png")
                        
        except Exception as e:
            logging.error(f"Error creating area comparison charts: {str(e)}")
            
        # === TAB 5: CORRELATION ANALYSIS ===
        try:
            # Calculate correlation between themes
            id_column = 'Record ID' if 'Record ID' in filtered_df.columns else filtered_df.columns[0]
            
            # Get top themes for correlation
            top_themes = filtered_df["Theme"].value_counts().head(10).index.tolist()
            
            # Create a binary pivot table
            theme_pivot = pd.crosstab(
                index=filtered_df[id_column], 
                columns=filtered_df['Theme'],
                values=filtered_df.get('Combined Score', filtered_df['Theme']),
                aggfunc='max'
            ).fillna(0)
            
            # Convert to binary
            theme_pivot = (theme_pivot > 0).astype(int)
            
            # Calculate correlation
            if len(theme_pivot.columns) > 1:  # Need at least 2 columns for correlation
                theme_corr = theme_pivot.corr()
                
                # Get only the top themes for clarity
                available_themes = [theme for theme in top_themes if theme in theme_corr.index]
                
                if available_themes:
                    top_theme_corr = theme_corr.loc[available_themes, available_themes]
                    
                    # Create a mapping dictionary for theme display names
                    theme_display_map = {theme: improved_truncate_text(theme, max_length=40) for theme in available_themes}
                    
                    # Format column and index labels
                    formatted_themes = [theme_display_map[theme] for theme in top_theme_corr.columns]
                    
                    # Create a heatmap of correlations
                    fig = px.imshow(
                        top_theme_corr.values,
                        color_continuous_scale=px.colors.diverging.RdBu_r,
                        color_continuous_midpoint=0,
                        labels=dict(x="Theme", y="Theme", color="Correlation"),
                        x=formatted_themes,
                        y=formatted_themes,
                        title="Theme Correlation Matrix",
                        text_auto=".2f"
                    )
                    
                    # Update layout
                    fig.update_layout(
                        width=1000,
                        height=800,
                        margin=dict(l=250, r=100, t=80, b=250)
                    )
                    
                    add_figure_to_zip(fig, f"theme_correlation_matrix_{timestamp}.png")
                    
                    # Create network graph
                    # Try different thresholds until we get a reasonable number of edges
                    for threshold in [0.6, 0.5, 0.4, 0.3, 0.2]:
                        G = nx.Graph()
                        
                        # Add nodes
                        for theme in available_themes:
                            G.add_node(theme, display_name=theme_display_map[theme])
                        
                        # Add edges
                        edge_count = 0
                        for i, theme1 in enumerate(available_themes):
                            for j, theme2 in enumerate(available_themes):
                                if i < j:  # Only process each pair once
                                    correlation = top_theme_corr.loc[theme1, theme2]
                                    if correlation >= threshold:
                                        G.add_edge(theme1, theme2, weight=correlation)
                                        edge_count += 1
                        
                        # If we have a reasonable number of edges, create the visualization
                        if edge_count > 0 and edge_count <= 20:
                            pos = nx.spring_layout(G, seed=42)
                            
                            # Create network visualization
                            edge_traces = []
                            
                            # Add edges
                            for edge in G.edges():
                                x0, y0 = pos[edge[0]]
                                x1, y1 = pos[edge[1]]
                                weight = G[edge[0]][edge[1]]['weight']
                                
                                edge_traces.append(
                                    go.Scatter(
                                        x=[x0, x1, None],
                                        y=[y0, y1, None],
                                        line=dict(width=weight*3, color=f'rgba(100,100,100,{weight})'),
                                        hoverinfo='none',
                                        mode='lines'
                                    )
                                )
                            
                            # Add nodes
                            node_x = []
                            node_y = []
                            node_text = []
                            node_size = []
                            
                            for node in G.nodes():
                                x, y = pos[node]
                                node_x.append(x)
                                node_y.append(y)
                                node_text.append(theme_display_map[node])
                                size = len(list(G.neighbors(node))) * 10 + 20
                                node_size.append(size)
                            
                            node_trace = go.Scatter(
                                x=node_x, 
                                y=node_y,
                                mode='markers+text',
                                text=node_text,
                                textposition="top center",
                                marker=dict(
                                    size=node_size,
                                    color='lightblue',
                                    line=dict(width=1)
                                )
                            )
                            
                            # Create figure
                            fig = go.Figure(
                                data=edge_traces + [node_trace],
                                layout=go.Layout(
                                    title=f'Theme Connection Network (r ≥ {threshold})',
                                    showlegend=False,
                                    hovermode='closest',
                                    margin=dict(b=20, l=5, r=5, t=80),
                                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                    width=800,
                                    height=800
                                )
                            )
                            
                            add_figure_to_zip(fig, f"theme_network_{timestamp}.png")
                            break  # We found a good threshold, no need to try lower ones
                
                # Create co-occurrence matrix
                co_occurrence_matrix = np.zeros((len(available_themes), len(available_themes)))
                
                # Count co-occurrences
                for doc_id in theme_pivot.index:
                    doc_themes = theme_pivot.columns[theme_pivot.loc[doc_id] == 1].tolist()
                    doc_themes = [t for t in doc_themes if t in available_themes]
                    
                    # Count pairs
                    for i, theme1 in enumerate(doc_themes):
                        idx1 = available_themes.index(theme1)
                        for theme2 in doc_themes:
                            idx2 = available_themes.index(theme2)
                            co_occurrence_matrix[idx1, idx2] += 1
                
                # Create a heatmap of co-occurrences
                fig = px.imshow(
                    co_occurrence_matrix,
                    labels=dict(x="Theme", y="Theme", color="Co-occurrences"),
                    x=[theme_display_map[theme] for theme in available_themes],
                    y=[theme_display_map[theme] for theme in available_themes],
                    title="Theme Co-occurrence Matrix",
                    color_continuous_scale="Viridis",
                    text_auto=".0f"
                )
                
                fig.update_layout(
                    width=1000,
                    height=800,
                    margin=dict(l=250, r=100, t=80, b=250)
                )
                
                add_figure_to_zip(fig, f"theme_cooccurrence_matrix_{timestamp}.png")
        except Exception as e:
            logging.error(f"Error creating correlation analysis charts: {str(e)}")
    
    # Reset buffer position
    zip_buffer.seek(0)
    
    # Check if zip is empty
    if image_count == 0:
        raise ValueError("No images were generated for the dashboard.")
    
    return zip_buffer.getvalue(), image_count
